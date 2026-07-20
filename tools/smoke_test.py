from __future__ import annotations

import py_compile
import hashlib
import ast
import io
import json
import re
import zipfile
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
APP = ROOT / "audit_app.py"
CHANGELOG = ROOT / "CHANGELOG_CUSTOMER.md"
PORTFOLIO = ROOT / "vendor_matrix_detailed.xlsx"
PRESENTATION_TEMPLATES = [
    ROOT / "static" / "audit_presentation_khalil.pptx",
    ROOT / "static" / "audit_presentation_btg.pptx",
]
PRESENTATION_QR_ASSETS = [
    ROOT / "static" / "presentation_khalil_qr.png",
    ROOT / "static" / "presentation_btg_qr.png",
]
PRESENTATION_COVER_ASSET = ROOT / "static" / "presentation_audit_cover.jpg"
BANK_DRAFT = ROOT / "samples" / "bank_audit_demo.json"


def assert_true(condition: bool, message: str) -> None:
    if not condition:
        raise AssertionError(message)


def read_text(path: Path) -> str:
    return path.read_text(encoding="utf-8")


def check_compile() -> None:
    py_compile.compile(str(APP), doraise=True)


def check_version() -> None:
    text = read_text(APP)
    match = re.search(r'APP_VERSION\s*=\s*"([^"]+)"', text)
    assert_true(match is not None, "APP_VERSION is missing")
    assert_true(match.group(1) == "12.24-dev", f"Unexpected APP_VERSION: {match.group(1)}")


def check_customer_changelog() -> None:
    text = read_text(CHANGELOG)
    forbidden = [
        "Telegram",
        "Gemini",
        "GEMINI",
        "TELEGRAM",
        "secrets",
        "prompt",
        "sales-only",
        "playbook",
    ]
    found = [word for word in forbidden if word in text]
    assert_true(not found, f"Customer changelog contains internal words: {found}")


def check_selectbox_contract() -> None:
    text = read_text(APP)
    tree = ast.parse(text)
    names = {
        "INDUSTRY_OPTIONS",
        "COUNTRY_CODE_OPTIONS",
        "NETWORK_TYPE_OPTIONS",
        "WIFI_TYPE_OPTIONS",
        "MAIL_SYSTEM_OPTIONS",
        "WEB_HOSTING_OPTIONS",
        "DRAFT_SELECTBOX_OPTIONS",
    }
    assignments = []
    for node in tree.body:
        if not isinstance(node, ast.Assign):
            continue
        if any(isinstance(target, ast.Name) and target.id in names for target in node.targets):
            assignments.append(node)
    namespace: dict[str, object] = {}
    exec(compile(ast.Module(body=assignments, type_ignores=[]), str(APP), "exec"), namespace)

    industries = namespace["INDUSTRY_OPTIONS"]
    required_industries = {
        "Финтех / Банки",
        "Страхование",
        "Здравоохранение / Медицинская организация",
        "Госсектор",
        "Квазигосударственный сектор",
        "КВОИКИ / Критическая инфраструктура",
        "Телеком / Связь",
        "Энергетика / Коммунальная инфраструктура",
        "Транспорт / Логистика",
        "Производство / АСУ ТП",
    }
    assert_true(required_industries.issubset(set(industries)), "Required industry options are missing")

    selectboxes = namespace["DRAFT_SELECTBOX_OPTIONS"]
    expected_keys = {
        "client_industry_select",
        "client_phone_code",
        "main_net_type",
        "back_net_type",
        "wf_type_sel",
        "mail_system",
        "web_hosting",
    }
    assert_true(set(selectboxes) == expected_keys, "Draft selectbox contract is incomplete")
    assert_true(
        len(re.findall(r"\.selectbox\(", text)) == len(expected_keys),
        "A form selectbox is missing from the draft compatibility contract",
    )
    assert_true("net_types = NETWORK_TYPE_OPTIONS" in text, "Network selectboxes do not use the shared contract")
    assert_true("country_codes = COUNTRY_CODE_OPTIONS" in text, "Country selectbox does not use the shared contract")

    normalizer = next(
        node
        for node in tree.body
        if isinstance(node, ast.FunctionDef) and node.name == "normalize_draft_selectbox_value"
    )
    exec(compile(ast.Module(body=[normalizer], type_ignores=[]), str(APP), "exec"), namespace)
    normalize = namespace["normalize_draft_selectbox_value"]
    arbitrary_list = ["Windows 11", "Linux"]
    assert_true(
        normalize("os_select", arbitrary_list) == (arbitrary_list, None),
        "Non-selectbox list values must pass through unchanged",
    )
    assert_true(
        normalize("client_industry_select", "Критическая инфраструктура")[0]
        == "КВОИКИ / Критическая инфраструктура",
        "Legacy critical-infrastructure value is not normalized",
    )
    assert_true(
        normalize("client_industry_select", "Новая отрасль")
        == ("Другое", "Новая отрасль"),
        "Unknown industry does not fall back to the custom option",
    )
    assert_true(
        normalize("web_hosting", "Устаревшее значение")[0] in namespace["WEB_HOSTING_OPTIONS"],
        "Invalid hosting value is not normalized",
    )


def find_header_row(ws, required_headers: set[str]) -> tuple[int, dict[str, int]]:
    for row_idx in range(1, min(ws.max_row, 20) + 1):
        values = [str(ws.cell(row=row_idx, column=col).value or "").strip() for col in range(1, ws.max_column + 1)]
        mapping = {value: idx + 1 for idx, value in enumerate(values) if value}
        if required_headers.issubset(mapping):
            return row_idx, mapping
    raise AssertionError(f"Required headers not found: {sorted(required_headers)}")


def check_portfolio() -> None:
    wb = load_workbook(PORTFOLIO, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    header_row, headers = find_header_row(ws, {"Vendor", "Distributor KZ", "Distributor Status"})
    vendor_col = headers["Vendor"]
    distributor_col = headers["Distributor KZ"]
    status_col = headers["Distributor Status"]

    verified = 0
    non_verified = 0
    for row_idx in range(header_row + 1, ws.max_row + 1):
        vendor = str(ws.cell(row=row_idx, column=vendor_col).value or "").strip()
        distributor = str(ws.cell(row=row_idx, column=distributor_col).value or "").strip()
        status = str(ws.cell(row=row_idx, column=status_col).value or "").strip().lower()
        if not vendor or not distributor:
            continue
        if status in {"подтверждено", "проверенный", "verified", "confirmed"}:
            verified += 1
        else:
            non_verified += 1

    assert_true(verified > 0, "No verified distributor rows found")
    assert_true(non_verified > 0, "No non-verified distributor rows found; filter cannot be validated")


def check_static_hooks() -> None:
    text = read_text(APP)
    assert_true("def build_ai_first_sales_opportunities" in text, "AI-first sales helper is missing")
    assert_true("last_report_risk_sources" in text and '"vendors": item.get("vendors", [])' in text, "Risk sources do not preserve vendors")
    assert_true('"area": item.get("_ai_area", "ИТ/ИБ")' in text, "Risk sources do not preserve IT/IB area")
    assert_true("build_ai_first_sales_opportunities(risk_sources, results, context)" in text, "Sales playbook does not use AI-first opportunities")
    assert_true("ensure_sales_playbook_priorities" in text, "Expert sales prioritization is missing")
    assert_true("def it_context_summary" in text and "ИТ-контекст" in text, "Client report IT context is missing")
    assert_true("def make_audit_presentation" in text, "Presentation generator is missing")
    assert_true("cached_presentation_bytes" in text, "Presentation download state is missing")
    assert_true("Скачать экспертный отчет (XLSX)" not in text, "Client XLSX download must stay hidden")
    assert_true("Скачать заключение по аудиту" in text, "Client presentation download is missing")
    assert_true("Скачать презентацию аудита (PPTX)" not in text, "Technical PPTX label must stay hidden")
    assert_true("st-key-presentation_download" in text, "Presentation download styling hook is missing")
    assert_true('div[data-testid="stElementContainer"]:has(.st-key-presentation_download)' in text, "Presentation download centering hook is missing")
    assert_true("st-key-presentation_generate" in text, "Presentation generation styling hook is missing")
    assert_true('key="presentation_generate"' in text, "Presentation generation button key is missing")
    assert_true('"suffix": ".pptx"' in text and "Audit_Presentation_" in text, "Telegram presentation attachment is missing")
    assert_true("AI quality gate rejected the customer presentation" in text, "Customer presentation is not blocked on AI failure")
    assert_true("Недостаточно подтвержденных рекомендаций для клиентской презентации" not in text, "Fixed recommendation threshold must stay removed")
    assert_true("Область для верификации" not in text, "Verification placeholder must not be emitted")
    assert_true("groq_prompt = f\"\"\"" in text, "Compact Groq prompt is missing")
    assert_true('"max_completion_tokens": 3400' in text, "Groq token budget is not capped below the free-tier TPM limit")
    assert_true('"response_format": {"type": "json_object"}' not in text, "Groq must use the local tolerant JSON parser")
    assert_true("generation_error_message" in text, "Persistent generation error state is missing")
    assert_true("def render_generation_failure_state" in text, "Stable generation failure screen is missing")
    assert_true('st.spinner("Производится глубокий анализ рисков...")' not in text, "Native spinner must not survive a failed generation")
    assert_true('min_items=1' in text and 'min_security_items=0' in text, "Groq quality gate must accept any confirmed recommendation count")
    assert_true("explicit_no_findings" in text, "A valid zero-recommendation AI result must be accepted")
    assert_true("if prepared_payload is not None:" in text, "Empty successful AI results must not be mistaken for failures")
    assert_true('"executive_summary": 2' in text and '"roadmap": 6' in text, "AI presentation narrative completeness check is missing")
    assert_true("часть презентационного материала будет дополнена" in text, "Partial AI narrative must preserve useful recommendations")
    assert_true("call_groq_with_rate_limit_retry(focus_it=True)" in text, "Fact-safe IT-focused Groq retry is missing")
    assert_true("confirmed_it_gap_topics(results)" in text, "Confirmed IT-gap detector is missing")
    assert_true("ai_it_gap_coverage(" in text, "AI IT-gap coverage gate is missing")
    assert_true("gemini_attempt_count >= 2" in text, "Gemini must receive a bounded second quality attempt")
    assert_true("stop_gemini = True" not in text, "Gemini must not be abandoned after its first incomplete response")
    assert_true("minimum_ai_items" in text, "Gemini quality threshold must adapt to confirmed gaps")
    assert_true("call_groq_with_rate_limit_retry" in text, "Groq rate-limit retry is missing")
    assert_true("retry_seconds + 1.0" in text, "Groq retry must respect the provider delay")
    assert_true('if str(item.get("source", "")).strip().lower() == "ии"' in text, "Customer presentation must use AI-authored risks only")
    assert_true("Сервис формирования экспертного заключения временно недоступен" in text, "Customer-safe generation error is missing")
    assert_true('replacements["__RECOMMENDATION_COUNT__"]' in text, "Presentation must support a variable recommendation count")
    assert_true("partial_recommendation_slide" in text, "Odd recommendation counts must use a single-card final slide")
    assert_true("entry = recommendation_by_key.get(roadmap_key)" in text, "Roadmap must exclude topics without a confirmed recommendation")
    assert_true("def staged_roadmap_action" in text, "Roadmap must provide assessment, pilot, and rollout stages")
    assert_true("roadmap_keys_by_phase" in text, "Roadmap must deduplicate technologies within each phase")
    assert_true("recover_complete_risk_objects(response_text)" in text, "Malformed AI JSON recovery is missing")


def check_partial_ai_json_recovery() -> None:
    app_tree = ast.parse(read_text(APP))
    recovery_node = next(
        node
        for node in app_tree.body
        if isinstance(node, ast.FunctionDef) and node.name == "recover_complete_risk_objects"
    )
    namespace = {"json": json, "re": re}
    exec(compile(ast.Module(body=[recovery_node], type_ignores=[]), str(APP), "exec"), namespace)
    malformed = """{
      "risks": [
        {"level":"HIGH","risk":"Первый подтвержденный риск","recommendation":"Выполнить полный первый комплекс мер и проверить измеримый результат."},
        {"level":"LOW","risk": },
        {"level":"MEDIUM","risk":"Второй подтвержденный риск","recommendation":"Выполнить полный второй комплекс мер и проверить измеримый результат."}
      ]
    }"""
    recovered = namespace["recover_complete_risk_objects"](malformed)
    assert_true(recovered is not None, "Malformed AI JSON was not recovered")
    assert_true(len(recovered["risks"]) == 2, "Valid AI recommendations around a broken item were lost")


def check_groq_payload_alias_normalization() -> None:
    app_tree = ast.parse(read_text(APP))
    normalize_node = next(
        node
        for node in app_tree.body
        if isinstance(node, ast.FunctionDef) and node.name == "normalize_ai_risks_payload"
    )
    count_node = next(
        node
        for node in app_tree.body
        if isinstance(node, ast.FunctionDef) and node.name == "count_ai_risk_candidates"
    )
    namespace = {"REGULATORY_CATALOG": {}}
    exec(
        compile(ast.Module(body=[normalize_node, count_node], type_ignores=[]), str(APP), "exec"),
        namespace,
    )
    payload = {
        "analysis": {
            "findings": [
                {
                    "title": "Недостаточный контроль привилегированных учетных записей",
                    "severity": "HIGH",
                    "domain": "ИБ",
                    "business_impact": "Компрометация административной учетной записи расширяет масштаб инцидента.",
                    "recommendation_steps": [
                        {"step": "Определить охват привилегированных учетных записей"},
                        {"step": "Провести пилот PAM и проверить сценарии доступа"},
                    ],
                    "metric": "100% критичных административных доступов заведены под контроль",
                }
            ]
        }
    }
    normalized = namespace["normalize_ai_risks_payload"](payload)
    assert_true(len(normalized) == 1, "Nested Groq finding aliases were not normalized")
    assert_true("Провести пилот PAM" in normalized[0]["recommendation"], "Groq action steps were lost")
    assert_true(normalized[0]["level"] == "HIGH", "Groq severity alias was lost")
    assert_true(namespace["count_ai_risk_candidates"](payload) == 1, "Raw Groq candidate count is incorrect")


def check_ai_narrative_risk_augmentation() -> None:
    app_tree = ast.parse(read_text(APP))
    nodes = {
        node.name: node
        for node in app_tree.body
        if isinstance(node, ast.FunctionDef)
        and node.name in {"risk_semantic_key", "augment_ai_risks_from_narrative"}
    }
    namespace = {"re": re}
    exec(
        compile(
            ast.Module(
                body=[nodes["risk_semantic_key"], nodes["augment_ai_risks_from_narrative"]],
                type_ignores=[],
            ),
            str(APP),
            "exec",
        ),
        namespace,
    )
    narrative = {
        "audit_observations": [
            {"title": "Привилегированные доступы", "text": "PAM в критичных системах не подтвержден."},
        ],
        "roadmap": [
            {"action": "Провести оценку привилегированных учетных записей и интегрировать PAM с SIEM.", "result": "Доступы учтены."},
            {"action": "Запустить пилотный NAC и проверить передачу событий в SIEM.", "result": "Устройства профилируются."},
            {"action": "Провести PoC IAM для управления жизненным циклом учетных записей.", "result": "Процесс согласован."},
        ],
    }
    initial = [{
        "risk": "Восстановление из резервных копий требует подтверждения",
        "recommendation": "Проводить регулярные тесты восстановления критичных сервисов и фиксировать RTO/RPO.",
    }]
    augmented = namespace["augment_ai_risks_from_narrative"](
        initial,
        narrative,
        {"PAM": "Нет", "NAC": "Нет", "IAM": "Нет"},
    )
    keys = {namespace["risk_semantic_key"](item) for item in augmented}
    assert_true(
        {"backup", "pam", "nac", "iam"}.issubset(keys),
        f"AI narrative did not restore all technical findings: {sorted(keys)}",
    )
    assert_true(
        namespace["risk_semantic_key"]({"risk": "Пилот NAC с передачей событий в SIEM"}) == "nac",
        "NAC must not be misclassified as SIEM when both technologies are mentioned",
    )
    assert_true(
        namespace["risk_semantic_key"]({"risk": "PAM с передачей событий в SIEM"}) == "pam",
        "PAM must not be misclassified as SIEM when both technologies are mentioned",
    )


def check_short_known_ai_risk_titles() -> None:
    app_tree = ast.parse(read_text(APP))
    wanted = {
        "risk_semantic_key", "is_truncated_ai_text",
        "canonical_ai_risk_title", "prepare_ai_risks_for_report",
    }
    nodes = {
        node.name: node
        for node in app_tree.body
        if isinstance(node, ast.FunctionDef) and node.name in wanted
    }
    namespace = {"re": re}
    exec(
        compile(
            ast.Module(
                body=[
                    nodes["risk_semantic_key"], nodes["is_truncated_ai_text"],
                    nodes["canonical_ai_risk_title"], nodes["prepare_ai_risks_for_report"],
                ],
                type_ignores=[],
            ),
            str(APP),
            "exec",
        ),
        namespace,
    )
    prepared = namespace["prepare_ai_risks_for_report"]([
        {
            "risk": "PAM",
            "recommendation": "Инвентаризировать привилегированные доступы, провести пилот и распространить контроль на критичные системы.",
        },
        {
            "risk": "NAC",
            "recommendation": "Провести пилот контроля допуска устройств, настроить профилирование и изоляцию неизвестных подключений.",
        },
    ])
    assert_true(len(prepared) == 2, "Known short PAM/NAC findings must survive the quality gate")
    assert_true(all(len(item["risk"]) >= 20 for item in prepared), "Short AI risk titles were not expanded")


def check_presentation_fact_guards() -> None:
    app_tree = ast.parse(read_text(APP))
    function_names = {
        "is_enabled",
        "risk_semantic_key",
        "control_confirmed_in_results",
        "risk_conflicts_with_answers",
        "network_segmentation_evidence",
        "enforce_audit_fact_policy",
    }
    nodes_by_name = {
        node.name: node
        for node in app_tree.body
        if isinstance(node, ast.FunctionDef) and node.name in function_names
    }
    namespace = {"re": re, "expand_regulatory_references": lambda value: str(value or "")}
    ordered_nodes = [nodes_by_name[name] for name in (
        "is_enabled",
        "risk_semantic_key",
        "control_confirmed_in_results",
        "risk_conflicts_with_answers",
        "network_segmentation_evidence",
        "enforce_audit_fact_policy",
    )]
    exec(compile(ast.Module(body=ordered_nodes, type_ignores=[]), str(APP), "exec"), namespace)
    conflicts = namespace["risk_conflicts_with_answers"]

    assert_true(
        conflicts(
            {"risk": "Устаревшие операционные системы требуют миграции"},
            {"ОС АРМ (Windows XP/Vista/7/8)": 0, "ОС Сервера (Windows Server 2008/2012 R2)": 0},
        ),
        "Zero legacy OS counts must block a legacy OS recommendation",
    )
    assert_true(
        conflicts({"risk": "Endpoint detection требует усиления EDR"}, {"EDR": "Check Point Harmony EDR"}),
        "Existing EDR must block an endpoint detection recommendation",
    )
    assert_true(
        conflicts(
            {"risk": "Публичным приложениям требуется WAF"},
            {"WAF": "Нет", "Примечание к разделу ИБ": "WAF есть и настроен на публичном контуре"},
        ),
        "A confirmed WAF in questionnaire notes must block a WAF recommendation",
    )
    assert_true(
        conflicts(
            {"risk": "Отсутствие резервного копирования критичных конфигураций"},
            {"Резервное копирование": "Veeam"},
        ),
        "Existing backup must block an absence-of-backup recommendation",
    )

    enforce = namespace["enforce_audit_fact_policy"]
    nac_item = enforce(
        {
            "level": "LOW",
            "risk": "Архитектура сетевой сегментации требует подтверждения",
            "recommendation": "Проверить VLAN и ACL.",
        },
        {"NAC": "Нет", "Маршрутизация": "OSPF", "Wi-Fi Точки доступа": 48},
        {"is_kvoiki": True, "has_personal_data": True},
    )
    assert_true(
        namespace["risk_semantic_key"](nac_item) == "nac" and nac_item["level"] == "MEDIUM",
        "Unknown segmentation must become a precise NAC admission-control recommendation",
    )
    assert_true(
        "lateral movement" not in nac_item["impact"].lower()
        and "vlan" in nac_item["description"].lower()
        and "не подтверждает отсутствие" in nac_item["description"].lower(),
        "NAC finding must not claim that segmentation or lateral movement was proven",
    )
    iam_item = enforce(
        {
            "level": "HIGH",
            "risk": "Отсутствие IAM приводит к избыточным привилегиям пользователей",
            "recommendation": "Включить RBAC в Active Directory.",
        },
        {"IAM": "Нет", "_user_count": 350},
        {"users": 350},
    )
    assert_true(
        "приема, перевода и увольнения" in iam_item["recommendation"].lower()
        and "active directory" not in iam_item["recommendation"].lower(),
        "IAM recommendation must cover the account lifecycle instead of only AD RBAC",
    )
    dlp_item = enforce(
        {"level": "LOW", "risk": "Отсутствие DLP", "recommendation": "Оценить DLP."},
        {"DLP": "Нет"},
        {"is_kvoiki": True, "has_personal_data": True},
    )
    assert_true(
        dlp_item["level"] == "HIGH"
        and dlp_item["risk"].endswith("персональных данных")
        and "DLP в анкете" in dlp_item["evidence"][0]
        and "виртуализац" not in " ".join(dlp_item["evidence"]).lower(),
        "KVOIKI DLP gap must have a complete title and DLP-specific evidence",
    )
    backup_item = enforce(
        {
            "level": "HIGH",
            "risk": "Недостаточная проверка восстановления резервных копий",
            "impact": "Невозможность восстановления данных и нарушение требований PD_LAW.",
            "recommendation": "Проверить восстановление.",
        },
        {"Резервное копирование": "Veeam Backup & Replication"},
        {"has_backup": True},
    )
    assert_true(
        "невозможность" not in backup_item["impact"].lower()
        and "Veeam Backup & Replication" in backup_item["description"],
        "Existing backup must be described as an unverified recovery capability, not as impossible recovery",
    )


def check_presentation_templates() -> None:
    required = {
        "{{COMPANY}}",
        "{{IT_SCORE}}",
        "{{SUMMARY_1}}",
        "{{RISK_1_TITLE}}",
        "{{RISK_6_TITLE}}",
        "{{STRENGTH_1}}",
        "{{THREAT_1_LABEL}}",
        "{{THREAT_1_VALUE}}",
        "{{THREAT_6_LABEL}}",
        "{{THREAT_6_VALUE}}",
        "{{COVERAGE_AVERAGE}}",
        "{{COVERAGE_INSIGHT}}",
        "{{REG_TITLE}}",
        "{{REG_APPLICABILITY}}",
        "{{REG_EXPECTATIONS}}",
        "{{REG_IMPLEMENTATION}}",
        "{{REG_ANCHORS}}",
        "{{FRAMEWORKS}}",
        "{{REC_1_TITLE}}",
        "{{REC_1_ACTION}}",
        "{{REC_1_EVIDENCE}}",
        "{{REC_1_LEGAL}}",
        "{{REC_1_METRIC}}",
        "{{REC_1_SOLUTION}}",
        "{{REC_1_VENDORS}}",
        "{{REC_8_TITLE}}",
        "{{REC_8_ACTION}}",
        "{{ROADMAP_1_1}}",
        "{{ROADMAP_1_1_RESULT}}",
        "{{ROADMAP_1_2_RESULT}}",
        "{{DECISION_1}}",
    }
    assert_true(PRESENTATION_COVER_ASSET.exists(), "Presentation cover image is missing")
    cover_hash = hashlib.sha256(PRESENTATION_COVER_ASSET.read_bytes()).hexdigest()
    for template, qr_asset in zip(PRESENTATION_TEMPLATES, PRESENTATION_QR_ASSETS):
        assert_true(template.exists(), f"Presentation template is missing: {template.name}")
        with zipfile.ZipFile(template, "r") as archive:
            slide_xml = "\n".join(
                archive.read(name).decode("utf-8")
                for name in archive.namelist()
                if name.startswith("ppt/slides/slide") and name.endswith(".xml")
            )
            media_hashes = {
                hashlib.sha256(archive.read(name)).hexdigest()
                for name in archive.namelist()
                if name.startswith("ppt/media/")
            }
        missing = sorted(token for token in required if token not in slide_xml)
        assert_true(not missing, f"{template.name} is missing placeholders: {missing}")
        qr_hash = hashlib.sha256(qr_asset.read_bytes()).hexdigest()
        assert_true(qr_hash in media_hashes, f"{template.name} does not embed its contact QR")
        assert_true(cover_hash in media_hashes, f"{template.name} does not embed the audit cover image")
    for qr_asset in PRESENTATION_QR_ASSETS:
        assert_true(qr_asset.exists(), f"Presentation QR is missing: {qr_asset.name}")
        assert_true(qr_asset.stat().st_size > 2000, f"Presentation QR is unexpectedly small: {qr_asset.name}")


def check_customer_output_normalization() -> None:
    app_tree = ast.parse(read_text(APP))
    wanted_assignments = {"REGULATORY_CATALOG", "INDUSTRY_REGULATORY_IDS", "INDUSTRY_FRAMEWORKS"}
    assignment_nodes = [
        node for node in app_tree.body
        if isinstance(node, ast.Assign)
        and any(isinstance(target, ast.Name) and target.id in wanted_assignments for target in node.targets)
    ]
    function_names = {
        "expand_regulatory_references",
        "sanitize_customer_roadmap_text",
        "risk_semantic_key",
        "industry_regulatory_profile",
    }
    functions = {
        node.name: node for node in app_tree.body
        if isinstance(node, ast.FunctionDef) and node.name in function_names
    }
    ordered = [
        *assignment_nodes,
        functions["expand_regulatory_references"],
        functions["sanitize_customer_roadmap_text"],
        functions["risk_semantic_key"],
        functions["industry_regulatory_profile"],
    ]
    namespace = {"re": re}
    exec(compile(ast.Module(body=ordered, type_ignores=[]), str(APP), "exec"), namespace)

    assert_true(
        "[PD_LAW]" not in namespace["expand_regulatory_references"]("Штрафы по [PD_LAW]"),
        "Internal regulatory IDs leaked into customer text",
    )
    bare_legal_text = namespace["expand_regulatory_references"](
        "Нарушение требований PD_LAW и FINANCE_IS."
    )
    assert_true(
        "PD_LAW" not in bare_legal_text and "FINANCE_IS" not in bare_legal_text,
        "Bare internal regulatory IDs leaked into customer text",
    )
    for industry, identifiers in namespace["INDUSTRY_REGULATORY_IDS"].items():
        expanded = namespace["expand_regulatory_references"](
            f"Профиль {industry}: " + ", ".join(identifiers)
        )
        assert_true(
            not any(re.search(rf"(?<![A-Za-z0-9_]){re.escape(identifier)}(?![A-Za-z0-9_])", expanded) for identifier in identifiers),
            f"Internal regulatory ID leaked for industry: {industry}",
        )
    assert_true(
        namespace["risk_semantic_key"]({
            "risk": "Привилегированные учетные записи требуют PAM",
            "recommendation": "Провести инвентаризацию администраторов",
        }) == "pam",
        "PAM risk was incorrectly classified as ITAM",
    )
    roadmap = namespace["sanitize_customer_roadmap_text"]("Закупить и начать пилот DLP (Forcepoint).")
    assert_true(
        "закупить" not in roadmap.lower() and "пилот" in roadmap.lower() and "forcepoint" not in roadmap.lower(),
        "DLP roadmap must use pilot-before-procurement and remain vendor-neutral",
    )
    pam_roadmap = namespace["sanitize_customer_roadmap_text"](
        "Развернуть пилотный PAM для администраторов Windows Server 2019/2022 (CyberArk)."
    )
    assert_true(
        "windows" not in pam_roadmap.lower()
        and "cyberark" not in pam_roadmap.lower()
        and "критич" in pam_roadmap.lower(),
        "PAM roadmap must cover the critical environment and remain vendor-neutral",
    )
    pam_scale = namespace["sanitize_customer_roadmap_text"](
        "Внедрить PAM на все серверы и масштабировать на платформы виртуализации."
    )
    assert_true(
        len(pam_scale) <= 140 and pam_scale.endswith("SIEM."),
        "Scaled PAM roadmap must remain complete within the presentation text budget",
    )
    nac_roadmap = namespace["sanitize_customer_roadmap_text"](
        "Внедрить выбранный NAC, настроить 1X, интегрировать с AD и R-Vision SIEM."
    )
    assert_true(
        "802.1X" in nac_roadmap and "r-vision" not in nac_roadmap.lower(),
        "NAC roadmap must use the correct protocol name and remain vendor-neutral",
    )
    pam_productless = namespace["sanitize_customer_roadmap_text"]("Развернуть PAM (без продукта).")
    assert_true(
        "без продукта" not in pam_productless.lower() and "пилот pam" in pam_productless.lower(),
        "Productless PAM placeholder leaked into the customer roadmap",
    )
    vm_roadmap = namespace["sanitize_customer_roadmap_text"](
        "Запустить базовый скан уязвимостей серверов с помощью OpenVAS."
    )
    assert_true(
        "openvas" not in vm_roadmap.lower() and "уязвим" in vm_roadmap.lower(),
        "Vulnerability roadmap must remain vendor-neutral",
    )
    backup_roadmap = namespace["sanitize_customer_roadmap_text"](
        "Внедрить регулярное тестирование восстановления из Veeam и документировать RTO/RPO."
    )
    assert_true(
        "veeam" not in backup_roadmap.lower()
        and "из и" not in backup_roadmap.lower()
        and "резервных копий" in backup_roadmap.lower(),
        "Vendor removal must not leave a broken backup roadmap sentence",
    )
    bank = namespace["industry_regulatory_profile"]("Финтех / Банки")
    kvoiki = namespace["industry_regulatory_profile"]("КВОИКИ / Критическая инфраструктура")
    assert_true(
        "BANK_IS" in bank["legal_ids"] and "KVOIKI_529" not in bank["legal_ids"],
        "Bank profile contains incorrect regulatory scope",
    )
    assert_true(
        "KVOIKI_529" in kvoiki["legal_ids"] and bank["legal_ids"] != kvoiki["legal_ids"],
        "Industry profiles do not produce distinct regulatory outputs",
    )


def check_dynamic_presentation_range() -> None:
    app_tree = ast.parse(read_text(APP))
    renderer_node = next(
        node
        for node in app_tree.body
        if isinstance(node, ast.FunctionDef) and node.name == "render_audit_presentation_template"
    )
    namespace = {"BytesIO": io.BytesIO, "re": re}
    exec(compile(ast.Module(body=[renderer_node], type_ignores=[]), str(APP), "exec"), namespace)
    renderer = namespace["render_audit_presentation_template"]

    template = PRESENTATION_TEMPLATES[0]
    with zipfile.ZipFile(template, "r") as archive:
        template_xml = "\n".join(
            archive.read(name).decode("utf-8")
            for name in archive.namelist()
            if name.endswith(".xml")
        )
    base_replacements = {
        token: "Проверка"
        for token in re.findall(r"\{\{([A-Z0-9_]+)\}\}", template_xml)
    }

    for recommendation_count in (0, 1, 7, 8, 9, 12, 15):
        replacements = dict(base_replacements)
        replacements["__RECOMMENDATION_COUNT__"] = recommendation_count
        replacements["__RISK_COUNT__"] = 1
        for index in range(1, recommendation_count + 2):
            for field in ("LEVEL", "TITLE", "ACTION", "SOLUTION", "VENDORS", "EVIDENCE", "LEGAL", "METRIC"):
                replacements[f"REC_{index}_{field}"] = f"Рекомендация {index}"
            replacements[f"REC_{index}_FILL"] = "#F4B400"
            replacements[f"REC_{index}_TEXT"] = "#1F2937"

        result = renderer(template, replacements)
        with zipfile.ZipFile(io.BytesIO(result), "r") as archive:
            presentation = archive.read("ppt/presentation.xml").decode("utf-8")
            content_types = archive.read("[Content_Types].xml").decode("utf-8")
            presentation_rels = archive.read("ppt/_rels/presentation.xml.rels").decode("utf-8")
            assert_true("<ns0:Types" not in content_types, "Content types XML uses a non-standard namespace prefix")
            assert_true("<ns0:Relationships" not in presentation_rels, "Presentation relationships use a non-standard namespace prefix")
            active_slide_count = presentation.count("<p:sldId ")
            expected_slide_count = 13 - 4 + ((recommendation_count + 1) // 2)
            assert_true(
                active_slide_count == expected_slide_count,
                f"Unexpected slide count for {recommendation_count} recommendations: "
                f"{active_slide_count} != {expected_slide_count}",
            )
            active_xml = "\n".join(
                archive.read(name).decode("utf-8")
                for name in archive.namelist()
                if name.startswith("ppt/slides/slide") and name.endswith(".xml")
            )
            assert_true("{{REC_" not in active_xml, f"Unresolved recommendation token at count {recommendation_count}")


def check_sample_drafts() -> None:
    if not BANK_DRAFT.exists():
        print("SKIP banking draft fixture is not present")
        return
    payload = json.loads(BANK_DRAFT.read_text(encoding="utf-8"))
    state = payload.get("state", {})
    assert_true(payload.get("schema") == "khalil-audit-draft-v1", "Unexpected banking draft schema")
    assert_true(state.get("client_industry_select") == "Финтех / Банки", "Banking sector is not selected")
    assert_true(state.get("mfa") is True, "Banking sample must confirm existing MFA")
    assert_true(state.get("pam") is False, "Banking sample must expose the PAM gap")
    arm_total = sum(int(state.get(f"arm_{name}", 0) or 0) for name in state.get("selected_os_arm", []))
    server_total = sum(int(state.get(f"fsrv_{name}", 0) or 0) for name in state.get("ms_srv_list", []))
    assert_true(arm_total == int(state.get("total_arm", 0)), "Banking ARM counts are inconsistent")
    assert_true(server_total == int(state.get("phys_srv", 0)) + int(state.get("virt_srv", 0)), "Banking server counts are inconsistent")


def main() -> None:
    checks = [
        check_compile,
        check_version,
        check_customer_changelog,
        check_selectbox_contract,
        check_portfolio,
        check_static_hooks,
        check_partial_ai_json_recovery,
        check_groq_payload_alias_normalization,
        check_ai_narrative_risk_augmentation,
        check_short_known_ai_risk_titles,
        check_presentation_fact_guards,
        check_presentation_templates,
        check_customer_output_normalization,
        check_dynamic_presentation_range,
        check_sample_drafts,
    ]
    for check in checks:
        check()
        print(f"OK {check.__name__}")
    print("SMOKE TEST PASSED")


if __name__ == "__main__":
    main()
