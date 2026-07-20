import streamlit as st
import streamlit.components.v1 as components
import pandas as pd
import os
import html
import base64
import json
import math
import re
import zlib
import threading
import time
import random
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from datetime import datetime

REQUEST_VERIFY = True


def configure_https_certificates():
    global REQUEST_VERIFY
    try:
        import certifi
        cert_path = certifi.where()
        os.environ.setdefault("SSL_CERT_FILE", cert_path)
        os.environ.setdefault("REQUESTS_CA_BUNDLE", cert_path)
        os.environ.setdefault("GRPC_DEFAULT_SSL_ROOTS_FILE_PATH", cert_path)
        REQUEST_VERIFY = cert_path
    except Exception:
        REQUEST_VERIFY = True


configure_https_certificates()

#----------ИИ-----------
# --- AI BLOCK START ---

def sanitize_for_ai(c_info, results):
    forbidden = [
        "Наименование компании",
        "Сайт компании",
        "Email",
        "ФИО контактного лица",
        "Должность",
        "Контактный телефон",
        "Город",
        "Компания",
        "Контакт",
        "Телефон",
        "client_",
        "name_",
        "email",
        "phone",
    ]
    safe_client = {
        "Сфера деятельности": c_info.get("Сфера деятельности", "")
    }
    safe_results = {
        k: v for k, v in results.items()
        if not any(f.lower() in str(k).lower() for f in forbidden)
    }
    return safe_client, safe_results

def load_vendor_matrix():
    try:
        df = pd.read_excel("Портфель для отчета.xlsx")

        vendors_text = ""

        for _, row in df.iterrows():
            row_text = " | ".join(
                [str(x) for x in row.values if pd.notna(x)]
            )
            vendors_text += row_text + "\n"

        return vendors_text

    except Exception as e:
        return f"Ошибка загрузки вендоров: {e}"


def load_vendor_names():
    try:
        detailed = load_detailed_vendor_names()
        if detailed:
            return detailed
        df = pd.read_excel("Портфель для отчета.xlsx", header=None)
        values = []
        for value in df.iloc[:, 0].dropna().tolist():
            vendor = str(value).strip()
            if vendor and vendor.lower() not in {"nan", "none"}:
                values.append(vendor)
        return list(dict.fromkeys(values))
    except Exception:
        return []


def get_app_secret(name, default=None):
    try:
        return st.secrets.get(name, default)
    except Exception:
        return default


APP_INSTANCE_DEFAULT = "Test"
APP_VERSION = "12.30-dev"


def get_app_instance_label():
    value = str(get_app_secret("APP_INSTANCE", APP_INSTANCE_DEFAULT) or "").strip()
    return value or APP_INSTANCE_DEFAULT


def redact_secret(text, *secrets):
    safe_text = str(text)
    for secret in secrets:
        if secret:
            safe_text = safe_text.replace(str(secret), "***")
    return safe_text


def find_node_exe():
    import shutil

    node_exe = get_app_secret("NODE_EXE", None)
    if node_exe:
        return node_exe

    node_exe = shutil.which("node")
    if node_exe:
        return node_exe

    local_node = os.path.expanduser(
        r"~\.cache\codex-runtimes\codex-primary-runtime\dependencies\node\bin\node.exe"
    )
    if os.path.exists(local_node):
        return local_node

    return None


def node_fetch_json(url, payload, timeout_seconds=25, env_extra=None):
    import subprocess
    import tempfile

    node_exe = find_node_exe()
    if not node_exe:
        raise RuntimeError("Node.js не найден для HTTPS-запроса.")

    node_script = r"""
const fs = require('fs');
const requestPath = process.argv[2];
const request = JSON.parse(fs.readFileSync(requestPath, 'utf8'));
const controller = new AbortController();
const timer = setTimeout(() => controller.abort(), request.timeoutMs || 25000);

(async () => {
  const headers = request.headers || {};
  let body;

  if (request.multipart) {
    body = new FormData();
    for (const [key, value] of Object.entries(request.fields || {})) {
      body.append(key, String(value));
    }
    for (const file of request.files || []) {
      const bytes = await fs.promises.readFile(file.path);
      body.append(file.field, new Blob([bytes]), file.filename);
    }
  } else {
    headers['Content-Type'] = 'application/json';
    body = JSON.stringify(request.body || {});
  }

  const response = await fetch(request.url, {
    method: request.method || 'POST',
    headers,
    body,
    signal: controller.signal,
  });
  const text = await response.text();
  clearTimeout(timer);

  if (!response.ok) {
    console.error(text.slice(0, 1500));
    process.exit(2);
  }

  process.stdout.write(text || '{}');
})().catch((error) => {
  clearTimeout(timer);
  console.error(`${error.name}: ${error.message}`);
  process.exit(1);
});
"""
    temp_paths = []
    try:
        with tempfile.NamedTemporaryFile("w", suffix=".json", delete=False, encoding="utf-8") as request_file:
            json.dump(payload, request_file, ensure_ascii=False)
            request_path = request_file.name
            temp_paths.append(request_path)
        with tempfile.NamedTemporaryFile("w", suffix=".cjs", delete=False, encoding="utf-8") as script_file:
            script_file.write(node_script)
            script_path = script_file.name
            temp_paths.append(script_path)

        env = os.environ.copy()
        env["NODE_TLS_REJECT_UNAUTHORIZED"] = "0"
        env["NODE_OPTIONS"] = f"{env.get('NODE_OPTIONS', '')} --no-warnings".strip()
        if env_extra:
            env.update(env_extra)

        completed = subprocess.run(
            [node_exe, script_path, request_path],
            capture_output=True,
            text=True,
            timeout=timeout_seconds + 5,
            env=env,
        )
        if completed.returncode != 0:
            stderr_lines = []
            for line in completed.stderr.splitlines():
                if "Assertion failed:" in line or "UV_HANDLE_CLOSING" in line:
                    continue
                if "NODE_TLS_REJECT_UNAUTHORIZED" in line:
                    continue
                stderr_lines.append(line)
            stderr_text = "\n".join(stderr_lines).strip()
            if len(stderr_text) > 1200:
                stderr_text = stderr_text[:1200] + "..."
            raise RuntimeError(stderr_text or "Node HTTPS request failed")
        return json.loads(completed.stdout or "{}")
    finally:
        for path in temp_paths:
            try:
                os.unlink(path)
            except OSError:
                pass


def telegram_send_node(token, method, fields, files=None, timeout_seconds=10):
    import requests

    upload_files = {}
    try:
        for file_item in files or []:
            field_name = file_item.get("field", "document")
            upload_files[field_name] = (
                file_item["filename"],
                BytesIO(file_item["bytes"]),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

        response = requests.post(
            f"https://api.telegram.org/bot{token}/{method}",
            data=fields,
            files=upload_files or None,
            timeout=timeout_seconds,
            verify=REQUEST_VERIFY,
        )
        response.raise_for_status()
        return response.json()
    except requests.exceptions.SSLError:
        response = requests.post(
            f"https://api.telegram.org/bot{token}/{method}",
            data=fields,
            files=upload_files or None,
            timeout=timeout_seconds,
            verify=False,
        )
        response.raise_for_status()
        return response.json()


def build_telegram_lead_text(client_info, final_score, sales_digest):
    ai_provider = str(st.session_state.get("ai_provider_used", "Не определен")).strip() or "Не определен"
    return (
        "🚨 Новый запрос на аудит!\n"
        f"📌 Приложение: {get_app_instance_label()}\n"
        f"🏢 Компания: {client_info.get('Наименование компании', '-')}\n"
        f"📍 Город: {client_info.get('Город', '-')}\n"
        f"📊 Сфера: {client_info.get('Сфера деятельности', '-')}\n"
        f"🌐 Сайт: {client_info.get('Сайт компании', '-')}\n"
        f"📧 Email: {client_info.get('Email', '-')}\n"
        f"📞 Телефон: {client_info.get('Контактный телефон', '-')}\n"
        f"👤 Контакт: {client_info.get('ФИО контактного лица', '-')}\n"
        f"💼 Должность: {client_info.get('Должность', '-')}\n"
        f"📊 Уровень зрелости: {final_score}%\n\n"
        f"🤖 Анализ: {ai_provider}\n\n"
        f"💡 Что предложить первым:\n{sales_digest}"
    )


def build_telegram_ai_failure_text(client_info, final_score, ai_error):
    safe_error = str(ai_error or "Неизвестная ошибка").strip()
    if len(safe_error) > 1800:
        safe_error = safe_error[:1800] + "..."

    return (
        "❌ Отчет заказчику не сформирован: ИИ-анализ временно недоступен\n"
        f"📌 Приложение: {get_app_instance_label()}\n"
        f"🏢 Компания: {client_info.get('Наименование компании', '-')}\n"
        f"📍 Город: {client_info.get('Город', '-')}\n"
        f"📊 Сфера: {client_info.get('Сфера деятельности', '-')}\n"
        f"🌐 Сайт: {client_info.get('Сайт компании', '-')}\n"
        f"📧 Email: {client_info.get('Email', '-')}\n"
        f"📞 Телефон: {client_info.get('Контактный телефон', '-')}\n"
        f"👤 Контакт: {client_info.get('ФИО контактного лица', '-')}\n"
        f"💼 Должность: {client_info.get('Должность', '-')}\n"
        f"📊 Уровень зрелости: {final_score}%\n\n"
        f"Внутренняя диагностика: {safe_error}"
    )


def build_telegram_generation_started_text(client_info, final_score):
    return (
        "🟢 Начато формирование аудита Khalil Audit\n"
        f"📌 Приложение: {get_app_instance_label()}\n"
        f"🏢 Компания: {client_info.get('Наименование компании', '-')}\n"
        f"📍 Город: {client_info.get('Город', '-')}\n"
        f"📊 Сфера: {client_info.get('Сфера деятельности', '-')}\n"
        f"🌐 Сайт: {client_info.get('Сайт компании', '-')}\n"
        f"📧 Email: {client_info.get('Email', '-')}\n"
        f"📞 Телефон: {client_info.get('Контактный телефон', '-')}\n"
        f"👤 Контакт: {client_info.get('ФИО контактного лица', '-')}\n"
        f"💼 Должность: {client_info.get('Должность', '-')}\n"
        f"📊 Предварительная зрелость: {final_score}%"
    )


def build_telegram_generation_error_text(client_info, final_score, error):
    safe_error = str(error or "Неизвестная ошибка").strip()
    if len(safe_error) > 1800:
        safe_error = safe_error[:1800] + "..."

    return (
        "🔴 Ошибка формирования аудита Khalil Audit\n"
        f"📌 Приложение: {get_app_instance_label()}\n"
        f"🏢 Компания: {client_info.get('Наименование компании', '-')}\n"
        f"📍 Город: {client_info.get('Город', '-')}\n"
        f"📧 Email: {client_info.get('Email', '-')}\n"
        f"📞 Телефон: {client_info.get('Контактный телефон', '-')}\n"
        f"👤 Контакт: {client_info.get('ФИО контактного лица', '-')}\n"
        f"📊 Предварительная зрелость: {final_score}%\n\n"
        f"Диагностика: {safe_error}"
    )


def send_internal_telegram_message(text, timeout_seconds=8):
    if not TOKEN or not CHAT_ID:
        return "Telegram не отправлен: не найдены TELEGRAM_TOKEN или TELEGRAM_CHAT_ID."

    try:
        telegram_send_node(
            TOKEN,
            "sendMessage",
            {"chat_id": CHAT_ID, "text": text},
            timeout_seconds=timeout_seconds
        )
        return "ok"
    except Exception as exc:
        return f"Telegram не отправлен: {redact_secret(exc, TOKEN)}"


def normalize_ai_risks_payload(payload):
    def repair_mojibake(value):
        text = str(value)
        suspicious = sum(text.count(marker) for marker in ("Р", "С", "Ð", "Ñ"))
        if suspicious < 3:
            return text

        try:
            repaired = text.encode("cp1251").decode("utf-8")
        except UnicodeError:
            return text

        original_badness = text.count("Р") + text.count("С") + text.count("Ð") + text.count("Ñ")
        repaired_badness = repaired.count("Р") + repaired.count("С") + repaired.count("Ð") + repaired.count("Ñ")
        return repaired if repaired_badness < original_badness else text

    def value_as_text(value, preferred_keys=()):
        if value is None:
            return ""
        if isinstance(value, str):
            return repair_mojibake(value).strip()
        if isinstance(value, (int, float, bool)):
            return str(value).strip()
        if isinstance(value, list):
            parts = [value_as_text(part, preferred_keys) for part in value]
            return "; ".join(part for part in parts if part)
        if isinstance(value, dict):
            for key in preferred_keys:
                text = value_as_text(value.get(key), preferred_keys)
                if text:
                    return text
            parts = [value_as_text(part, preferred_keys) for part in value.values()]
            return "; ".join(part for part in parts if part)
        return repair_mojibake(value).strip()

    def first_text(item, keys, preferred_keys=()):
        for key in keys:
            text = value_as_text(item.get(key), preferred_keys)
            if text:
                return text
        return ""

    def normalize_risk_item(item):
        if not isinstance(item, dict):
            return None

        risk = first_text(
            item,
            ("risk", "title", "finding", "issue", "gap", "name"),
            ("title", "risk", "text", "name"),
        )
        recommendation = first_text(
            item,
            (
                "recommendation", "recommendations", "action", "actions",
                "recommendation_steps", "remediation", "mitigation", "next_steps",
            ),
            ("action", "text", "step", "recommendation", "description"),
        )
        if not risk or not recommendation:
            return None

        vendors = item.get("vendors", [])
        legal_ids = item.get("legal_ids", [])
        frameworks = item.get("frameworks", [])
        evidence = item.get("evidence", [])
        if not isinstance(vendors, list):
            vendors = [vendors] if vendors else []
        if not isinstance(legal_ids, list):
            legal_ids = [legal_ids] if legal_ids else []
        if not isinstance(frameworks, list):
            frameworks = [frameworks] if frameworks else []
        if not isinstance(evidence, list):
            evidence = [evidence] if evidence else []

        normalized_legal_ids = [
            str(value).strip()
            for value in legal_ids
            if str(value).strip() in REGULATORY_CATALOG
        ][:3]

        return {
            "level": first_text(item, ("level", "severity", "priority")) or "MEDIUM",
            "area": first_text(item, ("area", "domain", "category")),
            "risk": risk,
            "description": first_text(
                item,
                ("description", "observation", "finding_details", "details", "context"),
                ("text", "description", "details"),
            ) or risk,
            "impact": first_text(
                item,
                ("impact", "business_impact", "consequence", "consequences"),
                ("text", "impact", "description"),
            ) or "Риск может привести к снижению устойчивости ИТ/ИБ процессов.",
            "recommendation": recommendation,
            "vendors": [repair_mojibake(value).strip() for value in vendors if repair_mojibake(value).strip()][:3],
            "legal_ids": normalized_legal_ids,
            "regulators": [REGULATORY_CATALOG[value]["short"] for value in normalized_legal_ids],
            "frameworks": [repair_mojibake(value).strip() for value in frameworks if repair_mojibake(value).strip()][:3],
            "evidence": [repair_mojibake(value).strip() for value in evidence if repair_mojibake(value).strip()][:3],
            "success_metric": first_text(
                item,
                ("success_metric", "metric", "target", "acceptance_criteria", "kpi"),
                ("text", "metric", "target", "value"),
            ),
        }

    def collect_from_list(value, source_area="ИИ"):
        normalized_items = []
        if not isinstance(value, list):
            return normalized_items
        for item in value:
            normalized = normalize_risk_item(item)
            if normalized:
                item_area = str(normalized.get("area", "")).strip().upper()
                normalized["_ai_area"] = item_area if item_area in {"ИТ", "ИБ"} else source_area
                normalized_items.append(normalized)
        return normalized_items

    if isinstance(payload, list):
        return [
            normalized
            for normalized in collect_from_list(payload)
            if normalized
        ]

    if isinstance(payload, dict):
        combined = []
        for key, source_area in (
            ("it_recommendations", "ИТ"),
            ("security_recommendations", "ИБ"),
            ("risks", "ИИ"),
            ("recommendations", "ИИ"),
            ("items", "ИИ"),
            ("data", "ИИ"),
        ):
            combined.extend(collect_from_list(payload.get(key), source_area))
        if combined:
            return combined

        for container_key in ("analysis", "audit", "report", "result", "findings"):
            nested = payload.get(container_key)
            if isinstance(nested, (dict, list)):
                combined.extend(normalize_ai_risks_payload(nested))
        if combined:
            return combined

        for key in ("risks", "recommendations", "items", "data"):
            value = payload.get(key)
            if isinstance(value, list):
                return [
                    normalized
                    for normalized in (normalize_risk_item(item) for item in value)
                    if normalized
                ]

    return []


def count_ai_risk_candidates(payload):
    if isinstance(payload, list):
        return sum(isinstance(item, dict) for item in payload)
    if not isinstance(payload, dict):
        return 0

    total = 0
    for key in (
        "it_recommendations", "security_recommendations", "risks",
        "recommendations", "items", "data", "findings",
    ):
        value = payload.get(key)
        if isinstance(value, list):
            total += sum(isinstance(item, dict) for item in value)
    for key in ("analysis", "audit", "report", "result"):
        value = payload.get(key)
        if isinstance(value, (dict, list)):
            total += count_ai_risk_candidates(value)
    return total


def normalize_ai_audit_narrative(payload):
    if not isinstance(payload, dict):
        return {}

    def clean_text(value):
        text = str(value or "").strip()
        suspicious = sum(text.count(marker) for marker in ("Р", "С", "Ð", "Ñ"))
        if suspicious >= 3:
            try:
                repaired = text.encode("cp1251").decode("utf-8")
                if repaired.count("Р") + repaired.count("С") < text.count("Р") + text.count("С"):
                    text = repaired
            except UnicodeError:
                pass
        return re.sub(r"\s+", " ", text).strip()

    def clean_list(values, limit=8):
        if not isinstance(values, list):
            return []
        cleaned = []
        for value in values:
            if isinstance(value, dict):
                text = clean_text(value.get("text") or value.get("summary") or value.get("recommendation"))
            else:
                text = clean_text(value)
            if text:
                cleaned.append(text)
            if len(cleaned) >= limit:
                break
        return cleaned

    def clean_observations(values, limit=6):
        if not isinstance(values, list):
            return []
        cleaned = []
        for value in values:
            if isinstance(value, dict):
                title = clean_text(value.get("title") or value.get("domain") or value.get("risk"))
                text = clean_text(value.get("text") or value.get("description") or value.get("recommendation"))
            else:
                title = "Наблюдение"
                text = clean_text(value)
            if title and text:
                cleaned.append((title[:80], text))
            if len(cleaned) >= limit:
                break
        return cleaned

    def clean_roadmap(values, limit=8):
        if not isinstance(values, list):
            return []
        rows = []
        for value in values:
            if not isinstance(value, dict):
                continue
            action = clean_text(value.get("action") or value.get("recommendation"))
            rationale = clean_text(value.get("rationale") or value.get("why") or value.get("impact"))
            if not action:
                continue
            rows.append({
                "phase": clean_text(value.get("phase") or "0-90 дней"),
                "priority": clean_text(value.get("priority") or "P2"),
                "domain": clean_text(value.get("domain") or "ИТ/ИБ"),
                "action": action,
                "rationale": rationale or "Мера снижает выявленный риск и повышает управляемость.",
                "owner": clean_text(value.get("owner") or "ИТ/ИБ"),
                "effort": clean_text(value.get("effort") or "Средняя"),
                "result": clean_text(
                    value.get("result") or value.get("outcome") or value.get("success_metric")
                ),
            })
            if len(rows) >= limit:
                break
        return rows

    return {
        "executive_summary": clean_list(payload.get("executive_summary"), limit=7),
        "audit_observations": clean_observations(payload.get("audit_observations"), limit=6),
        "management_decisions": clean_list(payload.get("management_decisions"), limit=6),
        "roadmap": clean_roadmap(payload.get("roadmap"), limit=8),
    }


def augment_ai_risks_from_narrative(items, narrative, results):
    """Recover AI-authored findings that the provider placed outside the risks array."""
    augmented = [dict(item) for item in items if isinstance(item, dict)]
    existing_keys = {risk_semantic_key(item) for item in augmented}
    allowed_keys = {
        "mfa", "iam", "pam", "nac", "dlp", "siem_soc", "patch",
        "endpoint_detection", "backup", "web_waf", "mail", "legacy_os",
        "it_monitoring", "change_management", "network_performance", "appsec", "dr",
    }
    profiles = {
        "mfa": ("ИБ", "Покрытие MFA требует расширения", "Компрометация пароля может открыть доступ к критичным ресурсам."),
        "iam": ("ИБ", "Жизненный цикл учетных записей не автоматизирован", "Несвоевременное создание, изменение или отзыв прав повышает риск избыточного доступа."),
        "pam": ("ИБ", "Привилегированные доступы требуют отдельного контроля", "Неконтролируемая административная сессия может затронуть несколько критичных систем."),
        "nac": ("ИБ", "Допуск устройств к сети не контролируется автоматически", "Неизвестное устройство может получить сетевой доступ до ручного обнаружения."),
        "dlp": ("ИБ", "Каналы передачи чувствительных данных требуют контроля", "Неконтролируемая передача данных повышает риск утечки и регуляторных последствий."),
        "siem_soc": ("ИБ", "Мониторинг событий и реагирование требуют развития", "Неполное покрытие источников увеличивает время обнаружения и расследования инцидентов."),
        "patch": ("ИТ", "Уязвимости и обновления требуют управляемого цикла", "Критичные уязвимости могут оставаться открытыми дольше согласованного срока."),
        "endpoint_detection": ("ИБ", "Защита конечных точек требует измеримого контроля", "Неполная телеметрия затрудняет обнаружение и расследование сложных атак."),
        "backup": ("ИТ", "Восстановление из резервных копий требует подтверждения", "Без тестов нельзя подтвердить достижение согласованных RTO и RPO."),
        "web_waf": ("ИБ", "Публичные приложения требуют прикладной защиты", "Атаки на веб-приложения могут затронуть данные и доступность цифровых сервисов."),
        "mail": ("ИБ", "Почтовый контур требует усиления защиты", "Фишинг остается вероятным каналом компрометации учетных записей."),
        "legacy_os": ("ИТ", "Операционные системы без поддержки требуют плана миграции", "Отсутствие обновлений повышает риск эксплуатации известных уязвимостей."),
        "it_monitoring": ("ИТ", "Мониторинг ИТ-сервисов требует централизации", "Позднее обнаружение деградации увеличивает продолжительность простоя."),
        "change_management": ("ИТ", "Изменения и конфигурации требуют формального процесса", "Несогласованные изменения повышают риск ошибок и простоев."),
        "network_performance": ("ИТ", "Сетевая архитектура требует подтверждения измерениями", "Без замеров нельзя обоснованно оценить емкость и отказоустойчивость каналов."),
        "appsec": ("ИБ", "Проверки безопасности необходимо встроить в релизы", "Уязвимости приложений и зависимостей могут попадать в продуктивную среду."),
        "dr": ("ИТ", "Аварийное восстановление требует формализованного сценария", "Неопределенные RTO/RPO повышают риск неприемлемого простоя критичных сервисов."),
    }
    control_by_key = {
        "mfa": "MFA", "iam": "IAM", "pam": "PAM", "nac": "NAC", "dlp": "DLP",
        "siem_soc": "SIEM/SOAR", "patch": "Patch Management", "endpoint_detection": "EDR/XDR/MDR",
        "backup": "Резервное копирование", "web_waf": "WAF", "mail": "Mail Security",
    }

    observations = narrative.get("audit_observations", []) if isinstance(narrative, dict) else []
    roadmap = narrative.get("roadmap", []) if isinstance(narrative, dict) else []
    candidates = {}
    for row in roadmap:
        if not isinstance(row, dict):
            continue
        action = str(row.get("action") or row.get("recommendation") or "").strip()
        if not action:
            continue
        key = risk_semantic_key({"risk": action, "recommendation": action})
        if key not in allowed_keys:
            continue
        candidates.setdefault(key, []).append(row)

    for key, rows in candidates.items():
        if key in existing_keys:
            continue
        area, title, impact = profiles[key]
        actions = []
        for row in rows[:3]:
            action = str(row.get("action") or "").strip()
            rationale = str(row.get("rationale") or "").strip()
            combined = ". ".join(part.rstrip(". ") for part in (action, rationale) if part)
            if combined:
                actions.append(combined + ".")
        recommendation = " ".join(actions)
        matching_observation = next(
            (
                str(row.get("text") or "").strip()
                for row in observations
                if isinstance(row, dict)
                and risk_semantic_key({"risk": row.get("title", ""), "description": row.get("text", "")}) == key
                and str(row.get("text") or "").strip()
            ),
            "",
        )
        control = control_by_key.get(key, title)
        evidence = [f"{control} в анкете: {results.get(control, 'не подтверждено')}"]
        success_metric = next(
            (str(row.get("result") or row.get("success_metric") or "").strip() for row in rows if str(row.get("result") or row.get("success_metric") or "").strip()),
            "Владелец, срок и измеримый критерий результата утверждены",
        )
        augmented.append({
            "area": area,
            "_ai_area": area,
            "level": "HIGH" if key in {"pam", "nac", "dlp", "mfa"} else "MEDIUM",
            "risk": title,
            "description": matching_observation or title,
            "impact": impact,
            "recommendation": recommendation,
            "evidence": evidence,
            "success_metric": success_metric,
            "vendors": [control],
            "legal_ids": [],
            "frameworks": [],
        })
        existing_keys.add(key)
    return augmented


def is_truncated_ai_text(value):
    text = str(value or "").strip()
    if not text:
        return True
    if len(text) < 20:
        return True
    unfinished_endings = (
        " и", " в", " с", " на", " для", " по", " от", " до", " при",
        "1.", "2.", "3.", " 1", " 2", " 3"
    )
    return text.endswith(unfinished_endings)


def canonical_ai_risk_title(semantic_key):
    titles = {
        "mfa": "Критичные доступы требуют полного покрытия MFA",
        "iam": "Жизненный цикл учетных записей не автоматизирован",
        "pam": "Привилегированные доступы требуют отдельного контроля",
        "nac": "Допуск устройств к сети не контролируется автоматически",
        "dlp": "Каналы передачи чувствительных данных требуют контроля",
        "siem_soc": "Мониторинг событий и реагирование требуют развития",
        "patch": "Уязвимости и обновления требуют управляемого цикла",
        "endpoint_detection": "Защита конечных точек требует измеримого контроля",
        "backup": "Восстановление из резервных копий требует подтверждения",
        "web_waf": "Публичные приложения требуют прикладной защиты",
        "mail": "Почтовый контур требует усиления защиты",
        "legacy_os": "Операционные системы без поддержки требуют плана миграции",
        "it_monitoring": "Мониторинг ИТ-сервисов требует централизации",
        "change_management": "Изменения и конфигурации требуют формального процесса",
        "network_performance": "Сетевая архитектура требует подтверждения измерениями",
        "appsec": "Проверки безопасности необходимо встроить в релизы",
        "dr": "Аварийное восстановление требует формализованного сценария",
    }
    return titles.get(str(semantic_key or ""), "")


def prepare_ai_risks_for_report(items, min_items=1):
    if not isinstance(items, list):
        return []

    prepared = []
    seen = set()
    for item in items:
        if not isinstance(item, dict):
            continue
        normalized_item = dict(item)
        semantic_key = risk_semantic_key(normalized_item)
        if is_truncated_ai_text(normalized_item.get("risk")):
            canonical_title = canonical_ai_risk_title(semantic_key)
            if not canonical_title:
                continue
            normalized_item["risk"] = canonical_title
        if is_truncated_ai_text(normalized_item.get("recommendation")):
            continue

        if not semantic_key or semantic_key in seen:
            continue
        prepared.append(normalized_item)
        seen.add(semantic_key)

    return prepared if len(prepared) >= min_items else []


def explain_ai_risk_rejections(items):
    reasons = {"not_object": 0, "short_risk": 0, "short_recommendation": 0, "duplicate": 0}
    seen = set()
    for item in items if isinstance(items, list) else []:
        if not isinstance(item, dict):
            reasons["not_object"] += 1
            continue
        semantic_key = risk_semantic_key(item)
        if is_truncated_ai_text(item.get("risk")) and not canonical_ai_risk_title(semantic_key):
            reasons["short_risk"] += 1
            continue
        if is_truncated_ai_text(item.get("recommendation")):
            reasons["short_recommendation"] += 1
            continue
        if not semantic_key or semantic_key in seen:
            reasons["duplicate"] += 1
            continue
        seen.add(semantic_key)
    return ", ".join(f"{key}={value}" for key, value in reasons.items() if value) or "нет"


def recover_complete_risk_objects(response_text):
    """Recover valid top-level risk objects from a partially malformed AI response."""
    text = str(response_text or "")
    risks_match = re.search(r'"risks"\s*:\s*\[', text, flags=re.IGNORECASE)
    if risks_match:
        array_start = text.find("[", risks_match.start())
    else:
        array_start = text.find("[")
    if array_start < 0:
        return None

    recovered = []
    object_start = None
    object_depth = 0
    in_string = False
    escaped = False
    for position in range(array_start + 1, len(text)):
        char = text[position]
        if in_string:
            if escaped:
                escaped = False
            elif char == "\\":
                escaped = True
            elif char == '"':
                in_string = False
            continue
        if char == '"':
            in_string = True
            continue
        if char == "{":
            if object_depth == 0:
                object_start = position
            object_depth += 1
            continue
        if char != "}" or object_depth == 0:
            continue
        object_depth -= 1
        if object_depth != 0 or object_start is None:
            continue
        candidate = text[object_start:position + 1]
        try:
            item = json.loads(candidate)
        except json.JSONDecodeError:
            object_start = None
            continue
        if isinstance(item, dict):
            recovered.append(item)
        object_start = None

    return {"risks": recovered} if recovered else None


def ai_quality_gate(items, min_items=6, min_security_items=3, min_it_items=3):
    prepared = prepare_ai_risks_for_report(items, min_items=1)
    if len(prepared) < min_items:
        return [], f"ИИ дал только {len(prepared)} пригодных пунктов из минимально ожидаемых {min_items}."

    security_markers = (
        "mfa", "edr", "xdr", "mdr", "epp", "siem", "soc", "pam", "dlp",
        "waf", "ids", "ips", "ztna", "mail", "почт", "уязв", "patch",
        "учет", "доступ", "endpoint", "мониторинг событий", "реагирован"
    )
    it_markers = (
        "backup", "резерв", "dr", "rto", "rpo", "сервер", "сеть", "схд",
        "виртуал", "мониторинг", "обновлен", "ос", "инфраструкт",
        "эксплуатац", "емкост", "производительност", "бизнес-систем"
    )

    security_count = 0
    it_count = 0
    weak_text_count = 0
    for item in prepared:
        item_area = str(item.get("area", item.get("_ai_area", ""))).strip().upper()
        combined = " ".join(
            str(item.get(field, ""))
            for field in ("risk", "description", "impact", "recommendation")
        ).lower()
        if item_area == "ИБ" or any(marker in combined for marker in security_markers):
            security_count += 1
        if item_area == "ИТ" or any(marker in combined for marker in it_markers):
            it_count += 1
        if len(str(item.get("recommendation", "")).strip()) < 90:
            weak_text_count += 1

    if weak_text_count > max(2, len(prepared) // 3):
        return [], "ИИ дал слишком короткие рекомендации; включены экспертные правила."
    if security_count < min_security_items:
        return [], "ИИ почти не покрыл ИБ-домены; включены экспертные правила."
    if it_count < min_it_items:
        return [], "ИИ почти не покрыл ИТ-инфраструктуру; включены экспертные правила."

    return prepared, ""


IT_GAP_LABELS = {
    "wifi_capacity": "емкость, покрытие и централизованное управление Wi-Fi",
    "network_performance": "каналы связи и отказоустойчивость сети",
    "virtualization": "ресурсный запас и capacity planning виртуализации",
    "storage": "емкость и производительность СХД",
    "it_monitoring": "единый мониторинг ИТ-сервисов и инфраструктуры",
    "itam": "CMDB, инвентаризация и жизненный цикл ИТ-активов",
    "change_management": "управление изменениями и планами отката",
    "dr": "RTO/RPO и регулярные тесты восстановления",
}


def confirmed_it_gap_topics(results):
    """Return only IT gaps explicitly supported by structured answers or notes."""
    def result_int(value):
        match = re.search(r"-?\d+(?:[.,]\d+)?", str(value or ""))
        if match:
            try:
                return int(float(match.group(0).replace(",", ".")))
            except (TypeError, ValueError):
                pass
        try:
            return int(float(value or 0))
        except (TypeError, ValueError):
            return 0

    gaps = {}
    notes = " ".join(
        str(value or "")
        for key, value in results.items()
        if "примечан" in str(key).lower()
    ).lower()

    users = result_int(results.get("_user_count"))
    access_points = result_int(results.get("WiFi Точки"))
    main_speed = result_int(results.get("_main_speed"))
    backup_speed = result_int(results.get("_back_speed"))
    wifi_controller = str(results.get("WiFi Контроллер", "Нет")).strip().lower()
    overloaded_wifi = users > 0 and access_points > 0 and users / access_points > 30
    weak_backup = main_speed > 0 and (backup_speed <= 0 or backup_speed / main_speed < 0.25)
    wifi_without_controller = (
        access_points >= 4
        and wifi_controller in {"", "нет", "no", "none"}
    )
    if overloaded_wifi or wifi_without_controller or any(
        marker in notes
        for marker in ("wi-fi перегруж", "wifi перегруж", "нестабильный роуминг", "радиообследован")
    ):
        wifi_facts = []
        if users and access_points:
            wifi_facts.append(f"{access_points} точек на {users} пользователей")
        if wifi_without_controller:
            wifi_facts.append("централизованный контроллер не указан")
        fact_suffix = f" ({'; '.join(wifi_facts)})" if wifi_facts else ""
        gaps["wifi_capacity"] = IT_GAP_LABELS["wifi_capacity"] + fact_suffix

    if weak_backup or any(
        marker in notes
        for marker in (
            "резервный канал слаб", "резервный канал недостаточ", "резервный канал перегруж",
            "failover не", "переключение тестируется нерегулярно", "отказоустойчивость канал не",
        )
    ):
        gaps["network_performance"] = IT_GAP_LABELS["network_performance"]

    if any(marker in notes for marker in (
        "capacity planning не", "запас вычислительной мощности недостаточ",
        "запас мощности недостаточ", "загрузка памяти достигает",
    )):
        gaps["virtualization"] = IT_GAP_LABELS["virtualization"]

    storage_fill = re.search(r"(?:схд|хранилищ\w*)[^.]{0,60}(\d{2,3})\s*%", notes)
    if (
        storage_fill and int(storage_fill.group(1)) >= 80
    ) or any(marker in notes for marker in (
        "прогноз исчерпания не", "порог расширения", "latency",
    )):
        gaps["storage"] = IT_GAP_LABELS["storage"]

    if any(marker in notes for marker in (
        "без единой панели", "не объединен в единый контур",
        "разрозненный мониторинг", "события мониторинга и изменения не связаны",
    )):
        gaps["it_monitoring"] = IT_GAP_LABELS["it_monitoring"]

    if any(marker in notes for marker in (
        "cmdb отсутств", "учет жизненного цикла оборудования ведется в нескольких",
        "единый реестр активов отсутств",
    )):
        gaps["itam"] = IT_GAP_LABELS["itam"]

    if any(marker in notes for marker in (
        "изменения согласуются в чат", "календарь изменений",
        "план отката используются не", "change management не",
    )):
        gaps["change_management"] = IT_GAP_LABELS["change_management"]

    continuity_gap = (
        ("rto" in notes or "rpo" in notes)
        and any(marker in notes for marker in ("не согласован", "не определен", "не утвержден"))
    ) or ("восстанов" in notes and "нерегуляр" in notes)
    if continuity_gap:
        gaps["dr"] = IT_GAP_LABELS["dr"]

    return gaps


def build_confirmed_it_gap_risks(results, context):
    """Convert questionnaire-confirmed IT gaps into fact-safe fallback findings."""
    gaps = confirmed_it_gap_topics(results)
    if not gaps:
        return []

    def number(value):
        match = re.search(r"-?\d+(?:[.,]\d+)?", str(value or ""))
        return float(match.group(0).replace(",", ".")) if match else 0.0

    users = int(number(results.get("_user_count")))
    access_points = int(number(results.get("WiFi Точки")))
    main_speed = number(results.get("_main_speed") or results.get("Интернет канал (осн)"))
    backup_speed = number(results.get("_back_speed") or results.get("Резервный канал"))
    templates = {
        "wifi_capacity": {
            "level": "HIGH" if users and access_points and users / access_points > 45 else "MEDIUM",
            "risk": "Емкость и централизованное управление Wi-Fi требуют усиления",
            "description": (
                f"В анкете указано {access_points} точек доступа для контура из {users} рабочих мест; "
                f"Wi-Fi контроллер: {results.get('WiFi Контроллер', 'Нет')}."
            ),
            "impact": "Перегрузка радиоэфира и отсутствие единого управления могут снижать качество связи, роуминга и видеоконференций.",
            "recommendation": "Провести радиообследование, определить требуемую плотность точек доступа и выполнить пилот централизованного WLAN-управления.",
            "success_metric": "Покрытие и емкость подтверждены радиообследованием; пиковая загрузка точек остается в целевых пределах",
            "vendors": ["Wi-Fi", "WLAN"],
        },
        "network_performance": {
            "level": "HIGH" if main_speed and (backup_speed <= 0 or backup_speed / main_speed < 0.1) else "MEDIUM",
            "risk": "Резервный канал не обеспечивает подтвержденную отказоустойчивость",
            "description": (
                f"Основной канал: {int(main_speed)} Mbit/s; резервный канал: "
                f"{int(backup_speed)} Mbit/s. Автоматическое переключение и независимость трасс требуют подтверждения."
            ),
            "impact": "При отказе основного канала критичные облачные сервисы, удаленная работа и коммуникации могут деградировать или стать недоступными.",
            "recommendation": "Определить требуемую резервную полосу и SLA, проверить независимость операторов и трасс, затем провести тест автоматического failover под рабочей нагрузкой.",
            "success_metric": "Резервный канал выдерживает согласованную критичную нагрузку; failover проходит в пределах SLA",
            "vendors": ["Network Equipment", "SD-WAN"],
        },
        "virtualization": {
            "level": "HIGH",
            "risk": "Виртуальной среде требуется подтвержденный запас ресурсов",
            "description": "Примечания анкеты указывают на недостаточный ресурсный запас или отсутствие формализованного capacity planning виртуальной среды.",
            "impact": "Недостаточный запас повышает вероятность деградации или простоя критичных сервисов при росте нагрузки и отказе хоста.",
            "recommendation": "Провести capacity-анализ, проверить сценарий отказа хоста и подготовить план расширения вычислительных ресурсов.",
            "success_metric": "Запас ресурсов подтвержден для отказа одного хоста и прогнозируемого роста нагрузки",
            "vendors": ["Virtualization"],
        },
        "storage": {
            "level": "HIGH",
            "risk": "Емкость и производительность СХД требуют управляемого плана развития",
            "description": "Анкета подтверждает высокий уровень заполнения, отсутствие утвержденного порога расширения или недостаточную глубину статистики производительности СХД.",
            "impact": "Исчерпание емкости или деградация производительности могут остановить зависимые бизнес-системы и усложнить восстановление.",
            "recommendation": "Провести health-check СХД, утвердить пороги расширения и план развития емкости, производительности и репликации.",
            "success_metric": "Запас емкости и производительности контролируется по утвержденным порогам и прогнозу роста",
            "vendors": ["Storage"],
        },
        "it_monitoring": {
            "level": "HIGH",
            "risk": "ИТ-мониторинг не объединен в единый эксплуатационный контур",
            "description": "Примечания анкеты подтверждают разрозненный контроль серверов, виртуализации, СХД или каналов без единой панели и SLA реакции.",
            "impact": "Команда позднее обнаруживает деградацию сервисов и не имеет единой картины доступности, производительности и емкости.",
            "recommendation": "Определить критичные сервисы и метрики, провести пилот единого ИТ-мониторинга и связать оповещения с ответственными и SLA реакции.",
            "success_metric": "Критичные сервисы имеют единые метрики, пороги, владельцев реакции и отчетность по SLA",
            "vendors": ["IT Monitoring", "NMS"],
        },
        "itam": {
            "level": "HIGH",
            "risk": "Учет ИТ-активов и сервисов не централизован",
            "description": "Примечания анкеты подтверждают отсутствие единой CMDB, каталога услуг или управляемого жизненного цикла ИТ-активов.",
            "impact": "Неполные данные об активах и зависимостях замедляют устранение инцидентов, изменения и планирование бюджета.",
            "recommendation": "Сформировать модель данных CMDB, назначить владельцев и связать активы с сервисами, изменениями и SLA.",
            "success_metric": "Не менее 95% критичных активов и сервисов имеют владельца, зависимости и актуальный статус",
            "vendors": ["ITAM", "ITSM", "CMDB"],
        },
        "change_management": {
            "level": "HIGH",
            "risk": "Управление изменениями не формализовано",
            "description": "Примечания анкеты подтверждают согласование изменений в чатах или неполное применение календаря изменений и планов отката.",
            "impact": "Непроверенные изменения повышают вероятность простоев и усложняют восстановление и расследование инцидентов.",
            "recommendation": "Ввести единый процесс регистрации, оценки риска, согласования, тестирования и отката продуктивных изменений.",
            "success_metric": "Все продуктивные изменения имеют владельца, согласование, тест и план отката",
            "vendors": ["ITSM", "Change Management", "CMDB"],
        },
        "dr": {
            "level": "HIGH",
            "risk": "RTO/RPO и регулярное тестирование восстановления не формализованы",
            "description": "Примечания анкеты подтверждают отсутствие согласованных RTO/RPO или нерегулярное полное тестовое восстановление.",
            "impact": "Наличие резервных копий не гарантирует восстановление критичных сервисов в приемлемые для бизнеса сроки.",
            "recommendation": "Согласовать RTO/RPO, провести контрольное восстановление и утвердить регулярные DR-учения с фиксацией результатов.",
            "success_metric": "Критичные сервисы проходят регулярный тест восстановления в пределах утвержденных RTO/RPO",
            "vendors": ["DR", "Backup"],
        },
    }

    findings = []
    for key in gaps:
        template = templates.get(key)
        if not template:
            continue
        findings.append({
            **template,
            "semantic_key": key,
            "_semantic_key": key,
            "_ai_area": "ИТ",
            "_source": "Базовые правила",
            "evidence": [templates[key]["description"]],
        })
    return findings


def ai_it_gap_coverage(items, expected_gaps):
    covered = set()
    coverage_markers = {
        "wifi_capacity": (
            "wi-fi", "wifi", "беспровод", "роуминг", "радиообслед", "точк доступа",
            "wlan", "контроллер", "ssid",
        ),
        "network_performance": (
            "резервн", "канал", "пропускн", "failover", "отказоустойчив",
            "маршрутиз", "wan", "sla",
        ),
        "virtualization": ("виртуал", "гипервиз", "vmware", "хост", "cpu", "ram"),
        "storage": ("схд", "storage", "raid", "iops", "latency", "емкост"),
        "it_monitoring": ("мониторинг", "наблюдаемост", "метрик", "nms", "доступност"),
        "itam": ("cmdb", "itam", "актив", "инвентаризац", "жизненн"),
        "change_management": ("изменен", "change", "откат", "согласован"),
        "dr": ("rto", "rpo", "восстанов", "dr", "аварийн"),
    }
    for item in items if isinstance(items, list) else []:
        if not isinstance(item, dict):
            continue
        covered.add(risk_semantic_key(item))
        text = " ".join(
            str(item.get(field, ""))
            for field in ("risk", "description", "impact", "recommendation", "success_metric")
        ).lower()
        for key, markers in coverage_markers.items():
            if key in expected_gaps and any(marker in text for marker in markers):
                covered.add(key)
    if "backup" in covered:
        covered.add("dr")
    matched = [key for key in expected_gaps if key in covered]
    missing = [key for key in expected_gaps if key not in covered]
    return matched, missing


def security_control_snapshot(results):
    controls = [
        "MFA",
        "EPP",
        "EDR",
        "XDR",
        "MDR",
        "DLP",
        "Mail Security",
        "WAF",
        "Anti-DDoS",
        "IDS/IPS",
        "NAC",
        "ZTNA",
        "IAM",
        "PAM",
        "SIEM",
        "SOAR",
        "NAD",
        "Patch Management",
        "SAST",
        "DAST",
    ]
    enabled = []
    missing = []
    for control in controls:
        value = results.get(control)
        if is_enabled(value):
            enabled.append(f"{control}: {value}")
        else:
            missing.append(control)
    return enabled, missing


def control_confirmed_in_results(results, control):
    if is_enabled(results.get(control)):
        return True
    note_values = [
        str(value)
        for key, value in results.items()
        if any(marker in str(key).lower() for marker in ("примеч", "комментар", "дополн"))
    ]
    notes = " ".join(note_values).lower()
    aliases = {
        "WAF": ("waf", "web application firewall", "fortiweb", "cloudflare waf", "imperva waf", "f5 asm"),
        "EDR": ("edr", "endpoint detection and response"),
        "MFA": ("mfa", "многофактор"),
    }.get(control, (control.lower(),))
    positive = ("есть", "включ", "использ", "внедр", "настро", "работает", "защищ")
    for alias in aliases:
        for match in re.finditer(re.escape(alias), notes):
            context = notes[max(0, match.start() - 45):match.end() + 45]
            if any(marker in context for marker in positive) and not any(
                marker in context for marker in ("нет ", "отсутств", "не внедр", "не настро")
            ):
                return True
    return False


def risk_conflicts_with_answers(item, results):
    key = risk_semantic_key(item)

    if key == "legacy_os":
        legacy_arm = int(results.get("ОС АРМ (Windows XP/Vista/7/8)", 0) or 0)
        legacy_servers = int(results.get("ОС Сервера (Windows Server 2008/2012 R2)", 0) or 0)
        if legacy_arm + legacy_servers == 0:
            return "No legacy operating systems reported"

    if key == "mfa" and is_enabled(results.get("MFA")):
        return "MFA already enabled"
    if key == "iam" and (is_enabled(results.get("IAM")) or is_enabled(results.get("IDM"))):
        return "IAM already enabled"
    if key == "pam" and is_enabled(results.get("PAM")):
        return "PAM already enabled"
    if key == "siem_soc" and is_enabled(results.get("SIEM")):
        return "SIEM already enabled"
    if key == "patch" and is_enabled(results.get("Patch Management")):
        return "Patch Management already enabled"
    if key == "web_waf" and control_confirmed_in_results(results, "WAF"):
        return "WAF already enabled"
    if key == "mail" and is_enabled(results.get("Mail Security")):
        return "Mail Security already enabled"
    if key == "dlp" and is_enabled(results.get("DLP")):
        return "DLP already enabled"
    if key == "backup" and is_enabled(results.get("Резервное копирование")):
        combined = " ".join(
            str(item.get(field, ""))
            for field in ("risk", "description", "recommendation")
        ).lower()
        absence_markers = (
            "отсутствие резервного копирования",
            "отсутствует резервное копирование",
            "резервные копии отсутств",
            "резервное копирование отсутств",
            "резервное копирование не внедр",
            "нет резервного копирования",
            "backup отсутств",
            "backup не внедр",
        )
        if any(marker in combined for marker in absence_markers):
            return "Backup already enabled"
    if key == "endpoint_detection" and any(
        control_confirmed_in_results(results, control)
        for control in ("EDR", "XDR", "MDR")
    ):
        return "Endpoint detection/response already enabled"

    return ""


def filter_ai_risks_by_answers(items, results):
    filtered = []
    skipped = []
    for item in items:
        conflict = risk_conflicts_with_answers(item, results)
        if conflict:
            skipped.append(f"{risk_semantic_key(item)}: {conflict}")
            continue
        filtered.append(item)
    return filtered, skipped


def normalize_site_domain(site):
    value = str(site or "").strip().lower()
    value = re.sub(r"^https?://", "", value)
    value = re.sub(r"^www\.", "", value)
    value = value.split("/")[0].split("?")[0].split("#")[0].strip()
    return value


def is_valid_domain(domain):
    if not domain or " " in domain or "." not in domain:
        return False

    return bool(re.match(r"^[a-z0-9][a-z0-9.-]*\.[a-z]{2,}$", domain))


def normalize_email(email):
    return str(email or "").strip().lower()


def is_valid_email(email):
    value = normalize_email(email)
    return bool(re.match(r"^[^@\s]+@[^@\s]+\.[^@\s]{2,}$", value))


def normalize_phone_number(phone):
    digits = re.sub(r"\D", "", str(phone or ""))
    if len(digits) == 10:
        return f"{digits[:3]} {digits[3:6]} {digits[6:8]} {digits[8:]}"
    if len(digits) == 9:
        return f"{digits[:3]} {digits[3:6]} {digits[6:]}"
    return str(phone or "").strip()


def widget_int(key):
    try:
        return int(st.session_state.get(key, 0) or 0)
    except (TypeError, ValueError):
        return 0


INDUSTRY_OPTIONS = [
    "Финтех / Банки",
    "Страхование",
    "Ритейл / E-commerce",
    "Здравоохранение / Медицинская организация",
    "Госсектор",
    "Квазигосударственный сектор",
    "КВОИКИ / Критическая инфраструктура",
    "Телеком / Связь",
    "Энергетика / Коммунальная инфраструктура",
    "Транспорт / Логистика",
    "Производство / АСУ ТП",
    "IT / Разработка",
    "Образование",
    "Услуги / Корпоративный сектор",
    "Другое",
]

COUNTRY_CODE_OPTIONS = [
    ("🇰🇿 +7", "+7"),
    ("🇷🇺 +7", "+7"),
    ("🇺🇿 +998", "+998"),
    ("🇰🇬 +996", "+996"),
    ("🇹🇯 +992", "+992"),
    ("🇦🇪 +971", "+971"),
    ("🇹🇷 +90", "+90"),
    ("🇦🇿 +994", "+994"),
]

NETWORK_TYPE_OPTIONS = [
    "Оптика", "RJ45 (Ethernet)", "Радиорелейная", "Спутник",
    "4G/5G", "Starlink", "ADSL/VDSL", "Нет",
]

WIFI_TYPE_OPTIONS = [
    "Wi-Fi 6/6E (802.11ax)",
    "Wi-Fi 5 (802.11ac)",
    "Wi-Fi 4 (802.11n)",
    "Другое",
]

MAIL_SYSTEM_OPTIONS = [
    "Exchange (On-Prem)", "Lotus", "Microsoft 365",
    "Google Workspace", "Собственный", "Нет",
]

WEB_HOSTING_OPTIONS = ["Собственный ЦОД", "Облако KZ", "Облако Global"]

DRAFT_SELECTBOX_OPTIONS = {
    "client_industry_select": [""] + INDUSTRY_OPTIONS,
    "client_phone_code": COUNTRY_CODE_OPTIONS,
    "main_net_type": NETWORK_TYPE_OPTIONS,
    "back_net_type": NETWORK_TYPE_OPTIONS,
    "wf_type_sel": WIFI_TYPE_OPTIONS,
    "mail_system": MAIL_SYSTEM_OPTIONS,
    "web_hosting": WEB_HOSTING_OPTIONS,
}


REGULATORY_CATALOG = {
    "PD_LAW": {
        "title": "Закон РК «О персональных данных и их защите» № 94-V",
        "short": "Закон РК о персональных данных № 94-V",
        "url": "https://adilet.zan.kz/rus/docs/Z1300000094",
        "scope": "Сбор, обработка, хранение и защита персональных данных.",
        "status": "Обязательное требование",
    },
    "PD_RULES": {
        "title": "Правила осуществления мер по защите персональных данных, № 32810",
        "short": "Правила защиты персональных данных № 32810",
        "url": "https://adilet.zan.kz/rus/docs/V2300032810",
        "scope": "Организационные и технические меры собственника, оператора и третьего лица.",
        "status": "Обязательное требование",
    },
    "INFORMATIZATION": {
        "title": "Закон РК «Об информатизации» № 418-V",
        "short": "Закон РК «Об информатизации» № 418-V",
        "url": "https://adilet.zan.kz/rus/docs/Z1500000418",
        "scope": "Защита объектов информатизации и специальные обязанности владельцев КВОИКИ.",
        "status": "Обязательное при применимости",
    },
    "UNIFIED_832": {
        "title": "Единые требования в области ИКТ и ИБ, постановление Правительства РК № 832",
        "short": "Единые требования ИКТ и ИБ № 832",
        "url": "https://adilet.zan.kz/rus/docs/P1600000832",
        "scope": "Государственный и квазигосударственный сектор, государственные интеграции и КВОИКИ.",
        "status": "Обязательное при применимости",
    },
    "KVOIKI_529": {
        "title": "Правила и критерии отнесения к КВОИКИ, постановление Правительства РК № 529",
        "short": "Критерии КВОИКИ № 529",
        "url": "https://adilet.zan.kz/rus/docs/P1600000529",
        "scope": "Определение и подтверждение статуса критически важного объекта ИК-инфраструктуры.",
        "status": "Обязательное для КВОИКИ",
    },
    "KVOIKI_MONITORING": {
        "title": "Правила мониторинга обеспечения ИБ объектов электронного правительства и КВОИКИ, № 17019",
        "short": "Правила мониторинга ИБ КВОИКИ № 17019",
        "url": "https://adilet.zan.kz/rus/docs/V1800017019",
        "scope": "Мониторинг событий, ответственный по ИБ, реагирование на инциденты и устранение уязвимостей для КВОИКИ.",
        "status": "Обязательное для КВОИКИ",
    },
    "FINANCE_IS": {
        "title": "Минимальные требования по обеспечению ИБ на финансовом рынке, № 38505",
        "short": "Требования ИБ финансового рынка № 38505",
        "url": "https://adilet.zan.kz/rus/docs/V2600038505",
        "scope": "Финансовые организации и регулируемые участники финансового рынка.",
        "status": "Отраслевое обязательное требование",
    },
    "BANK_IS": {
        "title": "Требования к ИБ банков и организаций, осуществляющих отдельные банковские операции, № 16772",
        "short": "Требования ИБ банков № 16772",
        "url": "https://adilet.zan.kz/rus/docs/V1800016772",
        "scope": "Банки, филиалы банков-нерезидентов и отдельные банковские операции.",
        "status": "Отраслевое обязательное требование",
    },
    "MEDICAL_DATA": {
        "title": "Правила сбора, обработки, хранения и защиты персональных медицинских данных, № 22550",
        "short": "Правила защиты медицинских данных № 22550",
        "url": "https://adilet.zan.kz/rus/docs/V2100022550",
        "scope": "Субъекты цифрового здравоохранения и персональные медицинские данные.",
        "status": "Отраслевое обязательное требование",
    },
}


def expand_regulatory_references(value):
    """Replace internal legal IDs in AI prose with customer-facing titles."""
    text = str(value or "")

    def replace_penalty(match):
        item = REGULATORY_CATALOG.get(match.group(1))
        if not item:
            return match.group(0)
        return f"Регуляторные последствия при нарушении требований: {item['short']}"

    text = re.sub(r"Штрафы\s+по\s*\[([A-Z0-9_]+)\]", replace_penalty, text, flags=re.IGNORECASE)

    def replace_requirements(match):
        item = REGULATORY_CATALOG.get(match.group(1))
        return f"требованиям документа «{item['short']}»" if item else match.group(0)

    text = re.sub(
        r"требованиям(?:\s+регулятор(?:а|ов))?\s*\[([A-Z0-9_]+)\]",
        replace_requirements,
        text,
        flags=re.IGNORECASE,
    )

    def replace_token(match):
        item = REGULATORY_CATALOG.get(match.group(1))
        return item["short"] if item else match.group(0)

    text = re.sub(r"\[([A-Z0-9_]+)\]", replace_token, text)

    def replace_requirement_group(match):
        prefix = match.group(1)
        identifiers = re.findall(r"[A-Z][A-Z0-9_]+", match.group(2))
        titles = [
            REGULATORY_CATALOG[identifier.upper()]["short"]
            for identifier in identifiers
            if identifier.upper() in REGULATORY_CATALOG
        ]
        return f"{prefix}: {'; '.join(titles)}" if titles else match.group(0)

    identifier_pattern = "|".join(
        re.escape(identifier)
        for identifier in sorted(REGULATORY_CATALOG, key=len, reverse=True)
    )
    if identifier_pattern:
        text = re.sub(
            rf"((?:нарушение|несоответствие)\s+требовани(?:й|ям))\s+"
            rf"((?:{identifier_pattern})(?:\s*(?:,|и)\s*(?:{identifier_pattern}))*)",
            replace_requirement_group,
            text,
            flags=re.IGNORECASE,
        )
        for identifier in sorted(REGULATORY_CATALOG, key=len, reverse=True):
            text = re.sub(
                rf"(?<![A-Za-z0-9_]){re.escape(identifier)}(?![A-Za-z0-9_])",
                REGULATORY_CATALOG[identifier]["short"],
                text,
                flags=re.IGNORECASE,
            )
    return text


def sanitize_customer_roadmap_text(value):
    """Keep roadmap vendor-neutral and enforce pilot-before-procurement wording."""
    text = expand_regulatory_references(value).strip()
    lowered = text.lower()

    if "nac" in lowered or "network access control" in lowered:
        if any(marker in lowered for marker in ("требован", "оценить", "обслед", "аудит", "выбрать решение")):
            return (
                "Описать типы подключений и требования к NAC; согласовать пилотный контур, "
                "сценарии 802.1X, профилирования и изоляции устройств."
            )
        if any(marker in lowered for marker in ("внедр", "развер", "пилот", "настро")):
            return (
                "Провести пилот NAC на Wi-Fi и одном проводном сегменте; проверить 802.1X, "
                "профилирование, контроль соответствия и изоляцию устройств."
            )

    if "edr" in lowered or "xdr" in lowered:
        return (
            "Провести пилот EDR/XDR на согласованной группе конечных точек и серверов; "
            "проверить покрытие, телеметрию, сценарии реагирования и передачу событий в SIEM."
        )

    if "soar" in lowered:
        return (
            "После стабилизации SIEM выбрать повторяемые сценарии реагирования и провести пилот "
            "SOAR с контролем времени обработки и доли ручных операций."
        )

    if "dlp" in lowered and any(marker in lowered for marker in ("закупить", "выбрать поставщика")):
        if "пилот" in lowered or "закупить" in lowered:
            return (
                "Провести ограниченный пилот DLP на согласованных каналах; по результатам "
                "подтвердить требования, модель внедрения и решение о масштабировании."
            )
        return "Определить каналы контроля, политики и измеримые критерии пилота DLP."

    if any(marker in lowered for marker in ("закупить", "закупка")) and any(
        marker in lowered for marker in ("пилот", "poc", "proof of concept")
    ):
        if "iam" in lowered:
            return (
                "Сформировать требования и провести PoC IAM на ограниченной группе пользователей; "
                "по результатам подтвердить архитектуру, интеграции и решение о масштабировании."
            )
        return (
            "Сформировать требования и провести ограниченный пилот решения; по измеримым результатам "
            "принять решение о закупке и масштабировании."
        )

    if "pam" in lowered or "привилегирован" in lowered:
        if any(marker in lowered for marker in (
            "рабочую группу", "собрать требования", "выбрать пилот", "аудит", "инвентар",
            "список ролей", "список привилегирован",
        )):
            return (
                "Сформировать рабочую группу, инвентаризировать привилегированные доступы ко всем "
                "критичным системам и выбрать пилотный сценарий PAM."
            )
        if "пилот" in lowered:
            return (
                "Провести пилот PAM на согласованном критичном контуре; проверить vault, контроль "
                "сессий, аварийный доступ и передачу событий в SIEM."
            )
        if any(marker in lowered for marker in ("полноцен", "масштаб", "все сервер", "внедрить pam")):
            return (
                "Расширить PAM на критичные системы, сетевое оборудование, базы данных и "
                "административные консоли; передавать события в SIEM."
            )
        return (
            "Провести пилот PAM на согласованном критичном контуре; проверить vault, контроль "
            "сессий, аварийный доступ и передачу событий в SIEM."
        )

    if "скан" in lowered and "уязв" in lowered:
        return (
            "Провести первичное сканирование уязвимостей согласованного ИТ-контура; "
            "зафиксировать критичные находки, владельцев и сроки устранения."
        )

    text = re.sub(
        r"\s*\((?:e\.?g\.?|например)[^)]*\)",
        "",
        text,
        flags=re.IGNORECASE,
    )
    text = re.sub(r"\s*\([^)]*(?:Cisco|CyberArk|Veeam|Fortinet|Check Point|Huawei|IBM|Splunk|ManageEngine|Broadcom|Forcepoint|OpenVAS|R-?Vision)[^)]*\)", "", text, flags=re.IGNORECASE)
    vendor_names = {
        "Cisco ISE", "CyberArk", "Veeam Backup", "Veeam", "Fortinet", "Check Point",
        "Huawei", "IBM", "Splunk", "ManageEngine", "Broadcom", "Forcepoint",
        "OpenVAS", "R-Vision", "R Vision", "Qualys", "Tenable", "Rapid7",
        "Zabbix", "Prometheus", "Microsoft", "Windows Server", "Duo",
    }
    try:
        vendor_names.update(load_detailed_vendor_names())
    except Exception:
        pass
    for vendor in sorted(vendor_names, key=len, reverse=True):
        text = re.sub(re.escape(vendor), "", text, flags=re.IGNORECASE)
    text = re.sub(r"\bBackup\b", "резервных копий", text, flags=re.IGNORECASE)
    text = re.sub(r"\s+с\s+помощью\s+(?:бесплатного\s+)?инструмента\b", "", text, flags=re.IGNORECASE)
    text = re.sub(
        r"восстановлени([ея])\s+из\s+(?=(?:и|с|для|по)\b)",
        r"восстановлени\1 из резервных копий ",
        text,
        flags=re.IGNORECASE,
    )
    text = re.sub(r"\(\s*\)", "", text)
    text = re.sub(r"\s*\([^)]*без\s+продукт[^)]*\)", "", text, flags=re.IGNORECASE)
    text = re.sub(r"(?<![\d.])\b1X\b", "802.1X", text, flags=re.IGNORECASE)
    text = re.sub(r"\s{2,}", " ", text)
    text = re.sub(r"\s+([,.;:])", r"\1", text)
    return text.strip(" .,-") + ("." if text.strip(" .,-") else "")


INDUSTRY_REGULATORY_IDS = {
    "Финтех / Банки": ["PD_LAW", "PD_RULES", "FINANCE_IS", "BANK_IS"],
    "Страхование": ["PD_LAW", "PD_RULES", "FINANCE_IS"],
    "Ритейл / E-commerce": ["PD_LAW", "PD_RULES"],
    "Здравоохранение / Медицинская организация": ["PD_LAW", "PD_RULES", "MEDICAL_DATA"],
    "Госсектор": ["PD_LAW", "PD_RULES", "INFORMATIZATION", "UNIFIED_832"],
    "Квазигосударственный сектор": ["PD_LAW", "PD_RULES", "INFORMATIZATION", "UNIFIED_832"],
    "КВОИКИ / Критическая инфраструктура": ["PD_LAW", "PD_RULES", "INFORMATIZATION", "UNIFIED_832", "KVOIKI_529", "KVOIKI_MONITORING"],
    "Телеком / Связь": ["PD_LAW", "PD_RULES", "INFORMATIZATION"],
    "Энергетика / Коммунальная инфраструктура": ["PD_LAW", "PD_RULES", "INFORMATIZATION", "KVOIKI_529"],
    "Транспорт / Логистика": ["PD_LAW", "PD_RULES", "INFORMATIZATION", "KVOIKI_529"],
    "Производство / АСУ ТП": ["PD_LAW", "PD_RULES", "INFORMATIZATION", "KVOIKI_529"],
}


INDUSTRY_FRAMEWORKS = {
    "Финтех / Банки": ["ISO/IEC 27001", "PCI DSS (при обработке карточных данных)", "GDPR (при обработке данных субъектов ЕЭЗ)"],
    "Страхование": ["ISO/IEC 27001", "NIST CSF", "GDPR (при обработке данных субъектов ЕЭЗ)"],
    "Ритейл / E-commerce": ["OWASP ASVS", "PCI DSS (при обработке карточных данных)", "GDPR (при обработке данных субъектов ЕЭЗ)", "ISO/IEC 27001"],
    "Здравоохранение / Медицинская организация": ["ISO/IEC 27001", "ISO 27799", "GDPR (при обработке данных субъектов ЕЭЗ)"],
    "Госсектор": ["СТ РК ISO/IEC 27002", "ISO/IEC 27001"],
    "Квазигосударственный сектор": ["СТ РК ISO/IEC 27002", "ISO/IEC 27001"],
    "КВОИКИ / Критическая инфраструктура": ["ISO/IEC 27001", "NIST CSF", "CIS Controls"],
    "Телеком / Связь": ["ISO/IEC 27001", "NIST CSF"],
    "Энергетика / Коммунальная инфраструктура": ["IEC 62443", "ISO/IEC 27001"],
    "Транспорт / Логистика": ["ISO/IEC 27001", "ISO 22301"],
    "Производство / АСУ ТП": ["IEC 62443", "ISO/IEC 27001"],
    "IT / Разработка": ["OWASP ASVS", "Secure SDLC", "ISO/IEC 27001"],
    "Образование": ["ISO/IEC 27001", "CIS Controls"],
    "Услуги / Корпоративный сектор": ["ISO/IEC 27001", "CIS Controls"],
}


def industry_regulatory_profile(industry):
    legal_ids = INDUSTRY_REGULATORY_IDS.get(industry, ["PD_LAW", "PD_RULES"])
    frameworks = list(INDUSTRY_FRAMEWORKS.get(industry, ["ISO/IEC 27001", "CIS Controls"]))
    return {
        "industry": industry or "Другое",
        "legal_ids": legal_ids,
        "laws": [REGULATORY_CATALOG[item] for item in legal_ids if item in REGULATORY_CATALOG],
        "frameworks": frameworks,
    }


def get_regulators_by_industry(industry):
    profile = industry_regulatory_profile(industry)
    legal_lines = [
        f"- [{item_id}] {REGULATORY_CATALOG[item_id]['title']}: {REGULATORY_CATALOG[item_id]['scope']}"
        for item_id in profile["legal_ids"]
        if item_id in REGULATORY_CATALOG
    ]
    framework_lines = [f"- {item}" for item in profile["frameworks"]]
    return (
        "ОБЯЗАТЕЛЬНЫЕ И УСЛОВНО ПРИМЕНИМЫЕ НОРМЫ РК:\n"
        + "\n".join(legal_lines)
        + "\n\nРЕКОМЕНДАТЕЛЬНЫЕ И ОТРАСЛЕВЫЕ СТАНДАРТЫ:\n"
        + "\n".join(framework_lines)
    )

def generate_rule_based_risks(results, context):

    risks = []
    users = context.get("users", 0)
    servers = context.get("servers", 0)

    has_critical_systems = context.get(
        "has_critical_systems",
        False
    )

    has_personal_data = context.get(
        "has_personal_data",
        False
    )

    has_public_web = context.get(
        "has_public_web",
        False
    )

    has_development = context.get(
        "has_development",
        False
    )

    large_company = context.get(
        "large_company",
        False
    )

    enterprise_company = context.get(
        "enterprise_company",
        False
    )

    users = results.get("_user_count", 0)

    # =========================
    # MFA / ACCESS CONTROL
    # =========================

    if results.get("MFA") == "Нет":
        access_scope = (
            "VPN, административных учетных записей, почты и критичных систем"
            if results.get("VPN") != "Нет"
            else "административных учетных записей, почты и критичных систем"
        )
        risks.append({
            "level": "CRITICAL",
            "risk": "Критичные доступы не защищены MFA",
            "description": f"В анкете не указана многофакторная аутентификация для {access_scope}.",
            "impact": "Компрометация одного пароля может привести к несанкционированному доступу к корпоративным системам.",
            "recommendation": "Внедрить MFA поэтапно: сначала для администраторов, VPN, почты и критичных бизнес-систем.",
            "regulators": ["ISO 27001", "NIST", "CIS Controls"],
            "vendors": ["Cisco Duo", "FortiAuthenticator", "Axidian"]
        })

    if results.get("VPN") != "Нет" and results.get("MFA") == "Нет":

        risks.append({
            "level": "CRITICAL",
            "risk": "Удаленный доступ без MFA",
            "description": "VPN-доступ реализован без многофакторной аутентификации.",
            "impact": "Высокий риск компрометации учетных записей и несанкционированного доступа.",
            "recommendation": "Внедрить MFA для VPN, административного доступа и критичных систем.",
            "regulators": ["ISO 27001", "NIST", "PCI DSS"],
            "vendors": ["Cisco Duo", "FortiAuthenticator", "Axidian"]
        })

        # =========================
    # NO SIEM
    # =========================

    if results.get("SIEM") == "Нет":

        severity = "LOW"
        recommendation = (
            "Для текущего масштаба инфраструктуры достаточно настроить "
            "централизованный сбор критичных журналов и регламент реакции. "
            "Полноценный SIEM целесообразно рассматривать при росте инфраструктуры "
            "или наличии регуляторных требований."
        )

        if (
            has_critical_systems
            or has_personal_data
            or large_company
        ):
            severity = "MEDIUM"
            recommendation = (
                "Рассмотреть MSSP/SOC или легковесный SIEM-сценарий для критичных "
                "систем, VPN, серверов и средств защиты."
            )

        if enterprise_company:
            severity = "HIGH"
            recommendation = (
                "Внедрить SIEM/SOC как целевую модель централизованного мониторинга "
                "с корреляцией событий, регламентами реагирования и регулярной "
                "аналитикой инцидентов."
            )

        risks.append({
            "level": severity,
            "risk": "Отсутствует централизованный мониторинг ИБ",
            "description": (
                "События безопасности не агрегируются "
                "в единой системе мониторинга."
            ),
            "impact": (
                "Увеличение времени обнаружения атак "
                "и расследования инцидентов."
            ),
            "recommendation": recommendation,
            "regulators": [
                "ISO 27001",
                "NIST CSF"
            ],
            "vendors": [
                "IBM QRadar",
                "Splunk",
                "R-Vision SIEM"
            ]
        })

    # =========================
    # EPP WITHOUT EDR
    # =========================

    if results.get("Антивирус") != "Нет" and results.get("EDR") == "Нет":

        risks.append({
            "level": "HIGH",
            "risk": "Endpoint-защита ограничена только антивирусом",
            "description": "Используется только базовая антивирусная защита без EDR/XDR-функциональности.",
            "impact": "Низкая эффективность обнаружения сложных атак, ransomware и lateral movement.",
            "recommendation": "Рассмотреть внедрение EDR/XDR платформы.",
            "regulators": ["NIST", "MITRE ATT&CK"],
            "vendors": ["CrowdStrike", "SentinelOne", "Trend Micro"]
        })

    if results.get("Антивирус") == "Нет" and users > 0:
        risks.append({
            "level": "CRITICAL",
            "risk": "Не указана базовая защита рабочих мест",
            "description": f"В анкете указано {users} АРМ, но не указано наличие EPP/антивирусной защиты.",
            "impact": "Рабочие станции остаются наиболее вероятной точкой входа для malware, ransomware и фишинговых атак.",
            "recommendation": "Внедрить централизованную EPP-защиту с едиными политиками, отчетностью и контролем статуса агентов.",
            "regulators": ["CIS Controls", "NIST CSF", "ISO 27001"],
            "vendors": ["Kaspersky", "Trend Micro", "Bitdefender", "Eset"]
        })

    # =========================
    # LARGE INFRA WITHOUT SEGMENTATION
    # =========================

    if users > 100 and results.get("Сегментация сети") == "Нет":

        risks.append({
            "level": "MEDIUM",
            "risk": "Требуется уточнить архитектуру сегментации сети",
            "description": "В анкете не описаны VLAN, ACL, политики FortiGate/NGFW и правила доступа между пользовательскими, серверными и гостевыми сегментами.",
            "impact": "Без подтвержденной схемы сегментации невозможно корректно оценить риск lateral movement; отсутствие NAC само по себе не доказывает отсутствие сегментации.",
            "recommendation": "Проверить текущие VLAN/ACL, политики NGFW, guest Wi-Fi и доступ к серверным сегментам; по результатам подготовить точечные меры усиления.",
            "regulators": ["ISO 27001", "NIST"],
            "vendors": ["Fortinet", "Cisco", "Huawei"]
        })

    # =========================
    # BACKUP RISKS
    # =========================

    if results.get("Резервное копирование") != "Нет" and results.get("Immutable Backup") == "Нет":

        risks.append({
            "level": "HIGH",
            "risk": "Backup не защищен от ransomware",
            "description": "Отсутствует immutable/offline backup.",
            "impact": "Риск уничтожения резервных копий при ransomware-инциденте.",
            "recommendation": "Внедрить immutable backup и air-gap копии.",
            "regulators": ["NIST", "ISO 27001"],
            "vendors": ["Veeam", "Commvault", "Rubrik"]
        })

    if results.get("Резервное копирование") == "Нет" and servers > 0:
        risks.append({
            "level": "CRITICAL",
            "risk": "Не указан контур резервного копирования серверов",
            "description": f"В инфраструктуре указано {servers} серверов, но резервное копирование не описано.",
            "impact": "При сбое, ошибке администратора или ransomware-инциденте восстановление сервисов может быть невозможным.",
            "recommendation": "Определить RPO/RTO, внедрить регулярный backup критичных систем и провести тестовое восстановление.",
            "regulators": ["ISO 27001", "NIST CSF", "CIS Controls"],
            "vendors": ["Veeam", "Commvault", "Veritas"]
        })

    if servers >= 10 and results.get("Резервное копирование") != "Нет" and results.get("DR") == "Нет":
        risks.append({
            "level": "HIGH",
            "risk": "Не описан план аварийного восстановления ИТ-сервисов",
            "description": f"В инфраструктуре указано {servers} серверов и backup-контур, но не описаны DR-сценарии, RTO/RPO и регулярные тесты восстановления.",
            "impact": "При сбое площадки, СХД или критичных виртуальных машин восстановление может занять дольше допустимого для бизнеса.",
            "recommendation": "Описать критичные сервисы, определить RTO/RPO, провести тест восстановления и подготовить поэтапный DR-план для ERP, CRM, почты и файловых сервисов.",
            "regulators": ["ISO 22301", "ISO 27001", "NIST CSF"],
            "vendors": ["Veeam", "Commvault", "Rubrik"]
        })

    if servers >= 10 and results.get("Мониторинг") == "Нет":
        risks.append({
            "level": "MEDIUM",
            "risk": "Не описан эксплуатационный мониторинг ИТ-инфраструктуры",
            "description": "В анкете есть серверы, сеть и бизнес-системы, но не указан единый мониторинг доступности, производительности и емкости.",
            "impact": "Инциденты по дискам, каналам, виртуальным машинам или бизнес-сервисам могут обнаруживаться постфактум по обращениям пользователей.",
            "recommendation": "Ввести мониторинг доступности и емкости для серверов, СХД, каналов связи, виртуализации и ключевых приложений; настроить пороги, ответственных и регулярный обзор трендов.",
            "regulators": ["ITIL", "ISO 20000", "ISO 27001"],
            "vendors": ["Zabbix", "PRTG", "ManageEngine"]
        })

    if results.get("Виртуализация") != "Нет" and results.get("Кластеризация") == "Нет":
        risks.append({
            "level": "MEDIUM",
            "risk": "Виртуализация требует проверки отказоустойчивости",
            "description": "В инфраструктуре используются виртуальные серверы, но не указаны кластеризация, правила распределения нагрузок и сценарии отказа хостов.",
            "impact": "Отказ одного физического хоста или ошибки распределения ресурсов могут повлиять сразу на несколько бизнес-сервисов.",
            "recommendation": "Проверить HA-настройки, резервы CPU/RAM/storage, правила размещения критичных VM и порядок обслуживания хостов без простоя сервисов.",
            "regulators": ["ITIL", "ISO 20000"],
            "vendors": ["VMware", "Veeam", "Zabbix"]
        })

    if results.get("СХД") != "Нет" and results.get("Мониторинг СХД") == "Нет":
        risks.append({
            "level": "LOW",
            "risk": "СХД: требуется проверить capacity/performance, без признаков аварийного риска",
            "description": "В анкете указаны RAID, диски, backup/snapshot-практики, но не раскрыты метрики утилизации, latency/IOPS и правила контроля емкости.",
            "impact": "Без метрик нельзя утверждать о проблемах надежности или производительности; корректнее провести health-check и подтвердить запас емкости.",
            "recommendation": "Проверить утилизацию, latency/IOPS, состояние RAID, snapshot-политики и связку с backup; по результатам дать рекомендации по capacity management.",
            "regulators": ["ITIL", "ISO 20000"],
            "vendors": ["Veeam"]
        })

    if has_public_web and results.get("WAF") == "Нет":
        risks.append({
            "level": "HIGH",
            "risk": "Публичные веб-сервисы не защищены WAF",
            "description": "В анкете указаны публичные веб-ресурсы, но не указана WAF-защита.",
            "impact": "Повышается риск атак на веб-приложения, утечки данных и нарушения доступности сервиса.",
            "recommendation": "Провести экспресс-оценку web-периметра и внедрить WAF/CDN-защиту для критичных публичных ресурсов.",
            "regulators": ["OWASP ASVS", "PCI DSS", "ISO 27001"],
            "vendors": ["Imperva", "F5", "Cloudflare", "Radware"]
        })

    if results.get("Mail Security") == "Нет":
        risks.append({
            "level": "MEDIUM",
            "risk": "Не указана специализированная защита электронной почты",
            "description": "В анкете отсутствует информация о mail security/anti-phishing защите.",
            "impact": "Фишинг и вредоносные вложения остаются одним из основных сценариев первичного компрометации.",
            "recommendation": "Проверить текущий почтовый контур и внедрить защиту от фишинга, вредоносных вложений и подмены отправителя.",
            "regulators": ["CIS Controls", "NIST CSF"],
            "vendors": ["Trend Micro", "Forcepoint", "Barracuda"]
        })

    if results.get("PAM") == "Нет" and (servers >= 10 or has_critical_systems):
        risks.append({
            "level": "HIGH",
            "risk": "Привилегированные учетные записи не выделены в отдельный контроль",
            "description": "Есть серверы или критичные системы, но не указано использование PAM.",
            "impact": "Компрометация администраторской учетной записи может привести к полному контролю над инфраструктурой.",
            "recommendation": "Ввести учет привилегированных доступов, vault для админ-паролей и контроль сессий для критичных систем.",
            "regulators": ["ISO 27001", "CIS Controls", "NIST"],
            "vendors": ["CyberArk", "Fudo", "Netwrix"]
        })

    if has_personal_data and results.get("DLP") == "Нет":
        risks.append({
            "level": "MEDIUM",
            "risk": "Не указан контроль утечек конфиденциальных данных",
            "description": "Компания обрабатывает чувствительные данные, но DLP/контроль каналов утечки не указан.",
            "impact": "Растет риск несанкционированной передачи персональных, клиентских или коммерческих данных.",
            "recommendation": "Определить критичные типы данных и внедрить контроль основных каналов: почта, web, USB, облачные хранилища.",
            "regulators": ["Закон РК о персональных данных", "ISO 27001"],
            "vendors": ["Forcepoint", "Zecurion", "Гарда", "Symantec"]
        })

    if has_development and results.get("SAST") == "Нет" and results.get("DAST") == "Нет":
        risks.append({
            "level": "MEDIUM",
            "risk": "Безопасность приложений не встроена в разработку",
            "description": "В анкете указана разработка, но не указаны SAST/DAST-проверки.",
            "impact": "Уязвимости приложений могут попадать в продуктив и становиться точкой компрометации публичных сервисов.",
            "recommendation": "Встроить базовые SAST/DAST-проверки в CI/CD или регулярный процесс релизной проверки.",
            "regulators": ["OWASP ASVS", "ISO 27001", "PCI DSS"],
            "vendors": ["Positive Technologies", "Checkmarx", "HCL"]
        })

    # =========================
    # PATCH MANAGEMENT
    # =========================

    old_os_detected = any([
        results.get("ОС АРМ (Windows XP/Vista/7/8)", 0) > 0,
        results.get("ОС Сервера (Windows Server 2008/2012 R2)", 0) > 0
    ])

    if results.get("Patch Management") == "Нет":

        severity = "MEDIUM"

        if old_os_detected:
            severity = "CRITICAL"

        elif users > 100:
            severity = "HIGH"

        risks.append({
            "level": severity,
            "risk": "Отсутствует централизованный Patch Management",
            "description": (
                "Обновления устанавливаются "
                "несистемно либо вручную."
            ),
            "impact": (
                "Высокая вероятность эксплуатации "
                "известных уязвимостей."
            ),
            "recommendation": (
                "Внедрить централизованное "
                "управление обновлениями."
            ),
            "regulators": [
                "CIS Controls",
                "NIST"
            ],
            "vendors": [
                "ManageEngine",
                "Ivanti",
                "Tanium",
                "Microsoft MECM"
            ]
        })
    # =========================
    # LEGACY OS
    # =========================

    legacy_workstations = (
        results.get(
            "ОС АРМ (Windows XP/Vista/7/8)",
            0
        )
    )

    legacy_servers = (
        results.get(
            "ОС Сервера (Windows Server 2008/2012 R2)",
            0
        )
    )

    if legacy_workstations > 0 or legacy_servers > 0:

        risks.append({

            "level": "CRITICAL",

            "risk": "Использование устаревших операционных систем",

            "description":
                f"Обнаружены устаревшие ОС. "
                f"АРМ: {legacy_workstations}, "
                f"Серверы: {legacy_servers}.",

            "impact":
                "Уязвимости больше не исправляются "
                "производителем.",

            "recommendation":
                "Разработать программу миграции "
                "на поддерживаемые версии ОС.",

            "regulators": [
                "ISO 27001",
                "CIS Controls"
            ],

            "vendors": [
                "Microsoft",
                "Red Hat",
                "VMware",
                "Citrix"
            ]
        })

    priority_order = {"CRITICAL": 1, "HIGH": 2, "MEDIUM": 3, "LOW": 4}
    unique_risks = []
    seen = set()
    for risk in risks:
        title = risk_semantic_key(risk)
        if title and title not in seen:
            unique_risks.append(risk)
            seen.add(title)

    return sorted(
        unique_risks,
        key=lambda item: priority_order.get(str(item.get("level", "")).upper(), 99)
    )
def ai_generate_risks_and_recs(c_info, results):
    import json
    import streamlit as st

    api_key = get_app_secret("GEMINI_API_KEY")
    groq_api_key = get_app_secret("GROQ_API_KEY")
    st.session_state.ai_analysis_succeeded = False

    if not api_key and not groq_api_key:
        return []

    try:
        model_name = get_app_secret("GEMINI_MODEL", "gemini-2.5-flash")
        groq_model = get_app_secret("GROQ_MODEL", "openai/gpt-oss-120b")
        groq_timeout = int(get_app_secret("GROQ_TIMEOUT_SECONDS", 55))
        ai_timeout = int(get_app_secret("GEMINI_TIMEOUT_SECONDS", 45))
        gemini_quality_attempts = int(get_app_secret("GEMINI_QUALITY_ATTEMPTS", 4))
        fallback_models = str(get_app_secret(
            "GEMINI_FALLBACK_MODELS",
            "gemini-2.5-flash-lite"
        ))
        model_candidates = []
        for candidate in [model_name, *fallback_models.split(",")]:
            candidate = str(candidate).strip()
            if candidate and candidate not in model_candidates:
                model_candidates.append(candidate)

        safe_client, safe_results = sanitize_for_ai(
            c_info,
            results
        )
        enabled_controls, missing_controls = security_control_snapshot(results)
        enabled_controls_text = "\n".join(enabled_controls) if enabled_controls else "Не указаны"
        missing_controls_text = ", ".join(missing_controls) if missing_controls else "Нет явных пробелов"
        confirmed_it_gaps = confirmed_it_gap_topics(results)
        confirmed_it_gaps_text = "\n".join(
            f"- {key}: {label}"
            for key, label in confirmed_it_gaps.items()
        ) or "- Явные ИТ-разрывы не подтверждены"

        vendor_context = ""

        regulator_context = get_regulators_by_industry(
            c_info.get("Сфера деятельности", "")
        )

        def summarize_for_ai(values, limit=70):
            summary = []
            priority_markers = (
                "арм", "ос", "сервер", "виртуал", "схд", "резерв", "backup",
                "маршрут", "канал", "wi-fi", "wifi", "точк", "контроллер", "примечан",
                "ngfw", "vpn", "epp", "edr", "xdr", "mdr",
                "dlp", "mail", "waf", "ddos", "ids", "nac", "ztna", "iam",
                "mfa", "pam", "siem", "soar", "уязв", "patch", "web", "разработ"
            )
            for key, value in values.items():
                if str(key).startswith("_"):
                    continue
                text_value = str(value).strip()
                if not text_value or text_value.lower() in {"нет", "none", "nan", "-"}:
                    continue
                key_text = str(key)
                normalized_key = key_text.lower()
                if priority_markers and not any(marker in normalized_key for marker in priority_markers):
                    continue
                summary.append(f"{key_text}: {text_value[:180]}")
                if len(summary) >= limit:
                    break
            if not summary:
                for key, value in values.items():
                    if str(key).startswith("_"):
                        continue
                    text_value = str(value).strip()
                    if text_value:
                        summary.append(f"{key}: {text_value[:180]}")
                    if len(summary) >= 35:
                        break
            return "\n".join(summary)

        ai_summary = summarize_for_ai(safe_results)

        prompt = f"""
Выступай как аудитор Big4, CISO, CTO, Enterprise Security Architect и руководитель ИТ-инфраструктуры.

ЗАДАЧА:
Провести профессиональный анализ ИТ и ИБ инфраструктуры компании на основании опросника.
Отчет должен включать не только кибербезопасность, но и ИТ-управляемость: сеть, серверы, виртуализацию, СХД, backup, DR, мониторинг, эксплуатационные процессы, бизнес-системы и разработку.

ВАЖНЫЕ ПРАВИЛА:

1. Не придумывай технологии без оснований в данных.

2. Не создавай риски только потому, что отсутствует продукт.

3. Оценивай бизнес-контекст компании.

4. Для каждой рекомендации сначала определяй КАТЕГОРИЮ решения:

Примеры:

Отсутствует SIEM
→ категория SIEM

Отсутствует Patch Management
→ категория Patch Management

Отсутствует Backup
→ категория Backup

Отсутствует CASB
→ категория CASB

Отсутствует ZTNA
→ категория ZTNA

Отсутствует SAST
→ категория Application Security

Устаревшие ОС
→ категория Migration Project

Устаревшие серверы
→ категория Infrastructure Modernization

5. Для категории решения сначала используй вендоров из каталога.

6. Если в каталоге отсутствуют подходящие решения,
разрешается использовать мировых лидеров рынка.

7. Никогда не предлагай:
- антивирусы
- EDR
- NGFW
- DLP

для устранения проблем устаревших операционных систем.

8. Для Legacy OS основная рекомендация:

- миграция
- модернизация
- виртуализация
- сегментация
- изоляция

9. Не предлагай SOAR как критическую необходимость
для компаний без зрелого SOC.

10. Не предлагай ZTNA как высокий риск без признаков удаленного доступа.

11. Для каждого риска указывай не более 3 наиболее релевантных решений.

12. Используй вендоров из каталога как приоритетных.

13. Если подходящего решения в каталоге нет,
подбери наиболее распространенные мировые решения.

14. Не дублируй риски.

15. Верни 10-12 наиболее важных рисков, если данных достаточно. Если реальных рисков меньше,
не растягивай отчет искусственно.

16. Приоритет риска определяй по шкале:

CRITICAL
HIGH
MEDIUM
LOW

17. Особое внимание уделяй:

- устаревшим ОС
- отсутствию резервного копирования
- отсутствию MFA
- отсутствию Patch Management
- отсутствию сегментации сети
- отсутствию мониторинга безопасности
- отсутствию защиты почты
- отсутствию контроля привилегированных учетных записей

18. Цель отчета - экспертная диагностика для клиента. Формулировки должны звучать как
профессиональное аудиторское заключение, а не как коммерческое предложение.

19. Рекомендации должны учитывать фактический масштаб инфраструктуры без ярлыков
"малая", "средняя" или "крупная" компания. Опирайся на число АРМ, серверов,
площадок, критичных сервисов, публичных приложений и доступные ресурсы эксплуатации.

20. Не перечисляй все отсутствующие технологии как риски. Выбирай только то, что логично
для размера, серверов, публичных сервисов, бизнес-систем, разработки и текущих средств защиты.

21. Для каждого риска дай конкретную, выполнимую рекомендацию и объясни, какой риск
она снижает.

22. Если в данных есть сеть, серверы, СХД, виртуализация, backup, бизнес-системы или разработка,
обязательно включи 3-5 ИТ-рекомендаций, а не только ИБ-рекомендации.

23. Не используй Microsoft как ИБ-вендора. Microsoft допустим только для тем ОС,
миграции Windows, Windows Server или обновления устаревших систем.

24. Не создавай несколько отдельных пунктов про EDR/XDR/MDR. Если тема одна,
объедини ее в один зрелый риск по защите рабочих мест и реагированию.

25. Не рекомендуй внедрение контроля, если он уже указан в блоке "УЖЕ ВНЕДРЕНО".
Если контроль есть, можно рекомендовать только улучшение покрытия, регламент,
контроль исключений или проверку эффективности, но нельзя писать, что контроль отсутствует.

26. Для каждого риска укажи 1-3 конкретных факта анкеты в evidence и один измеримый
критерий результата в success_metric.

27. legal_ids выбирай только из ID, перечисленных в регуляторном контексте. Не придумывай
законы, номера, статьи и обязательность. Если прямой нормы нет, верни пустой legal_ids.

28. ISO, NIST, CIS, ITIL, OWASP, IEC и PCI DSS указывай только в frameworks. Это стандарты
и методологии, а не регуляторы и не законы Республики Казахстан.

29. Закон устанавливает требуемую меру контроля, а не обязанность купить конкретный продукт.
Не формулируй коммерческое предложение как прямое требование закона.

30. Windows 10 не считай устаревшей ОС автоматически. Linux без номера версии означает
"версия и срок поддержки требуют подтверждения", а не legacy OS. Риск legacy OS допустим
только при явном количестве Windows XP/Vista/7/8/8.1 или Windows Server 2008/2012 R2.

31. Если резервное копирование уже указано, не пиши об его отсутствии. Допустимо рекомендовать
проверку RTO/RPO, immutable/offline-копий, резервирование конфигураций и тест восстановления,
но только как проверку зрелости существующего контура.

32. Для КВОИКИ и организаций с персональными или регулируемыми данными отсутствие DLP оценивай
как HIGH. Отсутствие NAC не доказывает отсутствие сегментации: формулируй отдельную меру по
идентификации, профилированию и допуску устройств к проводной и беспроводной сети.

33. При серверном или критичном контуре без PAM обязательно оцени контроль привилегированных
учетных записей. SOAR показывай только как следующий этап после стабилизации SIEM/SOC,
источников событий, сценариев корреляции и SLA реагирования.

УЖЕ ВНЕДРЕНО:
{enabled_controls_text}

НЕ УКАЗАНО В АНКЕТЕ:
{missing_controls_text}

ДАННЫЕ АУДИТА:

{safe_results}

СФЕРА ДЕЯТЕЛЬНОСТИ:

{c_info.get("Сфера деятельности", "")}

КАТАЛОГ РЕШЕНИЙ КОМПАНИИ:

{vendor_context}

РЕГУЛЯТОРНЫЕ ТРЕБОВАНИЯ:

{regulator_context}

Верни ТОЛЬКО JSON.

Формат:

[
  {{
    "level": "HIGH",
    "risk": "Название риска",
    "description": "Описание риска",
    "impact": "Последствия для бизнеса",
    "recommendation": "Что необходимо сделать",
    "evidence": [
      "Факт из анкеты 1",
      "Факт из анкеты 2"
    ],
    "success_metric": "Как проверить, что мера внедрена и работает",
    "vendors": [
      "Vendor1",
      "Vendor2",
      "Vendor3"
    ],
    "legal_ids": [
      "Только ID из регуляторного контекста, например PD_LAW"
    ],
    "frameworks": [
      "ISO/IEC 27001"
    ]
  }}
]
"""

        def gemini_url(active_model):
            return (
                "https://generativelanguage.googleapis.com/v1beta/models/"
                f"{active_model}:generateContent"
            )

        compact_prompt = f"""
Ты CISO/CTO-аудитор. По обезличенной анкете верни только JSON-объект для экспертного отчета ИТ и ИБ.
Без markdown, без пояснений вне JSON. Первый символ ответа должен быть {{.
JSON должен быть валидным: все строковые значения в одну строку, без переносов внутри кавычек.
Если нужно разделить мысль, используй точку с пробелом, а не перевод строки.

Строгий формат ответа:
{{
  "executive_summary": [
    "3 содержательных вывода для руководства без коммерческого тона"
  ],
  "audit_observations": [
    {{"title": "Короткий домен", "text": "Наблюдение аудитора, связанное с фактами анкеты"}}
  ],
  "it_recommendations": [
    {{
      "level": "HIGH",
      "risk": "Риск или зона незрелости ИТ",
      "description": "1-2 предложения с фактами анкеты",
      "impact": "Бизнес-последствие",
      "recommendation": "3 конкретных шага: быстрый шаг, проектный шаг, контроль результата",
      "evidence": ["Факт анкеты 1", "Факт анкеты 2"],
      "success_metric": "Измеримый критерий результата",
      "vendors": ["категория или решение"],
      "legal_ids": [],
      "frameworks": ["ITIL"]
    }}
  ],
  "security_recommendations": [
    {{
      "level": "HIGH",
      "risk": "Риск или зона незрелости ИБ",
      "description": "1-2 предложения с фактами анкеты",
      "impact": "Бизнес-последствие",
      "recommendation": "3 конкретных шага: быстрый шаг, проектный шаг, контроль результата",
      "evidence": ["Факт анкеты 1", "Факт анкеты 2"],
      "success_metric": "Измеримый критерий результата",
      "vendors": ["категория или решение"],
      "legal_ids": ["Только ID из регуляторного контекста"],
      "frameworks": ["ISO/IEC 27001"]
    }}
  ],
  "management_decisions": [
    "4-6 решений, которые руководитель может утвердить сразу"
  ],
  "roadmap": [
    {{"phase": "0-30 дней", "priority": "P1", "domain": "ИТ/ИБ", "action": "Не более 120 символов", "rationale": "Почему это важно", "owner": "ИТ/ИБ", "effort": "Низкая|Средняя|Высокая", "result": "Измеримый эффект этапа, не более 100 символов"}}
  ]
}}

Правила:
- учитывай масштаб инфраструктуры;
- верни только подтвержденные рекомендации, обычно 2-4 по ИТ и 2-4 по ИБ; не создавай пункты ради количества;
- roadmap должен содержать ровно 6 объектов: по 2 для фаз 0-30, 31-60 и 61-90 дней;
- management_decisions должен содержать ровно 4 конкретных решения;
- ИТ-рекомендации должны быть про сеть, каналы, серверы, виртуализацию, СХД, backup/DR, мониторинг, patch/change/capacity management, бизнес-системы и разработку;
- минимум 4 ИТ-рекомендации должны быть самостоятельными и не сводиться к ИБ-продуктам;
- ИБ-рекомендации должны учитывать уже внедренные контроли и не повторять отсутствующий контроль, если он есть;
- каждое наблюдение должно связывать минимум два факта анкеты, если это возможно;
- избегай примитивных фраз уровня "внедрить продукт"; сначала опиши управленческую/техническую меру, затем только категорию решения;
- не повторяй один и тот же домен разными словами: MFA, удаленный доступ и учетные записи объединяй в один сильный риск, если нет отдельного факта;
- не повторяй EDR/XDR/MDR несколькими пунктами; объединяй в один риск по endpoint detection/response;
- не называй организацию или инфраструктуру малой, средней или крупной; используй только объективные показатели анкеты;
- не предлагай тяжелый SIEM/SOAR как первоочередной проект без достаточного числа источников событий, владельца процесса и ресурса на эксплуатацию;
- оценивай связки: учетные записи, сеть, серверы, резервное копирование, почта, уязвимости, мониторинг и реагирование;
- делай рекомендации экспертными: что сделать, зачем, какой ожидаемый эффект;
- для каждой рекомендации дай порядок внедрения: быстрый шаг, проектный шаг, контроль результата;
- legacy OS закрывай миграцией, изоляцией, сегментацией, обновлением, а не EDR/DLP;
- не перечисляй все отсутствующие продукты как риски;
- не путай категории: MFA не закрывается PAM, уязвимости не закрываются SIEM, устаревшие ОС не закрываются EDR.
- OSPF, BGP и статическая маршрутизация описывают маршрутизацию и не доказывают наличие или отсутствие сетевой сегментации;
- отсутствие NAC или ZTNA не доказывает отсутствие VLAN, ACL, VRF, зон NGFW и межсегментных политик;
- если сегментация прямо не описана, формулируй "архитектура сегментации требует подтверждения", а не "сегментация отсутствует";
- DLP защищает данные от утечек и никогда не является решением для сетевой сегментации;
- не используй Microsoft как ИБ-вендора; Microsoft допустим только для ОС/миграции Windows/Windows Server.
- строго не пиши "отсутствует", "не внедрено" или "не указано" про контроль из списка "Уже внедрено";
- если MFA есть в списке "Уже внедрено", не создавай риск по отсутствию MFA.
- legal_ids выбирай только из идентификаторов в блоке "Регуляторный контекст"; не придумывай законы, номера документов и статьи;
- ISO, NIST, CIS, ITIL, OWASP и PCI DSS указывай только в frameworks, а не в legal_ids;
- наличие правовой обязанности не означает обязанность купить конкретный продукт: связывай норму с требуемой мерой контроля, а производителя показывай только как вариант реализации;
- evidence должен содержать конкретные факты анкеты, а success_metric — проверяемый результат без рекламных формулировок.
- Windows 10 не является legacy OS по умолчанию; Linux без версии требует уточнения, а не вывода об устаревании.
- если backup указан, не создавай риск его отсутствия; оценивай только RTO/RPO, immutable/offline-копии и тест восстановления.
- для КВОИКИ с персональными данными отсутствие DLP имеет уровень HIGH.
- при наличии серверов и критичных систем без PAM обязательно оцени PAM; SOAR указывай как этап развития SIEM/SOC, а не отдельный срочный проект.
- отсутствие NAC описывай как отсутствие автоматизированного контроля допуска устройств, а не как доказательство отсутствия сегментации.
- в roadmap не указывай производителей и названия продуктов; используй только классы технологий и управленческие действия.
- result каждого объекта roadmap должен описывать измеримый результат именно его action, а не всей фазы целиком.
- для новых решений соблюдай порядок: требования и критерии -> ограниченный пилот -> решение о закупке и масштабировании.
- каждый подтвержденный ИТ-разрыв ниже оформляй самостоятельным объектом risks; не объединяй Wi-Fi с резервированием WAN-канала.

Отрасль: {c_info.get("Сфера деятельности", "-")}

Уже внедрено:
{enabled_controls_text}

Не указано:
{missing_controls_text}

Ключевые данные анкеты:
{ai_summary}

Регуляторный контекст:
{regulator_context[:1200]}

Подтвержденные ИТ-разрывы, которые нельзя игнорировать:
{confirmed_it_gaps_text}
"""
        fallback_payload = {
            "contents": [
                {
                    "role": "user",
                    "parts": [{"text": compact_prompt}]
                }
            ],
            "generationConfig": {
                "temperature": 0.15,
                "maxOutputTokens": 6144,
            },
        }

        minimal_prompt = f"""
Ты CISO/CTO-аудитор. Верни только валидный JSON без markdown.
Корневые поля: executive_summary, audit_observations, it_recommendations,
security_recommendations, management_decisions, roadmap.
Каждый подтвержденный ИТ-разрыв оформи самостоятельной законченной рекомендацией.
Не объединяй Wi-Fi с резервированием WAN. Не предлагай уже внедренные контроли.
Для каждой рекомендации укажи level, risk, description, impact, recommendation,
evidence, success_metric, vendors, legal_ids и frameworks.

Подтвержденные ИТ-разрывы:
{confirmed_it_gaps_text}

Уже внедрено:
{enabled_controls_text}

Факты анкеты:
{ai_summary[:4200]}
"""
        minimal_payload = {
            "contents": [
                {
                    "role": "user",
                    "parts": [{"text": minimal_prompt}]
                }
            ],
            "generationConfig": {
                "temperature": 0.05,
                "maxOutputTokens": 4096,
            },
        }

        line_prompt = f"""
Ты CISO/CTO-аудитор. Верни 8-10 строк без JSON и без markdown. Покрой ИТ и ИБ, не только endpoint.
Формат каждой строки строго:
LEVEL | RISK | DESCRIPTION | IMPACT | RECOMMENDATION

LEVEL только CRITICAL, HIGH, MEDIUM или LOW.
Не используй символ | внутри полей.
Каждое поле в одну строку.
Не используй Microsoft как ИБ-вендора. Не дублируй EDR/XDR/MDR отдельными строками.
Не пиши, что отсутствует контроль из списка "Уже внедрено".

Уже внедрено:
{enabled_controls_text}

Не указано:
{missing_controls_text}

Данные:
{ai_summary[:3600]}

Подтвержденные ИТ-разрывы, каждый из которых нужен отдельной строкой:
{confirmed_it_gaps_text}
"""
        line_payload = {
            "contents": [
                {
                    "role": "user",
                    "parts": [{"text": line_prompt}]
                }
            ],
            "generationConfig": {
                "temperature": 0.05,
                "maxOutputTokens": 4096,
            },
        }

        def focused_line_payload(title, target_count, focus):
            focused_prompt = f"""
Ты CISO/CTO-аудитор. Верни ровно {target_count} строк без JSON и без markdown.
Формат каждой строки строго:
LEVEL | RISK | DESCRIPTION | IMPACT | RECOMMENDATION

LEVEL только CRITICAL, HIGH, MEDIUM или LOW.
Не используй символ | внутри полей.
Каждое поле в одну строку, без нумерации.
Каждая рекомендация должна быть законченной фразой: быстрый шаг; проектный шаг; контроль результата.
Не используй Microsoft как ИБ-вендора. Microsoft допустим только для ОС/Windows/Windows Server.
Не дублируй EDR/XDR/MDR отдельными строками.

Раздел: {title}
Фокус анализа: {focus}

Отрасль: {c_info.get("Сфера деятельности", "-")}

Ключевые данные анкеты:
{ai_summary}
"""
            return {
                "contents": [
                    {
                        "role": "user",
                        "parts": [{"text": focused_prompt[:5200]}]
                    }
                ],
                "generationConfig": {
                    "temperature": 0.05,
                    "maxOutputTokens": 3072,
                },
            }

        focused_payloads = (
            (
                "line",
                focused_line_payload(
                    "ИТ-инфраструктура и эксплуатационная зрелость",
                    4,
                    "сеть, каналы связи, серверы, виртуализация, СХД, backup, DR, мониторинг, бизнес-системы, разработка"
                )
            ),
            (
                "line",
                focused_line_payload(
                    "Информационная безопасность",
                    5,
                    "MFA, endpoint detection/response, vulnerability/patch management, mail security, WAF, PAM, SIEM/SOC, DLP"
                )
            ),
        )

        def extract_gemini_error(response):
            try:
                payload = response.json()
                message = payload.get("error", {}).get("message", response.text)
                status = payload.get("error", {}).get("status", "")
                return f"HTTP {response.status_code} {status}: {message}"
            except Exception:
                return f"HTTP {response.status_code}: {response.text[:1200]}"

        def call_gemini(request_payload, active_model):
            active_url = gemini_url(active_model)
            response_payload = None
            if os.name != "nt":
                import requests

                def gemini_post(verify):
                    return requests.post(
                        active_url,
                        params={"key": api_key},
                        json=request_payload,
                        timeout=ai_timeout,
                        verify=verify,
                    )

                try:
                    response = gemini_post(REQUEST_VERIFY)
                except requests.exceptions.SSLError:
                    response = gemini_post(False)
                if not response.ok:
                    raise RuntimeError(extract_gemini_error(response))
                return response.json()

            if response_payload is None:
                response_payload = node_fetch_json(
                    f"{active_url}?key={api_key}",
                    {
                        "url": f"{active_url}?key={api_key}",
                        "method": "POST",
                        "body": request_payload,
                        "timeoutMs": ai_timeout * 1000,
                    },
                    timeout_seconds=ai_timeout,
                )
            return response_payload

        def gemini_response_text(response_payload):
            parts = (
                response_payload
                .get("candidates", [{}])[0]
                .get("content", {})
                .get("parts", [])
            )
            return "\n".join(
                str(part.get("text", ""))
                for part in parts
                if part.get("text")
            ).strip()

        def gemini_empty_reason(response_payload):
            candidate = response_payload.get("candidates", [{}])[0]
            finish_reason = candidate.get("finishReason", "UNKNOWN")
            prompt_feedback = response_payload.get("promptFeedback", {})
            return f"пустой ответ Gemini; finishReason={finish_reason}; promptFeedback={prompt_feedback}"

        gemini_retry_count = int(get_app_secret("GEMINI_RETRY_COUNT", 1))
        gemini_retry_delay = float(get_app_secret("GEMINI_RETRY_DELAY_SECONDS", 2.0))

        def is_gemini_quota_exhausted(error_text):
            lowered = str(error_text or "").lower()
            return (
                "resource_exhausted" in lowered
                or "quota exceeded" in lowered
                or "http 429" in lowered
            )

        def should_retry_gemini_error(error_text):
            lowered = str(error_text or "").lower()
            return any(
                marker in lowered
                for marker in (
                    "503",
                    "unavailable",
                    "high demand",
                    "пустой ответ gemini",
                    "finishreason=unknown",
                )
            )

        def call_gemini_with_retries(request_payload, active_model):
            last_error = None
            for attempt in range(gemini_retry_count + 1):
                try:
                    response_payload = call_gemini(request_payload, active_model)
                    response_text = gemini_response_text(response_payload)
                    if response_text:
                        return response_payload, response_text

                    last_error = gemini_empty_reason(response_payload)
                except Exception as exc:
                    last_error = redact_secret(exc, api_key)

                if attempt < gemini_retry_count and should_retry_gemini_error(last_error):
                    time.sleep(gemini_retry_delay * (attempt + 1))
                    continue

                break

            raise RuntimeError(str(last_error or "Gemini did not return text"))

        def repair_json_text(text):
            cleaned = []
            in_string = False
            escaped = False
            for char in text:
                if in_string and char in "\r\n\t":
                    cleaned.append(" ")
                    escaped = False
                    continue
                cleaned.append(char)
                if escaped:
                    escaped = False
                elif char == "\\":
                    escaped = True
                elif char == '"':
                    in_string = not in_string

            repaired = "".join(cleaned).strip()
            array_start = repaired.find("[")
            object_start = repaired.find("{")
            if object_start >= 0 and (array_start < 0 or object_start < array_start):
                start = object_start
                end = repaired.rfind("}")
            else:
                start = array_start
                end = repaired.rfind("]")
            if start >= 0:
                repaired = repaired[start:end + 1] if end > start else repaired[start:]

            open_strings = repaired.count('"') % 2
            if open_strings:
                repaired += '"'

            open_braces = repaired.count("{") - repaired.count("}")
            open_brackets = repaired.count("[") - repaired.count("]")
            if open_braces > 0:
                repaired += "}" * open_braces
            if open_brackets > 0:
                repaired += "]" * open_brackets

            repaired = re.sub(r",\s*([}\]])", r"\1", repaired)
            return repaired

        def parse_ai_response_text(response_text):
            response_text = re.sub(r"^```(?:json)?", "", response_text).strip()
            response_text = re.sub(r"```$", "", response_text).strip()
            variants = [response_text]
            start = response_text.find("[")
            end = response_text.rfind("]")
            if start >= 0 and end > start:
                variants.append(response_text[start:end + 1])
            object_start = response_text.find("{")
            object_end = response_text.rfind("}")
            if object_start >= 0 and object_end > object_start:
                variants.append(response_text[object_start:object_end + 1])
            variants.append(repair_json_text(response_text))

            last_error = None
            for variant in variants:
                try:
                    return json.loads(variant)
                except json.JSONDecodeError as exc:
                    last_error = exc
            recovered_payload = recover_complete_risk_objects(response_text)
            if recovered_payload:
                return recovered_payload
            raise last_error

        def parse_line_response(response_text):
            rows = []
            for raw_line in response_text.splitlines():
                line = raw_line.strip().lstrip("-•0123456789. ")
                if "|" not in line:
                    continue
                parts = [part.strip() for part in line.split("|")]
                if len(parts) < 5:
                    continue
                level, risk, description, impact, recommendation = parts[:5]
                if level.upper() not in {"CRITICAL", "HIGH", "MEDIUM", "LOW"}:
                    continue
                rows.append({
                    "level": level.upper(),
                    "risk": risk,
                    "description": description,
                    "impact": impact,
                    "recommendation": recommendation,
                    "vendors": [],
                    "regulators": [],
                })
            return rows

        def accept_ai_payload(parsed_payload, provider_label, model_label, errors):
            ai_narrative = normalize_ai_audit_narrative(parsed_payload)
            narrative_requirements = {
                "executive_summary": 2,
                "audit_observations": 2,
                "management_decisions": 4,
                "roadmap": 6,
            }
            missing_narrative = [
                f"{field}<{minimum}"
                for field, minimum in narrative_requirements.items()
                if len(ai_narrative.get(field, [])) < minimum
            ]
            roadmap_phase_counts = {
                phase: sum(phase in str(item.get("phase", "")) for item in ai_narrative.get("roadmap", []))
                for phase in ("0-30", "31-60", "61-90")
            }
            if any(count < 2 for count in roadmap_phase_counts.values()):
                missing_narrative.append("roadmap phases<2")
            if missing_narrative:
                errors.append(
                    f"{provider_label}: часть презентационного материала будет дополнена "
                    f"экспертным движком ({', '.join(missing_narrative)})"
                )
            raw_candidate_count = count_ai_risk_candidates(parsed_payload)
            normalized_payload = normalize_ai_risks_payload(parsed_payload)
            normalized_payload = augment_ai_risks_from_narrative(
                normalized_payload,
                ai_narrative,
                results,
            )
            normalized_candidate_count = len(normalized_payload)
            recommendation_fields = ("risks", "it_recommendations", "security_recommendations")
            present_recommendation_fields = [
                field
                for field in recommendation_fields
                if isinstance(parsed_payload, dict) and isinstance(parsed_payload.get(field), list)
            ]
            explicit_no_findings = bool(present_recommendation_fields) and not normalized_payload and all(
                not parsed_payload.get(field) for field in present_recommendation_fields
            )
            if explicit_no_findings and not confirmed_it_gaps:
                st.session_state.ai_last_error = ""
                st.session_state.ai_model_used = model_label
                st.session_state.ai_provider_used = provider_label
                st.session_state.ai_audit_narrative = ai_narrative
                st.session_state.ai_analysis_succeeded = True
                return []
            normalized_payload, skipped_ai_items = filter_ai_risks_by_answers(
                normalized_payload,
                results,
            )
            fact_checked_count = len(normalized_payload)
            if skipped_ai_items:
                errors.append(
                    f"{provider_label}: отброшены противоречивые пункты: "
                    + "; ".join(skipped_ai_items[:5])
                )

            if provider_label == "Groq":
                prepared_payload, quality_error = ai_quality_gate(
                    normalized_payload,
                    min_items=1,
                    min_security_items=0,
                    min_it_items=0,
                )
            else:
                minimum_ai_items = max(
                    1,
                    min(4, math.ceil(len(confirmed_it_gaps) * 0.5)),
                ) if confirmed_it_gaps else 1
                prepared_payload, quality_error = ai_quality_gate(
                    normalized_payload,
                    min_items=minimum_ai_items,
                    min_security_items=0,
                    min_it_items=0,
                )
            if not prepared_payload:
                rejection_details = explain_ai_risk_rejections(normalized_payload)
                errors.append(
                    f"{provider_label}: "
                    f"{quality_error or 'нет пригодных законченных рекомендаций'} "
                    f"Диагностика приема: raw={raw_candidate_count}, "
                    f"normalized={normalized_candidate_count}, fact_checked={fact_checked_count}; "
                    f"rejected={rejection_details}."
                )
                return None

            matched_it_gaps, missing_it_gaps = ai_it_gap_coverage(
                prepared_payload,
                confirmed_it_gaps,
            )
            minimum_it_gap_coverage = max(
                1,
                math.ceil(len(confirmed_it_gaps) * 0.7),
            ) if confirmed_it_gaps else 0
            mandatory_it_gaps = {
                key for key in ("wifi_capacity", "network_performance")
                if key in confirmed_it_gaps
            }
            missing_mandatory_it_gaps = mandatory_it_gaps.intersection(missing_it_gaps)
            if (
                len(matched_it_gaps) < minimum_it_gap_coverage
                or missing_mandatory_it_gaps
            ):
                errors.append(
                    f"{provider_label}: ИТ-анализ покрыл {len(matched_it_gaps)} из "
                    f"{len(confirmed_it_gaps)} подтвержденных тем; не покрыты: "
                    + ", ".join(confirmed_it_gaps.get(key, key) for key in missing_it_gaps)
                )
                return None

            st.session_state.ai_last_error = ""
            st.session_state.ai_model_used = model_label
            st.session_state.ai_provider_used = provider_label
            st.session_state.ai_audit_narrative = ai_narrative
            st.session_state.ai_analysis_succeeded = True
            return prepared_payload

        def call_groq_once(focus_it=False):
            focus_instruction = ""
            if focus_it:
                focus_instruction = """
Предыдущий ответ был отклонен, потому что предлагал уже внедренные ИБ-контроли.
Сформируй рекомендации ТОЛЬКО по подтвержденным ИТ-разрывам из фактов и примечаний:
производительность и управляемость Wi-Fi, емкость каналов и failover, мониторинг,
capacity planning серверов/виртуализации/СХД, CMDB/ITAM, change management,
SLA, RTO/RPO, тесты восстановления и эксплуатационные процессы. Не упоминай
внедрение DLP, SIEM, EDR/XDR, MFA, PAM, WAF, NAC, patch management и других
контролей, если они перечислены в блоке «Уже внедрено». Не создавай ИБ-риск
только ради баланса доменов.
""".strip()
            groq_prompt = f"""
Ты senior-аудитор ИТ и ИБ. Проанализируй только факты обезличенной анкеты.
Верни от 4 до 6 законченных подтвержденных рекомендаций по ИТ и ИБ, если в блоке
«Не указано» есть не менее четырех применимых разрывов. Каждый технический домен,
который используется в roadmap, обязательно представь отдельным объектом в risks.
Не добавляй искусственные пункты ради количества. Если ответ не помещается, сократи число
пунктов, но обязательно заверши JSON и каждую выданную рекомендацию.

{focus_instruction}

Подтвержденные ИТ-разрывы для обязательного покрытия:
{confirmed_it_gaps_text}

Верни только валидный JSON-объект со следующими корневыми полями:
- executive_summary: 3 коротких содержательных вывода для руководителя;
- audit_observations: 3 объекта с полями title и text;
- management_decisions: 4 конкретных управленческих решения;
- roadmap: 6 объектов, по 2 на фазы "0-30 дней", "31-60 дней", "61-90 дней".
  Поля: phase, priority, domain, action, rationale, owner, effort, result.
  action не длиннее 120 символов, result не длиннее 100 символов и объясняет
  измеримый эффект именно этой фазы;
- risks: массив подтвержденных рекомендаций.

Каждый элемент risks должен содержать поля:
area (только "ИТ" или "ИБ"), level (CRITICAL/HIGH/MEDIUM/LOW), risk,
description, impact, recommendation, evidence (массив строк), success_metric,
vendors (массив строк), legal_ids (массив строк), frameworks (массив строк).
Дополнительные поля допустимы, но не заменяют перечисленные обязательные поля.

Для каждой рекомендации:
- свяжи риск минимум с одним конкретным фактом анкеты;
- укажи бизнес-последствие;
- дай три шага через точку с запятой: быстрый шаг, проектный шаг, контроль результата;
- добавь измеримый success_metric;
- не объявляй контроль отсутствующим, если он есть в блоке "Уже внедрено";
- перед выдачей ответа удали любой пункт, который предлагает внедрить уже включенный контроль;
- не превращай каждый отсутствующий продукт в отдельный риск;
- не путай MFA с PAM, управление уязвимостями с SIEM, DLP с сегментацией;
- OSPF/BGP не доказывают наличие или отсутствие сегментации;
- legacy OS закрывай миграцией, обновлением и компенсирующими мерами;
- Windows 10 не считай legacy OS автоматически; Linux без версии означает необходимость уточнения версии и поддержки;
- если backup уже указан, не создавай риск его отсутствия: оценивай RTO/RPO, immutable/offline-копии и тест восстановления;
- для КВОИКИ с персональными данными отсутствие DLP оценивай как HIGH;
- отсутствие NAC формулируй как разрыв автоматизированного допуска и профилирования устройств, а не как отсутствие VLAN/ACL;
- при серверном и критичном контуре без PAM обязательно оцени привилегированные доступы;
- SOAR показывай только как следующий этап развития SIEM/SOC после стабилизации источников, сценариев и SLA;
- в roadmap не указывай производителей или продукты; result каждого объекта должен относиться только к его action;
- для нового решения сначала требования и пилот, затем решение о закупке и масштабировании;
- не используй Microsoft как ИБ-вендора, кроме миграции Windows/Windows Server;
- legal_ids выбирай только из переданного регуляторного контекста.

Отрасль: {c_info.get("Сфера деятельности", "-")}

Уже внедрено:
{enabled_controls_text[:1800]}

Не указано:
{missing_controls_text[:1200]}

Ключевые данные анкеты:
{ai_summary[:5200]}

Регуляторный контекст:
{regulator_context[:1000]}
""".strip()
            groq_payload = {
                "model": groq_model,
                "messages": [
                    {
                        "role": "system",
                        "content": (
                            "Ты senior-аудитор ИТ и ИБ. Анализируй только факты "
                            "обезличенной анкеты, не выдумывай отсутствующие проблемы "
                            "и возвращай только валидный JSON-объект с массивом risks."
                        ),
                    },
                    {"role": "user", "content": groq_prompt},
                ],
                "temperature": 0.05,
                "max_completion_tokens": 3400,
                "reasoning_effort": "low",
            }
            groq_url = "https://api.groq.com/openai/v1/chat/completions"
            headers = {"Authorization": f"Bearer {groq_api_key}"}

            if os.name != "nt":
                import requests

                def groq_post(verify):
                    return requests.post(
                        groq_url,
                        headers=headers,
                        json=groq_payload,
                        timeout=groq_timeout,
                        verify=verify,
                    )

                try:
                    response = groq_post(REQUEST_VERIFY)
                except requests.exceptions.SSLError:
                    response = groq_post(False)
                if not response.ok:
                    try:
                        detail = response.json().get("error", {}).get("message", response.text)
                    except Exception:
                        detail = response.text
                    raise RuntimeError(f"HTTP {response.status_code}: {str(detail)[:1200]}")
                response_payload = response.json()
            else:
                response_payload = node_fetch_json(
                    groq_url,
                    {
                        "url": groq_url,
                        "method": "POST",
                        "headers": headers,
                        "body": groq_payload,
                        "timeoutMs": groq_timeout * 1000,
                    },
                    timeout_seconds=groq_timeout,
                )

            response_text = (
                response_payload.get("choices", [{}])[0]
                .get("message", {})
                .get("content", "")
            )
            if not str(response_text).strip():
                raise RuntimeError("Groq вернул пустой ответ")
            return parse_ai_response_text(str(response_text))

        def call_groq_with_rate_limit_retry(focus_it=False):
            try:
                return call_groq_once(focus_it=focus_it)
            except Exception as exc:
                error_text = str(exc)
                if "HTTP 429" not in error_text and "rate limit" not in error_text.lower():
                    raise
                retry_match = re.search(
                    r"try again in\s+([0-9]+(?:\.[0-9]+)?)s",
                    error_text,
                    flags=re.IGNORECASE,
                )
                retry_seconds = float(retry_match.group(1)) if retry_match else 35.0
                time.sleep(min(50.0, max(5.0, retry_seconds + 1.0)))
                return call_groq_once(focus_it=focus_it)

        ai_errors = []
        payload_attempts = (
            ("json", fallback_payload),
            ("json", minimal_payload),
            ("line", line_payload),
        )
        gemini_attempt_count = 0
        if api_key:
            # Exhaust the primary Gemini model and its compact recovery formats
            # before moving to another Gemini model, and only then fall back to Groq.
            for active_model in model_candidates:
                if gemini_attempt_count >= gemini_quality_attempts:
                    break
                for response_format, request_payload in payload_attempts:
                    if gemini_attempt_count >= gemini_quality_attempts:
                        break
                    gemini_attempt_count += 1
                    try:
                        response_payload, response_text = call_gemini_with_retries(
                            request_payload,
                            active_model,
                        )
                        parsed_payload = (
                            parse_line_response(response_text)
                            if response_format == "line"
                            else parse_ai_response_text(response_text)
                        )
                        prepared_payload = accept_ai_payload(
                            parsed_payload,
                            "Gemini",
                            active_model,
                            ai_errors,
                        )
                        if prepared_payload is not None:
                            return prepared_payload
                    except Exception as exc:
                        safe_error = redact_secret(exc, api_key)
                        ai_errors.append(f"Gemini/{active_model}: {safe_error}")

        if groq_api_key:
            try:
                parsed_payload = call_groq_with_rate_limit_retry()
                prepared_payload = accept_ai_payload(
                    parsed_payload,
                    "Groq",
                    groq_model,
                    ai_errors,
                )
                if prepared_payload is not None:
                    return prepared_payload

                parsed_payload = call_groq_with_rate_limit_retry(focus_it=True)
                prepared_payload = accept_ai_payload(
                    parsed_payload,
                    "Groq",
                    groq_model,
                    ai_errors,
                )
                if prepared_payload is not None:
                    return prepared_payload
            except Exception as exc:
                ai_errors.append(
                    f"Groq/{groq_model}: {redact_secret(exc, groq_api_key)}"
                )

        raise ValueError(
            "ИИ-провайдеры не дали пригодный ответ. " + " | ".join(ai_errors[-8:])
        )

    except Exception as e:
        safe_error = redact_secret(e, api_key)
        safe_error = redact_secret(safe_error, groq_api_key)
        st.session_state.ai_last_error = safe_error
        st.session_state.ai_provider_used = "Нет ответа"
        st.session_state.ai_audit_narrative = {}
        st.session_state.ai_analysis_succeeded = False
        return []



# --- AI BLOCK END ---

# --- 1. НАСТРОЙКИ СТРАНИЦЫ ---
st.set_page_config(page_title="Аудит ИТ и ИБ 2026", layout="wide", page_icon="🛡️")

def inject_audit_design():
    st.markdown("""
    <style>
    :root {
        --audit-bg: #f6f7f9;
        --audit-panel: #ffffff;
        --audit-border: #d8dee8;
        --audit-text: #151922;
        --audit-muted: #667085;
        --audit-accent: #0f766e;
        --audit-accent-soft: #dff3ef;
        --audit-warn: #b45309;
        --audit-warn-soft: #fff7ed;
        --audit-risk: #b42318;
        --audit-risk-soft: #fff1f0;
        --audit-ink: #1f2937;
    }

    .stApp {
        background: var(--audit-bg);
        color: var(--audit-text);
    }

    section[data-testid="stSidebar"] {
        background: #fbfcfe;
        border-right: 1px solid var(--audit-border);
    }

    div[data-testid="stHeader"] {
        background: rgba(246, 247, 249, 0.92);
        backdrop-filter: blur(10px);
    }

    .block-container {
        padding-top: 2.2rem;
        padding-bottom: 7rem;
        max-width: 1380px;
    }

    h1, h2, h3 {
        letter-spacing: 0;
    }

    .audit-hero {
        background: var(--audit-panel);
        border: 1px solid var(--audit-border);
        border-radius: 8px;
        padding: 22px 24px;
        margin: 0 0 14px 0;
        box-shadow: 0 10px 30px rgba(16, 24, 40, 0.05);
    }

    .audit-brand-hero {
        display: grid;
        grid-template-columns: minmax(0, 1fr) minmax(300px, 380px);
        gap: 26px;
        align-items: center;
    }

    .audit-kicker {
        color: var(--audit-accent);
        font-size: 12px;
        font-weight: 700;
        text-transform: uppercase;
        letter-spacing: 0;
        margin-bottom: 8px;
    }

    .audit-title {
        color: var(--audit-text);
        font-size: 34px;
        font-weight: 760;
        line-height: 1.12;
        margin: 0 0 8px 0;
    }

    .audit-subtitle {
        color: var(--audit-muted);
        font-size: 15px;
        line-height: 1.55;
        max-width: 900px;
        margin: 0;
    }

    .audit-start-grid {
        display: grid;
        grid-template-columns: repeat(3, minmax(0, 1fr));
        gap: 10px;
        margin-top: 18px;
        max-width: 920px;
    }

    .audit-start-card {
        background: #f8fafc;
        border: 1px solid var(--audit-border);
        border-radius: 8px;
        padding: 12px 13px;
        min-height: 86px;
    }

    .audit-start-card strong {
        display: block;
        color: var(--audit-text);
        font-size: 13px;
        margin-bottom: 5px;
    }

    .audit-start-card span {
        color: var(--audit-muted);
        display: block;
        font-size: 12px;
        line-height: 1.45;
    }

    .audit-logo-lockup {
        border-left: 1px solid var(--audit-border);
        padding-left: 24px;
        text-align: center;
    }

    .audit-logo-lockup img {
        max-width: 320px;
        max-height: 120px;
        width: auto;
        height: auto;
        object-fit: contain;
    }

    .audit-logo-lockup .brand-name {
        color: var(--audit-text);
        font-size: 13px;
        font-weight: 760;
        margin-top: 10px;
    }

    .audit-logo-lockup .brand-signature {
        color: var(--audit-accent);
        font-size: 12px;
        font-weight: 760;
        margin-top: 3px;
    }

    .audit-section {
        background: var(--audit-panel);
        border: 1px solid var(--audit-border);
        border-radius: 8px;
        padding: 16px 18px;
        margin: 24px 0 14px 0;
    }

    .audit-anchor {
        display: block;
        position: relative;
        top: -84px;
        visibility: hidden;
    }

    .audit-section .eyebrow {
        color: var(--audit-muted);
        font-size: 12px;
        font-weight: 700;
        text-transform: uppercase;
        margin-bottom: 4px;
    }

    .audit-section .title {
        color: var(--audit-text);
        font-size: 22px;
        font-weight: 760;
        margin-bottom: 4px;
    }

    .audit-section .body {
        color: var(--audit-muted);
        font-size: 14px;
        margin: 0;
    }

    .metric-card {
        background: var(--audit-panel);
        border: 1px solid var(--audit-border);
        border-radius: 8px;
        padding: 14px 14px 12px 14px;
        min-height: 96px;
    }

    .metric-card .label {
        color: var(--audit-muted);
        font-size: 12px;
        font-weight: 700;
        text-transform: uppercase;
        margin-bottom: 8px;
    }

    .metric-card .value {
        color: var(--audit-text);
        font-size: 26px;
        font-weight: 760;
        line-height: 1.05;
    }

    .metric-card .hint {
        color: var(--audit-muted);
        font-size: 12px;
        margin-top: 6px;
    }

    .domain-row {
        display: grid;
        grid-template-columns: repeat(5, minmax(0, 1fr));
        gap: 10px;
        margin: 12px 0 18px 0;
    }

    .domain-card {
        background: var(--audit-panel);
        border: 1px solid var(--audit-border);
        border-radius: 8px;
        padding: 12px;
        min-height: 86px;
    }

    .domain-card strong {
        display: block;
        color: var(--audit-text);
        font-size: 13px;
        margin-bottom: 10px;
    }

    .domain-score {
        color: var(--audit-accent);
        font-size: 24px;
        font-weight: 760;
    }

    .risk-chip {
        border-radius: 8px;
        padding: 10px 12px;
        border: 1px solid var(--audit-border);
        background: var(--audit-panel);
        color: var(--audit-ink);
        margin-bottom: 8px;
        font-size: 13px;
        line-height: 1.45;
    }

    .risk-chip.critical {
        background: var(--audit-risk-soft);
        border-color: #fecdca;
    }

    .risk-chip.warn {
        background: var(--audit-warn-soft);
        border-color: #fed7aa;
    }

    .analysis-teaser {
        background: var(--audit-panel);
        border: 1px solid var(--audit-border);
        border-radius: 8px;
        padding: 16px 18px;
        margin: 20px 0 12px 0;
        box-shadow: 0 10px 26px rgba(16, 24, 40, 0.04);
    }

    .analysis-teaser-head {
        display: flex;
        align-items: flex-start;
        justify-content: space-between;
        gap: 18px;
        margin-bottom: 14px;
    }

    .analysis-teaser-title {
        color: var(--audit-text);
        font-size: 19px;
        font-weight: 760;
        line-height: 1.25;
        margin-bottom: 4px;
    }

    .analysis-teaser-copy {
        color: var(--audit-muted);
        font-size: 13px;
        line-height: 1.45;
        max-width: 760px;
    }

    .analysis-pill {
        background: var(--audit-accent-soft);
        color: #0f5f59;
        border: 1px solid #b9e5de;
        border-radius: 999px;
        padding: 7px 10px;
        font-size: 12px;
        font-weight: 760;
        white-space: nowrap;
    }

    .teaser-grid {
        display: grid;
        grid-template-columns: repeat(3, minmax(0, 1fr));
        gap: 10px;
        margin-bottom: 14px;
    }

    .teaser-card {
        background: #f8fafc;
        border: 1px solid var(--audit-border);
        border-radius: 8px;
        padding: 12px;
        min-height: 86px;
    }

    .teaser-card .label {
        color: var(--audit-muted);
        font-size: 12px;
        font-weight: 700;
        margin-bottom: 7px;
    }

    .teaser-card .value {
        color: var(--audit-text);
        font-size: 23px;
        font-weight: 760;
        line-height: 1.05;
    }

    .teaser-card .hint {
        color: var(--audit-muted);
        font-size: 12px;
        margin-top: 6px;
    }

    .teaser-columns {
        display: grid;
        grid-template-columns: 1.1fr 0.9fr;
        gap: 12px;
    }

    .teaser-columns .label {
        color: var(--audit-muted);
        font-size: 12px;
        font-weight: 700;
        margin-bottom: 8px;
    }

    .mini-signal {
        background: #ffffff;
        border: 1px solid var(--audit-border);
        border-radius: 8px;
        padding: 10px 12px;
        margin-bottom: 8px;
        color: var(--audit-ink);
        font-size: 13px;
        line-height: 1.45;
    }

    .mini-signal.critical {
        border-color: #fecdca;
        background: var(--audit-risk-soft);
    }

    .mini-signal.warn {
        border-color: #fed7aa;
        background: var(--audit-warn-soft);
    }

    .mini-domain {
        display: flex;
        justify-content: space-between;
        gap: 10px;
        border-bottom: 1px solid #eef2f6;
        padding: 9px 0;
        color: var(--audit-ink);
        font-size: 13px;
    }

    .mini-domain:last-child {
        border-bottom: 0;
    }

    .mini-domain strong {
        color: var(--audit-text);
    }

    .analysis-teaser-note {
        color: var(--audit-muted);
        font-size: 12px;
        margin-top: 10px;
    }

    .analysis-status-panel {
        background: var(--audit-panel);
        border: 1px solid var(--audit-border);
        border-left: 4px solid var(--audit-accent);
        border-radius: 8px;
        padding: 16px 18px;
        margin-bottom: 16px;
        color: var(--audit-ink);
    }

    .analysis-status-title {
        color: var(--audit-text);
        font-size: 17px;
        font-weight: 760;
        margin-bottom: 6px;
    }

    .page-lock-note {
        color: var(--audit-risk);
        font-size: 13px;
        font-weight: 700;
        margin-top: 8px;
    }

    .analysis-log {
        background: #f8fafc;
        color: #344054;
        font-family: ui-monospace, SFMono-Regular, Menlo, Consolas, monospace;
        padding: 13px 14px;
        border: 1px solid var(--audit-border);
        min-height: 110px;
        border-radius: 8px;
        margin-bottom: 16px;
        font-size: 13px;
        line-height: 1.55;
    }

    .facts-panel {
        background: #fff8ed;
        border: 1px solid #fed7aa;
        border-left: 4px solid var(--audit-warn);
        border-radius: 8px;
        color: #1d2939;
        padding: 16px 18px;
        margin: 12px 0 18px 0;
        line-height: 1.55;
    }

    .facts-panel strong {
        color: #111827;
    }

    .section-feedback {
        background: #fff7ed;
        border: 1px solid #fed7aa;
        border-left: 4px solid var(--audit-warn);
        border-radius: 8px;
        color: #1d2939;
        margin: 12px 0 16px 0;
        padding: 12px 14px;
    }

    .section-feedback.ok {
        background: #f0fdf9;
        border-color: #99f6e4;
        border-left-color: var(--audit-accent);
    }

    .section-feedback-title {
        color: var(--audit-text);
        font-size: 13px;
        font-weight: 760;
        margin-bottom: 6px;
    }

    .section-feedback ul {
        margin: 0;
        padding-left: 18px;
    }

    .section-feedback li {
        font-size: 13px;
        line-height: 1.45;
        margin: 3px 0;
    }

    .sidebar-step {
        display: flex;
        align-items: center;
        gap: 8px;
        padding: 7px 0;
        color: #344054;
        font-size: 13px;
    }

    .sidebar-step-link {
        color: #344054 !important;
        text-decoration: none !important;
        border-radius: 6px;
        margin: 0 -6px;
        padding: 0 6px;
    }

    .sidebar-step-link:hover {
        background: #f2f4f7;
        color: var(--audit-text) !important;
    }

    .sidebar-dot {
        width: 9px;
        height: 9px;
        border-radius: 50%;
        flex: 0 0 9px;
    }

    .sidebar-dot.green {
        background: var(--audit-accent);
    }

    .sidebar-dot.red {
        background: var(--audit-risk);
    }

    .sidebar-dot.gray {
        background: #98a2b3;
    }

    .sidebar-error-link {
        display: block;
        color: #7a2e0e !important;
        text-decoration: none !important;
        background: #fff7ed;
        border: 1px solid #fed7aa;
        border-radius: 8px;
        padding: 8px 9px;
        margin: 6px 0;
        font-size: 12px;
        line-height: 1.35;
    }

    .sidebar-error-link:hover {
        border-color: var(--audit-warn);
        background: #ffedd5;
    }

    .sidebar-error-link span {
        display: block;
        color: #9a3412;
        font-weight: 760;
        margin-bottom: 2px;
    }

    .validation-panel {
        background: #fff7ed;
        border: 1px solid #fed7aa;
        border-left: 4px solid var(--audit-warn);
        border-radius: 8px;
        padding: 14px 16px;
        margin: 12px 0 18px 0;
    }

    .validation-panel-title {
        color: var(--audit-text);
        font-size: 16px;
        font-weight: 760;
        margin-bottom: 4px;
    }

    .validation-panel-copy {
        color: #7a2e0e;
        font-size: 13px;
        line-height: 1.45;
        margin-bottom: 12px;
    }

    .validation-error-grid {
        display: grid;
        grid-template-columns: repeat(2, minmax(0, 1fr));
        gap: 8px;
    }

    .validation-error-link {
        color: #7a2e0e !important;
        text-decoration: none !important;
        background: #ffffff;
        border: 1px solid #fed7aa;
        border-radius: 8px;
        padding: 10px 12px;
        font-size: 13px;
        line-height: 1.4;
    }

    .validation-error-link:hover {
        border-color: var(--audit-warn);
        background: #fffbeb;
    }

    .validation-error-link strong {
        display: block;
        color: #9a3412;
        font-size: 12px;
        margin-bottom: 3px;
    }


    .st-key-floating_draft_save {
        position: fixed;
        left: 50%;
        bottom: 18px;
        transform: translateX(-50%);
        z-index: 9999;
        width: min(560px, calc(100vw - 32px));
        padding: 0;
        background: transparent;
        border: 0;
        box-shadow: none;
    }

    .st-key-floating_draft_save .element-container,
    .st-key-floating_draft_save iframe {
        display: block !important;
        width: 100% !important;
        min-width: 0 !important;
        height: 96px !important;
        min-height: 96px !important;
        margin: 0 !important;
    }

    .st-key-presentation_download,
    div[data-testid="stElementContainer"]:has(.st-key-presentation_download) {
        width: min(460px, 100%) !important;
        max-width: 460px !important;
        margin: 18px auto 10px auto !important;
        box-sizing: border-box;
    }

    .st-key-presentation_download [data-testid="stDownloadButton"] {
        width: 100% !important;
        max-width: 100% !important;
    }

    .st-key-presentation_download [data-testid="stDownloadButton"] button {
        width: 100% !important;
        max-width: 100% !important;
        box-sizing: border-box;
        min-height: 58px;
        background: var(--audit-accent);
        color: #ffffff;
        border: 1px solid var(--audit-accent);
        border-radius: 8px;
        font-size: 16px;
        font-weight: 760;
        box-shadow: 0 12px 28px rgba(15, 118, 110, 0.24);
        transition: background 160ms ease, color 160ms ease, box-shadow 160ms ease, transform 160ms ease;
    }

    .st-key-presentation_download [data-testid="stDownloadButton"] button:hover {
        background: #0b5f59;
        color: #ffffff;
        border-color: #0b5f59;
        box-shadow: 0 15px 34px rgba(15, 118, 110, 0.30);
        transform: translateY(-2px);
    }

    .st-key-presentation_download [data-testid="stDownloadButton"] button:active {
        transform: translateY(0);
        box-shadow: 0 7px 16px rgba(15, 118, 110, 0.20);
    }

    .st-key-presentation_download [data-testid="stDownloadButton"] button:focus-visible {
        outline: 3px solid rgba(15, 118, 110, 0.24);
        outline-offset: 3px;
    }

    .st-key-presentation_download [data-testid="stDownloadButton"] button p {
        display: flex;
        align-items: center;
        justify-content: center;
        gap: 10px;
    }

    .st-key-presentation_download [data-testid="stDownloadButton"] button p::before {
        content: "↓";
        display: inline-flex;
        align-items: center;
        justify-content: center;
        width: 30px;
        height: 30px;
        border-radius: 50%;
        background: rgba(255, 255, 255, 0.16);
        font-size: 18px;
        font-weight: 700;
        line-height: 1;
    }

    .st-key-presentation_generate,
    div[data-testid="stElementContainer"]:has(.st-key-presentation_generate) {
        width: min(460px, 100%) !important;
        max-width: 460px !important;
        margin: 22px auto 18px auto !important;
        box-sizing: border-box;
    }

    .st-key-presentation_generate [data-testid="stButton"] {
        width: 100% !important;
        max-width: 100% !important;
    }

    .st-key-presentation_generate [data-testid="stButton"] button {
        width: 100% !important;
        max-width: 100% !important;
        min-height: 58px;
        box-sizing: border-box;
        background: var(--audit-accent);
        color: #ffffff;
        border: 1px solid var(--audit-accent);
        border-radius: 8px;
        font-size: 16px;
        font-weight: 760;
        box-shadow: 0 12px 28px rgba(15, 118, 110, 0.24);
        transition: background 160ms ease, border-color 160ms ease, box-shadow 160ms ease, transform 160ms ease;
    }

    .st-key-presentation_generate [data-testid="stButton"] button:hover {
        background: #0b5f59;
        color: #ffffff;
        border-color: #0b5f59;
        box-shadow: 0 15px 34px rgba(15, 118, 110, 0.30);
        transform: translateY(-2px);
    }

    .st-key-presentation_generate [data-testid="stButton"] button:active {
        transform: translateY(0);
        box-shadow: 0 7px 16px rgba(15, 118, 110, 0.20);
    }

    .st-key-presentation_generate [data-testid="stButton"] button:focus-visible {
        outline: 3px solid rgba(15, 118, 110, 0.24);
        outline-offset: 3px;
    }

    .st-key-presentation_generate [data-testid="stButton"] button:disabled {
        background: #d8e0e7;
        color: #6b7280;
        border-color: #d8e0e7;
        box-shadow: none;
        transform: none;
    }

    .st-key-presentation_generate [data-testid="stButton"] button p {
        display: flex;
        align-items: center;
        justify-content: center;
        gap: 10px;
    }

    .st-key-presentation_generate [data-testid="stButton"] button p::before {
        content: "→";
        display: inline-flex;
        align-items: center;
        justify-content: center;
        width: 30px;
        height: 30px;
        border-radius: 50%;
        background: rgba(255, 255, 255, 0.16);
        font-size: 18px;
        line-height: 1;
    }

    @media (max-width: 900px) {
        .audit-title {
            font-size: 26px;
        }

        .audit-brand-hero {
            grid-template-columns: 1fr;
        }

        .audit-logo-lockup {
            border-left: 0;
            border-top: 1px solid var(--audit-border);
            padding-left: 0;
            padding-top: 14px;
            text-align: center;
        }

        .audit-logo-lockup img {
            max-width: min(100%, 330px);
            max-height: 126px;
        }

        .audit-start-grid,
        .teaser-grid,
        .teaser-columns {
            grid-template-columns: 1fr;
        }

        .analysis-teaser-head {
            display: block;
        }

        .analysis-pill {
            display: inline-block;
            margin-top: 10px;
        }

        .domain-row {
            grid-template-columns: 1fr;
        }

        .validation-error-grid {
            grid-template-columns: 1fr;
        }

        .st-key-floating_draft_save {
            bottom: 12px;
            width: calc(100vw - 24px);
        }
    }
    </style>
    """, unsafe_allow_html=True)


def get_logo_data_uri(path="logo.png"):
    if not os.path.exists(path):
        return ""

    with open(path, "rb") as logo_file:
        encoded = base64.b64encode(logo_file.read()).decode("ascii")

    return f"data:image/png;base64,{encoded}"


def render_app_header():
    logo_uri = get_logo_data_uri()
    logo_html = (
        f'<img src="{logo_uri}" alt="Khalil Trade">'
        if logo_uri
        else '<strong>Khalil Trade</strong>'
    )

    st.markdown(f"""
    <div class="audit-hero audit-brand-hero">
        <div>
            <div class="audit-kicker">Технический ИТ- и ИБ-аудит</div>
            <div class="audit-title">Опросник технического аудита ИТ и ИБ</div>
            <p class="audit-subtitle">
                Структурированный сбор данных для экспертной оценки инфраструктуры,
                зрелости защиты и подготовки экспертного отчета и презентации.
            </p>
            <div class="audit-start-grid">
                <div class="audit-start-card">
                    <strong>7-10 минут на заполнение</strong>
                    <span>Начните с компании и конечных точек, остальные блоки включайте по факту наличия.</span>
                </div>
                <div class="audit-start-card">
                    <strong>Живая оценка по ходу анкеты</strong>
                    <span>Навигатор показывает готовность, слабые места и разделы, которые требуют внимания.</span>
                </div>
                <div class="audit-start-card">
                    <strong>Отчет и презентация для обсуждения</strong>
                    <span>На выходе формируется краткая управленческая презентация с выводами, решениями и планом действий.</span>
                </div>
            </div>
        </div>
        <div class="audit-logo-lockup">
            {logo_html}
            <div class="brand-name">Khalil Audit System {APP_VERSION}</div>
            <div class="brand-signature">by Ivan Rudoy</div>
        </div>
    </div>
    """, unsafe_allow_html=True)


def render_anchor(anchor_id):
    if anchor_id:
        st.markdown(
            f'<span id="{html.escape(anchor_id)}" class="audit-anchor"></span>',
            unsafe_allow_html=True
        )


def render_section_marker(kicker, title, body, anchor_id=None):
    render_anchor(anchor_id)
    st.markdown(f"""
    <div class="audit-section">
        <div class="eyebrow">{html.escape(kicker)}</div>
        <div class="title">{html.escape(title)}</div>
        <p class="body">{html.escape(body)}</p>
    </div>
    """, unsafe_allow_html=True)


def get_section_errors(validation_errors, *markers):
    selected = []
    marker_values = [str(marker).lower() for marker in markers]

    for error in validation_errors:
        error_text = str(error)
        normalized = error_text.lower()
        if any(marker in normalized for marker in marker_values):
            selected.append(error_text)

    return list(dict.fromkeys(selected))


def get_error_anchor(error):
    normalized = str(error).lower()
    marker_map = [
        (("общая информация", "обязательные поля"), "section-company"),
        (("сайт", "email", "телефон"), "section-company"),
        (("арм", "ос на арм"), "section-endpoints"),
        (("маршрутизац", "маршрутизатор", "коммутатор", "core", "распредел", "wi-fi", "ngfw"), "section-network"),
        (("сервер", "хост", "резервного копирования"), "section-servers"),
        (("схд", "raid"), "section-storage"),
        (("название и версию", "версию"), "section-systems"),
        (("производитель", "иб"), "section-security"),
        (("разработчик", "языки разработки"), "section-dev"),
    ]

    for markers, anchor_id in marker_map:
        if any(marker in normalized for marker in markers):
            return anchor_id

    return "section-company"


def render_validation_summary(validation_errors):
    unique_errors = list(dict.fromkeys(validation_errors))
    if not unique_errors:
        return

    items = "".join(
        f'<a class="validation-error-link" href="#{html.escape(get_error_anchor(error))}">'
        f'<strong>Исправить</strong>{html.escape(error)}'
        f'</a>'
        for error in unique_errors
    )

    st.markdown(f"""
    <div class="validation-panel">
        <div class="validation-panel-title">Отчет пока нельзя сформировать</div>
        <div class="validation-panel-copy">
            Осталось исправить {len(unique_errors)} пункт(ов). Нажмите на карточку, чтобы перейти к нужному разделу анкеты.
        </div>
        <div class="validation-error-grid">{items}</div>
    </div>
    """, unsafe_allow_html=True)


def render_section_feedback(title, errors, enabled=True):
    if not enabled:
        return

    if errors:
        items = "".join(
            f"<li>{html.escape(error)}</li>"
            for error in errors
        )
        st.markdown(f"""
        <div class="section-feedback">
            <div class="section-feedback-title">{html.escape(title)}: что исправить</div>
            <ul>{items}</ul>
        </div>
        """, unsafe_allow_html=True)
    else:
        st.markdown(f"""
        <div class="section-feedback ok">
            <div class="section-feedback-title">{html.escape(title)}: раздел выглядит корректно</div>
            <ul><li>Можно переходить к следующему блоку или уточнить детали в примечании.</li></ul>
        </div>
        """, unsafe_allow_html=True)


def draft_safe_value(value):
    if isinstance(value, (str, int, float, bool)) or value is None:
        return value

    if isinstance(value, tuple):
        return [draft_safe_value(item) for item in value]

    if isinstance(value, list):
        return [draft_safe_value(item) for item in value]

    if isinstance(value, dict):
        return {
            str(key): draft_safe_value(item)
            for key, item in value.items()
            if draft_safe_value(item) is not None
        }

    return None


DRAFT_SYSTEM_KEYS = {
    "draft_upload",
    "draft_expander_download",
    "floating_draft_download",
    "arm_clear_os_counts",
    "srv_fill_single_os",
    "srv_fill_remaining",
    "srv_clear_os_counts",
    "cached_report_bytes",
    "report_ready",
    "generation_active",
    "draft_link_ready",
    "draft_link_notice",
    "_draft_query_marker",
    "_arm_os_selection_mode",
}

DRAFT_FORBIDDEN_WIDGET_KEYS = {
    "draft_expander_download",
    "floating_draft_download",
    "arm_clear_os_counts",
    "srv_fill_single_os",
    "srv_fill_remaining",
    "srv_clear_os_counts",
}

SECURITY_CHECKBOX_KEYS = {
    "epp", "edr", "xdr", "mdr",
    "dlp", "mail_sec", "casb",
    "waf", "ddos", "nad", "ids", "nac", "ztna",
    "sast", "dast",
    "iam", "mfa", "pam",
    "siem", "soar",
    "vuln", "patch",
}


def coerce_draft_bool(value):
    if isinstance(value, bool):
        return value

    if isinstance(value, (int, float)):
        return value != 0

    if isinstance(value, str):
        return value.strip().lower() in {"true", "1", "yes", "y", "да", "вкл", "включено"}

    return False


def is_draft_system_key(key):
    key_text = str(key)
    action_markers = (
        "_clear_",
        "_fill_",
        "_download",
        "_generate",
        "_apply",
        "_button",
    )
    return (
        key_text in DRAFT_SYSTEM_KEYS
        or any(marker in key_text for marker in action_markers)
        or key_text.startswith("FormSubmitter:")
        or key_text.startswith("floating_")
    )


def is_forbidden_widget_state_key(key):
    key_text = str(key)
    return (
        key_text in DRAFT_FORBIDDEN_WIDGET_KEYS
        or key_text.startswith("FormSubmitter:")
    )


def clear_forbidden_widget_state():
    for key in list(st.session_state.keys()):
        if is_forbidden_widget_state_key(key):
            st.session_state.pop(key, None)


def collect_draft_state():
    draft_state = {}

    for key, value in st.session_state.items():
        if is_draft_system_key(key):
            continue

        safe_value = draft_safe_value(value)
        if safe_value is not None:
            draft_state[str(key)] = safe_value

    return {
        "schema": "khalil-audit-draft-v1",
        "saved_at": datetime.now().isoformat(timespec="seconds"),
        "state": draft_state,
    }


def build_draft_download():
    draft_payload = collect_draft_state()
    draft_bytes = json.dumps(
        draft_payload,
        ensure_ascii=False,
        indent=2,
    ).encode("utf-8")
    draft_filename = f"khalil_audit_draft_{datetime.now():%Y%m%d_%H%M}.json"

    return draft_bytes, draft_filename


def encode_draft_token(payload):
    raw = json.dumps(
        payload,
        ensure_ascii=False,
        separators=(",", ":"),
    ).encode("utf-8")
    compressed = zlib.compress(raw, level=9)
    return base64.urlsafe_b64encode(compressed).decode("ascii").rstrip("=")


def decode_draft_token(token):
    padding = "=" * (-len(token) % 4)
    compressed = base64.urlsafe_b64decode((token + padding).encode("ascii"))
    raw = zlib.decompress(compressed)
    return json.loads(raw.decode("utf-8"))


def normalize_draft_selectbox_value(key, value):
    if key not in DRAFT_SELECTBOX_OPTIONS:
        return value, None

    aliases = {
        "client_industry_select": {
            "Финансы / Банки": "Финтех / Банки",
            "Банки": "Финтех / Банки",
            "Медицина": "Здравоохранение / Медицинская организация",
            "Медицинское учреждение": "Здравоохранение / Медицинская организация",
            "Критическая инфраструктура": "КВОИКИ / Критическая инфраструктура",
            "Квазигоссектор": "Квазигосударственный сектор",
        },
        "main_net_type": {
            "Оптоволокно": "Оптика",
            "Радиоканал": "Радиорелейная",
        },
        "back_net_type": {
            "Оптоволокно": "Оптика",
            "Радиоканал": "Радиорелейная",
        },
        "web_hosting": {
            "ЦОД KZ": "Собственный ЦОД",
            "Собственный дата-центр": "Собственный ЦОД",
        },
    }

    if key == "client_phone_code" and isinstance(value, list):
        value = tuple(value)

    value = aliases.get(key, {}).get(value, value)
    options = DRAFT_SELECTBOX_OPTIONS.get(key)
    if options is None or value in options:
        return value, None

    if key == "client_industry_select" and str(value or "").strip():
        return "Другое", str(value).strip()

    defaults = {
        "client_phone_code": COUNTRY_CODE_OPTIONS[0],
        "main_net_type": "Нет",
        "back_net_type": "Нет",
        "wf_type_sel": WIFI_TYPE_OPTIONS[0],
        "mail_system": "Нет",
        "web_hosting": WEB_HOSTING_OPTIONS[0],
    }
    return defaults.get(key, options[0]), None


def apply_draft_state(payload):
    state = payload.get("state", payload)
    if not isinstance(state, dict):
        raise ValueError("Файл черновика не содержит блок state.")

    applied = 0

    normalized_state = dict(state)
    for key, value in state.items():
        if is_draft_system_key(key):
            continue
        value, custom_industry = normalize_draft_selectbox_value(key, value)
        if custom_industry:
            normalized_state["client_industry_other"] = custom_industry
        normalized_state[key] = value

    for key, value in normalized_state.items():
        if is_draft_system_key(key):
            continue
        if key in SECURITY_CHECKBOX_KEYS:
            value = coerce_draft_bool(value)
        st.session_state[key] = value
        applied += 1

    return applied


def restore_draft_from_query():
    token = st.query_params.get("draft")
    if not token:
        return

    if isinstance(token, list):
        token = token[0] if token else ""

    if not token:
        return

    marker = f"draft:{token[:24]}:{len(token)}"
    if st.session_state.get("_draft_query_marker") == marker:
        return

    try:
        payload = decode_draft_token(token)
        applied = apply_draft_state(payload)
        st.session_state["_draft_query_marker"] = marker
        st.session_state["draft_notice"] = (
            f"Черновик из ссылки применен: восстановлено полей {applied}."
        )
        st.query_params.clear()
        st.rerun()
    except Exception as exc:
        st.error(f"Не удалось применить черновик из ссылки: {exc}")


def render_floating_draft_save():
    draft_bytes, draft_filename = build_draft_download()
    draft_token = encode_draft_token(collect_draft_state())
    draft_text_json = json.dumps(draft_bytes.decode("utf-8"))
    draft_filename_json = json.dumps(draft_filename)
    draft_token_json = json.dumps(draft_token)
    can_share_json = "true" if len(draft_token) <= 6000 else "false"

    with st.container(key="floating_draft_save"):
        components.html(f"""
        <style>
        * {{
            box-sizing: border-box;
        }}

        .floating-draft-panel {{
            width: 100%;
            padding: 12px 14px 10px 14px;
            background: rgba(255, 255, 255, 0.97);
            border: 1px solid rgba(208, 213, 221, 0.92);
            border-radius: 22px;
            box-shadow: 0 18px 44px rgba(16, 24, 40, 0.16);
            backdrop-filter: blur(10px);
            font-family: "Inter", "Segoe UI", sans-serif;
        }}

        .floating-draft-actions {{
            display: grid;
            grid-template-columns: 1.35fr 1fr;
            gap: 12px;
            align-items: center;
        }}

        .floating-draft-action {{
            width: 100%;
            height: 52px;
            border-radius: 999px;
            border: 1px solid transparent;
            color: #ffffff;
            cursor: pointer;
            display: flex;
            align-items: center;
            justify-content: center;
            font-size: 16px;
            font-weight: 760;
            line-height: 1;
            margin: 0;
            padding: 0 18px;
            white-space: nowrap;
        }}

        .floating-draft-action.save {{
            background: #f97316;
            border-color: #ea580c;
            box-shadow: 0 12px 28px rgba(234, 88, 12, 0.22);
        }}

        .floating-draft-action.save:hover {{
            background: #ea580c;
            border-color: #c2410c;
        }}

        .floating-draft-action.share {{
            background: #0f766e;
            border-color: #0f766e;
            box-shadow: 0 12px 28px rgba(15, 118, 110, 0.18);
        }}

        .floating-draft-action.share:hover {{
            background: #0d5f59;
            border-color: #0d5f59;
        }}

        .floating-draft-action:disabled {{
            background: #98a2b3;
            border-color: #98a2b3;
            cursor: not-allowed;
            box-shadow: none;
        }}

        .floating-draft-hint {{
            color: #475467;
            font-size: 11px;
            line-height: 1.25;
            margin-top: 8px;
            text-align: center;
            text-shadow: 0 1px 2px rgba(255, 255, 255, 0.85);
        }}

        @media (max-width: 460px) {{
            .floating-draft-panel {{
                padding: 10px 10px 8px 10px;
                border-radius: 18px;
            }}

            .floating-draft-actions {{
                gap: 8px;
            }}

            .floating-draft-action {{
                height: 46px;
                font-size: 13px;
                padding: 0 10px;
            }}
        }}
        </style>
        <div class="floating-draft-panel">
            <div class="floating-draft-actions">
                <button class="floating-draft-action save" id="floating-save" type="button">
                    Сохранить черновик
                </button>
                <button class="floating-draft-action share" id="floating-share" type="button">
                    Поделиться
                </button>
            </div>
            <div class="floating-draft-hint" id="floating-draft-status">
                Сохраните JSON-файл или поделитесь ссылкой на заполненную анкету.
            </div>
        </div>
        <script>
        const draftText = {draft_text_json};
        const draftFilename = {draft_filename_json};
        const draftToken = {draft_token_json};
        const canShareDraft = {can_share_json};
        const saveButton = document.getElementById("floating-save");
        const shareButton = document.getElementById("floating-share");
        const status = document.getElementById("floating-draft-status");

        function getParentUrl() {{
            try {{
                return window.parent.location.href;
            }} catch (error) {{
                return document.referrer || window.location.href;
            }}
        }}

        function buildShareUrl() {{
            const url = new URL(getParentUrl());
            url.searchParams.set("draft", draftToken);
            return url.toString();
        }}

        function saveDraftFile() {{
            const blob = new Blob([draftText], {{
                type: "application/json;charset=utf-8"
            }});
            const href = URL.createObjectURL(blob);
            const link = document.createElement("a");
            link.href = href;
            link.download = draftFilename;
            document.body.appendChild(link);
            link.click();
            link.remove();
            window.setTimeout(() => URL.revokeObjectURL(href), 1000);
            status.textContent = "JSON-черновик скачан. Его можно загрузить в блоке «Черновик анкеты».";
        }}

        async function copyShareUrl(shareUrl) {{
            if (navigator.clipboard && window.isSecureContext) {{
                await navigator.clipboard.writeText(shareUrl);
                status.textContent = "Ссылка скопирована. Передайте ее коллеге.";
                return;
            }}

            window.prompt("Скопируйте ссылку на заполненную анкету", shareUrl);
            status.textContent = "Скопируйте ссылку из окна браузера.";
        }}

        async function shareDraft() {{
            if (!canShareDraft) {{
                status.textContent = "Анкета слишком большая для ссылки. Скачайте JSON-черновик.";
                return;
            }}

            const shareUrl = buildShareUrl();
            const shareData = {{
                title: "Khalil Audit System",
                text: "Заполненная анкета Khalil Audit System by Ivan Rudoy",
                url: shareUrl
            }};

            try {{
                if (navigator.share) {{
                    await navigator.share(shareData);
                    status.textContent = "Анкета готова к отправке.";
                    return;
                }}

                await copyShareUrl(shareUrl);
            }} catch (error) {{
                if (error && error.name === "AbortError") {{
                    status.textContent = "Отправка отменена.";
                    return;
                }}

                try {{
                    await copyShareUrl(shareUrl);
                }} catch (copyError) {{
                    window.prompt("Скопируйте ссылку на заполненную анкету", shareUrl);
                    status.textContent = "Скопируйте ссылку из окна браузера.";
                }}
            }}
        }}

        if (!canShareDraft) {{
            shareButton.disabled = true;
            shareButton.title = "Анкета слишком большая для ссылки";
        }}

        saveButton.addEventListener("click", saveDraftFile);
        shareButton.addEventListener("click", shareDraft);
        </script>
        """, height=96)


def render_draft_share_button(
    draft_token,
    label="Поделиться заполненной анкетой",
    status_text="Откроет меню отправки или скопирует ссылку.",
    height=82,
    compact=False,
):
    token_json = json.dumps(draft_token)
    label_json = json.dumps(label)
    status_json = json.dumps(status_text)
    status_markup = (
        '<div class="draft-share-status" id="draft-share-status"></div>'
        if status_text
        else '<div class="draft-share-status sr-only" id="draft-share-status"></div>'
    )
    border_radius = "999px" if compact else "8px"
    min_height = "42px" if compact else "38px"
    font_size = "14px" if compact else "14px"

    components.html(f"""
    <style>
    .draft-share-wrap {{
        font-family: "Inter", "Segoe UI", sans-serif;
        width: 100%;
    }}

    .draft-share-button {{
        width: 100%;
        min-height: {min_height};
        border: 1px solid #0f766e;
        border-radius: {border_radius};
        background: #0f766e;
        color: #ffffff;
        cursor: pointer;
        font-size: {font_size};
        font-weight: 650;
        line-height: 1.2;
    }}

    .draft-share-button:hover {{
        background: #0d5f59;
        border-color: #0d5f59;
    }}

    .draft-share-status {{
        color: #667085;
        font-size: 11px;
        line-height: 1.35;
        margin-top: 7px;
        text-align: center;
    }}

    .sr-only {{
        position: absolute;
        width: 1px;
        height: 1px;
        padding: 0;
        margin: -1px;
        overflow: hidden;
        clip: rect(0, 0, 0, 0);
        white-space: nowrap;
        border: 0;
    }}
    </style>
    <div class="draft-share-wrap">
        <button class="draft-share-button" id="draft-share-button" type="button">
        </button>
        {status_markup}
    </div>
    <script>
    const draftToken = {token_json};
    const buttonLabel = {label_json};
    const initialStatus = {status_json};
    const button = document.getElementById("draft-share-button");
    const status = document.getElementById("draft-share-status");
    button.textContent = buttonLabel;
    if (status) {{
        status.textContent = initialStatus;
    }}

    function getParentUrl() {{
        try {{
            return window.parent.location.href;
        }} catch (error) {{
            return document.referrer || window.location.href;
        }}
    }}

    function buildShareUrl() {{
        const url = new URL(getParentUrl());
        url.searchParams.set("draft", draftToken);
        return url.toString();
    }}

    async function copyShareUrl(shareUrl) {{
        if (navigator.clipboard && window.isSecureContext) {{
            await navigator.clipboard.writeText(shareUrl);
            status.textContent = "Ссылка скопирована. Передайте ее коллеге.";
            return;
        }}

        window.prompt("Скопируйте ссылку на заполненную анкету", shareUrl);
        status.textContent = "Скопируйте ссылку из окна браузера.";
    }}

    button.addEventListener("click", async () => {{
        const shareUrl = buildShareUrl();
        const shareData = {{
            title: "Khalil Audit System",
            text: "Заполненная анкета Khalil Audit System by Ivan Rudoy",
            url: shareUrl
        }};

        try {{
            if (navigator.share) {{
                await navigator.share(shareData);
                status.textContent = "Анкета готова к отправке.";
                return;
            }}

            await copyShareUrl(shareUrl);
        }} catch (error) {{
            if (error && error.name === "AbortError") {{
                status.textContent = "Отправка отменена.";
                return;
            }}

            try {{
                await copyShareUrl(shareUrl);
            }} catch (copyError) {{
                window.prompt("Скопируйте ссылку на заполненную анкету", shareUrl);
                status.textContent = "Скопируйте ссылку из окна браузера.";
            }}
        }}
    }});
    </script>
    """, height=height)


def render_draft_tools():
    if st.session_state.get("draft_notice"):
        st.success(st.session_state.pop("draft_notice"))

    with st.expander("Черновик анкеты: сохранить или продолжить заполнение", expanded=False):
        st.caption(
            "Сохраните JSON-черновик или создайте ссылку, если анкету нужно передать коллеге "
            "и продолжить заполнение на другом компьютере. Черновик содержит введенные ответы, "
            "поэтому передавайте его только доверенному получателю."
        )

        draft_bytes, draft_filename = build_draft_download()
        draft_token = encode_draft_token(collect_draft_state())

        col_save, col_link, col_load = st.columns(3)
        with col_save:
            st.download_button(
                "Скачать черновик JSON",
                data=draft_bytes,
                file_name=draft_filename,
                mime="application/json",
                use_container_width=True,
                key="draft_expander_download",
            )

        with col_link:
            if len(draft_token) > 6000:
                st.warning(
                    "Анкета уже слишком большая для надежной ссылки. "
                    "Скачайте JSON-черновик и передайте файл коллеге."
                )
            else:
                render_draft_share_button(draft_token)

        with col_load:
            uploaded_draft = st.file_uploader(
                "Загрузить черновик JSON",
                type=["json"],
                key="draft_upload",
                help="После применения страница перезагрузится и подставит сохраненные ответы.",
            )
            if uploaded_draft and st.button("Применить черновик", use_container_width=True):
                try:
                    payload = json.loads(uploaded_draft.getvalue().decode("utf-8-sig"))
                    applied = apply_draft_state(payload)
                    st.session_state["draft_notice"] = f"Черновик применен: восстановлено полей {applied}."
                    st.rerun()
                except Exception as exc:
                    st.error(f"Не удалось применить черновик: {exc}")


def render_generation_guard(active):
    action = "install" if active else "remove"
    components.html(f"""
    <script>
    const target = window.parent;
    const message = "Экспертная презентация формируется. Это может занять до 4 минут. Не закрывайте и не обновляйте страницу.";

    if (!target.__khalilBeforeUnloadHandler) {{
      target.__khalilBeforeUnloadHandler = (event) => {{
        event.preventDefault();
        event.returnValue = message;
        return message;
      }};
    }}

    if ("{action}" === "install" && !target.__khalilBeforeUnloadActive) {{
      target.addEventListener("beforeunload", target.__khalilBeforeUnloadHandler);
      target.__khalilBeforeUnloadActive = true;
    }}

    if ("{action}" === "remove" && target.__khalilBeforeUnloadActive) {{
      target.removeEventListener("beforeunload", target.__khalilBeforeUnloadHandler);
      target.__khalilBeforeUnloadActive = false;
    }}
    </script>
    """, height=0)


def render_generation_live_panel(stage_title, active_step=0):
    steps = [
        "Нормализация анкеты",
        "Сопоставление масштаба инфраструктуры",
        "Расчет зрелости ИТ/ИБ",
        "Формирование экспертного заключения",
        "Сборка презентации и плана действий",
        "Отправка результата",
    ]
    step_items = []
    for idx, step in enumerate(steps):
        state_class = "done" if idx < active_step else "active" if idx == active_step else ""
        step_items.append(f'<li class="{state_class}">{html.escape(step)}</li>')

    components.html(f"""
    <style>
    .gen-panel {{
        font-family: Inter, system-ui, -apple-system, BlinkMacSystemFont, "Segoe UI", sans-serif;
        border: 1px solid #c7d7fe;
        border-left: 4px solid #13877c;
        border-radius: 10px;
        background: #ffffff;
        padding: 16px 18px;
        color: #101828;
        box-shadow: 0 10px 30px rgba(16, 24, 40, 0.08);
    }}
    .gen-head {{
        display: flex;
        align-items: center;
        gap: 12px;
        margin-bottom: 10px;
    }}
    .gen-pulse {{
        width: 13px;
        height: 13px;
        border-radius: 50%;
        background: #13877c;
        box-shadow: 0 0 0 rgba(19, 135, 124, 0.5);
        animation: pulse 1.4s infinite;
        flex: 0 0 13px;
    }}
    @keyframes pulse {{
        0% {{ box-shadow: 0 0 0 0 rgba(19, 135, 124, 0.5); }}
        70% {{ box-shadow: 0 0 0 12px rgba(19, 135, 124, 0); }}
        100% {{ box-shadow: 0 0 0 0 rgba(19, 135, 124, 0); }}
    }}
    .gen-title {{
        font-size: 17px;
        font-weight: 760;
    }}
    .gen-sub {{
        color: #475467;
        font-size: 13px;
        line-height: 1.45;
        margin-bottom: 12px;
    }}
    .gen-time {{
        display: inline-block;
        color: #0f5f59;
        background: #e6fffb;
        border: 1px solid #99f6e4;
        border-radius: 999px;
        padding: 5px 9px;
        font-size: 12px;
        font-weight: 760;
        margin-bottom: 12px;
    }}
    .gen-steps {{
        display: grid;
        grid-template-columns: repeat(3, minmax(0, 1fr));
        gap: 8px;
        margin: 0;
        padding: 0;
        list-style: none;
    }}
    .gen-steps li {{
        border: 1px solid #e4e7ec;
        border-radius: 8px;
        padding: 8px 10px;
        color: #667085;
        background: #f8fafc;
        font-size: 12px;
        line-height: 1.3;
    }}
    .gen-steps li.done {{
        background: #f0fdf9;
        border-color: #99f6e4;
        color: #0f5f59;
        font-weight: 700;
    }}
    .gen-steps li.active {{
        background: #fff7ed;
        border-color: #fed7aa;
        color: #9a3412;
        font-weight: 760;
    }}
    </style>
    <div class="gen-panel">
        <div class="gen-head">
            <div class="gen-pulse"></div>
            <div class="gen-title">{html.escape(stage_title)}</div>
        </div>
        <div class="gen-sub">
            Формирование может занять до 4 минут. Если счетчик ниже идет, браузер не завис: отчет собирается на сервере.
        </div>
        <div class="gen-time">Прошло: <span id="gen-elapsed">00:00</span></div>
        <ul class="gen-steps">{''.join(step_items)}</ul>
    </div>
    <script>
    const startedAt = Date.now();
    const target = document.getElementById("gen-elapsed");
    setInterval(() => {{
        const total = Math.floor((Date.now() - startedAt) / 1000);
        const minutes = String(Math.floor(total / 60)).padStart(2, "0");
        const seconds = String(total % 60).padStart(2, "0");
        if (target) target.textContent = `${{minutes}}:${{seconds}}`;
    }}, 1000);
    </script>
    """, height=275)


inject_audit_design()

# Якорь для принудительного перехода в начало страницы
st.markdown("<div id='top'></div>", unsafe_allow_html=True)
# --- НАСТРОЙКИ TELEGRAM (из Secrets) ---
TOKEN = get_app_secret("TELEGRAM_TOKEN")
CHAT_ID = get_app_secret("TELEGRAM_CHAT_ID")

restore_draft_from_query()
clear_forbidden_widget_state()

render_app_header()
render_floating_draft_save()

# --- ИНСТРУКЦИЯ ДЛЯ ПОЛЬЗОВАТЕЛЯ ---
with st.expander("Инструкция по заполнению", expanded=False):
    st.markdown("""
    ### Как пройти анкету

    1. **Начните с обязательного минимума:** заполните общую информацию о компании и блок **Конечные точки (АРМ)**. У заказчика почти всегда есть рабочие станции, поэтому этот блок не отключается и нужен для получения отчета.
    2. **Включайте только реальные разделы:** сеть, серверы, виртуализация, СХД, внутренние системы, ИБ и разработка открываются тумблерами. Если раздел включен, его вложенные обязательные поля нужно заполнить.
    3. **Смотрите навигатор слева:** красная точка означает, что раздел требует внимания, зеленая - раздел выглядит заполненным, серая - раздел отключен и не участвует в анкете.
    4. **Исправляйте по подсказкам на месте:** под каждым крупным блоком есть панель “что исправить”. Она показывает конкретные недостающие поля именно для этого раздела.
    5. **Используйте предварительную аналитику:** блок “Предварительная аналитика” показывает сводку аудита и быстрые улучшения еще до формирования финального отчета. Полную версию можно раскрыть ниже.
    6. **Сохраняйте черновик при совместном заполнении:** скачайте JSON-черновик и передайте его коллеге. Он сможет загрузить файл в этом же блоке и продолжить заполнение с сохраненных ответов.
    7. **Не закрывайте страницу во время отчета:** формирование экспертного отчета может занять до 4 минут. Пока идет генерация, не обновляйте и не закрывайте вкладку.

    Поля “Примечание” необязательны, но помогают эксперту точнее описать контекст, ограничения и планы развития инфраструктуры.
    """)

render_draft_tools()

data = {}
client_info = {}
validation_errors = []
score = 0
# --- ШАПКА: ОБЩАЯ ИНФОРМАЦИЯ ---
render_section_marker(
    "01 / КОМПАНИЯ",
    "Общая информация",
    "Контекст компании, отрасль и контактные данные для экспертного отчета.",
    anchor_id="section-company"
)
col_h1, col_h2 = st.columns(2)

with col_h1:
    client_info['Город'] = st.text_input("Город*", key="client_city", help="Укажите город фактического нахождения головного офиса.")
    industry_options = INDUSTRY_OPTIONS
    selected_ind = st.selectbox(
        "Сфера деятельности компании*",
        [""] + industry_options,
        format_func=lambda x: "Выберите сферу..." if x == "" else x,
        key="client_industry_select",
        help="Отрасль влияет на профиль угроз и регуляторные требования."
    )

    if selected_ind == "Другое":
        industry = st.text_input("Укажите вашу сферу деятельности*", key="client_industry_other", help="Введите отрасль вручную")
    else:
        industry = selected_ind

    client_info['Сфера деятельности'] = industry
    client_info['Наименование компании'] = st.text_input("Наименование компании*", key="client_company_name", help="Официальное или сокращенное название юрлица.")

    site_input = st.text_input("Сайт компании*", key="site_field", placeholder="example.kz", help="Можно указать example.kz или https://example.kz - система приведет адрес к домену.")
    clean_domain = normalize_site_domain(site_input)
    client_info['Сайт компании'] = clean_domain
    if site_input and clean_domain != site_input.strip().lower():
        st.caption(f"Будет использован домен: {clean_domain}")

    custom_email_mode = st.checkbox("Email отличается от сайта", key="client_custom_email_mode", help="Отметьте, если корпоративная почта находится на другом домене.")

    if custom_email_mode:
        custom_email = st.text_input("Email контактного лица*", key="client_email_custom", help="Личный корпоративный email для отправки результатов.")
        client_info['Email'] = normalize_email(custom_email)
    else:
        if clean_domain and "." in clean_domain:
            st.write("Email контактного лица (логин)*")
            e_col1, e_col2 = st.columns([2, 3])
            with e_col1:
                email_prefix = st.text_input("Логин", placeholder="info", label_visibility="collapsed", key="email_pre", help="Только часть адреса до символа @")
            with e_col2:
                st.markdown(f"<div style='padding-top: 5px; font-weight: bold; color: #1F4E78;'>@{clean_domain}</div>", unsafe_allow_html=True)
            client_info['Email'] = normalize_email(f"{email_prefix}@{clean_domain}") if email_prefix else ""
        else:
            client_info['Email'] = ""

with col_h2:
    client_info['ФИО контактного лица'] = st.text_input("ФИО контактного лица*", key="client_contact_name", help="С кем наш эксперт сможет обсудить детали отчета.")
    client_info['Должность'] = st.text_input("Должность*", key="client_contact_role", help="Например: ИТ-Директор, Системный администратор, CEO.")

    st.write("Контактный телефон*")
    p_col1, p_col2 = st.columns([1, 2])
    country_codes = COUNTRY_CODE_OPTIONS
    selected_code = p_col1.selectbox("Код", country_codes, format_func=lambda x: x[0], label_visibility="collapsed", key="client_phone_code")
    phone_num = p_col2.text_input("Номер", placeholder="777 777 77 77", label_visibility="collapsed", key="client_phone_number", help="Телефон для оперативной связи.")
    normalized_phone = normalize_phone_number(phone_num)
    if phone_num and normalized_phone != phone_num.strip():
        st.caption(f"Номер будет записан как: {selected_code[1]} {normalized_phone}")
    client_info['Контактный телефон'] = f"{selected_code[1]} {normalized_phone}" if phone_num else ""

if not all([client_info.get('Город'), client_info.get('Наименование компании'), client_info.get('Сфера деятельности'), client_info.get('Сайт компании'), client_info.get('Email'), client_info.get('ФИО контактного лица'), client_info.get('Должность'), phone_num]):
    validation_errors.append("Заполните все обязательные поля в блоке 'Общая информация'")
elif not is_valid_domain(client_info.get('Сайт компании')):
    validation_errors.append("Укажите корректный сайт компании")
elif not is_valid_email(client_info.get('Email')):
    validation_errors.append("Укажите корректный email контактного лица")
elif phone_num and len(re.sub(r"\D", "", phone_num)) < 7:
    validation_errors.append("Укажите корректный контактный телефон")

render_section_feedback(
    "Общая информация",
    get_section_errors(validation_errors, "общая информация", "сайт", "email", "телефон")
)

st.divider()

# --- БЛОК 1: ИНФОРМАЦИОННЫЕ ТЕХНОЛОГИИ ---
render_section_marker(
    "02 / ИНФРАСТРУКТУРА",
    "Информационные технологии",
    "АРМ, сеть, серверы, хранение данных и внутренние бизнес-системы."
)

render_anchor("section-endpoints")
st.subheader("1.1. Конечные точки (АРМ)")
total_arm = st.number_input("Общее количество АРМ (шт)", min_value=0, step=1, key="total_arm", help="Общее число ПК, ноутбуков и тонких клиентов в организации.")
data['1.1. Всего АРМ'] = total_arm
selected_os_arm = st.multiselect("ОС на АРМ", ["Windows XP/Vista/7/8", "Windows 10", "Windows 11", "Linux", "macOS", "Другое"], key="selected_os_arm", help="Выберите все типы операционных систем, используемых сотрудниками.")

sum_os_arm = 0
if selected_os_arm:
    arm_keys = [f"arm_{os_item}" for os_item in selected_os_arm]
    if len(selected_os_arm) == 1:
        only_os = selected_os_arm[0]
        st.session_state[f"arm_{only_os}"] = total_arm
        st.session_state["_arm_os_selection_mode"] = "single"
        st.caption(f"Выбрана одна ОС, поэтому количество автоматически равно общему числу АРМ: {total_arm}.")
    else:
        if st.session_state.get("_arm_os_selection_mode") == "single":
            for key in arm_keys:
                st.session_state[key] = 0
            st.session_state["_arm_os_selection_mode"] = "multi"
            st.rerun()

        st.session_state["_arm_os_selection_mode"] = "multi"

    if len(selected_os_arm) > 1 and total_arm > 0:
        if st.button("Очистить распределение ОС", use_container_width=True, key="arm_clear_os_counts"):
            for key in arm_keys:
                st.session_state[key] = 0
            st.rerun()
else:
    st.session_state["_arm_os_selection_mode"] = "empty"

for os_item in selected_os_arm:
    val = st.number_input(f"Кол-во на {os_item}", min_value=0, step=1, key=f"arm_{os_item}", help=f"Укажите точное или примерное число устройств с {os_item}.")
    data[f"ОС АРМ ({os_item})"] = val
    sum_os_arm += val

data['1.1. Примечание'] = st.text_area("Примечание к разделу 1.1", placeholder="Напр.: планируем обновление Windows 10 до 11 в Q3", key="note_1_1")

if total_arm <= 0:
    validation_errors.append("Укажите количество АРМ")
elif not selected_os_arm:
    validation_errors.append("Выберите ОС на АРМ")
elif sum_os_arm != total_arm:
    st.warning(f"⚠️ Ошибка: Сумма по ОС ({sum_os_arm}) должна быть равна общему количеству АРМ ({total_arm}).")
    validation_errors.append("Несовпадение количества АРМ и ОС")

render_section_feedback(
    "Конечные точки",
    get_section_errors(validation_errors, "арм", "ос на арм")
)

st.write("---")
render_anchor("section-network")
st.subheader("1.2. Сетевая инфраструктура")
main_speed, back_speed, ap_cnt = 0, 0, 0
selected_routing = []
ngfw_vendor = "Нет"
wifi_enabled = False
wifi_ctrl_enabled = False
net_active = st.toggle("Своя сетевая инфраструктура", key="net_toggle", help="Активируйте, если организация самостоятельно управляет сетевым оборудованием.")

if net_active:
    net_types = NETWORK_TYPE_OPTIONS
    routing_types = ["Статическая", "RIP", "OSPF", "EIGRP", "BGP", "IS-IS"]

    col_net1, col_net2 = st.columns(2)
    with col_net1:
        st.write("Основной канал")
        main_net_kwargs = {"key": "main_net_type", "help": "Технология подключения основного интернет-канала."}
        if "main_net_type" not in st.session_state:
            main_net_kwargs["index"] = 7
        main_type = st.selectbox("Тип (основной)", net_types, **main_net_kwargs)
        main_speed = st.number_input("Скорость основного (Mbit/s)", min_value=0, step=10, key="main_net_speed")
        data['1.2.1. Основной канал'] = f"{main_type} ({main_speed} Mbit/s)"
    with col_net2:
        st.write("Резервный канал")
        back_net_kwargs = {"key": "back_net_type", "help": "Наличие и тип независимого резервного канала."}
        if "back_net_type" not in st.session_state:
            back_net_kwargs["index"] = net_types.index("Нет")
        back_type = st.selectbox("Тип (резервный)", net_types, **back_net_kwargs)
        entered_back_speed = st.number_input(
            "Скорость резервного (Mbit/s)",
            min_value=0,
            step=10,
            key="back_net_speed",
            disabled=back_type == "Нет",
        )
        back_speed = 0 if back_type == "Нет" else entered_back_speed
        data['1.2.2. Резервный канал'] = (
            "Нет" if back_type == "Нет" else f"{back_type} ({back_speed} Mbit/s)"
        )

    st.write("Логика сети")
    selected_routing = st.multiselect("Тип маршрутизации*", routing_types, key="routing_sel", help="Протоколы динамической маршрутизации, используемые в сети.")
    data['1.2.3. Маршрутизация'] = ", ".join(selected_routing)
    if not selected_routing:
        validation_errors.append("Выберите тип маршрутизации")

    st.write("Активное сетевое оборудование")
    c_net1, c_net2, c_net3 = st.columns(3)
    with c_net1:
        if st.checkbox("Маршрутизаторы", key="router_chk", help="Устройства для связи разных сетей и выхода в интернет."):
            r_count = st.number_input("Кол-во маршрутизаторов", min_value=0, step=1, key="router_cnt")
            data['1.2.4. Маршрутизаторы'] = f"Да ({r_count} шт)"
            if r_count == 0: validation_errors.append("Укажите количество маршрутизаторов")
    with c_net2:
        if st.checkbox("Коммутаторы L2", key="swl2_chk", help="Управляемые или неуправляемые коммутаторы уровня доступа."):
            sw2_count = st.number_input("Кол-во коммутаторов L2", min_value=0, step=1, key="swl2_cnt")
            data['1.2.5. Коммутаторы L2'] = f"Да ({sw2_count} шт)"
            if sw2_count == 0: validation_errors.append("Укажите количество коммутаторов L2")
    with c_net3:
        if st.checkbox("Коммутаторы L3", key="swl3_chk", help="Коммутаторы с функциями маршрутизации (ядро или агрегация)."):
            sw3_count = st.number_input("Кол-во коммутаторов L3", min_value=0, step=1, key="swl3_cnt")
            data['1.2.6. Коммутаторы L3'] = f"Да ({sw3_count} шт)"
            if sw3_count == 0: validation_errors.append("Укажите количество коммутаторов L3")

    st.write("Уровни сети")
    l_col1, l_col2, l_col3 = st.columns(3)
    with l_col1:
        if st.checkbox("Ядро (Core)", key="net_core", help="Центральная часть сети, обеспечивающая максимальную скорость."):
            core_v = st.text_input("Основной производитель (Core)", key="core_vendor", help="Например: Cisco, Huawei, Juniper, MikroTik.")
            data['Уровень сети Ядро'] = core_v
            if not core_v: validation_errors.append("Укажите производителя Core-уровня")
    with l_col2:
        if st.checkbox("Уровень распределения", key="net_dist", help="Связующее звено между ядром и уровнем доступа."):
            dist_v = st.text_input("Основной производитель (Dist)", key="dist_vendor")
            data['Уровень сети Распределение'] = dist_v
            if not dist_v: validation_errors.append("Укажите производителя уровня распределения")
    with l_col3:
        if st.checkbox("Уровень доступа", key="net_acc", help="Уровень, к которому подключаются конечные пользователи."):
            acc_v = st.text_input("Основной производитель (Access)", key="acc_vendor")
            data['Уровень сети Доступ'] = acc_v
            if not acc_v: validation_errors.append("Укажите производителя уровня доступа")

    wifi_enabled = st.checkbox("Wi-Fi", key="wifi_toggle", help="Наличие корпоративной беспроводной сети.")
    if wifi_enabled:
        w_col1, w_col2, w_col3 = st.columns(3)
        with w_col1:
            wifi_ctrl_enabled = st.checkbox("Контроллер", key="wifi_ctrl", help="Централизованное управление точками доступа (аппаратное или программное).")
            if wifi_ctrl_enabled:
                wc_v = st.text_input("Производитель/модель контроллера", key="wc_vendor")
                data['Wi-Fi Контроллер'] = wc_v
                if not wc_v: validation_errors.append("Укажите модель Wi-Fi контроллера")
            else:
                data['Wi-Fi Контроллер'] = "Нет"
        with w_col2:
            ap_cnt = st.number_input("Количество точек доступа (шт)", min_value=0, step=1, key="ap_cnt", help="Общее число активных Wi-Fi точек.")
            data['Wi-Fi Точки доступа'] = ap_cnt
            if ap_cnt == 0: validation_errors.append("Укажите количество точек доступа Wi-Fi")
        with w_col3:
            wf_types = WIFI_TYPE_OPTIONS
            data['Wi-Fi Тип'] = st.selectbox("Тип Wi-Fi", wf_types, key="wf_type_sel", help="Преимущественный стандарт беспроводной связи.")

    if st.checkbox("Межсетевой экран (NGFW)", key="ngfw_chk", help="Многофункциональные шлюзы безопасности (FortiGate, UserGate, CheckPoint и т.д.)."):
        ngfw_vendor = st.text_input("Производитель (NGFW)", key="ngfw_v")
        data['1.2.7. NGFW'] = f"Да ({ngfw_vendor if ngfw_vendor else 'не указан'})"
        if not ngfw_vendor: validation_errors.append("Укажите производителя NGFW")
        score += 20

    data['1.2. Примечание'] = st.text_area("Примечание к разделу 1.2", placeholder="Особенности топологии сети...", key="note_1_2")

render_section_feedback(
    "Сетевая инфраструктура",
    get_section_errors(
        validation_errors,
        "маршрутизац",
        "маршрутизатор",
        "коммутатор",
        "core",
        "распредел",
        "доступ",
        "wi-fi",
        "ngfw"
    ),
    enabled=net_active
)

st.write("---")
render_anchor("section-servers")
st.subheader("1.3. Серверы и Виртуализация")
sum_os_srv = 0
phys_count = 0
virt_count = 0
server_active = st.toggle("Серверы и виртуализация", key="server_toggle", help="Включите, если у заказчика есть физические серверы, виртуальные машины или платформы виртуализации.")
v_n_b = "Нет"
selected_os_srv = []
selected_virt_sys = []

if server_active:
    col_s1, col_s2 = st.columns(2)
    with col_s1:
        phys_count = st.number_input("Количество физических серверов", min_value=0, step=1, key="phys_srv", help="Количество 'железных' серверов в серверной или ЦОД.")
        data['1.3.1. Физические серверы'] = phys_count
    with col_s2:
        virt_count = st.number_input("Количество виртуальных серверов", min_value=0, step=1, key="virt_srv", help="Суммарное количество виртуальных машин (VM).")
        data['1.3.2. Виртуальные серверы'] = virt_count

    if phys_count == 0 and virt_count == 0:
        validation_errors.append("Укажите количество физических или виртуальных серверов")

    s_os_list = ["Windows Server 2008/2012 R2", "Windows Server 2016", "Windows Server 2019", "Windows Server 2022", "Linux", "Unix", "Другое"]
    selected_os_srv = st.multiselect("Выберите ОС серверов", s_os_list, key="ms_srv_list", help="Операционные системы, установленные на серверах.")
    if selected_os_srv:
        srv_os_keys = [f"fsrv_{os_s}" for os_s in selected_os_srv]
        current_srv_os_total = sum(widget_int(key) for key in srv_os_keys)
        server_os_target = max(phys_count, virt_count)
        server_os_remaining = server_os_target - current_srv_os_total

        if server_os_target > 0:
            srv_quick_cols = st.columns(2)
            if len(selected_os_srv) == 1:
                only_srv_os = selected_os_srv[0]
                with srv_quick_cols[0]:
                    if st.button(f"Все {server_os_target} на {only_srv_os}", use_container_width=True, key="srv_fill_single_os"):
                        st.session_state[f"fsrv_{only_srv_os}"] = server_os_target
                        st.rerun()
            elif server_os_remaining > 0:
                target_srv_os = selected_os_srv[-1]
                with srv_quick_cols[0]:
                    if st.button(f"Добавить остаток {server_os_remaining} к {target_srv_os}", use_container_width=True, key="srv_fill_remaining"):
                        st.session_state[f"fsrv_{target_srv_os}"] = widget_int(f"fsrv_{target_srv_os}") + server_os_remaining
                        st.rerun()
            with srv_quick_cols[1]:
                if st.button("Очистить распределение серверных ОС", use_container_width=True, key="srv_clear_os_counts"):
                    for key in srv_os_keys:
                        st.session_state[key] = 0
                    st.rerun()

        for os_s in selected_os_srv:
            val_os = st.number_input(f"Кол-во на {os_s}", min_value=0, key=f"fsrv_{os_s}")
            data[f"ОС Сервера ({os_s})"] = val_os
            sum_os_srv += val_os

    if virt_count > 0 and sum_os_srv < virt_count:
        st.warning(f"⚠️ Ошибка: Количество ОС ({sum_os_srv}) должно быть больше или равно количеству виртуальных серверов ({virt_count}).")
        validation_errors.append("Недостаточное количество ОС для серверов")

    selected_virt_sys = st.multiselect("Выберите системы виртуализации", ["VMware", "Hyper-V", "Proxmox", "KVM", "Другое", "Нет"], key="virt_sys_list", help="Технологии управления виртуальной инфраструктурой.")
    if selected_virt_sys and "Нет" not in selected_virt_sys:
        for v_sys in selected_virt_sys:
            v_h_cnt = st.number_input(f"Количество хостов {v_sys}", min_value=0, step=1, key=f"fv_cnt_{v_sys}", help=f"Сколько физических серверов (нод) в кластере {v_sys}?")
            data[f"Система виртуализации ({v_sys})"] = v_h_cnt
            if v_h_cnt == 0: validation_errors.append(f"Укажите количество хостов для {v_sys}")

    if st.checkbox("Резервное копирование", key="ib_backup", help="Наличие специализированного ПО для бэкапа (Veeam, Commvault, Veritas и т.д.)."):
        v_n_b = st.text_input("Вендор Резервного копирования", key="vn_backup", help="Укажите название используемого продукта.")
        data["Резервное копирование"] = v_n_b
        if not v_n_b: validation_errors.append("Укажите вендора резервного копирования")
        score += 20

    data['1.3. Примечание'] = st.text_area("Примечание к разделу 1.3", placeholder="Специфика серверного парка...", key="note_1_3")

render_section_feedback(
    "Серверы и виртуализация",
    get_section_errors(
        validation_errors,
        "сервер",
        "хост",
        "резервного копирования"
    ),
    enabled=server_active
)

st.write("---")
render_anchor("section-storage")
st.subheader("1.4. Системы хранения данных (СХД)")
st_media_sel = []
cnt_hdd = 0
cnt_ssd = 0
raid_selected = []
storage_active = st.toggle("Есть собственная СХД", key="storage_toggle")
if storage_active:
    st_media_sel = st.multiselect(
        "Типы носителей",
        ["HDD (NL-SAS / SATA)", "SSD (SATA / SAS)", "NVMe", "SCM"],
        key="st_media"
    )
    data['1.4.1. Типы носителей'] = ", ".join(st_media_sel) if st_media_sel else "Не указано"

    col_pct1, col_pct2 = st.columns(2)
    with col_pct1:
        cnt_hdd = st.number_input("Количество дисков HDD", min_value=0, step=1, key="cnt_hdd")
        data['1.4.2. Кол-во HDD'] = cnt_hdd
    with col_pct2:
        cnt_ssd = st.number_input("Количество дисков SSD", min_value=0, step=1, key="cnt_ssd")
        data['1.4.3. Кол-во SSD'] = cnt_ssd

    if st_media_sel and (cnt_hdd + cnt_ssd == 0):
        st.info("ℹ️ Укажите количество дисков для СХД.")
        validation_errors.append("Не указано количество дисков СХД")

    col_chk1, col_chk2 = st.columns(2)
    with col_chk1:
        hybrid = st.checkbox("Используется гибридная СХД", key="hybrid_st")
        data['1.4.4. Гибридная СХД'] = "Да" if hybrid else "Нет"
    with col_chk2:
        allflash = st.checkbox("Есть All-Flash массивы", key="allflash_st")
        data['1.4.5. All-Flash'] = "Да" if allflash else "Нет"
        if allflash: score += 5

    raid_selected = st.multiselect(
        "Используемые RAID-группы",
        ["RAID 0", "RAID 1", "RAID 5", "RAID 6", "RAID 10", "RAID 50", "RAID 60", "JBOD"],
        key="raid_list"
    )
    data['1.4.6. RAID-группы'] = ", ".join(raid_selected) if raid_selected else "Не указано"

    if not raid_selected:
        validation_errors.append("Не указаны RAID-группы СХД")
    if "RAID 0" in raid_selected or "JBOD" in raid_selected:
        score -= 10

    data['1.4. Примечание'] = st.text_area("Примечание к разделу 1.4", placeholder="SAN/NAS, replication, snapshot, DR-site, tiering и т.д.", key="note_1_4")

render_section_feedback(
    "Системы хранения данных",
    get_section_errors(validation_errors, "схд", "raid"),
    enabled=storage_active
)

st.write("---")
render_anchor("section-systems")
st.subheader("1.5. Внутренние Информационные системы")
is_active = st.toggle("ИС организации", key="is_toggle", help="Бизнес-приложения и корпоративные сервисы.")
if is_active:
    is_types = {
        "ERP": "erp", "CRM": "crm", "HelpDesk/ServiceDesk": "sd",
        "СЭД (Документооборот)": "sed", "HRM (Кадры)": "hrm",
        "BI (Аналитика)": "bi", "WMS (Склад)": "wms", "Учет (Бухгалтерия)": "acc"
    }

    m_opts = MAIL_SYSTEM_OPTIONS
    m_sys = st.selectbox("Почтовая система", m_opts, key="mail_system", help="Где физически и логически располагается ваша электронная почта.")

    if m_sys in ["Exchange (On-Prem)", "Lotus"]:
        m_ver = st.text_input(f"Версия {m_sys}*", key="mail_version_input", help="Например: 2016 CU23 или v9.0.1.")
        data['1.5.1. Почтовая система'] = f"{m_sys} (v.{m_ver})"
        if not m_ver: validation_errors.append(f"Укажите версию {m_sys}")
    else:
        data['1.5.1. Почтовая система'] = m_sys

    for label, ks in is_types.items():
        if st.checkbox(label, key=f"is_chk_{ks}"):
            c_is1, c_is2 = st.columns(2)
            with c_is1:
                name_is = st.text_input(f"Название продукта {label}*", key=f"name_{ks}")
            with c_is2:
                ver_is = st.text_input(f"Версия {label}*", key=f"ver_{ks}")
            data[f"ИС {label}"] = f"{name_is} (v.{ver_is})"
            if not name_is or not ver_is:
                validation_errors.append(f"Укажите название и версию для {label}")

    data['1.5. Примечание'] = st.text_area("Примечание к разделу 1.5", placeholder="Дополнительные ИС...", key="note_1_5")

render_section_feedback(
    "Внутренние информационные системы",
    get_section_errors(validation_errors, "версию", "название и версию"),
    enabled=is_active
)

st.divider()

# --- БЛОК 2: ИНФОРМАЦИОННАЯ БЕЗОПАСНОСТЬ ---
render_section_marker(
    "03 / БЕЗОПАСНОСТЬ",
    "Информационная безопасность",
    "Текущие средства защиты, мониторинг, доступ, приложения и управление уязвимостями.",
    anchor_id="section-security"
)

enable_security_kwargs = {"key": "enable_security"}
if "enable_security" not in st.session_state:
    enable_security_kwargs["value"] = False
enable_security = st.toggle("Включить блок ИБ", **enable_security_kwargs)

# Инициализация переменных ИБ
epp, epp_v, edr, edr_v, xdr, xdr_v, mdr, mdr_v = False, "", False, "", False, "", False, ""
dlp, dlp_v, mail_sec, mail_v, casb, casb_v = False, "", False, "", False, ""
waf, waf_v, ddos, ddos_v, ids, ids_v, nac, nac_v, ztna, ztna_v = False, "", False, "", False, "", False, "", False, ""
sast, sast_v, dast, dast_v = False, "", False, ""
iam, iam_v, mfa, mfa_v, pam, pam_v = False, "", False, "", False, ""
siem, siem_v, soar, soar_v = False, "", False, ""
vuln, vuln_v, patch, patch_v, nad, nad_v = False, "", False, "", False, ""

def security_product(label, checkbox_key, vendor_label, vendor_key):
    enabled = st.checkbox(label, key=checkbox_key) is True
    if enabled:
        return enabled, st.text_input(vendor_label, key=vendor_key)
    st.session_state.pop(vendor_key, None)
    return enabled, ""

if enable_security:
    errors = []

    # =========================
    # Защита конечных устройств
    # =========================
    st.markdown("#### Защита конечных устройств")
    col1, col2 = st.columns(2)
    with col1:
        epp, epp_v = security_product("EPP (антивирусная защита)", "epp", "Производитель EPP", "epp_v")
        data['Блок 2. EPP'] = epp_v if epp else "Нет"

        edr, edr_v = security_product("EDR (обнаружение и реагирование)", "edr", "Производитель EDR", "edr_v")
        data['Блок 2. EDR'] = edr_v if edr else "Нет"
    with col2:
        xdr, xdr_v = security_product("XDR (расширенная защита)", "xdr", "Производитель XDR", "xdr_v")
        data['Блок 2. XDR'] = xdr_v if xdr else "Нет"

        mdr, mdr_v = security_product("MDR (внешний мониторинг)", "mdr", "Провайдер MDR", "mdr_v")
        data['Блок 2. MDR'] = mdr_v if mdr else "Нет"

    # =========================
    # Защита данных
    # =========================
    st.markdown("#### Защита данных")
    col1, col2 = st.columns(2)
    with col1:
        dlp, dlp_v = security_product("DLP (предотвращение утечек)", "dlp", "Производитель DLP", "dlp_v")
        data['Блок 2. DLP'] = dlp_v if dlp else "Нет"

        mail_sec, mail_v = security_product("Mail Security (защита почты)", "mail_sec", "Производитель Mail Security", "mail_v")
        data['Блок 2. Mail Security'] = mail_v if mail_sec else "Нет"
    with col2:
        casb, casb_v = security_product("CASB (контроль облаков)", "casb", "Производитель CASB", "casb_v")
        data['Блок 2. CASB'] = casb_v if casb else "Нет"

    # =========================
    # Сетевая безопасность
    # =========================
    st.markdown("#### Сетевая безопасность")
    col1, col2 = st.columns(2)
    with col1:
        waf, waf_v = security_product("WAF (защита веб-приложений)", "waf", "Производитель WAF", "waf_v")
        data['Блок 2. WAF'] = waf_v if waf else "Нет"

        ddos, ddos_v = security_product("Anti-DDoS (защита от атак)", "ddos", "Производитель Anti-DDoS", "ddos_v")
        data['Блок 2. Anti-DDoS'] = ddos_v if ddos else "Нет"

        nad, nad_v = security_product("NAD (Network Attack Discovery)", "nad", "Производитель NAD", "nad_v")
        data['Блок 2. NAD'] = nad_v if nad else "Нет"
    with col2:
        ids, ids_v = security_product("IDS/IPS (сетевые атаки)", "ids", "Производитель IDS/IPS", "ids_v")
        data['Блок 2. IDS/IPS'] = ids_v if ids else "Нет"

        nac, nac_v = security_product("NAC (контроль доступа)", "nac", "Производитель NAC", "nac_v")
        data['Блок 2. NAC'] = nac_v if nac else "Нет"

        ztna, ztna_v = security_product("ZTNA (Zero Trust доступ)", "ztna", "Производитель ZTNA", "ztna_v")
        data['Блок 2. ZTNA'] = ztna_v if ztna else "Нет"

    # =========================
    # Безопасность приложений
    # =========================
    st.markdown("#### Безопасность приложений")
    col1, col2 = st.columns(2)
    with col1:
        sast, sast_v = security_product("SAST (анализ кода)", "sast", "Производитель SAST", "sast_v")
        data['Блок 2. SAST'] = sast_v if sast else "Нет"
    with col2:
        dast, dast_v = security_product("DAST (тестирование приложений)", "dast", "Производитель DAST", "dast_v")
        data['Блок 2. DAST'] = dast_v if dast else "Нет"

    # =========================
    # Управление доступом
    # =========================
    st.markdown("#### Управление доступом")
    col1, col2 = st.columns(2)
    with col1:
        iam, iam_v = security_product("IAM (учетные записи)", "iam", "Производитель IAM", "iam_v")
        data['Блок 2. IAM'] = iam_v if iam else "Нет"

        mfa, mfa_v = security_product("MFA (многофакторная аутентификация)", "mfa", "Производитель MFA", "mfa_v")
        data['Блок 2. MFA'] = mfa_v if mfa else "Нет"
    with col2:
        pam, pam_v = security_product("PAM (привилегированный доступ)", "pam", "Производитель PAM", "pam_v")
        data['Блок 2. PAM'] = pam_v if pam else "Нет"

    # =========================
    # SOC
    # =========================
    st.markdown("#### Мониторинг и реагирование")
    col1, col2 = st.columns(2)
    with col1:
        siem, siem_v = security_product("SIEM (мониторинг событий)", "siem", "Производитель SIEM", "siem_v")
        data['Блок 2. SIEM'] = siem_v if siem else "Нет"
    with col2:
        soar, soar_v = security_product("SOAR (автоматизация)", "soar", "Производитель SOAR", "soar_v")
        data['Блок 2. SOAR'] = soar_v if soar else "Нет"

    # =========================
    # ДОПОЛНИТЕЛЬНО
    # =========================
    st.markdown("#### Дополнительно")
    col1, col2 = st.columns(2)
    with col1:
        vuln, vuln_v = security_product("Сканер уязвимостей", "vuln", "Производитель сканера", "vuln_v")
        data['Блок 2. Сканер уязвимостей'] = vuln_v if vuln else "Нет"
    with col2:
        patch, patch_v = security_product("Patch Management (управление обновлениями)", "patch", "Производитель Patch Management", "patch_v")
        data['Блок 2. Patch Management'] = patch_v if patch else "Нет"

    # Валидация
    ib_items = [
        ("EPP", epp, epp_v), ("EDR", edr, edr_v), ("XDR", xdr, xdr_v),
        ("MDR", mdr, mdr_v), ("DLP", dlp, dlp_v), ("Mail Security", mail_sec, mail_v),
        ("CASB", casb, casb_v), ("WAF", waf, waf_v), ("Anti-DDoS", ddos, ddos_v),
        ("IDS/IPS", ids, ids_v), ("NAC", nac, nac_v), ("ZTNA", ztna, ztna_v),
        ("SAST", sast, sast_v), ("DAST", dast, dast_v),
        ("IAM", iam, iam_v), ("MFA", mfa, mfa_v), ("PAM", pam, pam_v),
        ("SIEM", siem, siem_v), ("SOAR", soar, soar_v),
        ("Сканер уязвимостей", vuln, vuln_v),
        ("Patch Management", patch, patch_v),
        ("NAD", nad, nad_v)
    ]
    for name, enabled, vendor in ib_items:
        if enabled and not vendor:
            errors.append(f"Не указан производитель: {name}")

    if errors:
        st.error("Заполните обязательные поля в блоке ИБ:")
        for e in errors:
            st.write(f"- {e}")
        validation_errors.extend(errors)

render_section_feedback(
    "Информационная безопасность",
    get_section_errors(validation_errors, "производитель", "иб"),
    enabled=enable_security
)

# --- БЛОК 3: WEB-РЕСУРСЫ ---
render_section_marker(
    "04 / ЦИФРОВАЯ ПОВЕРХНОСТЬ",
    "Веб-ресурсы",
    "Публичная поверхность, фронтенд-стек и особенности размещения.",
    anchor_id="section-web"
)
web_active = st.toggle("Веб-ресурсы", key="web_toggle")
if web_active:
    data['3.1. Хостинг'] = st.selectbox("Хостинг", WEB_HOSTING_OPTIONS, key="web_hosting")
    data['3.2. Frontend'] = st.multiselect("Frontend серверы", ["Nginx", "Apache", "IIS", "LiteSpeed", "Cloudflare"], key="web_frontend")
    data['Примечание (Web)'] = st.text_area("Примечания по Web", placeholder="Стек...", key="note_web")

st.divider()

# --- БЛОК 4: РАЗРАБОТКА ---
render_section_marker(
    "05 / РАЗРАБОТКА",
    "Разработка",
    "Команда разработки, языки, CI/CD и дополнительный технологический контекст.",
    anchor_id="section-dev"
)
dev_count = 0
sel_langs = []
cicd_active = False
dev_active = st.toggle("Разработка", key="dev_toggle")
if dev_active:
    col_d1, col_d2 = st.columns(2)
    with col_d1:
        dev_count = st.number_input("Кол-во разработчиков*", min_value=0, key="dev_cnt_f")
        data['4.1. Разработчики'] = dev_count
        cicd_active = st.checkbox("Используется CI/CD", key="cicd_f")
        data['4.2. CICD'] = "Да" if cicd_active else "Нет"
        if dev_count == 0: validation_errors.append("Укажите количество разработчиков")
    with col_d2:
        lang_list = ["Python", "JavaScript/TypeScript", "Java", "C# / .NET", "PHP", "Go", "C++", "Swift/Kotlin", "Другое"]
        sel_langs = st.multiselect("Языки программирования*", lang_list, key="langs_f")
        if not sel_langs:
            validation_errors.append("Выберите языки разработки")
            data['4.3. Языки разработки'] = "Не указаны"
        elif "Другое" in sel_langs:
            other_l = st.text_input("Укажите другие языки", key="other_langs_f")
            data['4.3. Языки разработки'] = f"{', '.join([l for l in sel_langs if l != 'Другое'])}, {other_l}"
        else:
            data['4.3. Языки разработки'] = ", ".join(sel_langs)
    data['Блок 4. Примечание'] = st.text_area("Примечание к разделу Разработка", placeholder="Стек, фреймворки...", key="note_dev")

render_section_feedback(
    "Разработка",
    get_section_errors(validation_errors, "разработчик", "языки разработки"),
    enabled=dev_active
)
#----Подготовка---
def get_maturity_level(score):
    if score <= 20:
        return "Начальный", "🔴"
    elif score <= 40:
        return "Базовый", "🟠"
    elif score <= 60:
        return "Развивающийся", "🟡"
    elif score <= 80:
        return "Управляемый", "🟢"
    else:
        return "Оптимальный", "🔵"


def calculate_weighted_security_score(enabled, controls):
    if not enabled:
        return 0

    available_weight = sum(max(0, weight) for _, _, weight in controls)
    if available_weight <= 0:
        return 0

    earned_weight = 0
    for is_enabled, vendor, weight in controls:
        if is_enabled and str(vendor).strip():
            earned_weight += max(0, weight)

    # Анкета подтверждает наличие средств защиты, но не качество настройки,
    # покрытие активов и эффективность реагирования. Поэтому самооценка без
    # проверки артефактов не может означать абсолютные 100% зрелости.
    normalized_score = round((earned_weight / available_weight) * 92)
    return min(92, max(0, normalized_score))


def calculate_it_maturity_score(
    total_arm,
    selected_os_arm,
    sum_os_arm,
    net_active,
    main_speed,
    back_speed,
    selected_routing,
    ap_cnt,
    ngfw_vendor,
    server_active,
    phys_count,
    virt_count,
    selected_virt_sys,
    backup_vendor,
    storage_active,
    st_media_sel,
    cnt_hdd,
    cnt_ssd,
    raid_selected,
    systems_active,
    web_active,
    dev_active,
    dev_count,
    sel_langs,
    cicd_active,
    wifi_enabled=False,
    wifi_ctrl_enabled=False,
    operational_notes=None,
):
    earned = 0.0
    available = 20.0

    # Размер парка задает сложность эксплуатации, но сам по себе не является зрелостью.
    if total_arm > 0:
        earned += 4
    if selected_os_arm and sum_os_arm == total_arm:
        earned += 10
    legacy_markers = ("xp", "vista", "windows 7", "windows 8")
    if selected_os_arm and not any(
        marker in str(os_name).lower()
        for os_name in selected_os_arm
        for marker in legacy_markers
    ):
        earned += 6

    if net_active:
        available += 24
        earned += 4 if main_speed > 0 else 0
        earned += 6 if back_speed > 0 else 0
        earned += 5 if selected_routing else 0
        earned += 3 if ap_cnt > 0 else 0
        earned += 6 if str(ngfw_vendor).strip().lower() not in {"", "нет", "no", "none"} else 0

    if server_active:
        available += 24
        earned += 5 if (phys_count + virt_count) > 0 else 0
        earned += 5 if virt_count > 0 and selected_virt_sys else 0
        earned += 10 if str(backup_vendor).strip().lower() not in {"", "нет", "no", "none"} else 0
        earned += 4 if phys_count > 0 and virt_count > 0 else 0

    if storage_active:
        available += 14
        earned += 3 if (cnt_hdd + cnt_ssd) > 0 else 0
        earned += 3 if st_media_sel else 0
        earned += 6 if raid_selected else 0
        earned += 2 if cnt_ssd > 0 else 0

    if systems_active:
        available += 6
        earned += 4

    if web_active:
        available += 4
        earned += 2

    if dev_active:
        available += 8
        earned += 2 if dev_count > 0 else 0
        earned += 2 if sel_langs else 0
        earned += 4 if cicd_active else 0

    score = round((earned / available) * 100) if available else 0
    penalty = 0

    if net_active and main_speed > 0 and total_arm > 0:
        bandwidth_per_arm = main_speed / total_arm
        if bandwidth_per_arm < 1:
            penalty += 5
        elif bandwidth_per_arm < 2:
            penalty += 2

    if net_active and main_speed > 0:
        reserve_ratio = back_speed / main_speed if back_speed > 0 else 0
        if reserve_ratio == 0:
            penalty += 10
        elif reserve_ratio < 0.1:
            penalty += 7
        elif reserve_ratio < 0.25:
            penalty += 4

    if net_active and wifi_enabled and ap_cnt > 0 and total_arm > 0:
        endpoints_per_ap = total_arm / ap_cnt
        if endpoints_per_ap > 50:
            penalty += 7
        elif endpoints_per_ap > 30:
            penalty += 4
        if ap_cnt >= 4 and not wifi_ctrl_enabled:
            penalty += 6

    notes = " ".join(str(value or "") for value in (operational_notes or [])).lower()
    if any(marker in notes for marker in (
        "без единой панели", "не объединен в единый контур",
        "разрозненный мониторинг", "разрозненные системы мониторинга",
    )):
        penalty += 4
    if any(marker in notes for marker in (
        "capacity planning не", "запас вычислительной мощности недостаточ",
        "запас мощности недостаточ", "перегружена виртуализац",
    )):
        penalty += 4
    if (
        ("rto" in notes or "rpo" in notes)
        and any(marker in notes for marker in ("не согласован", "не определен", "не утвержден"))
    ) or ("восстанов" in notes and "нерегуляр" in notes):
        penalty += 4
    if any(marker in notes for marker in (
        "cmdb отсутств", "изменения согласуются в чат",
        "календарь изменений" , "план отката используются не",
    )):
        penalty += 4

    storage_fill = re.search(r"(?:схд|хранилищ\w*)[^.]{0,50}(\d{2,3})\s*%", notes)
    if storage_fill and int(storage_fill.group(1)) >= 80:
        penalty += 4

    # Наличие оборудования и ПО не равно зрелости эксплуатации. Верхняя граница
    # отражает отсутствие проверки SLA, ITSM, capacity и DR-артефактов в анкете.
    return min(90, max(0, score - penalty))


def build_context(results, client_info):

    users = int(results.get("_user_count", 0))

    industry = str(
        client_info.get(
            "Сфера деятельности",
            ""
        )
    )

    context = {}

    # ==================================
    # COMPANY SIZE
    # ==================================

    context["users"] = users

    context["small_company"] = users < 50

    context["medium_company"] = (
        50 <= users < 250
    )

    context["large_company"] = (
        250 <= users < 1000
    )

    context["enterprise_company"] = (
        users >= 1000
    )

    # ==================================
    # INDUSTRY
    # ==================================

    context["industry"] = industry

    context["is_finance"] = (
        "Финтех" in industry
        or "Банк" in industry
        or "Страхование" in industry
    )

    context["is_gov"] = (
        "Госсектор" in industry
        or "Квазигосударственный" in industry
    )

    context["is_kvoiki"] = (
        "КВОИКИ" in industry
        or "Критическая инфраструктура" in industry
    )

    context["is_healthcare"] = (
        "Здравоохранение" in industry
        or "Медицин" in industry
    )

    context["is_telecom"] = (
        "Телеком" in industry
        or "Связь" in industry
    )

    context["is_industrial"] = any(marker in industry for marker in (
        "Производство", "АСУ ТП", "Энергетика", "Коммунальная", "Транспорт"
    ))

    context["is_it_company"] = (
        "IT" in industry
        or "Разработка" in industry
    )

    context["is_retail"] = (
        "Ритейл" in industry
        or "E-commerce" in industry
    )

    context["is_manufacturing"] = (
        "Производство" in industry
    )

    # ==================================
    # Бизнес-системы
    # ==================================

    result_text = str(results)

    context["has_erp"] = (
        "ERP" in result_text
    )

    context["has_crm"] = (
        "CRM" in result_text
    )

    context["has_accounting"] = (
        "Учет" in result_text
        or "Бухгалтерия" in result_text
    )

    context["has_hrm"] = (
        "HRM" in result_text
    )

    context["has_document_flow"] = (
        "СЭД" in result_text
    )

    context["has_mail"] = (
        results.get(
            "1.5.1. Почтовая система",
            "Нет"
        ) != "Нет"
    )

    # ==================================
    # CRITICAL DATA
    # ==================================

    context["has_personal_data"] = any([
        context["has_hrm"],
        context["has_crm"],
        context["has_accounting"]
    ])

    context["has_critical_systems"] = any([
        context["has_erp"],
        context["has_crm"],
        context["has_accounting"]
    ])

    # ==================================
    # Разработка
    # ==================================

    context["has_development"] = (
        "4.1. Разработчики" in results
    )

    dev_count = int(
        results.get(
            "4.1. Разработчики",
            0
        )
    )

    context["developers"] = dev_count

    context["large_dev_team"] = (
        dev_count >= 10
    )

    # ==================================
    # WEB
    # ==================================

    context["has_public_web"] = (
        "3.1. Хостинг" in results
    )

    # ==================================
    # Инфраструктура
    # ==================================

    context["servers"] = int(
        results.get(
            "Серверы (вирт)",
            0
        )
    ) + int(
        results.get(
            "Серверы (физ)",
            0
        )
    )

    context["has_virtualization"] = (
        context["servers"] > 0
    )

    context["has_backup"] = (
        results.get(
            "Резервное копирование",
            "Нет"
        ) != "Нет"
    )

    # ==================================
    # Безопасность
    # ==================================

    context["has_ngfw"] = (
        results.get(
            "NGFW",
            "Нет"
        ) != "Нет"
    )

    context["has_siem"] = (
        results.get(
            "SIEM",
            "Нет"
        ) != "Нет"
    )

    context["has_edr"] = (
        results.get(
            "EDR",
            "Нет"
        ) != "Нет"
    )

    context["has_patch_management"] = (
        results.get(
            "Patch Management",
            "Нет"
        ) != "Нет"
    )

    context["has_mfa"] = (
        results.get(
            "MFA",
            "Нет"
        ) != "Нет"
    )

    return context


def russian_count(value, one, few, many):
    number = int(value or 0)
    last_two = number % 100
    last = number % 10
    if 11 <= last_two <= 14:
        word = many
    elif last == 1:
        word = one
    elif 2 <= last <= 4:
        word = few
    else:
        word = many
    return f"{number} {word}"


def infrastructure_profile(context):
    users = context.get("users", 0)
    servers = context.get("servers", 0)
    users_text = russian_count(users, "АРМ", "АРМ", "АРМ")
    servers_text = russian_count(servers, "сервер", "сервера", "серверов")

    if context.get("enterprise_company"):
        return "Распределенный ИТ-контур", (
            f"{users_text}, {servers_text}. Требуется формальная модель "
            "управления ИБ, мониторинга, резервирования и регулярной отчетности."
        )

    if context.get("large_company"):
        return "Многоуровневый ИТ-контур", (
            f"{users_text}, {servers_text}. Важно стандартизировать процессы, "
            "сегментацию, управление доступом и контроль обновлений."
        )

    if context.get("medium_company"):
        return "ИТ-контур с серверной инфраструктурой", (
            f"{users_text}, {servers_text}. Приоритеты определяются по критичности сервисов, "
            "управляемости рабочих мест, восстановлению и контролю доступа."
        )

    return "ИТ-контур", (
        f"{users_text}, {servers_text}. Приоритеты определяются по критичности сервисов, "
        "операционным ограничениям и измеримому эффекту мер."
    )


def it_context_summary(results, context):
    assets = []
    if context.get("users", 0):
        assets.append(f"{context.get('users', 0)} АРМ")
    if context.get("servers", 0):
        assets.append(russian_count(context.get("servers", 0), "сервер", "сервера", "серверов"))
    if is_enabled(results.get("Виртуализация")):
        assets.append("виртуализация")
    if is_enabled(results.get("СХД")):
        assets.append("СХД")
    if context.get("has_public_web"):
        assets.append("публичные web-сервисы")
    if context.get("has_critical_systems"):
        assets.append("критичные бизнес-системы")
    if context.get("has_development"):
        assets.append("разработка")

    process_focus = []
    if not is_enabled(results.get("Мониторинг")):
        process_focus.append("эксплуатационный мониторинг")
    if not is_enabled(results.get("Patch Management")):
        process_focus.append("управление обновлениями")
    if context.get("servers", 0) and not is_enabled(results.get("DR")):
        process_focus.append("DR/RTO/RPO")
    if is_enabled(results.get("Резервное копирование")):
        process_focus.append("проверка восстановления")
    elif context.get("servers", 0):
        process_focus.append("резервное копирование")

    assets_text = ", ".join(assets) if assets else "масштаб ИТ не раскрыт"
    focus_text = ", ".join(dict.fromkeys(process_focus)) if process_focus else "поддержание текущей модели эксплуатации"
    return assets_text, focus_text


def build_contextual_roadmap(results, context, domain_scores, risks):
    roadmap = []

    def add(phase, priority, domain, action, rationale, owner="ИТ/ИБ", effort="Средняя"):
        roadmap.append({
            "phase": phase,
            "priority": priority,
            "domain": domain,
            "action": action,
            "rationale": rationale,
            "owner": owner,
            "effort": effort,
        })

    is_small = context.get("small_company", False)
    is_medium = context.get("medium_company", False)
    is_large = context.get("large_company", False)
    is_enterprise = context.get("enterprise_company", False)
    has_servers = context.get("servers", 0) > 0
    has_public_web = context.get("has_public_web", False)
    has_development = context.get("has_development", False)
    has_critical_systems = context.get("has_critical_systems", False)

    if results.get("MFA") == "Нет":
        add(
            "0-30 дней",
            "P1",
            "Доступ",
            "Включить MFA для администраторов, VPN, почты и критичных систем.",
            "Быстро снижает риск компрометации учетных записей без тяжелого проекта.",
            effort="Низкая",
        )

    if results.get("Резервное копирование") == "Нет" and has_servers:
        add(
            "0-30 дней",
            "P1",
            "Устойчивость",
            "Организовать регулярное резервное копирование серверов и критичных данных.",
            "Для серверной инфраструктуры backup является базовым условием восстановления после сбоя или ransomware.",
            effort="Средняя",
        )
    elif results.get("Резервное копирование") != "Нет" and results.get("Immutable Backup") == "Нет":
        add(
            "31-60 дней",
            "P2",
            "Устойчивость",
            "Добавить immutable/offline-копии и проверить сценарий восстановления.",
            "Обычные резервные копии могут быть удалены или зашифрованы при атаке.",
            effort="Средняя",
        )

    legacy_arm = results.get("ОС АРМ (Windows XP/Vista/7/8)", 0)
    legacy_srv = results.get("ОС Сервера (Windows Server 2008/2012 R2)", 0)
    if legacy_arm or legacy_srv:
        add(
            "0-30 дней",
            "P1",
            "Инфраструктура",
            "Составить план миграции или изоляции устаревших ОС.",
            f"Обнаружены устаревшие ОС: АРМ {legacy_arm}, серверы {legacy_srv}.",
            effort="Высокая",
        )

    if results.get("Patch Management") == "Нет":
        phase = "31-60 дней" if is_small else "0-30 дней"
        add(
            phase,
            "P2",
            "Управляемость",
            "Ввести централизованный контроль обновлений для АРМ и серверов.",
            "Чем больше инфраструктура, тем выше риск ручного и несогласованного обновления.",
            effort="Средняя",
        )

    if results.get("EDR") == "Нет":
        action = (
            "Выбрать управляемый EDR/MDR-сервис для рабочих мест."
            if is_small or is_medium
            else "Внедрить EDR/XDR с централизованной политикой и реагированием."
        )
        add(
            "31-60 дней",
            "P2",
            "Endpoint",
            action,
            "Обычного антивируса недостаточно для расследования сложных атак и lateral movement.",
            effort="Средняя",
        )

    if results.get("SIEM") == "Нет":
        if is_enterprise:
            action = "Запустить проект SIEM/SOC с корреляцией событий и регламентами реагирования."
            phase = "61-90 дней"
            priority = "P2"
            effort = "Высокая"
        elif is_large or has_critical_systems:
            action = "Подключить MSSP/SOC или легковесный SIEM для критичных систем и средств защиты."
            phase = "61-90 дней"
            priority = "P3"
            effort = "Средняя"
        else:
            action = "Настроить сбор критичных журналов и назначить ответственного за разбор инцидентов."
            phase = "31-60 дней"
            priority = "P3"
            effort = "Низкая"

        add(
            phase,
            priority,
            "Мониторинг",
            action,
            "Мониторинг должен соответствовать масштабу: для малой инфраструктуры важнее управляемый минимум, чем тяжелая SIEM-платформа.",
            effort=effort,
        )

    if has_public_web and results.get("WAF") == "Нет":
        add(
            "31-60 дней",
            "P2",
            "Веб",
            "Оценить необходимость WAF/CDN-защиты для публичных веб-ресурсов.",
            "Публичные сервисы увеличивают поверхность атаки и требуют отдельного контроля.",
            effort="Средняя",
        )

    if has_development and results.get("SAST") == "Нет" and results.get("DAST") == "Нет":
        add(
            "61-90 дней",
            "P3",
            "Разработка",
            "Встроить базовые SAST/DAST-проверки в процесс разработки.",
            "Для собственной разработки риски приложений лучше выявлять до публикации.",
            owner="Разработка/ИБ",
            effort="Средняя",
        )

    weakest_domains = [
        domain
        for domain, score in sorted(domain_scores.items(), key=lambda item: item[1])
        if score < 50
    ][:2]
    for domain in weakest_domains:
        if not any(item["domain"] == domain for item in roadmap):
            add(
                "61-90 дней",
                "P3",
                domain,
                f"Разработать целевой план улучшения домена: {domain}.",
                "Домен имеет низкую оценку и требует отдельного набора мер.",
                effort="Средняя",
            )

    risk_titles = {item.get("risk") for item in risks[:3]}
    if risk_titles and not roadmap:
        add(
            "0-30 дней",
            "P1",
            "Приоритизация",
            "Разобрать ключевые риски и назначить ответственных за их устранение.",
            "Анкета выявила риски, которые требуют управленческого решения.",
            effort="Низкая",
        )

    if not roadmap:
        add(
            "31-60 дней",
            "P3",
            "Развитие",
            "Провести контрольную проверку политик, backup и доступа администраторов.",
            "Даже при хорошем базовом уровне полезна регулярная проверка устойчивости.",
            effort="Низкая",
        )

    priority_order = {"P1": 1, "P2": 2, "P3": 3}
    phase_order = {"0-30 дней": 1, "31-60 дней": 2, "61-90 дней": 3}
    return sorted(
        roadmap,
        key=lambda item: (
            phase_order.get(item["phase"], 99),
            priority_order.get(item["priority"], 99),
        )
    )[:10]


def pick_catalog_vendors(catalog, keywords, fallback):
    selected = []
    for vendor in catalog:
        vendor_lower = vendor.lower()
        if any(keyword.lower() in vendor_lower for keyword in keywords):
            selected.append(vendor)

    if not selected:
        selected = fallback

    return list(dict.fromkeys(selected))[:4]


def normalize_vendor_key(value):
    return re.sub(r"[^a-zа-я0-9]+", " ", str(value or "").lower()).strip()


def clean_vendor_display_name(value):
    vendor = str(value or "").strip()
    fixes = {
        "imperva": "Imperva",
        "huawei": "Huawei",
        "splunc": "Splunk",
        "mccafee": "McAfee",
    }
    return fixes.get(vendor.lower(), vendor)


DETAILED_VENDOR_MATRIX_FILE = "vendor_matrix_detailed.xlsx"


def load_detailed_vendor_names():
    try:
        if not os.path.exists(DETAILED_VENDOR_MATRIX_FILE):
            return []
        df = pd.read_excel(DETAILED_VENDOR_MATRIX_FILE)
        if df.empty or "Vendor" not in df.columns:
            return []
        values = []
        for value in df["Vendor"].dropna().tolist():
            vendor = clean_vendor_display_name(value)
            if vendor and vendor.lower() not in {"nan", "none"}:
                values.append(vendor)
        return list(dict.fromkeys(values))
    except Exception:
        return []


def load_detailed_solution_vendor_map():
    try:
        if not os.path.exists(DETAILED_VENDOR_MATRIX_FILE):
            return {}
        df = pd.read_excel(DETAILED_VENDOR_MATRIX_FILE)
        if df.empty or "Vendor" not in df.columns:
            return {}

        category_map = {}
        for _, row in df.iterrows():
            vendor = clean_vendor_display_name(row.get("Vendor"))
            if not vendor:
                continue
            for column in df.columns:
                if column == "Vendor":
                    continue
                marker = str(row.get(column) or "").strip()
                if marker == "+":
                    category_map.setdefault(str(column).strip(), []).append(vendor)

        return {
            category: list(dict.fromkeys(vendors))
            for category, vendors in category_map.items()
        }
    except Exception:
        return {}


def normalize_portfolio_header(value):
    return normalize_vendor_key(value)


def split_portfolio_list(value):
    if value is None:
        return []
    parts = re.split(r"[,;|\n]+", str(value))
    return [part.strip() for part in parts if part and part.strip()]


def load_verified_distributor_map():
    try:
        if not os.path.exists(DETAILED_VENDOR_MATRIX_FILE):
            return {}
        df = pd.read_excel(DETAILED_VENDOR_MATRIX_FILE)
        if df.empty or "Vendor" not in df.columns:
            return {}

        normalized_columns = {
            normalize_portfolio_header(column): column
            for column in df.columns
        }
        distributor_column = None
        status_column = None
        for key, column in normalized_columns.items():
            if (
                key in {
                "distributor", "distributors", "distributor name",
                "дистрибьютор", "дистрибьюторы", "поставщик"
                }
                or (
                    key.startswith("distributor ")
                    and "source" not in key
                    and "status" not in key
                )
            ):
                distributor_column = column
            if key in {
                "distributor status", "status distributor",
                "статус дистрибьютора", "статус"
            }:
                status_column = column

        if not distributor_column or not status_column:
            return {}

        verified_statuses = {
            "verified", "approved", "trusted", "checked",
            "проверенный", "проверено", "подтвержденный", "подтверждено",
        }
        distributor_map = {}
        for _, row in df.iterrows():
            status = normalize_portfolio_header(row.get(status_column))
            if status not in verified_statuses:
                continue
            vendor = clean_vendor_display_name(row.get("Vendor"))
            distributors = split_portfolio_list(row.get(distributor_column))
            if vendor and distributors:
                distributor_map.setdefault(vendor, [])
                distributor_map[vendor].extend(distributors)

        return {
            vendor: list(dict.fromkeys(distributors))
            for vendor, distributors in distributor_map.items()
        }
    except Exception:
        return {}


def verified_distributors_for_vendors(vendors_text):
    distributor_map = load_verified_distributor_map()
    if not distributor_map:
        return "-"

    def collect_distributors(vendor_values):
        values = []
        for vendor in vendor_values:
            normalized_vendor = normalize_vendor_key(vendor)
            for known_vendor, distributors in distributor_map.items():
                if normalize_vendor_key(known_vendor) == normalized_vendor:
                    values.append(f"{known_vendor}: {', '.join(distributors)}")
        return values

    vendors = split_portfolio_list(vendors_text)
    values = collect_distributors(vendors)
    if not values and vendors:
        inferred_vendors = manufacturers_for_report_item({
            "vendors": vendors,
            "risk": str(vendors_text),
            "description": str(vendors_text),
        })
        values = collect_distributors(split_portfolio_list(inferred_vendors))

    return "\n".join(list(dict.fromkeys(values))) or "-"


def portfolio_vendors_for_report_item(item):
    inferred = manufacturers_for_report_item(item)
    if inferred and inferred != "-":
        return inferred

    existing_values = item.get("vendors", [])
    if not isinstance(existing_values, list):
        existing_values = [existing_values] if existing_values else []

    distributor_map = load_verified_distributor_map()
    direct_matches = []
    for value in existing_values:
        normalized_value = normalize_vendor_key(value)
        for known_vendor, distributors in distributor_map.items():
            if normalize_vendor_key(known_vendor) == normalized_value:
                direct_matches.append(known_vendor)

    if direct_matches:
        return ", ".join(list(dict.fromkeys(direct_matches))[:8])

    return ", ".join(str(value).strip() for value in existing_values if str(value).strip()) or "-"


def portfolio_vendors_by_categories(categories, preferred=None, exclude=None, gap_text=None, limit=6):
    detailed_matrix = load_detailed_solution_vendor_map()
    preferred = preferred or []
    exclude_keys = {
        normalize_vendor_key(value)
        for value in (exclude or [])
    }
    values = []
    for category in categories:
        values.extend(detailed_matrix.get(category, []))

    filtered = []
    for vendor in values:
        if normalize_vendor_key(vendor) in exclude_keys:
            continue
        filtered.append(vendor)

    preferred_keys = [normalize_vendor_key(value) for value in preferred]

    def sort_key(vendor):
        key = normalize_vendor_key(vendor)
        return preferred_keys.index(key) if key in preferred_keys else len(preferred_keys)

    filtered = sorted(list(dict.fromkeys(filtered)), key=sort_key)
    if filtered:
        return ", ".join(filtered[:limit])

    return gap_text or "Нет подходящего производителя в матрице"


def normalize_report_vendor_values(item):
    values = item.get("vendors", [])
    if not isinstance(values, list):
        values = [values] if values else []
    return [str(value).strip() for value in values if str(value).strip()]


def solution_categories_for_report_item(item):
    key = risk_semantic_key(item)
    text = normalize_vendor_key(" ".join([
        str(item.get("risk", "")),
        str(item.get("description", "")),
        str(item.get("recommendation", "")),
        " ".join(normalize_report_vendor_values(item)),
    ]))

    if "ids" in text or "ips" in text or "intrusion" in text:
        return "IDS/IPS; NGFW IPS-профили; мониторинг сетевых атак"

    categories_by_key = {
        "mfa": "MFA / Conditional Access",
        "iam": "IAM / Identity Governance / управление жизненным циклом учетных записей",
        "legacy_os": "Миграция на поддерживаемую ОС; изоляция legacy-сегмента",
        "siem_soc": "SOC / MSSP; SIEM; SOAR как этап 2",
        "change_management": "ITSM / Change Management / CMDB",
        "patch": "Vulnerability Management; Patch Management",
        "endpoint_detection": "EDR / XDR / MDR",
        "backup": "Backup; immutable-копии; восстановление после ransomware",
        "web_waf": "WAF / CDN / Web Application Security",
        "pam": "PAM; vault; контроль привилегированных сессий",
        "dlp": "DLP / Data Security",
        "mail": "Mail Security / Anti-Phishing",
        "segmentation": "Сегментация сети; VLAN / ACL; NGFW policies",
        "nac": "NAC / Network Access Control; профилирование устройств",
        "it_monitoring": "IT-мониторинг; capacity management",
        "virtualization": "Виртуализация; lifecycle management",
        "storage": "СХД health-check; capacity management",
        "dr": "Disaster Recovery; RTO/RPO-планирование",
        "appsec": "SAST / DAST / SCA",
        "business_systems": "Обследование бизнес-систем; интеграционный аудит",
        "wifi_capacity": "Обследование Wi-Fi; централизованное WLAN-управление; оптимизация радиопокрытия",
        "network_performance": "Резервирование WAN; SD-WAN; балансировка каналов; контроль SLA",
        "itam": "ITAM / SAM / управление лицензиями",
    }
    return categories_by_key.get(key, "Уточнить класс решения по результатам пресейла")


def portfolio_manufacturers_for_report_item(item):
    key = risk_semantic_key(item)
    text = normalize_vendor_key(" ".join([
        str(item.get("risk", "")),
        str(item.get("description", "")),
        str(item.get("recommendation", "")),
        " ".join(normalize_report_vendor_values(item)),
    ]))

    if "ids" in text or "ips" in text or "intrusion" in text:
        return portfolio_vendors_by_categories(
            ["IDS/IPS", "NGFW"],
            preferred=["Fortinet", "Check Point", "Palo Alto", "Forcepoint"],
            gap_text="Нет отдельной категории IDS/IPS в матрице; проверить NGFW-портфель",
        )

    if key == "legacy_os":
        return portfolio_vendors_by_categories(
            ["Operating Systems"],
            preferred=["Microsoft"],
            gap_text="Нет производителя ОС в матрице",
            limit=1,
        )

    if key == "it_monitoring":
        return portfolio_vendors_by_categories(
            ["ITSM"],
            preferred=["ManageEngine", "Broadcom (Symantec)"],
            exclude=["Ivanti"],
            gap_text="Нет производителя мониторинга в матрице",
            limit=2,
        )

    category_map = {
        "mfa": (["MFA", "IAM"], [], ["ManageEngine"]),
        "iam": (["IAM", "IGA"], ["ManageEngine", "Wallix", "CyberArk"], []),
        "siem_soc": (["SOC", "SIEM"], ["Fortinet", "ManageEngine", "Splunk", "IBM", "Rapid7", "Palo Alto"], []),
        "change_management": (["ITSM"], ["ManageEngine", "Ivanti", "Broadcom (Symantec)"], []),
        "patch": (["VM", "Patch Management"], ["Qualys", "Tenable", "Rapid7", "Ivanti"], []),
        "endpoint_detection": (["EDR", "XDR"], ["Fortinet", "Check Point", "CrowdStrike", "Trend Micro"], []),
        "backup": (["Backup"], [], []),
        "web_waf": (["WAF"], ["Cloudflare", "Fortinet", "Check Point", "F5", "Imperva"], []),
        "pam": (["PAM"], ["Wallix", "CyberArk", "BeyondTrust"], ["ManageEngine"]),
        "dlp": (["DLP"], [], []),
        "mail": (["Email", "Mail Security"], ["Check Point", "Fortinet", "Trend Micro", "Forcepoint"], []),
        "segmentation": (["NAC", "NGFW", "Network Equipment"], ["Fortinet", "Cisco", "Huawei", "Check Point"], []),
        "nac": (["NAC"], ["Fortinet", "Cisco", "Check Point"], []),
        "virtualization": (["Virtualization"], [], []),
        "storage": (["Storage", "Backup"], [], []),
        "dr": (["Backup", "DR"], [], []),
        "appsec": (["SAST", "DAST", "SCA", "VM"], ["Qualys", "Checkmarx", "HCL AppScan"], []),
        "wifi_capacity": (["Network Equipment", "Wireless", "Wi-Fi"], ["Cisco", "Huawei", "Fortinet"], []),
        "network_performance": (["Network Equipment", "SD-WAN", "Monitoring", "NMS"], ["Fortinet", "Cisco", "Huawei", "Check Point", "ManageEngine"], ["Broadcom (Symantec)"]),
        "itam": (["ITAM", "ITSM"], ["ManageEngine", "Ivanti"], []),
    }

    if key in category_map:
        categories, preferred, exclude = category_map[key]
        return portfolio_vendors_by_categories(
            categories,
            preferred=preferred,
            exclude=exclude,
            gap_text="Нет подходящего производителя в матрице",
            limit=4 if key in {"segmentation", "nac"} else 6,
        )

    return manufacturers_for_report_item(item)


def load_solution_vendor_map():
    solution_aliases = {
        "DLP": ("dlp", "утеч", "защита данных"),
        "Шифрование и маскирование данных": ("шифр", "маскир"),
        "Системы контроля доступа и управления привилегиями": (
            "pam", "iam", "privileged", "привилег", "учетн"
        ),
        "Архивирование и резервное копирование": (
            "backup", "резерв", "архив", "восстановлен", "rto", "rpo"
        ),
        "Защита баз данных и аудит": ("database", "баз дан", "db audit"),
        "NGFW": ("ngfw", "firewall", "межсет", "периметр"),
        "Сетевое оборудование": ("сетевое оборудование", "vlan", "маршрут", "коммутатор", "switch", "wifi", "wi fi"),
        "IDS/IPS": ("ids", "ips"),
        "WAF": ("waf", "web application security", "веб прилож", "owasp"),
        "AntiDDos": ("antiddos", "anti ddos", "anti-ddos", "ddos"),
        "Защита облаков": ("cloud", "облак", "casb"),
        "Cyber Risk Management": ("vulnerability", "уязв", "cve", "scanner", "сканер"),
        "Мониторинг и логирование": ("siem", "soar", "soc", "логирование", "журнал"),
        "Защита почты": ("mail security", "почт", "email", "фишинг"),
        "CASB, разработка и защита контейнеров": (
            "casb", "container", "контейнер", "sast", "dast", "appsec", "разработ"
        ),
        "MDM": ("mdm",),
        "Antiransomeware/EDR": ("edr", "xdr", "mdr", "epp", "endpoint", "конечн"),
        "Мультифаторная аутификация": ("mfa", "2fa", "многофактор"),
        "ITSM/CMDB": (
            "itsm", "cmdb", "change management", "configuration management",
            "управление изменениями", "управление конфигурациями",
            "учет конфигураций", "учета конфигураций", "система учета конфигураций"
        ),
        "Миграция и виртуализация": (
            "миграция ос", "миграция", "замена оборудования",
            "виртуализация", "virtualization", "migration project"
        ),
    }
    detailed_matrix = load_detailed_solution_vendor_map()
    if detailed_matrix:
        solutions = list(solution_aliases)
        solution_categories = {
            solutions[0]: ("DLP",),
            solutions[1]: ("Encryption",),
            solutions[2]: ("PAM", "IGA"),
            solutions[3]: ("Backup", "Archiving"),
            solutions[4]: ("DAM/DB Security", "DB Security", "Data Discovery", "Data Classification"),
            solutions[5]: ("NGFW",),
            solutions[6]: ("Network Equipment",),
            solutions[7]: ("IDS/IPS",),
            solutions[8]: ("WAF",),
            solutions[9]: ("Anti-DDoS",),
            solutions[10]: ("Cloud Security", "CASB", "CNAPP", "CWPP"),
            solutions[11]: ("VM", "ASM", "Cyber Risk"),
            solutions[12]: ("SIEM", "SOAR", "UEBA"),
            solutions[13]: ("Email Security", "Email", "Corporate Email"),
            solutions[14]: ("CASB", "CWPP", "CNAPP"),
            solutions[15]: ("MDM", "MDM/UEM"),
            solutions[16]: ("EDR", "XDR", "AV"),
            solutions[17]: ("MFA",),
            solutions[18]: ("ITSM", "ITAM"),
            solutions[19]: ("Servers", "Storage", "Operating Systems", "Virtualization"),
        }
        vendor_map = {}
        for solution, categories in solution_categories.items():
            vendors = []
            for category in categories:
                vendors.extend(detailed_matrix.get(category, []))
            vendor_map[solution] = list(dict.fromkeys(vendors))
        return vendor_map, solution_aliases

    solution_vendor_keywords = {
        "DLP": ("symantec", "forcepoint", "zecurion", "ibatyr", "гарда"),
        "Шифрование и маскирование данных": ("symantec", "thales", "imperva", "гарда"),
        "Системы контроля доступа и управления привилегиями": (
            "cyberark", "axidian", "netwrix", "fudo"
        ),
        "Архивирование и резервное копирование": ("veeam", "commvault", "veritas"),
        "Защита баз данных и аудит": ("imperva", "гарда"),
        "NGFW": ("check point", "palo alto", "fortinet", "cisco", "huawei"),
        "Сетевое оборудование": ("cisco", "huawei", "juniper", "h3c", "hp", "dell"),
        "IDS/IPS": ("trend micro", "forcepoint", "check point", "fortinet"),
        "WAF": ("imperva", "f5", "cloudflare", "radware", "a10", "check point"),
        "AntiDDos": ("f5", "radware", "check point", "barracuda"),
        "Защита облаков": ("check point", "palo alto", "fortinet", "crowdstrike", "trend micro"),
        "Cyber Risk Management": ("tenable", "qualys", "positive", "harmony"),
        "Мониторинг и логирование": ("positive", "касперского", "netwrix", "splunc", "r vision", "qazsiem", "mccafee"),
        "Защита почты": ("trend micro", "barracuda", "forcepoint", "fortinet"),
        "CASB, разработка и защита контейнеров": (
            "check point", "palo alto", "fortinet", "forcepoint", "qualys", "hcl", "symantec", "trend micro"
        ),
        "MDM": ("kaspersky", "bitdefender", "eset"),
        "Antiransomeware/EDR": (
            "trend micro", "kaspersky", "check point", "symantec", "bitdefender", "fortinet", "eset", "crowdstrike"
        ),
        "Мультифаторная аутификация": ("eset", "axidian", "thales"),
        "ITSM/CMDB": (),
        "Миграция и виртуализация": (),
    }

    catalog = [
        clean_vendor_display_name(vendor)
        for vendor in load_vendor_names()
        if str(vendor).strip()
    ]
    vendor_map = {}

    for solution, keywords in solution_vendor_keywords.items():
        matched = []
        for vendor in catalog:
            normalized_vendor = normalize_vendor_key(vendor)
            if any(normalize_vendor_key(keyword) in normalized_vendor for keyword in keywords):
                matched.append(vendor)
        vendor_map[solution] = list(dict.fromkeys(matched))

    return vendor_map, solution_aliases


def manufacturers_for_report_item(item):
    vendor_map, solution_aliases = load_solution_vendor_map()
    existing_values = item.get("vendors", [])
    if not isinstance(existing_values, list):
        existing_values = [existing_values] if existing_values else []

    solution_text = normalize_vendor_key(" ".join(str(value) for value in existing_values))
    fallback_text = normalize_vendor_key(" ".join([
        str(item.get("risk", "")),
        str(item.get("description", "")),
    ]))

    manufacturers = []
    matched_solution = False

    detailed_matrix = load_detailed_solution_vendor_map()
    for category, vendors in detailed_matrix.items():
        if normalize_vendor_key(category) in solution_text:
            manufacturers.extend(vendors)
            matched_solution = True

    for solution, aliases in solution_aliases.items():
        if any(normalize_vendor_key(alias) in solution_text for alias in aliases):
            manufacturers.extend(vendor_map.get(solution, []))
            matched_solution = True

    if not matched_solution:
        for solution, aliases in solution_aliases.items():
            if any(normalize_vendor_key(alias) in fallback_text for alias in aliases):
                manufacturers.extend(vendor_map.get(solution, []))

    all_known_vendors = []
    for vendors in vendor_map.values():
        all_known_vendors.extend(vendors)

    for value in existing_values:
        normalized_value = normalize_vendor_key(value)
        for known_vendor in all_known_vendors:
            if normalize_vendor_key(known_vendor) == normalized_value:
                manufacturers.append(known_vendor)

    return ", ".join(list(dict.fromkeys(manufacturers))[:8]) or "-"


def build_sales_opportunities(results, context, roadmap_items):
    catalog = load_vendor_names()
    opportunities = []

    def add(priority, problem, offer, trigger, vendors, next_step):
        opportunities.append({
            "priority": priority,
            "problem": problem,
            "offer": offer,
            "trigger": trigger,
            "vendors": ", ".join(vendors),
            "next_step": next_step,
            "source": "Базовые правила",
        })

    if results.get("MFA") == "Нет":
        add(
            "P1",
            "Нет MFA для критичных доступов",
            "Проект MFA/IAM для администраторов, почты, VPN и критичных систем",
            "В анкете не указана многофакторная аутентификация.",
            pick_catalog_vendors(catalog, ["axidian", "cisco", "forti"], ["Cisco Duo", "FortiAuthenticator", "Axidian"]),
            "Уточнить текущий IdP/почтовую платформу и предложить пилот MFA без привязки к PAM-проекту.",
        )

    if results.get("EDR") == "Нет":
        add(
            "P1",
            "Есть EPP, но нет EDR/XDR",
            "Endpoint Detection & Response / MDR как развитие текущей защиты рабочих мест",
            "EPP закрывает базовую защиту, но не дает полноценного расследования атак.",
            pick_catalog_vendors(catalog, ["crowdstrike", "kaspersky", "trend", "bitdefender", "eset"], ["CrowdStrike", "Kaspersky", "Trend Micro"]),
            "Показать разницу EPP vs EDR на примерах ransomware/lateral movement.",
        )

    if results.get("Patch Management") == "Нет":
        add(
            "P1",
            "Нет централизованного Patch Management",
            "Проект управления обновлениями и уязвимостями",
            "Ручное обновление не масштабируется на текущий парк АРМ и серверов.",
            pick_catalog_vendors(catalog, ["tenable", "qualys", "positive", "kaspersky"], ["Tenable", "Qualys", "Positive Technologies"]),
            "Предложить инвентаризацию и отчет по критичным CVE за 1-2 недели.",
        )

    if results.get("SIEM") == "Нет" and not context.get("small_company"):
        add(
            "P2",
            "Нет централизованного мониторинга событий ИБ",
            "SIEM/SOC или MSSP-мониторинг для критичных систем",
            "Масштаб инфраструктуры и бизнес-системы требуют управляемого мониторинга.",
            pick_catalog_vendors(catalog, ["r-vision", "qazsiem", "r‑vision"], ["R-Vision SIEM", "QazSIEM"]),
            "Обсудить минимальный scope: NGFW, серверы, AD/учетки, EPP, backup.",
        )

    if results.get("WAF") == "Нет" and context.get("has_public_web"):
        add(
            "P2",
            "Публичные веб-ресурсы без WAF",
            "WAF/CDN/DDoS-защита для интернет-магазина и личного кабинета",
            "Публичная поверхность увеличивает риск атак на приложения и доступность.",
            pick_catalog_vendors(catalog, ["imperva", "f5", "cloudflare", "radware", "a10", "barracuda"], ["Imperva", "F5", "Cloudflare"]),
            "Предложить экспресс-аудит web-периметра и пилот WAF/CDN.",
        )

    if results.get("Резервное копирование") == "Нет":
        add(
            "P1",
            "Не указан backup-контур",
            "Резервное копирование и восстановление критичных сервисов",
            "Без backup бизнес не сможет гарантированно восстановиться после сбоя/ransomware.",
            pick_catalog_vendors(catalog, ["veeam", "commvault", "veritas"], ["Veeam", "Commvault", "Veritas"]),
            "Запросить RPO/RTO и предложить дизайн backup-политик.",
        )
    elif results.get("Immutable Backup") == "Нет":
        add(
            "P2",
            "Backup есть, но не указана защита копий от ransomware",
            "Immutable/offline backup и тест восстановления",
            "Обычные backup-копии могут быть удалены или зашифрованы атакующим.",
            pick_catalog_vendors(catalog, ["veeam", "commvault", "veritas"], ["Veeam", "Commvault", "Veritas"]),
            "Продать assessment текущего backup и сценарий контрольного восстановления.",
        )

    if context.get("servers", 0) >= 10 and results.get("DR") == "Нет":
        add(
            "P2",
            "Не описан DR-план для критичных ИТ-сервисов",
            "DR assessment, RTO/RPO-дизайн и тест восстановления",
            "Есть серверы, виртуализация и бизнес-системы, но не указан формализованный DR-сценарий.",
            pick_catalog_vendors(catalog, ["veeam", "commvault", "rubrik", "veritas"], ["Veeam", "Commvault", "Rubrik"]),
            "Запросить список критичных сервисов и предложить воркшоп по RTO/RPO с контрольным восстановлением.",
        )

    if context.get("servers", 0) >= 10 and results.get("Мониторинг") == "Нет":
        add(
            "P2",
            "Нет единого эксплуатационного мониторинга ИТ",
            "Мониторинг серверов, сети, СХД, виртуализации и бизнес-сервисов",
            "Инциденты по емкости, доступности и производительности могут обнаруживаться после влияния на пользователей.",
            pick_catalog_vendors(catalog, ["zabbix", "prtg", "manageengine"], ["Zabbix", "PRTG", "ManageEngine"]),
            "Предложить быстрый health-check мониторинга и пилот с 10-15 критичными объектами.",
        )

    if results.get("Виртуализация") != "Нет" and results.get("Кластеризация") == "Нет":
        add(
            "P3",
            "Нужно проверить отказоустойчивость виртуализации",
            "Аудит HA/DRS, резервов CPU/RAM/storage и размещения критичных VM",
            "Виртуальные серверы обслуживают бизнес-системы, но отказ хоста может затронуть сразу несколько сервисов.",
            pick_catalog_vendors(catalog, ["vmware", "veeam", "zabbix"], ["VMware", "Veeam", "Zabbix"]),
            "Предложить инфраструктурный аудит виртуализации и карту рисков по критичным VM.",
        )

    if results.get("СХД") != "Нет" and results.get("Мониторинг СХД") == "Нет":
        add(
            "P3",
            "Не описан capacity management СХД",
            "Контроль емкости, производительности и snapshot-политик СХД",
            "Рост данных или деградация RAID/latency может повлиять на ERP, CRM и файловые сервисы.",
            pick_catalog_vendors(catalog, ["zabbix", "prtg", "manageengine"], ["Zabbix", "PRTG", "ManageEngine"]),
            "Запросить модель СХД, текущую утилизацию и предложить capacity/performance assessment.",
        )

    if results.get("PAM") == "Нет" and (context.get("servers", 0) >= 10 or context.get("has_critical_systems")):
        add(
            "P2",
            "Нет PAM для привилегированных учетных записей",
            "Privileged Access Management",
            "Есть серверы/критичные системы, но не указан контроль администраторского доступа.",
            pick_catalog_vendors(catalog, ["cyberark", "fudo", "netwrix"], ["CyberArk", "Fudo", "Netwrix"]),
            "Предложить обследование админ-доступов и пилот vault/session recording.",
        )

    if not opportunities and roadmap_items:
        first = roadmap_items[0]
        add(
            first["priority"],
            "Требуется приоритизация улучшений",
            "Экспертный воркшоп и технический пресейл по roadmap",
            first["rationale"],
            catalog[:4] or ["Khalil Trade"],
            "Назначить встречу с ИТ/ИБ и пройти roadmap по шагам.",
        )

    priority_order = {"P1": 1, "P2": 2, "P3": 3}
    return sorted(opportunities, key=lambda item: priority_order.get(item["priority"], 99))[:10]


def result_contains_any(results, markers):
    text = normalize_vendor_key(" ".join(str(value) for value in results.values()))
    return any(normalize_vendor_key(marker) in text for marker in markers)


def sales_override_for_item(item, results, context):
    combined = normalize_vendor_key(" ".join([
        str(item.get("risk", "")),
        str(item.get("description", "")),
        str(item.get("impact", "")),
        str(item.get("recommendation", "")),
        " ".join(str(value) for value in item.get("vendors", []) if value)
        if isinstance(item.get("vendors"), list)
        else str(item.get("vendors", "")),
    ]))

    has_fortinet = result_contains_any(results, ["fortinet", "fortigate", "forti"])
    has_cloudflare = result_contains_any(results, ["cloudflare"])
    has_m365 = result_contains_any(results, ["microsoft 365", "office 365", "m365", "exchange online"])

    def has(*markers):
        return any(normalize_vendor_key(marker) in combined for marker in markers)

    if has("устаревш", "legacy", "windows xp", "windows 7", "windows 8", "windows server 2008", "2012 r2"):
        return {
            "priority": "P1",
            "problem": "Устаревшие ОС требуют миграции, а не компенсации ИБ-продуктами",
            "offer": (
                "Основной проект - обновление или вывод устаревших Windows/Windows Server. "
                "Если часть систем нельзя быстро обновить, параллельно провести vulnerability assessment, "
                "изоляцию сегмента и контроль исключений до даты миграции."
            ),
            "trigger": item.get("impact") or "Устаревшие ОС не получают полноценные исправления и не должны закрываться покупкой EDR/DLP вместо миграции.",
            "vendors": portfolio_vendors_by_categories(
                ["Operating Systems", "VM"],
                preferred=["Microsoft", "Qualys", "Tenable", "Rapid7"],
            ),
            "next_step": "Уточнить количество legacy-хостов, зависимые приложения, ограничения миграции и предложить план: обновление, изоляция, сканирование уязвимостей, контроль срока вывода.",
            "source": item.get("source", "ИИ"),
        }

    if has("mfa", "многофактор", "2fa"):
        return {
            "priority": "P1",
            "problem": "MFA для критичных доступов",
            "offer": "Запустить MFA-проект для Microsoft 365, администраторов, VPN/удаленного доступа и критичных бизнес-систем; PAM рассматривать следующим этапом, а не вместо MFA.",
            "trigger": item.get("impact") or "Microsoft 365 и критичные доступы без MFA дают высокий риск компрометации учетных записей.",
            "vendors": portfolio_vendors_by_categories(
                ["MFA"],
                exclude=["ManageEngine"],
                gap_text="Нет корректного MFA-вендора в матрице: добавьте FortiAuthenticator / Microsoft Entra ID / Cisco Duo",
            ),
            "next_step": "Сначала проверить Microsoft 365/Entra ID, FortiGate SSL VPN, администраторские учетные записи и исключения; затем предложить быстрый MFA-пилот.",
            "source": item.get("source", "ИИ"),
        }

    if has("vulnerability", "уязв", "cve", "scanner", "сканер", "patch management"):
        return {
            "priority": "P1" if context.get("users", 0) >= 100 else "P2",
            "problem": "Нужен управляемый цикл уязвимостей и обновлений",
            "offer": "Vulnerability Management: инвентаризация активов, регулярное сканирование, SLA на критичные CVE и отчет по исключениям.",
            "trigger": item.get("impact") or "Для парка 100+ АРМ и серверов ручной контроль CVE быстро теряет управляемость.",
            "vendors": portfolio_vendors_by_categories(
                ["VM"],
                preferred=["Qualys", "Tenable", "Rapid7"],
            ),
            "next_step": "Предложить быстрый assessment на внешнем периметре, серверах и рабочих местах, затем показать топ критичных CVE и план закрытия.",
            "source": item.get("source", "ИИ"),
        }

    if has("mail security", "почт", "email", "фишинг", "phishing"):
        return {
            "priority": "P1" if has_m365 else "P2",
            "problem": "Защита облачной почты и anti-phishing",
            "offer": "Усилить Microsoft 365 почту отдельным mail security/anti-phishing контуром: URL/attachment sandboxing, impersonation protection, DMARC-контроль и обучение пользователей.",
            "trigger": item.get("impact") or "Microsoft 365 часто становится первой точкой атаки через фишинг и компрометацию учетных записей.",
            "vendors": portfolio_vendors_by_categories(
                ["Email"],
                preferred=["Check Point", "Fortinet", "Trend Micro", "Forcepoint"],
            ),
            "next_step": "Проверить текущие политики M365, SPF/DKIM/DMARC, статистику фишинга и предложить пилот защиты почты.",
            "source": item.get("source", "ИИ"),
        }

    if has("waf", "web application", "веб прилож", "owasp"):
        return {
            "priority": "P2",
            "problem": item.get("risk") or "Публичные веб-сервисы требуют WAF",
            "offer": "Провести экспресс-оценку web-периметра и текущих Cloudflare-политик; по результатам сравнить усиление действующего контура с FortiWeb/F5/Imperva и выбрать архитектуру WAF/CDN под критичные приложения.",
            "trigger": item.get("impact") or "Интернет-магазин и личный кабинет формируют публичную поверхность атаки.",
            "vendors": portfolio_vendors_by_categories(
                ["WAF"],
                preferred=["Fortinet", "Check Point", "Cloudflare", "F5", "Imperva"],
                limit=4,
            ),
            "next_step": "Снять список публичных доменов, текущие Cloudflare-политики, критичные URL и предложить WAF health-check.",
            "source": item.get("source", "ИИ"),
        }

    if has("ids", "ips", "сетевых атак", "сетевыми атаками"):
        return {
            "priority": "P3",
            "problem": item.get("risk") or "IDS/IPS требует уточнения в рамках текущего NGFW",
            "offer": "Проверить, используются ли IPS-профили на FortiGate/NGFW и есть ли потребность в отдельном IDS/IPS-контуре.",
            "trigger": item.get("impact") or "В матрице портфеля нет отдельной категории IDS/IPS, поэтому нельзя подставлять произвольных производителей.",
            "vendors": portfolio_vendors_by_categories(
                ["IDS/IPS"],
                gap_text="Нет категории IDS/IPS в матрице: добавьте производителей или используйте NGFW/IPS текущего стека",
            ),
            "next_step": "Уточнить модели NGFW, включенные IPS-профили, сегменты инспекции и текущие события блокировок.",
            "source": item.get("source", "ИИ"),
        }

    if has("edr", "xdr", "endpoint", "конечн"):
        return {
            "priority": "P2",
            "problem": item.get("risk") or "Endpoint-защита требует обнаружения и реагирования",
            "offer": "EDR/XDR-пилот на критичных группах пользователей и серверах с регламентом реагирования и метриками MTTD/MTTR.",
            "trigger": item.get("impact") or "EPP снижает базовый malware-риск, но не закрывает расследование сложных атак.",
            "vendors": portfolio_vendors_by_categories(
                ["EDR", "XDR"],
                preferred=["Fortinet", "Check Point", "CrowdStrike", "Trend Micro"],
                limit=4,
            ),
            "next_step": "Сравнить текущий EPP, определить пилотную группу и показать сценарии ransomware/lateral movement.",
            "source": item.get("source", "ИИ"),
        }

    if has("siem", "soc", "mssp", "мониторинг событий"):
        return {
            "priority": "P2",
            "problem": item.get("risk") or "Нужен управляемый мониторинг событий ИБ",
            "offer": "Начать с MSSP/SOC или легкого SIEM-scope: FortiGate, Windows Server/AD, Microsoft 365, EPP/EDR, backup и web.",
            "trigger": item.get("impact") or "Для 120 АРМ и 22 серверов реалистичнее начать с управляемого мониторинга, а не тяжелого enterprise-проекта.",
            "vendors": portfolio_vendors_by_categories(
                ["SIEM", "SOAR", "UEBA"],
                preferred=["Fortinet", "IBM", "Splunk"],
                limit=4,
            ),
            "next_step": "Согласовать минимальный scope источников логов и показать формат ежемесячного отчета по инцидентам.",
            "source": item.get("source", "ИИ"),
        }

    if has("sast", "dast", "appsec", "разработ"):
        return {
            "priority": "P3",
            "problem": item.get("risk") or "Безопасность разработки требует уточнения",
            "offer": "AppSec assessment: определить реальные языки, CI/CD, публичные приложения и только после этого выбирать SAST/DAST/SCA.",
            "trigger": item.get("impact") or "Без фактов о процессе разработки нельзя подбирать общих ИБ-вендоров вместо AppSec-инструментов.",
            "vendors": portfolio_vendors_by_categories(
                ["SAST", "DAST", "SCA", "Application Security"],
                preferred=["Checkmarx", "HCL AppScan", "Positive Technologies", "Qualys"],
                gap_text="Нет AppSec-вендоров в матрице: добавьте SAST/DAST/SCA производителей",
            ),
            "next_step": "Уточнить наличие разработки, репозитории, CI/CD и критичные приложения; затем предложить пилот SAST/DAST.",
            "source": item.get("source", "ИИ"),
        }

    if has("схд", "storage", "raid", "snapshot", "iops", "latency"):
        return {
            "priority": "P3",
            "problem": "СХД: требуется проверка capacity/performance, без вывода о проблеме надежности",
            "offer": "Не продавать замену СХД без фактов. Предложить health-check: утилизация, latency/IOPS, состояние RAID, snapshot-политики, backup integration.",
            "trigger": "В анкете есть RAID, диски, Veeam/Snapshots; доказательств отказов или нехватки производительности нет.",
            "vendors": portfolio_vendors_by_categories(
                ["Backup", "Storage"],
                preferred=["Veeam"],
            ),
            "next_step": "Запросить модель СХД, текущую утилизацию, метрики latency/IOPS, расписание snapshot и результаты тестов восстановления.",
            "source": item.get("source", "ИИ"),
        }

    if has("wifi", "wi fi", "wi-fi", "nac", "сегментац", "vlan", "acl"):
        return {
            "priority": "P3",
            "problem": "Требуется уточнить архитектуру сегментации сети и Wi-Fi",
            "offer": "Не трактовать отсутствие NAC как отсутствие сегментации. Проверить VLAN/ACL/FortiGate policies, guest Wi-Fi, доступ AP и правила между сегментами.",
            "trigger": "В анкете есть UniFi/Wi-Fi 6/FortiGate/Cisco/Huawei, но нет фактов о VLAN/ACL и межсегментных политиках.",
            "vendors": portfolio_vendors_by_categories(
                ["NGFW", "Network Equipment", "NAC"],
                preferred=["Fortinet", "Cisco", "Huawei"],
            ),
            "next_step": "Попросить схему VLAN, правила FortiGate, SSID/guest Wi-Fi и список критичных сегментов.",
            "source": item.get("source", "ИИ"),
        }

    if has("pam", "привилег"):
        return {
            "priority": "P3",
            "problem": item.get("risk") or "PAM для привилегированных учетных записей",
            "offer": "PAM рассматривать после MFA: инвентаризация админов, vault, session recording и регулярный пересмотр прав.",
            "trigger": item.get("impact") or "PAM важен для серверов и критичных систем, но коммерчески должен идти после закрытия MFA.",
            "vendors": portfolio_vendors_by_categories(
                ["PAM"],
                preferred=["Wallix", "CyberArk", "BeyondTrust"],
                exclude=["ManageEngine"],
            ),
            "next_step": "Сначала закрыть MFA, затем собрать список привилегированных учетных записей и предложить PAM-пилот.",
            "source": item.get("source", "ИИ"),
        }

    if has("change management", "configuration management", "cmdb", "управление изменениями", "конфигурац"):
        return {
            "priority": "P3",
            "problem": item.get("risk") or "Управление изменениями требует формализации",
            "offer": "Проводить как процессный ITSM/CMDB воркшоп, а не как первоочередную продуктовую продажу.",
            "trigger": item.get("impact") or "Без интервью нельзя утверждать, что процесс отсутствует; можно только проверить зрелость.",
            "vendors": portfolio_vendors_by_categories(
                ["ITSM", "ITAM"],
                preferred=["ManageEngine"],
                gap_text="Нет ITSM/CMDB-вендора в матрице",
            ),
            "next_step": "Уточнить, где ведутся заявки/изменения/активы, и предложить короткий ITSM maturity workshop.",
            "source": item.get("source", "ИИ"),
        }

    return None


def build_ai_first_sales_opportunities(risk_sources, results=None, context=None):
    if not isinstance(risk_sources, list):
        return []

    results = results or {}
    context = context or {}
    opportunities = []
    priority_order = {"P1": 1, "P2": 2, "P3": 3}
    level_priority = {
        "Критический": "P1",
        "Высокий": "P1",
        "Средний": "P2",
        "Низкий": "P3",
        "CRITICAL": "P1",
        "HIGH": "P1",
        "MEDIUM": "P2",
        "LOW": "P3",
    }

    for item in risk_sources:
        if not isinstance(item, dict):
            continue
        if item.get("source") != "ИИ":
            continue

        override = sales_override_for_item(item, results, context)
        if override:
            opportunities.append(override)
            continue

        vendors_text = portfolio_vendors_for_report_item(item)

        risk = str(item.get("risk", "")).strip()
        recommendation = str(item.get("recommendation", "")).strip()
        if not risk or not recommendation:
            continue

        impact = str(item.get("impact") or item.get("description") or "").strip()
        area = str(item.get("area") or "ИТ/ИБ").strip()
        priority = level_priority.get(str(item.get("level", "")).strip(), "P2")

        opportunities.append({
            "priority": priority,
            "problem": risk,
            "offer": recommendation,
            "trigger": impact or "Вывод сформирован по данным анкеты и экспертному анализу.",
            "vendors": vendors_text or "-",
            "next_step": (
                f"Согласовать с заказчиком факты по домену {area}, подтвердить владельца риска, "
                "оценить текущие ограничения и подготовить короткий план внедрения с бюджетным диапазоном."
            ),
            "source": "ИИ",
        })

    return sorted(opportunities, key=lambda item: priority_order.get(item["priority"], 99))[:10]


def ensure_sales_playbook_priorities(opportunities, results, context):
    opportunities = [item for item in opportunities if isinstance(item, dict)]

    def existing_has(*markers):
        text = normalize_vendor_key(" ".join(
            " ".join(str(item.get(field, "")) for field in ("problem", "offer", "trigger", "vendors"))
            for item in opportunities
        ))
        return any(normalize_vendor_key(marker) in text for marker in markers)

    def add_if_missing(markers, pseudo_item):
        if not existing_has(*markers):
            item = sales_override_for_item(
                {**pseudo_item, "source": "Экспертная приоритизация"},
                results,
                context
            )
            if item:
                opportunities.append(item)

    if results.get("MFA") == "Нет":
        add_if_missing(
            ("mfa", "многофактор"),
            {
                "risk": "MFA отсутствует для критичных доступов",
                "impact": "Компрометация учетных записей остается одним из наиболее вероятных сценариев атаки.",
                "recommendation": "Включить MFA для Microsoft 365, VPN, администраторов и критичных систем.",
                "vendors": ["MFA"],
            }
        )

    if results.get("Patch Management") == "Нет" or results.get("ОС АРМ (Windows XP/Vista/7/8)", 0) or results.get("ОС Сервера (Windows Server 2008/2012 R2)", 0):
        add_if_missing(
            ("vulnerability", "уязв", "cve", "patch"),
            {
                "risk": "Vulnerability Management и контроль обновлений",
                "impact": "Без регулярного сканирования и SLA критичные CVE могут оставаться открытыми.",
                "recommendation": "Запустить vulnerability assessment и процесс устранения критичных уязвимостей.",
                "vendors": ["Vulnerability Management"],
            }
        )

    if results.get("Mail Security") == "Нет" and result_contains_any(results, ["Microsoft 365", "Office 365", "Exchange"]):
        add_if_missing(
            ("mail security", "почт", "фишинг"),
            {
                "risk": "Microsoft 365 требует усиления защиты почты",
                "impact": "Фишинг и impersonation остаются основным входом в атаку.",
                "recommendation": "Проверить SPF/DKIM/DMARC и запустить пилот защиты облачной почты.",
                "vendors": ["Mail Security"],
            }
        )

    if results.get("WAF") == "Нет" and context.get("has_public_web"):
        add_if_missing(
            ("waf", "web application", "веб"),
            {
                "risk": "Публичные web-сервисы требуют WAF",
                "impact": "Интернет-магазин и личный кабинет увеличивают поверхность атаки.",
                "recommendation": "Проверить текущий Cloudflare и рассмотреть WAF/FortiWeb/F5/Imperva.",
                "vendors": ["WAF"],
            }
        )

    if results.get("EDR") == "Нет":
        add_if_missing(
            ("edr", "xdr", "endpoint"),
            {
                "risk": "Endpoint-защита требует EDR/XDR",
                "impact": "EPP не закрывает расследование сложных атак и lateral movement.",
                "recommendation": "Провести EDR/XDR-пилот на критичных группах пользователей и серверов.",
                "vendors": ["EDR/XDR"],
            }
        )

    if results.get("SIEM") == "Нет" and not context.get("small_company"):
        add_if_missing(
            ("siem", "soc", "mssp"),
            {
                "risk": "SIEM/SOC/MSSP для критичных источников",
                "impact": "Разрозненные журналы увеличивают время обнаружения и расследования.",
                "recommendation": "Начать с MSSP/SOC или минимального SIEM-scope.",
                "vendors": ["SIEM"],
            }
        )

    if results.get("PAM") == "Нет" and (context.get("servers", 0) >= 10 or context.get("has_critical_systems")):
        add_if_missing(
            ("pam", "привилег"),
            {
                "risk": "PAM для привилегированных учетных записей",
                "impact": "Администраторские доступы требуют отдельного контроля, но после MFA.",
                "recommendation": "После MFA провести инвентаризацию админов и PAM-пилот.",
                "vendors": ["PAM"],
            }
        )

    priority_order = {"P1": 1, "P2": 2, "P3": 3}
    seen = set()
    cleaned = []
    for item in sorted(opportunities, key=lambda row: priority_order.get(row.get("priority"), 99)):
        key = risk_semantic_key({
            "risk": item.get("problem", ""),
            "description": item.get("trigger", ""),
            "recommendation": item.get("offer", ""),
        })
        if key and key in seen:
            continue
        cleaned.append(item)
        seen.add(key)
        if len(cleaned) >= 10:
            break
    return cleaned


def sales_account_guidance(item):
    key = risk_semantic_key({
        "risk": item.get("problem", item.get("risk", "")),
        "description": item.get("trigger", item.get("description", "")),
        "impact": item.get("business_value", item.get("impact", "")),
        "recommendation": item.get("offer", item.get("recommendation", "")),
    })
    profiles = {
        "legacy_os": {
            "business_value": "Снизить риск простоя и зависимости от неподдерживаемых платформ, зафиксировав контролируемый план вывода legacy-систем.",
            "stakeholders": "ИТ-директор; владелец зависимого бизнес-приложения; инфраструктурная команда; финансы/закупки",
            "qualification": "Количество legacy-узлов, зависимые приложения, допустимое окно миграции, владелец и целевая дата вывода.",
            "meeting_goal": "Согласовать реестр legacy-систем и выбрать первый контур для миграции или временной изоляции.",
        },
        "mfa": {
            "business_value": "Снизить вероятность захвата учетных записей без изменения основных бизнес-процессов.",
            "stakeholders": "ИТ-директор; руководитель ИБ; владельцы Microsoft 365/AD/VPN; служба поддержки",
            "qualification": "IdP, охват пользователей, критичные доступы, исключения, способы аутентификации и требования к пользовательскому опыту.",
            "meeting_goal": "Определить 1-2 критичных контура для быстрого MFA-пилота и критерии его успешности.",
        },
        "patch": {
            "business_value": "Сделать технический долг по уязвимостям измеримым и сократить время закрытия критичных CVE.",
            "stakeholders": "ИТ-директор; руководители серверной и endpoint-команд; ИБ; владельцы критичных приложений",
            "qualification": "Полнота инвентаризации, текущие окна обновлений, SLA по CVE, исключения и ответственные за remediation.",
            "meeting_goal": "Согласовать scope assessment и формат отчета, по которому будет принято решение о платформе или сервисе.",
        },
        "endpoint_detection": {
            "business_value": "Сократить время обнаружения и локализации атак на рабочих местах и серверах.",
            "stakeholders": "Руководитель ИБ; endpoint-команда; серверная команда; Service Desk; SOC/MSSP",
            "qualification": "Текущий EPP, покрытие агентов, сценарии реагирования, MTTD/MTTR, пилотная группа и доступный ресурс мониторинга.",
            "meeting_goal": "Выбрать пилотную группу и 3-4 сценария проверки EDR/MDR без дублирования действующего EPP.",
        },
        "mail": {
            "business_value": "Снизить риск фишинга, подмены отправителя и компрометации облачной почты.",
            "stakeholders": "Руководитель ИБ; владелец Microsoft 365/почты; Service Desk; HR/обучение",
            "qualification": "SPF/DKIM/DMARC, статистика фишинга, текущие политики M365, VIP-пользователи и процесс разбора писем.",
            "meeting_goal": "Подтвердить сценарии атак на почту и согласовать ограниченный пилот с измеримыми метриками.",
        },
        "siem_soc": {
            "business_value": "Сделать обнаружение и расследование инцидентов управляемым на критичных источниках событий.",
            "stakeholders": "Руководитель ИБ; ИТ-директор; владельцы AD/NGFW/серверов/endpoint; SOC или MSSP",
            "qualification": "Критичные источники, объем событий, use cases, режим дежурства, SLA реагирования и доступный операционный ресурс.",
            "meeting_goal": "Определить минимальный scope мониторинга и выбрать модель: собственная эксплуатация, SOC или MSSP.",
        },
        "web_waf": {
            "business_value": "Снизить риск простоя и компрометации публичных клиентских сервисов.",
            "stakeholders": "Владелец e-commerce/продукта; ИБ; разработка; DevOps; инфраструктура",
            "qualification": "Домены, критичные URL/API, текущие Cloudflare-политики, OWASP-события, требования к доступности и владельцы приложений.",
            "meeting_goal": "Согласовать web health-check и перечень приложений для пилота WAF/CDN.",
        },
        "segmentation": {
            "business_value": "Подтвердить фактическую изоляцию критичных контуров до обсуждения продукта или проекта изменений.",
            "stakeholders": "Сетевой архитектор; владелец NGFW; ИБ; владельцы серверного, пользовательского и гостевого контуров",
            "qualification": "Схема VLAN/VRF, ACL, зоны NGFW, матрица потоков, guest Wi-Fi и правила доступа к критичным сегментам.",
            "meeting_goal": "Провести архитектурную сверку и отделить подтвержденные разрывы от предположений анкеты.",
        },
        "it_monitoring": {
            "business_value": "Сократить время обнаружения деградаций и перевести доступность сервисов в измеримые показатели.",
            "stakeholders": "ИТ-директор; инфраструктурная команда; Service Desk; владельцы ERP/CRM и других критичных сервисов",
            "qualification": "Карта сервисов, текущие инструменты, пороги, on-call, SLA, отчеты по доступности и capacity.",
            "meeting_goal": "Выбрать 3-5 критичных сервисов и согласовать monitoring assessment с целевыми метриками.",
        },
    }
    return profiles.get(key, {
        "business_value": "Перевести выявленную зону риска в проверяемый план с владельцем, сроком и критерием результата.",
        "stakeholders": "ИТ-директор; руководитель ИБ; технический владелец; владелец затронутого бизнес-процесса",
        "qualification": "Текущий процесс, фактический разрыв, бизнес-влияние, владелец, срок, бюджетный контур и критерий выбора.",
        "meeting_goal": "Подтвердить проблему и согласовать следующий артефакт: assessment, воркшоп, пилот или расчет проекта.",
    })


def build_sales_conversation_pack(c_info, results, context, roadmap_items, opportunities):
    company = c_info.get("Наименование компании", "клиент")
    users = context.get("users", 0)
    servers = context.get("servers", 0)
    _, profile_text = infrastructure_profile(context)

    pains = []

    def add_pain(priority, pain, evidence, commercial_angle, discovery_question):
        pains.append({
            "priority": priority,
            "pain": pain,
            "evidence": evidence,
            "commercial_angle": commercial_angle,
            "discovery_question": discovery_question,
        })

    legacy_arm = int(results.get("ОС АРМ (Windows XP/Vista/7/8)", 0) or 0)
    legacy_servers = int(results.get("ОС Сервера (Windows Server 2008/2012 R2)", 0) or 0)
    if legacy_arm or legacy_servers:
        add_pain(
            "P1",
            "Неподдерживаемые операционные системы создают технический долг с фиксированным сроком вывода.",
            f"В анкете указано: {russian_count(legacy_arm, 'legacy-АРМ', 'legacy-АРМ', 'legacy-АРМ')} и {russian_count(legacy_servers, 'legacy-сервер', 'legacy-сервера', 'legacy-серверов')}.",
            "Программа миграции legacy ОС с временной изоляцией и контролем исключений.",
            "Какие приложения или оборудование удерживают legacy ОС и какая дата вывода реалистична для каждого узла?"
        )

    if results.get("MFA") == "Нет":
        add_pain(
            "P1",
            "Компрометация учетных записей остается одним из самых быстрых сценариев инцидента.",
            "В анкете не указана MFA для критичных доступов.",
            "MFA/IAM-пилот для администраторов, VPN, почты и критичных систем.",
            "Где сейчас находятся самые критичные учетные записи: почта, VPN, AD, ERP/CRM, облака?"
        )

    if results.get("Резервное копирование") == "Нет":
        add_pain(
            "P1",
            "Нет подтвержденного сценария восстановления после сбоя или ransomware.",
            f"Серверов указано: {servers}; backup-контур не указан.",
            "Backup assessment, дизайн RPO/RTO и внедрение резервного копирования.",
            "Какие сервисы бизнес должен восстановить первыми и за какое время?"
        )
    elif results.get("Immutable Backup") == "Нет":
        add_pain(
            "P2",
            "Backup есть, но его устойчивость к ransomware не подтверждена.",
            "Не указаны immutable/offline-копии и регулярный тест восстановления.",
            "Immutable backup, контрольное восстановление, регламент RTO/RPO.",
            "Когда последний раз проводили тестовое восстановление и кто подписал результат?"
        )

    if results.get("Patch Management") == "Нет":
        add_pain(
            "P1",
            "Риск эксплуатации известных уязвимостей растет быстрее, чем команда успевает обновлять вручную.",
            f"В анкете указано {russian_count(users, 'АРМ', 'АРМ', 'АРМ')} и {russian_count(servers, 'сервер', 'сервера', 'серверов')}; централизованный patch management не указан.",
            "Инвентаризация, vulnerability/patch management, регулярный отчет по критичным CVE.",
            "Как сейчас принимается решение, какие обновления ставить срочно, а какие можно отложить?"
        )

    if results.get("EDR") == "Нет":
        add_pain(
            "P1",
            "Команда может не увидеть сложную атаку на рабочих местах до влияния на бизнес.",
            "EDR/XDR/MDR не указаны; EPP без расследования и реагирования закрывает только базовый уровень.",
            "EDR/MDR-пилот на критичных группах пользователей и серверов.",
            "Есть ли сейчас возможность понять цепочку атаки: пользователь, файл, процесс, сеть, сервер?"
        )

    if results.get("SIEM") == "Нет" and not context.get("small_company"):
        add_pain(
            "P2",
            "События ИБ разрознены, поэтому инциденты сложно обнаруживать и расследовать.",
            "SIEM/SOC не указан при наличии инфраструктуры, требующей централизованного мониторинга.",
            "MSSP/SOC или поэтапное подключение SIEM для минимального critical scope.",
            "Какие источники логов сейчас реально просматриваются ежедневно, а какие только после инцидента?"
        )

    if context.get("has_public_web") and results.get("WAF") == "Нет":
        add_pain(
            "P2",
            "Публичные веб-сервисы увеличивают поверхность атаки и риск простоя.",
            "Публичный web-контур есть, WAF/CDN-защита не указана.",
            "Экспресс-аудит web-периметра, WAF/CDN/DDoS-пилот.",
            "Какие публичные сервисы приносят выручку или обслуживают клиентов напрямую?"
        )

    if context.get("servers", 0) >= 10 and results.get("Мониторинг") == "Нет":
        add_pain(
            "P2",
            "Инфраструктурные сбои могут обнаруживаться после жалоб пользователей.",
            "Есть серверный контур, но эксплуатационный мониторинг не указан.",
            "Мониторинг серверов, сети, СХД, виртуализации и бизнес-сервисов.",
            "Какие показатели сейчас отслеживаются: доступность, диски, latency, backup jobs, сервисы?"
        )

    if not pains:
        add_pain(
            "P3",
            "Анкета не показывает явных критичных разрывов, но нужен экспертный разбор целевой модели.",
            profile_text,
            "Пресейл-воркшоп по roadmap и проверка приоритетов.",
            "Какая зона сейчас больше всего беспокоит бизнес: доступность, безопасность, производительность или compliance?"
        )

    primary_opportunity = opportunities[0] if opportunities else {
        "offer": pains[0]["commercial_angle"],
        "next_step": "Провести экспертный воркшоп и подтвердить приоритет.",
    }
    primary_guidance = sales_account_guidance(primary_opportunity)
    primary_key = risk_semantic_key({
        "risk": primary_opportunity.get("problem", ""),
        "description": primary_opportunity.get("trigger", ""),
        "recommendation": primary_opportunity.get("offer", ""),
    })
    primary_pain = next(
        (
            pain for pain in pains
            if risk_semantic_key({
                "risk": pain.get("pain", ""),
                "description": pain.get("evidence", ""),
                "recommendation": pain.get("commercial_angle", ""),
            }) == primary_key
        ),
        pains[0],
    )

    call_script = [
        {
            "stage": "Подготовка",
            "goal": "Войти в разговор с гипотезой, а не с каталогом продуктов.",
            "talk_track": (
                f"До звонка: проверить роль контакта, выбрать одну P1-гипотезу, подготовить 2 факта из анкеты "
                f"и определить желаемый следующий шаг. Зафиксировано: {russian_count(users, 'АРМ', 'АРМ', 'АРМ')}, "
                f"{russian_count(servers, 'сервер', 'сервера', 'серверов')}."
            )
        },
        {
            "stage": "Открытие и разрешение",
            "goal": "Согласовать формат и получить право задавать вопросы.",
            "talk_track": (
                f"Мы изучили анкету {company}. В ней указано {russian_count(users, 'АРМ', 'АРМ', 'АРМ')} и "
                f"{russian_count(servers, 'сервер', 'сервера', 'серверов')}. "
                "Предлагаю за 20 минут проверить две гипотезы из отчета, понять их влияние на бизнес "
                "и решить, нужен ли отдельный технический разбор. Подходит такой формат?"
            )
        },
        {
            "stage": "Подтверждение факта",
            "goal": "Отделить факт анкеты от предположения.",
            "talk_track": (
                f"Первая гипотеза: {primary_pain['pain']} Основание: {primary_pain['evidence']} "
                "Что из этого подтверждается, а какие компенсирующие меры не попали в анкету?"
            )
        },
        {
            "stage": "Бизнес-влияние",
            "goal": "Связать техническую тему с приоритетом заказчика.",
            "talk_track": (
                f"Если этот сценарий реализуется, какой эффект будет наиболее чувствительным: простой, "
                f"потеря данных, влияние на клиентов или нагрузка на команду? Для нашей гипотезы ценность такова: "
                f"{primary_guidance['business_value']}"
            )
        },
        {
            "stage": "Текущий подход",
            "goal": "Понять процесс, инструменты и ограничения.",
            "talk_track": (
                f"Как сейчас управляется эта зона, кто отвечает за результат и по какой метрике видно, "
                f"что контроль работает? Для квалификации важно уточнить: {primary_guidance['qualification']}"
            )
        },
        {
            "stage": "Почему сейчас",
            "goal": "Проверить срочность без давления.",
            "talk_track": (
                "Есть ли событие, которое задает срок: аудит, инцидент, продление лицензий, новый сервис, "
                "миграция, требование руководства или бюджетный цикл?"
            )
        },
        {
            "stage": "Гипотеза решения",
            "goal": "Предложить следующий шаг, не преждевременную спецификацию.",
            "talk_track": (
                f"С учетом подтвержденных фактов логичный первый шаг: {primary_opportunity.get('offer', pains[0]['commercial_angle'])} "
                "Предлагаю сначала согласовать scope и критерии результата, а затем сравнивать архитектуру и производителей."
            )
        },
        {
            "stage": "Карта решения",
            "goal": "Выявить ЛПР, технических участников и порядок закупки.",
            "talk_track": (
                f"Кого нужно подключить, чтобы подтвердить архитектуру и принять решение? Обычно здесь участвуют: "
                f"{primary_guidance['stakeholders']}. Как у вас устроен технический выбор и бюджетное согласование?"
            )
        },
        {
            "stage": "Закрытие и recap",
            "goal": "Зафиксировать конкретный взаимный следующий шаг.",
            "talk_track": (
                f"Предлагаю следующий шаг: {primary_guidance['meeting_goal']} Со своей стороны мы подготовим "
                "повестку и список данных; со стороны заказчика нужны технический владелец и подтверждение фактов. "
                "После встречи фиксируем решение: пилот, assessment, расчет проекта или закрытие гипотезы."
            )
        },
    ]

    questions = []
    seen_questions = set()

    def add_question(topic, question, purpose="Подтвердить факт и влияние до выбора решения."):
        normalized = question.strip().lower()
        if normalized in seen_questions:
            return
        questions.append((topic, question, purpose))
        seen_questions.add(normalized)

    for pain in pains[:4]:
        add_question(pain["priority"] + " / " + pain["commercial_angle"].split(",")[0][:28], pain["discovery_question"])

    add_question("Приоритет бизнеса", "Какой сценарий наиболее чувствителен: простой сервиса, потеря данных, влияние на клиентов или ручная нагрузка на команду?", "Определить бизнес-драйвер инициативы.")
    add_question("Критичные сервисы", "Какие 3 системы нужно защищать и восстанавливать первыми, если случится инцидент?", "Определить scope и владельцев сервисов.")
    add_question("Почему сейчас", "Какое событие задает срок: аудит, инцидент, продление, новый проект, требование руководства или бюджетный цикл?", "Проверить срочность и реальный trigger сделки.")
    add_question("Процесс решения", "Кто подтверждает технические требования, кто утверждает бюджет и кто может остановить проект?", "Построить карту влияния и ЛПР.")
    add_question("Критерии выбора", "По каким критериям будет принято решение: функциональность, интеграция, сервис, локальная экспертиза, сроки или стоимость владения?", "Понять критерии оценки и конкурентную позицию.")
    add_question("Бюджетный контур", "Инициатива уже включена в бюджет или сначала нужен assessment и обоснование для руководства?", "Выбрать корректный коммерческий следующий шаг.")

    if results.get("Patch Management") == "Нет":
        add_question("Patch / CVE", "Есть ли сейчас отчет, какие критичные CVE открыты на АРМ и серверах дольше допустимого срока?")
    if results.get("EDR") == "Нет":
        add_question("Endpoint", "Если на АРМ сработает подозрительный процесс, сможете ли вы восстановить цепочку: пользователь, файл, процесс, сеть, сервер?")
    if results.get("SIEM") == "Нет" and not context.get("small_company"):
        add_question("Логи / SOC", "Какие источники логов вы готовы подключить первыми: NGFW, AD/учетки, серверы, EPP, backup или почту?")
    if results.get("MFA") == "Нет":
        add_question("MFA / IAM", "Где MFA можно включить быстрее всего без ломки процессов: VPN, почта, администраторы, облака или бизнес-системы?")
    if results.get("Резервное копирование") != "Нет":
        add_question("Backup", "Когда последний раз делали тестовое восстановление и какой результат можно показать руководству?")
    if context.get("has_public_web"):
        add_question("Web", "Какие публичные приложения критичны для выручки или клиентского сервиса, и кто владелец их доступности?")
    if context.get("has_development"):
        add_question("Разработка", "Где в процессе релиза можно поставить security gate: зависимости, SAST, DAST или ручной review?")

    questions = questions[:16]

    objections = [
        (
            "У нас уже есть антивирус/NGFW",
            "Признать действующую инвестицию и уточнить покрытие: задача не заменить продукт, а проверить телеметрию, сценарии реагирования, исключения и измеримый результат.",
            "Какие сценарии текущий стек уже закрывает и какой отчет это подтверждает?"
        ),
        (
            "Сейчас нет бюджета",
            "Не спорить с бюджетом. Разделить обязательный результат и способ его достижения; предложить assessment с ограниченным scope для расчета эффекта и бюджета следующего цикла.",
            "Какой артефакт поможет защитить инициативу перед руководством: расчет риска, roadmap, TCO или результаты пилота?"
        ),
        (
            "Полноценный SIEM для нашего контура избыточен",
            "Согласиться, что продукт не является целью. Вернуться к требуемому результату: какие критичные события нужно видеть, кто реагирует и в какой срок; затем сравнить минимальный scope и MSSP.",
            "Какие 5-7 сценариев обнаружения действительно важны и кто сегодня их разбирает?"
        ),
        (
            "Все работает, инцидентов не было",
            "Не оспаривать опыт клиента. Перевести разговор к проверяемости: тест восстановления, отчет по критичным CVE, покрытие MFA и время реакции на контрольный сценарий.",
            "Какой последний тест или отчет подтверждает, что критичный сценарий будет обнаружен и восстановлен в целевой срок?"
        ),
        (
            "У нас уже есть действующий поставщик",
            "Не атаковать поставщика. Уточнить, где именно есть разрыв: покрытие, эксплуатация, лицензии, интеграция или сервис; предложить независимую проверку или дополнение существующего стека.",
            "Что вы хотели бы улучшить в текущем решении, не меняя его без необходимости?"
        ),
        (
            "У команды нет ресурса на новый проект",
            "Предложить поэтапный scope, сервисную модель и четкое разделение ответственности; оценить нагрузку клиента до пилота.",
            "Какие операции вы готовы оставить внутри, а какие разумно передать партнеру или MSSP?"
        ),
    ]

    next_steps = []
    for item in opportunities[:5]:
        guidance = sales_account_guidance(item)
        next_steps.append({
            "priority": item["priority"],
            "step": item["next_step"],
            "offer": item["offer"],
            "stakeholders": guidance["stakeholders"],
            "meeting_goal": guidance["meeting_goal"],
            "seller_artifact": "Повестка встречи, список входных данных и одностраничная гипотеза ценности без преждевременной спецификации.",
            "success_criteria": "Подтверждены факт, бизнес-влияние, владелец, срок и следующий взаимный шаг: assessment, пилот, расчет проекта или закрытие гипотезы.",
        })
    if not next_steps and roadmap_items:
        for item in roadmap_items[:3]:
            next_steps.append({
                "priority": item["priority"],
                "step": item["action"],
                "offer": "Экспертный воркшоп по roadmap",
                "stakeholders": "ИТ-директор; руководитель ИБ; технические владельцы затронутых систем",
                "meeting_goal": "Подтвердить приоритет, владельца и измеримый результат инициативы.",
                "seller_artifact": "Повестка, карта вопросов и проект плана работ.",
                "success_criteria": "Клиент подтвердил приоритет и согласовал следующий созвон с техническими владельцами.",
            })

    return {
        "pains": pains[:8],
        "call_script": call_script,
        "questions": questions,
        "objections": objections,
        "next_steps": next_steps,
    }


def build_expert_conclusion(results, context, final_score, domain_scores, roadmap_items):
    profile_title, _ = infrastructure_profile(context)
    users = context.get("users", 0)
    servers = context.get("servers", 0)
    weak_domains = [
        domain
        for domain, score in sorted(domain_scores.items(), key=lambda item: item[1])
        if score < 50
    ]
    p1_actions = [
        item["action"]
        for item in roadmap_items
        if item["priority"] == "P1"
    ][:3]

    conclusion = [
        (
            f"Профиль ИТ-контура: {profile_title.lower()} "
            f"({russian_count(users, 'АРМ', 'АРМ', 'АРМ')}, {russian_count(servers, 'сервер', 'сервера', 'серверов')}). "
            "Рекомендации сформированы с учетом фактического состава: "
            "для текущего профиля приоритет отдается мерам, которые быстро повышают управляемость, "
            "восстановимость и контроль доступа без избыточного внедрения enterprise-платформ."
        )
    ]

    if final_score < 35:
        conclusion.append(
            "Текущий уровень зрелости ИБ оценивается как начальный: ключевые защитные процессы "
            "фрагментарны, а эффективность реагирования зависит от ручных действий специалистов."
        )
    elif final_score < 60:
        conclusion.append(
            "Текущий уровень зрелости ИБ оценивается как базовый/развивающийся: отдельные средства "
            "защиты уже используются, но отсутствует целостная операционная модель контроля, "
            "обновлений, мониторинга и реагирования."
        )
    else:
        conclusion.append(
            "Текущий уровень зрелости ИБ достаточен для базового контроля, однако дальнейшее развитие "
            "должно быть направлено на устойчивость процессов, измеримость и регулярную проверку защиты."
        )

    if weak_domains:
        conclusion.append(
            "Наиболее слабые домены: "
            + ", ".join(weak_domains[:3])
            + ". Именно они должны лечь в основу первоочередного плана улучшений."
        )

    if results.get("MFA") == "Нет":
        conclusion.append(
            "Отсутствие MFA повышает риск компрометации учетных записей. Для бизнеса это один из "
            "самых быстрых и экономически оправданных шагов по снижению вероятности инцидента."
        )

    if results.get("EPP") != "Нет" and results.get("EDR") == "Нет":
        conclusion.append(
            "Наличие EPP закрывает базовую антивирусную защиту, но не решает задачу расследования "
            "сложных атак. Для текущего масштаба разумнее рассматривать EDR/MDR как следующий этап, "
            "а не как замену существующему EPP."
        )

    if results.get("SIEM") == "Нет":
        if context.get("small_company"):
            conclusion.append(
                "Полноценный SIEM не является первоочередной инвестицией при текущем составе источников. "
                "Практичнее начать со сбора критичных журналов, ответственного за разбор событий "
                "и понятного регламента реагирования."
            )
        else:
            conclusion.append(
                "Для заданного масштаба требуется развитие централизованного мониторинга: "
                "минимум - MSSP/SOC или легковесный SIEM для критичных систем, максимум - "
                "полноценная корреляция событий и регламенты реагирования."
            )

    if p1_actions:
        conclusion.append(
            "Первоочередные действия на 0-30 дней: " + "; ".join(p1_actions) + "."
        )

    return conclusion[:7]


def is_enabled(value):

    if value is None:
        return False

    value = str(value).strip().lower()

    if value in ["", "нет", "none", "false", "-", "n/a"]:
        return False

    return True

def calculate_domain_scores(results):

    domains = {
        "Сетевая безопасность": 0,
        "Защита конечных точек": 0,
        "Идентификация и доступ": 0,
        "Мониторинг и SOC": 0,
        "Резервное копирование": 0,
        "Инфраструктура": 0
    }

    # =========================
    # Сетевая безопасность
    # =========================

    if is_enabled(results.get("NGFW")):
        domains["Сетевая безопасность"] += 25

    if is_enabled(results.get("WAF")):
        domains["Сетевая безопасность"] += 15

    if is_enabled(results.get("Anti-DDoS")):
        domains["Сетевая безопасность"] += 15

    if is_enabled(results.get("VPN")):
        domains["Сетевая безопасность"] += 10

    if is_enabled(results.get("NAC")):
        domains["Сетевая безопасность"] += 20

    if is_enabled(results.get("Сегментация сети")):
        domains["Сетевая безопасность"] += 15

    # =========================
    # Защита конечных устройств
    # =========================

    if is_enabled(results.get("Антивирус")):
        domains["Защита конечных точек"] += 20

    if is_enabled(results.get("EDR")):
        domains["Защита конечных точек"] += 40

    if is_enabled(results.get("Patch Management")):
        domains["Защита конечных точек"] += 20

    if is_enabled(results.get("MDM")):
        domains["Защита конечных точек"] += 10

    if is_enabled(results.get("Device Control")):
        domains["Защита конечных точек"] += 10

    # =========================
    # IAM
    # =========================

    if is_enabled(results.get("MFA")):
        domains["Идентификация и доступ"] += 35

    if is_enabled(results.get("IDM")):
        domains["Идентификация и доступ"] += 25

    if is_enabled(results.get("PAM")):
        domains["Идентификация и доступ"] += 25

    if is_enabled(results.get("SSO")):
        domains["Идентификация и доступ"] += 15

    # =========================
    # MONITORING
    # =========================

    if is_enabled(results.get("SIEM")):
        domains["Мониторинг и SOC"] += 40

    if is_enabled(results.get("SOC")):
        domains["Мониторинг и SOC"] += 30

    if is_enabled(results.get("NAD")):
        domains["Мониторинг и SOC"] += 15

    if is_enabled(results.get("Threat Intelligence")):
        domains["Мониторинг и SOC"] += 15

    # =========================
    # BACKUP
    # =========================

    if is_enabled(results.get("Резервное копирование")):
        domains["Резервное копирование"] += 30

    if is_enabled(results.get("Immutable Backup")):
        domains["Резервное копирование"] += 30

    if is_enabled(results.get("DR")):
        domains["Резервное копирование"] += 20

    if is_enabled(results.get("Air-Gap Backup")):
        domains["Резервное копирование"] += 20

    # =========================
    # INFRA
    # =========================

    if is_enabled(results.get("Виртуализация")):
        domains["Инфраструктура"] += 25

    if is_enabled(results.get("СХД")):
        domains["Инфраструктура"] += 25

    if is_enabled(results.get("Мониторинг")):
        domains["Инфраструктура"] += 20

    if is_enabled(results.get("Резервный канал")):
        domains["Инфраструктура"] += 15

    if is_enabled(results.get("Кластеризация")):
        domains["Инфраструктура"] += 15

    # Ограничение 100%
    for k in domains:
        domains[k] = min(domains[k], 100)

    return domains


def risk_level_label(level):
    labels = {
        "CRITICAL": "Критический",
        "HIGH": "Высокий",
        "MEDIUM": "Средний",
        "LOW": "Низкий",
        "КРИТИЧЕСКИЙ": "Критический",
        "ВЫСОКИЙ": "Высокий",
        "СРЕДНИЙ": "Средний",
        "НИЗКИЙ": "Низкий",
    }
    return labels.get(str(level).upper(), str(level))


def risk_source_label(source):
    if str(source).lower() in {"ai", "ии", "gemini"}:
        return "ИИ"
    return "Базовые правила"


def risk_semantic_key(item):
    explicit_key = str(
        item.get("semantic_key") or item.get("_semantic_key") or ""
    ).strip()
    if explicit_key:
        return explicit_key

    title = str(item.get("risk", "")).strip().lower()
    if "сегментац" in title and not any(marker in title for marker in ("nac", "network access control")):
        return "segmentation"

    title_buckets = [
        ("wifi_capacity", ("wi-fi", "wifi", "wlan", "беспроводн", "роуминг", "точек доступа")),
        ("network_performance", ("канал связи", "канала связи", "резервный канал", "wan", "failover")),
        ("storage", ("схд", "storage", "дисков", "хранилищ")),
        ("virtualization", ("виртуальн", "гипервизор", "vmware", "hyper-v")),
        ("dr", ("rto", "rpo", "аварийн", "drp", "disaster recovery")),
        ("it_monitoring", ("мониторинг ит", "мониторинга ит", "мониторинг инфраструктур", "наблюдаемост")),
        ("change_management", ("управления изменениями", "управление изменениями", "change management")),
    ]
    for key, markers in title_buckets:
        if any(marker in title for marker in markers):
            return key

    text = " ".join(
        str(item.get(field, ""))
        for field in ("risk", "description", "impact", "recommendation")
    ).lower()
    text = re.sub(r"[\u2010-\u2015\u2212]", "-", text)

    buckets = [
        ("mfa", ("mfa", "многофактор", "2fa", "двухфактор")),
        ("legacy_os", ("legacy", "устаревш", "windows xp", "windows vista", "windows 7", "windows 8", "2008", "2012 r2")),
        ("pam", ("pam", "привилегирован", "администраторск")),
        ("iam", ("iam", "identity and access management", "централизованному управлению учетными", "централизованное управление учетными", "управление идентификацией", "жизненный цикл учетных записей")),
        ("nac", ("nac", "контроль подключения устройств", "контроль доступа устройств к сети", "network access control")),
        ("dlp", ("dlp", "утеч", "эксфильтрац", "data loss")),
        ("siem_soc", ("siem", "soc", "soar", "мониторинг событий", "централизованный мониторинг")),
        ("wifi_capacity", (
            "wi-fi", "wifi", "wlan", "беспроводн", "роуминг", "радиообслед",
            "точек доступа", "точки доступа", "контроллер беспровод",
        )),
        ("network_performance", (
            "масштабируемость сетевой", "производительность сетевой", "сетевая топология",
            "конфигурации маршрутизации",
            "резервный канал", "пропускная способность", "failover",
        )),
        ("itam", (
            "программными активами", "жизненным циклом", "управление активами",
            "лицензи", "инвентаризац", "cmdb", "учет активов",
        )),
        ("change_management", ("управления изменениями", "управление изменениями", "change management", "изменениями и конфигурациями")),
        ("patch", ("patch", "обновлен", "cve", "уязвим")),
        ("endpoint_detection", ("edr", "xdr", "endpoint", "рабочих мест", "lateral movement")),
        ("backup", ("backup", "резерв", "immutable", "ransomware")),
        ("dr", ("dr", "аварийн", "rto", "rpo", "восстановлен")),
        ("web_waf", ("waf", "web", "веб", "owasp", "публичн")),
        ("segmentation", ("сегментац", "vlan", "lateral")),
        ("mail", ("mail", "почт", "фишинг")),
        ("virtualization", ("виртуализац", "гипервизор", "vm", "хост")),
        ("storage", ("схд", "storage", "raid", "snapshot", "iops")),
        ("it_monitoring", ("эксплуатационный мониторинг", "доступности", "производительности", "capacity")),
        ("appsec", ("sast", "dast", "appsec", "разработ", "безопасность прилож")),
        ("business_systems", ("erp", "crm", "бизнес-систем")),
    ]

    for key, markers in buckets:
        if any(marker in text for marker in markers):
            return key

    return re.sub(r"\s+", " ", title)


def network_segmentation_evidence(results):
    network_values = []
    for key, value in results.items():
        key_text = str(key).lower()
        if any(marker in key_text for marker in ("1.2", "сеть", "маршрут", "ngfw", "wifi", "wi-fi", "коммут")):
            network_values.append(f"{key}: {value}")
    text = " ".join(network_values).lower()
    negative_markers = (
        "нет сегментации", "сегментация отсутствует", "плоская сеть", "flat network",
        "единый vlan", "один vlan для всех",
    )
    positive_markers = (
        "vlan", "vrf", "acl", "межсегмент", "сегментац", "dmz",
        "зоны ngfw", "firewall zone",
    )
    if any(marker in text for marker in negative_markers):
        return "absent"
    if any(marker in text for marker in positive_markers):
        return "present"
    return "unknown"


def neutralize_company_scale_language(value):
    text = str(value or "")
    replacements = {
        "маленькая компания": "компания с указанным ИТ-контуром",
        "малая компания": "компания с указанным ИТ-контуром",
        "средняя компания": "компания с указанным ИТ-контуром",
        "крупная компания": "компания с указанным ИТ-контуром",
        "малая инфраструктура": "инфраструктура указанного масштаба",
        "средняя инфраструктура": "инфраструктура указанного масштаба",
        "крупная инфраструктура": "инфраструктура указанного масштаба",
        "малого масштаба": "указанного масштаба",
    }
    for source, target in replacements.items():
        text = re.sub(re.escape(source), target, text, flags=re.IGNORECASE)
    return text


def sanitize_ai_audit_narrative(narrative, results):
    if not isinstance(narrative, dict):
        return {}

    segmentation_status = network_segmentation_evidence(results)
    legacy_reported = (
        int(results.get("ОС АРМ (Windows XP/Vista/7/8)", 0) or 0)
        + int(results.get("ОС Сервера (Windows Server 2008/2012 R2)", 0) or 0)
    ) > 0

    def clean_text(value):
        text = expand_regulatory_references(neutralize_company_scale_language(value))
        lowered = text.lower()
        if not legacy_reported and any(marker in lowered for marker in (
            "устаревш", "legacy", "windows 10", "linux-сервер", "linux сервер",
        )):
            return (
                "Для парка рабочих мест и серверов требуется подтвердить версии, сроки поддержки "
                "и единый цикл устранения уязвимостей; Windows 10 и Linux без номера версии "
                "не классифицируются как устаревшие автоматически."
            )
        if (
            segmentation_status == "unknown"
            and "сегментац" in lowered
            and any(marker in lowered for marker in ("отсутств", "недостаточ", "ospf", "nac", "ztna"))
        ):
            return (
                "Архитектура сетевой сегментации не подтверждена данными анкеты. "
                "OSPF описывает маршрутизацию, а отсутствие NAC/ZTNA не доказывает отсутствие "
                "VLAN, ACL, VRF или межсегментных политик; требуется анализ схемы и правил NGFW."
            )
        return text

    cleaned = dict(narrative)
    cleaned["executive_summary"] = [
        clean_text(item) for item in narrative.get("executive_summary", []) if str(item).strip()
    ]
    cleaned["management_decisions"] = [
        clean_text(item) for item in narrative.get("management_decisions", []) if str(item).strip()
    ]
    cleaned["audit_observations"] = [
        {
            "title": clean_text(item.get("title", "Наблюдение")),
            "text": clean_text(item.get("text", "")),
        }
        for item in narrative.get("audit_observations", [])
        if isinstance(item, dict) and str(item.get("text", "")).strip()
    ]
    cleaned["roadmap"] = [
        {
            **item,
            "action": sanitize_customer_roadmap_text(item.get("action", "")),
            "rationale": expand_regulatory_references(item.get("rationale", "")),
            "result": sanitize_customer_roadmap_text(item.get("result", "")),
        }
        for item in narrative.get("roadmap", [])
        if isinstance(item, dict) and str(item.get("action", "")).strip()
    ]
    return cleaned


def enforce_audit_fact_policy(item, results, context):
    """Apply narrow fact and priority guards without rewriting the AI conclusion."""
    normalized = dict(item)
    for field in ("risk", "description", "impact", "recommendation"):
        if field in normalized:
            normalized[field] = expand_regulatory_references(normalized[field])
    evidence = normalized.get("evidence", [])
    if isinstance(evidence, list):
        normalized["evidence"] = [expand_regulatory_references(value) for value in evidence]
    key = risk_semantic_key(normalized)

    if key == "segmentation" and network_segmentation_evidence(results) != "absent" and not is_enabled(results.get("NAC")):
        normalized.update({
            "level": "MEDIUM",
            "risk": "Подключение устройств к сети не контролируется централизованно",
            "description": (
                "В анкете NAC не указан. Это не доказывает отсутствие VLAN или ACL, но означает, "
                "что допуск проводных, беспроводных и неизвестных устройств к сети требует отдельной проверки."
            ),
            "impact": (
                "Неидентифицированное или несоответствующее политике устройство может получить доступ "
                "к корпоративной сети до ручного обнаружения."
            ),
            "recommendation": (
                "Составить матрицу типов подключений; провести пилот NAC на Wi-Fi и одном проводном сегменте; "
                "настроить профилирование, проверку соответствия и изоляцию неизвестных устройств."
            ),
            "evidence": [
                f"NAC в анкете: {results.get('NAC', 'Нет')}",
                f"Точки доступа Wi-Fi: {results.get('Wi-Fi Точки доступа', results.get('Количество точек доступа', 'не указано'))}",
            ],
            "success_metric": "100% подключений идентифицируются; неизвестные устройства изолируются",
            "vendors": ["NAC"],
        })
        key = "nac"

    if key == "nac" and not is_enabled(results.get("NAC")):
        normalized.update({
            "level": "MEDIUM",
            "risk": "Допуск устройств к сети не контролируется автоматически",
            "description": (
                "В анкете NAC не указан. Это не подтверждает отсутствие VLAN, ACL или межсетевого "
                "экранирования, но указывает на отсутствие автоматической проверки устройств при подключении."
            ),
            "impact": (
                "Неизвестное или несоответствующее политике устройство может получить сетевой доступ "
                "до ручного обнаружения и изоляции."
            ),
            "recommendation": (
                "Описать типы проводных и беспроводных подключений; провести пилот NAC на Wi-Fi и одном "
                "проводном сегменте; проверить профилирование, контроль соответствия и изоляцию устройств."
            ),
            "evidence": [
                f"NAC в анкете: {results.get('NAC', 'Нет')}",
                f"Точки доступа Wi-Fi: {results.get('Wi-Fi Точки доступа', results.get('Количество точек доступа', 'не указано'))}",
            ],
            "success_metric": "Все подключения идентифицируются; неизвестные устройства автоматически изолируются",
            "vendors": ["NAC"],
        })

    if key == "iam" and not is_enabled(results.get("IAM")):
        user_count = int(context.get("users", results.get("_user_count", 0)) or 0)
        normalized.update({
            "level": "MEDIUM",
            "risk": "Жизненный цикл учетных записей не автоматизирован",
            "description": (
                f"В анкете IAM не указан для контура из {user_count} рабочих мест. "
                "Порядок создания, изменения, регулярного пересмотра и отзыва прав требует подтверждения."
            ),
            "impact": (
                "Задержка отзыва доступа и накопление избыточных прав повышают вероятность "
                "несанкционированного доступа к бизнес-системам."
            ),
            "recommendation": (
                "Описать процессы приема, перевода и увольнения; определить владельцев ролей и согласования; "
                "провести PoC IAM на выбранных подразделениях и критичных системах."
            ),
            "evidence": [
                f"IAM в анкете: {results.get('IAM', 'Нет')}",
                f"Рабочие места: {user_count}",
            ],
            "success_metric": "Учетные записи имеют владельца; создание, изменение и отзыв прав выполняются по SLA",
            "vendors": ["IAM"],
        })

    if key == "dlp" and not is_enabled(results.get("DLP")) and (
        context.get("is_kvoiki") or context.get("has_personal_data")
    ):
        normalized.update({
            "level": "HIGH",
            "risk": "Отсутствие DLP повышает риск утечки персональных данных",
            "description": (
                "В анкете DLP не указан, при этом организация обрабатывает персональные данные. "
                "Контроль каналов передачи и правил обращения с чувствительными данными требует подтверждения."
            ),
            "impact": (
                "Неконтролируемая передача данных через почту, веб, облачные сервисы или съемные носители "
                "может привести к утечке и регуляторным последствиям."
            ),
            "recommendation": (
                "Определить категории данных, каналы контроля и критерии успеха; провести ограниченный "
                "пилот DLP; по результатам выбрать архитектуру и масштабировать подтвержденные политики."
            ),
            "evidence": [
                f"DLP в анкете: {results.get('DLP', 'Нет')}",
                "Обработка персональных данных: указана",
            ],
            "success_metric": "Политики DLP контролируют согласованные каналы передачи персональных данных",
            "vendors": ["DLP"],
        })

    if key == "backup" and context.get("has_backup"):
        normalized.update({
            "risk": "Восстановление из резервных копий не подтверждено регулярными тестами",
            "description": (
                f"В анкете указан действующий backup-контур: "
                f"{results.get('Резервное копирование', 'не указано')}. "
                "При этом целевые RTO/RPO и результаты контрольного восстановления не зафиксированы."
            ),
            "impact": (
                "Без регулярной проверки нельзя подтвердить, что критичные сервисы будут восстановлены "
                "в согласованные с бизнесом сроки и с допустимой потерей данных."
            ),
            "recommendation": (
                "Согласовать RTO/RPO по критичным сервисам; утвердить сценарии контрольного "
                "восстановления; регулярно проводить тесты и фиксировать фактическое время и полноту восстановления."
            ),
            "evidence": [
                f"Резервное копирование: {results.get('Резервное копирование', 'указано')}",
                "RTO/RPO и результаты тестового восстановления в анкете не зафиксированы",
            ],
            "success_metric": (
                "Все критичные сервисы проходят тест восстановления в пределах утвержденных RTO/RPO"
            ),
            "vendors": ["Backup"],
        })

    if key == "siem_soc" and not is_enabled(results.get("SOAR")):
        recommendation = str(normalized.get("recommendation", "")).strip()
        if "soar" not in recommendation.lower():
            normalized["recommendation"] = (
                recommendation.rstrip(". ")
                + ". После стабилизации источников, сценариев и SLA перейти к SOAR-автоматизации повторяемых операций реагирования."
            ).strip()

    return normalized


def professionalize_risk_item(item, results, context):
    source = item.get("_source", "Базовые правила")
    key = risk_semantic_key(item)
    users = context.get("users", results.get("_user_count", 0))
    servers = context.get("servers", 0)
    legacy_arm = results.get("ОС АРМ (Windows XP/Vista/7/8)", 0)
    legacy_srv = results.get("ОС Сервера (Windows Server 2008/2012 R2)", 0)

    if key == "segmentation" and network_segmentation_evidence(results) != "absent":
        normalized = dict(item)
        normalized.update({
            "level": "LOW",
            "risk": "Архитектура сетевой сегментации требует подтверждения",
            "description": (
                "Анкета описывает маршрутизацию и сетевое оборудование, но не содержит схемы VLAN/VRF, "
                "ACL и межсегментных политик. OSPF не является признаком наличия или отсутствия сегментации, "
                "а отсутствие NAC/ZTNA само по себе не подтверждает плоскую сеть."
            ),
            "impact": (
                "Без схемы и правил фильтрации нельзя достоверно оценить возможность бокового перемещения "
                "между пользовательскими, серверными, гостевыми и критичными сегментами."
            ),
            "recommendation": (
                "Запросить актуальную схему VLAN/VRF и матрицу потоков; проверить ACL и политики NGFW между "
                "сегментами; по результатам зафиксировать подтвержденные разрывы и только затем формировать проект улучшений."
            ),
            "vendors": [],
            "regulators": ["CIS Controls", "ISO 27001"],
            "_source": source,
        })
        return normalized

    if risk_source_label(source) == "ИИ":
        normalized = dict(item)
        normalized["_source"] = source
        for field in ("risk", "description", "impact", "recommendation"):
            if field in normalized:
                normalized[field] = neutralize_company_scale_language(normalized[field])
        normalized.setdefault("level", "MEDIUM")
        normalized.setdefault("description", normalized.get("risk", "Риск требует дополнительного уточнения."))
        normalized.setdefault("impact", "Риск может повлиять на устойчивость ИТ/ИБ процессов.")
        normalized.setdefault("recommendation", "Уточнить текущий процесс, назначить ответственного и определить измеримый план улучшений.")
        if not normalized.get("vendors"):
            normalized["vendors"] = []
        if not normalized.get("regulators"):
            normalized["regulators"] = ["ISO 27001"]
        return normalized

    profiles = {
        "legacy_os": {
            "level": "CRITICAL",
            "risk": "Устаревшие операционные системы требуют миграции или изоляции",
            "description": f"В анкете указаны устаревшие ОС: АРМ - {legacy_arm}, серверы - {legacy_srv}. Такие системы не получают полноценные исправления безопасности и не должны рассматриваться как обычные рабочие места.",
            "impact": "Уязвимости устаревших ОС повышают риск компрометации, невозможности установки актуальных агентов защиты и остановки связанных бизнес-процессов.",
            "recommendation": "Составить реестр устаревших ОС; подготовить план миграции на поддерживаемые версии Microsoft Windows/Windows Server; если часть систем нельзя обновить быстро, временно изолировать их и контролировать уязвимости до вывода из эксплуатации.",
            "vendors": ["Microsoft", "Qualys", "Tenable", "Rapid7"],
            "regulators": ["CIS Controls", "ISO 27001", "NIST CSF"],
        },
        "endpoint_detection": {
            "level": "HIGH",
            "risk": "Защита рабочих мест не обеспечивает полноценное обнаружение и реагирование",
            "description": f"В инфраструктуре указано {users} АРМ, при этом EDR/XDR/MDR не описаны. Базовый EPP снижает риск массового malware, но не закрывает расследование сложных атак и lateral movement.",
            "impact": "Инцидент на рабочей станции может дольше оставаться незамеченным, распространяться на учетные записи и серверы, а расследование будет зависеть от ручного сбора данных.",
            "recommendation": "Проверить фактическое покрытие EPP и актуальность агентов; выбрать пилот EDR/MDR для критичных групп пользователей и серверов; внедрить регламент реагирования с метриками MTTD/MTTR и контрольной проверкой сценариев ransomware.",
            "vendors": ["CrowdStrike", "SentinelOne", "Trend Micro"],
            "regulators": ["CIS Controls", "NIST CSF", "MITRE ATT&CK"],
        },
        "mfa": {
            "level": "HIGH",
            "risk": "Критичные доступы не защищены многофакторной аутентификацией",
            "description": "В анкете не указана MFA для администраторов, почты, VPN/удаленного доступа и критичных бизнес-систем.",
            "impact": "Компрометация одного пароля может дать атакующему доступ к корпоративной почте, административным консолям или бизнес-приложениям.",
            "recommendation": "Включить MFA сначала для администраторов, почты и удаленного доступа; затем распространить на ERP/CRM и критичные облачные сервисы; контролировать исключения и ежемесячно выгружать отчет по пользователям без MFA.",
            "vendors": ["Cisco Duo", "FortiAuthenticator", "Axidian"],
            "regulators": ["ISO 27001", "CIS Controls", "NIST CSF"],
        },
        "siem_soc": {
            "level": "HIGH" if context.get("enterprise_company") else "MEDIUM",
            "risk": "Мониторинг событий ИБ не покрывает критичные источники",
            "description": "Не описан централизованный сбор и анализ событий от NGFW, серверов, учетных записей, endpoint-защиты, почты и публичных сервисов.",
            "impact": "Атаки и нарушения политик могут обнаруживаться поздно, а расследование будет затруднено из-за разрозненных журналов.",
            "recommendation": "Начать с минимального SOC/MSSP-scope: NGFW, AD/учетные записи, серверы, EPP/EDR, backup и почта; настроить 8-12 базовых сценариев корреляции; ежемесячно разбирать инциденты и качество источников логов.",
            "vendors": ["IBM QRadar", "Splunk", "R-Vision SIEM"],
            "regulators": ["ISO 27001", "NIST CSF", "CIS Controls"],
        },
        "web_waf": {
            "level": "HIGH",
            "risk": "Публичные веб-сервисы требуют специализированной защиты приложений",
            "description": "В анкете указаны интернет-магазин и личный кабинет, но WAF-защита и прикладные правила контроля атак не описаны.",
            "impact": "Публичные приложения остаются под риском OWASP-атак, бот-активности, утечки данных и нарушения доступности клиентских сервисов.",
            "recommendation": "Провести экспресс-оценку web-периметра; проверить текущие Cloudflare-политики; включить WAF/CDN или FortiWeb/F5/Imperva для публичных ресурсов с профилем под приложение.",
            "vendors": ["Cloudflare", "Fortinet", "F5", "Imperva"],
            "regulators": ["OWASP ASVS", "PCI DSS", "ISO 27001"],
        },
        "pam": {
            "level": "HIGH",
            "risk": "Привилегированные учетные записи не выделены в отдельный контур контроля",
            "description": f"В инфраструктуре указано {servers} серверов и критичные бизнес-системы, но PAM или сопоставимый контроль административных доступов не описан.",
            "impact": "Компрометация одной администраторской учетной записи может привести к изменению конфигураций, отключению защитных средств или доступу к критичным данным.",
            "recommendation": "Провести инвентаризацию привилегированных учетных записей; внедрить vault и контроль сессий для критичных систем; настроить регулярный пересмотр прав и запрет постоянных локальных админов.",
            "vendors": ["CyberArk", "Fudo", "Netwrix"],
            "regulators": ["ISO 27001", "CIS Controls", "NIST CSF"],
        },
        "mail": {
            "level": "MEDIUM",
            "risk": "Почтовый контур требует отдельной anti-phishing защиты",
            "description": "В анкете указана корпоративная почта, но специализированная защита от фишинга, вредоносных вложений и подмены отправителя не описана.",
            "impact": "Фишинговая атака может стать первичной точкой компрометации учетных записей, рабочих мест и облачных сервисов.",
            "recommendation": "Проверить SPF/DKIM/DMARC и политики Microsoft 365; внедрить mail security для вложений, URL и impersonation; проводить регулярные фишинг-симуляции и разбор результатов.",
            "vendors": ["Check Point", "Fortinet", "Trend Micro", "Forcepoint"],
            "regulators": ["CIS Controls", "NIST CSF"],
        },
        "patch": {
            "level": "HIGH" if users > 100 else "MEDIUM",
            "risk": "Управление обновлениями не выглядит централизованным",
            "description": f"Для парка {users} АРМ и {servers} серверов не описан управляемый процесс patch management и контроль устранения критичных CVE.",
            "impact": "Известные уязвимости могут оставаться открытыми дольше допустимого окна, повышая риск компрометации через типовые exploit-сценарии.",
            "recommendation": "Ввести ежемесячный цикл обновлений; отдельно контролировать критичные CVE с SLA; формировать отчет по покрытию АРМ, серверов и исключений после каждого окна обновлений.",
            "vendors": ["ManageEngine", "Ivanti", "Tanium"],
            "regulators": ["CIS Controls", "NIST CSF", "ISO 27001"],
        },
        "dr": {
            "level": "HIGH",
            "risk": "Не описан план аварийного восстановления критичных ИТ-сервисов",
            "description": "В анкете есть серверы, виртуализация, backup и бизнес-системы, но не описаны RTO/RPO, DR-сценарии и регулярные тесты восстановления.",
            "impact": "При сбое площадки, СХД или критичных виртуальных машин восстановление может занять дольше допустимого для бизнеса.",
            "recommendation": "Определить перечень критичных сервисов и целевые RTO/RPO; провести тест восстановления ERP/CRM/почты; оформить DR-runbook и назначить периодичность контрольных учений.",
            "vendors": ["Veeam", "Commvault", "Rubrik"],
            "regulators": ["ISO 22301", "ISO 27001", "NIST CSF"],
        },
        "it_monitoring": {
            "level": "MEDIUM",
            "risk": "Эксплуатационный мониторинг ИТ-инфраструктуры требует формализации",
            "description": "В анкете есть сеть, серверы, СХД и бизнес-системы, но не описан единый мониторинг доступности, производительности и емкости.",
            "impact": "Инциденты по дискам, каналам связи, виртуальным машинам или бизнес-сервисам могут обнаруживаться постфактум по обращениям пользователей.",
            "recommendation": "Настроить мониторинг серверов, сетевых устройств, СХД, виртуализации и ключевых приложений; определить пороги и ответственных; ежемесячно анализировать тренды емкости и доступности.",
            "vendors": ["Zabbix", "PRTG", "ManageEngine"],
            "regulators": ["ITIL", "ISO 20000", "ISO 27001"],
        },
        "virtualization": {
            "level": "MEDIUM",
            "risk": "Отказоустойчивость виртуализации требует отдельной проверки",
            "description": "Используются виртуальные серверы, но в анкете не описаны HA/DRS-настройки, резервы ресурсов и правила размещения критичных VM.",
            "impact": "Отказ одного хоста или нехватка ресурсов может затронуть сразу несколько бизнес-сервисов.",
            "recommendation": "Проверить HA-настройки и резервы CPU/RAM/storage; разделить критичные VM по хостам; оформить регламент обслуживания гипервизоров без простоя сервисов.",
            "vendors": ["VMware", "Veeam", "Zabbix"],
            "regulators": ["ITIL", "ISO 20000"],
        },
        "storage": {
            "level": "LOW",
            "risk": "СХД требует health-check, а не вывода о проблеме надежности",
            "description": "В анкете указаны RAID, диски и snapshot/backup-практики, но не раскрыты метрики утилизации, latency/IOPS и запас емкости.",
            "impact": "Без фактических метрик нельзя утверждать о проблемах производительности или надежности; корректнее подтвердить запас и правила контроля.",
            "recommendation": "Проверить утилизацию, latency/IOPS, состояние RAID, snapshot-политики и связку с backup; по результатам определить, нужен ли capacity management или расширение.",
            "vendors": ["Veeam"],
            "regulators": ["ITIL", "ISO 20000"],
        },
    }

    profile = profiles.get(key)
    if profile:
        normalized = {**item, **profile}
        normalized["_source"] = source
        return normalized

    normalized = dict(item)
    normalized["_source"] = source
    normalized.setdefault("level", "MEDIUM")
    normalized.setdefault("description", normalized.get("risk", "Риск требует дополнительного уточнения."))
    normalized.setdefault("impact", "Риск может повлиять на устойчивость ИТ/ИБ процессов.")
    normalized.setdefault("recommendation", "Уточнить текущий процесс, назначить ответственного и определить измеримый план улучшений.")
    if not normalized.get("vendors"):
        normalized["vendors"] = []
    if not normalized.get("regulators"):
        normalized["regulators"] = ["ISO 27001"]
    return normalized


def align_report_vendors(item, results, context):
    normalized = dict(item)
    key = risk_semantic_key(normalized)
    facts_text = normalize_vendor_key(" ".join(str(value) for value in results.values()))
    has_fortinet = any(marker in facts_text for marker in ("fortinet", "fortigate", "forti"))
    has_cloudflare = "cloudflare" in facts_text

    vendor_profiles = {
        "legacy_os": ["Microsoft", "Qualys", "Tenable", "Rapid7"],
        "mfa": ["Fortinet", "Microsoft", "Cisco"],
        "patch": ["Qualys", "Tenable", "Rapid7", "Ivanti"],
        "mail": ["Check Point", "Fortinet", "Trend Micro", "Forcepoint"],
        "endpoint_detection": ["Fortinet", "Check Point", "CrowdStrike", "Trend Micro"],
        "siem_soc": ["Fortinet", "IBM", "Splunk"],
        "pam": ["Wallix", "CyberArk", "BeyondTrust"],
        "appsec": ["Checkmarx", "HCL AppScan", "Positive Technologies", "Qualys"],
        "storage": ["Veeam"],
        "segmentation": ["Fortinet", "Cisco", "Huawei"],
        "nac": ["Fortinet", "Cisco", "Check Point"],
    }

    if key == "web_waf":
        vendors = []
        if has_cloudflare:
            vendors.append("Cloudflare")
        if has_fortinet:
            vendors.append("Fortinet")
        vendors.extend(["F5", "Imperva"])
        normalized["vendors"] = list(dict.fromkeys(vendors))
        return normalized

    if key in vendor_profiles:
        normalized["vendors"] = vendor_profiles[key]

    return normalized


def build_report_risk_set(c_info, results, context):
    rule_risks = generate_rule_based_risks(
        results,
        context
    )
    rule_risks = [
        {**item, "_source": "Базовые правила"}
        for item in rule_risks
        if isinstance(item, dict)
    ]

    ai_risks = ai_generate_risks_and_recs(
        c_info,
        results
    )
    ai_succeeded = bool(st.session_state.get("ai_analysis_succeeded"))
    ai_used = isinstance(ai_risks, list) and any(isinstance(item, dict) for item in ai_risks)
    st.session_state.ai_used_in_last_report = ai_succeeded

    combined_risks = []
    if ai_used:
        combined_risks.extend([
            {**item, "_source": "ИИ"}
            for item in ai_risks
            if isinstance(item, dict)
        ])
    combined_risks.extend(build_confirmed_it_gap_risks(results, context))
    combined_risks.extend(rule_risks)

    priority_order = {"CRITICAL": 1, "HIGH": 2, "MEDIUM": 3, "LOW": 4}
    unique_risks = []
    seen_risks = set()
    skipped_conflicting_risks = []
    for item in combined_risks:
        conflict = risk_conflicts_with_answers(item, results)
        if conflict:
            skipped_conflicting_risks.append(
                f"{risk_semantic_key(item)}: {conflict}"
            )
            continue
        item = enforce_audit_fact_policy(item, results, context)
        item = professionalize_risk_item(item, results, context)
        item = align_report_vendors(item, results, context)
        semantic_key = risk_semantic_key(item)
        if not semantic_key or semantic_key in seen_risks:
            continue
        item = dict(item)
        item["semantic_key"] = semantic_key
        unique_risks.append(item)
        seen_risks.add(semantic_key)

    st.session_state.last_report_skipped_conflicts = skipped_conflicting_risks

    sorted_risks = sorted(
        unique_risks,
        key=lambda item: priority_order.get(str(item.get("level", "")).upper(), 99)
    )
    required_it_risks = [
        item
        for item in sorted_risks
        if item.get("_ai_area") == "ИТ"
    ][:4]

    report_risks = []
    report_keys = set()
    for item in [*required_it_risks, *sorted_risks]:
        key = risk_semantic_key(item)
        if not key or key in report_keys:
            continue
        report_risks.append(item)
        report_keys.add(key)
        if len(report_risks) >= 12:
            break

    st.session_state.last_report_risk_sources = [
        {
            "level": risk_level_label(item.get("level", "MEDIUM")),
            "risk": item.get("risk", "Риск"),
            "description": item.get("description", "-"),
            "impact": item.get("impact", "-"),
            "recommendation": item.get("recommendation", "-"),
            "vendors": item.get("vendors", []),
            "area": item.get("_ai_area", "ИТ/ИБ"),
            "source": risk_source_label(item.get("_source")),
            "legal_ids": item.get("legal_ids", []),
            "frameworks": item.get("frameworks", []),
            "evidence": item.get("evidence", []),
            "success_metric": item.get("success_metric", ""),
            "semantic_key": risk_semantic_key(item),
        }
        for item in report_risks
    ]

    return report_risks, ai_used


def build_audit_observations(results, context, domain_scores, report_risks):
    users = context.get("users", 0)
    servers = context.get("servers", 0)
    weak_domains = [
        domain
        for domain, score in sorted(domain_scores.items(), key=lambda item: item[1])
        if score < 50
    ][:4]

    observations = [
        (
            "Масштаб и управляемость",
            f"В анкете зафиксировано {russian_count(users, 'АРМ', 'АРМ', 'АРМ')} и "
            f"{russian_count(servers, 'сервер', 'сервера', 'серверов')}; такой состав требует формализованных процессов "
            "обновлений, мониторинга, резервного копирования и контроля изменений."
        )
    ]

    if weak_domains:
        observations.append((
            "Слабые домены защиты",
            "Наиболее слабые зоны по анкете: " + ", ".join(weak_domains) + ". Их нужно закрывать как программу улучшений, а не отдельными разрозненными закупками."
        ))

    if results.get("Резервное копирование") != "Нет":
        observations.append((
            "Устойчивость и восстановление",
            "Backup-контур заявлен, но для управленческого уровня важно проверить RTO/RPO, immutable/offline-копии и факт регулярного тестового восстановления."
        ))

    if context.get("has_public_web"):
        observations.append((
            "Публичная поверхность",
            "Интернет-магазин и личный кабинет формируют внешний периметр риска: защиту web-приложений, журналирование и реагирование нужно выделить в отдельный поток работ."
        ))

    if context.get("has_development"):
        observations.append((
            "Разработка и изменения",
            "Наличие внутренней разработки означает, что часть рисков нужно закрывать до публикации: через SAST/DAST, контроль релизов и требования безопасности к CI/CD."
        ))

    if report_risks:
        observations.append((
            "Фокус первых действий",
            "Первые управленческие решения должны закрыть доступы, устаревшие ОС, endpoint detection, patch management и web-периметр."
        ))

    return observations[:6]


def build_target_operating_model(results, context):
    users = context.get("users", 0)
    servers = context.get("servers", 0)
    access_text = (
        "Контроль покрытия MFA для администраторов, почты, удаленного доступа и критичных систем; регулярный пересмотр исключений и привилегий."
        if is_enabled(results.get("MFA"))
        else "MFA для администраторов, почты, удаленного доступа и критичных систем; регулярный пересмотр исключений и привилегий."
    )
    endpoint_text = (
        f"Единое покрытие EPP/EDR для {users} АРМ с контролем статуса агентов, инцидентов и реакции на ransomware-сценарии."
        if not any(is_enabled(results.get(control)) for control in ("EDR", "XDR", "MDR"))
        else f"Контроль покрытия endpoint-защиты для {users} АРМ: статус агентов, качество телеметрии, сценарии реагирования и метрики MTTD/MTTR."
    )
    return [
        ("Доступы", access_text),
        ("Рабочие места", endpoint_text),
        ("Инфраструктура", f"Мониторинг серверов, виртуализации, СХД и каналов связи для {servers} серверов с порогами, владельцами и отчетом по доступности."),
        ("Восстановление", "Проверенные RTO/RPO, immutable/offline backup и регулярный тест восстановления критичных бизнес-систем."),
        ("Публичные сервисы", "WAF/CDN, журналирование web-событий, разбор блокировок и связка с процессом реагирования."),
        ("Процессы", "Ежемесячный цикл patch management, управление изменениями, контроль исключений и регулярный управленческий отчет по рискам."),
    ]


def build_management_decisions(results, context):
    decisions = [
        "Утвердить владельцев направлений: доступы, рабочие места, инфраструктура, backup/DR, web-периметр.",
        "Провести пилот EDR/MDR на критичных группах пользователей и серверах, затем принять решение о масштабировании.",
        "Определить минимальный SOC/MSSP-scope: NGFW, AD/учетки, серверы, endpoint, backup, почта и web.",
        "Согласовать регулярный отчет для руководства: остаточные риски, выполненные меры, исключения, SLA закрытия критичных уязвимостей.",
    ]
    if is_enabled(results.get("MFA")):
        decisions.insert(
            1,
            "Запустить 30-дневный план: проверить покрытие MFA по критичным доступам, закрыть устаревшие ОС, формализовать patch management и проверить backup-восстановление."
        )
    else:
        decisions.insert(
            1,
            "Запустить 30-дневный план: включить MFA для критичных доступов, закрыть устаревшие ОС, формализовать patch management и проверить backup-восстановление."
        )
    if context.get("has_development"):
        decisions.append("Добавить требования AppSec в процесс релизов: SAST/DAST, проверка зависимостей и критерии допуска в продуктив.")
    return decisions[:6]


# --- Презентация по результатам аудита ---
def presentation_text(value, limit=180):
    text = str(value or "")
    suspicious = sum(text.count(marker) for marker in ("Р", "С", "Ð", "Ñ"))
    if suspicious >= 3:
        try:
            repaired = text.encode("cp1251").decode("utf-8")
            original_badness = sum(text.count(marker) for marker in ("Р", "С", "Ð", "Ñ"))
            repaired_badness = sum(repaired.count(marker) for marker in ("Р", "С", "Ð", "Ñ"))
            if repaired_badness < original_badness:
                text = repaired
        except UnicodeError:
            pass
    text = re.sub(r"\s+", " ", text).strip(" .;-")
    if not text:
        return "Не указано"
    if len(text) <= limit:
        return text
    candidate = text[:limit]
    sentence_end = max(candidate.rfind(". "), candidate.rfind("! "), candidate.rfind("? "))
    if sentence_end >= max(48, limit // 2):
        return candidate[: sentence_end + 1].strip()
    shortened = candidate.rsplit(" ", 1)[0].rstrip(" ,;:.-")
    return f"{shortened}."


def presentation_brand_key():
    return "btg" if "btg" in get_app_instance_label().lower() else "khalil"


def presentation_regulatory_summary(industry, regulatory_profile):
    profiles = {
        "Финтех / Банки": (
            "Для банков требования ИБ закреплены отдельными отраслевыми нормами",
            "Выбран банковский профиль. Требования финансового рынка применяются наряду с правилами защиты персональных данных.",
            "Нужно подтвердить управление доступом, журналирование, реагирование, непрерывность и контроль критичных информационных систем.",
        ),
        "Страхование": (
            "Для страхового сектора действуют требования финансового рынка и защиты данных",
            "Выбран страховой профиль, поэтому общие требования к персональным данным дополняются отраслевыми требованиями финансового рынка.",
            "Нужно подтвердить управление доступом, событиями ИБ, инцидентами, резервированием и непрерывностью ключевых сервисов.",
        ),
        "Здравоохранение / Медицинская организация": (
            "Медицинские данные требуют отдельного режима защиты",
            "Выбран профиль медицинской организации. Помимо общих правил защиты данных применяются требования к персональным медицинским данным.",
            "Нужно подтвердить разграничение доступа, учет действий, сохранность, резервирование и порядок реагирования на инциденты.",
        ),
        "Госсектор": (
            "Для государственного сектора требования ИКТ и ИБ являются частью обязательного контура",
            "Выбран профиль государственного сектора. Применимость единых требований определяется статусом систем и государственными интеграциями.",
            "Нужно подтвердить документацию по ИБ, управление доступом, журналирование, мониторинг, реагирование и восстановление.",
        ),
        "Квазигосударственный сектор": (
            "Для квазигосударственного сектора применимость требований зависит от роли систем и интеграций",
            "Выбран квазигосударственный профиль. Единые требования применяются с учетом статуса организации, систем и государственных интеграций.",
            "Нужно подтвердить границы применимости, документацию по ИБ, журналирование, реагирование и восстановление.",
        ),
        "КВОИКИ / Критическая инфраструктура": (
            "Для КВОИКИ мониторинг, реагирование и восстановление входят в обязательный контур",
            "В анкете выбран профиль КВОИКИ. Включение конкретных объектов в официальный перечень и границы применимости подтверждаются документально.",
            "Нужно подтвердить ответственного по ИБ, мониторинг событий, план реагирования, устранение уязвимостей, восстановление и взаимодействие с ОЦИБ/НКЦИБ.",
        ),
    }
    title, applicability, expectations = profiles.get(industry, (
        "Регуляторные требования определяются отраслью, данными и ролью информационных систем",
        f"Выбран профиль «{industry or 'Другое'}». Базово применяются требования по защите персональных данных; дополнительные нормы требуют подтверждения.",
        "Нужно подтвердить состав регулируемых данных и систем, владельцев контролей, управление доступом, реагирование и восстановление.",
    ))
    legal_ids = regulatory_profile.get("legal_ids", [])
    anchor_priority = [
        "KVOIKI_MONITORING", "UNIFIED_832", "BANK_IS", "FINANCE_IS",
        "MEDICAL_DATA", "INFORMATIZATION", "KVOIKI_529", "PD_RULES", "PD_LAW",
    ]
    anchors = [
        REGULATORY_CATALOG[item_id]["short"]
        for item_id in anchor_priority
        if item_id in legal_ids and item_id in REGULATORY_CATALOG
    ][:4]
    return {
        "title": title,
        "applicability": applicability,
        "expectations": expectations,
        "implementation": "Конкретное основание указано рядом с каждой мерой на следующих слайдах. Норма определяет требуемый контроль, а продукт выбирается после архитектурной проработки.",
        "anchors": "; ".join(anchors),
    }


def presentation_legal_basis(semantic_key, regulatory_profile):
    legal_ids = set(regulatory_profile.get("legal_ids", [])) if regulatory_profile else set()
    priority_by_key = {
        "siem_soc": ["KVOIKI_MONITORING", "UNIFIED_832", "BANK_IS", "FINANCE_IS", "INFORMATIZATION"],
        "it_monitoring": ["KVOIKI_MONITORING", "UNIFIED_832", "BANK_IS", "INFORMATIZATION"],
        "patch": ["KVOIKI_MONITORING", "UNIFIED_832", "BANK_IS", "FINANCE_IS"],
        "backup": ["UNIFIED_832", "BANK_IS", "FINANCE_IS", "MEDICAL_DATA", "INFORMATIZATION"],
        "mfa": ["UNIFIED_832", "BANK_IS", "FINANCE_IS", "MEDICAL_DATA", "PD_RULES"],
        "iam": ["UNIFIED_832", "BANK_IS", "FINANCE_IS", "MEDICAL_DATA", "PD_RULES"],
        "pam": ["UNIFIED_832", "BANK_IS", "FINANCE_IS", "PD_RULES"],
    }
    default_priority = [
        "UNIFIED_832", "KVOIKI_MONITORING", "BANK_IS", "FINANCE_IS",
        "MEDICAL_DATA", "INFORMATIZATION", "PD_RULES", "PD_LAW", "KVOIKI_529",
    ]
    selected = [
        item_id
        for item_id in priority_by_key.get(semantic_key, default_priority)
        if item_id in legal_ids and item_id in REGULATORY_CATALOG
    ][:2]
    return "; ".join(REGULATORY_CATALOG[item_id]["short"] for item_id in selected)


def presentation_action_text(value, limit=165):
    def complete_sentence(sentence):
        sentence = str(sentence or "").strip(" .;-")
        words = sentence.split()
        incomplete_tail = {"и", "или", "а", "но", "с", "со", "для", "на", "по", "в", "во", "к", "из", "под"}
        while words and words[-1].lower().strip(".,:;()") in incomplete_tail:
            words.pop()
        sentence = " ".join(words)
        if sentence and sentence[-1] not in ".!?":
            sentence += "."
        return sentence

    text = presentation_text(value, 10000)
    text = re.sub(r"\s*\((?:например|напр\.)[^)]*\)", "", text, flags=re.IGNORECASE)
    text = re.sub(r"\s+\d+\.\s*$", "", text).strip()
    numbered_parts = [
        part.strip(" .;-")
        for part in re.split(r"(?:^|\s)\d+\.\s*", text)
        if part.strip(" .;-")
    ]
    parts = numbered_parts if len(numbered_parts) > 1 else [
        part.strip(" .;-")
        for part in re.split(r";\s*", text)
        if part.strip(" .;-")
    ]

    selected = []
    for part in parts:
        sentence = complete_sentence(part)
        if not sentence:
            continue
        candidate = " ".join([*selected, sentence])
        if len(candidate) > limit:
            if selected:
                break
            clauses = [
                clause.strip(" .;-")
                for clause in re.split(r"[,;:]\s*", sentence)
                if clause.strip(" .;-")
            ]
            compact = []
            for clause in clauses:
                clause_candidate = ", ".join([*compact, clause])
                if compact and len(clause_candidate) > limit - 1:
                    break
                if not compact and len(clause_candidate) > limit - 1:
                    words = clause.split()
                    while words and len(" ".join(words)) > limit - 2:
                        words.pop()
                    clause_candidate = " ".join(words)
                if clause_candidate:
                    compact = clause_candidate.split(", ")
            sentence = complete_sentence(", ".join(compact))
            if not sentence:
                continue
            candidate = sentence
        if selected and len(candidate) > limit:
            break
        selected.append(sentence)
        if len(candidate) >= limit * 0.72:
            break
    return " ".join(selected) or complete_sentence(presentation_text(text, limit))


def presentation_title_text(value, limit=88):
    """Shorten a title at a semantic boundary without leaving a broken phrase."""
    text = re.sub(r"\s+", " ", str(value or "")).strip(" .;-:")
    if len(text) <= limit:
        return text
    clauses = [part.strip(" .;-:") for part in re.split(r"[,;:]\s*", text) if part.strip(" .;-:")]
    if clauses and len(clauses[0]) <= limit:
        return clauses[0]
    words = text.split()
    incomplete_tail = {
        "и", "или", "а", "но", "с", "со", "для", "на", "по", "в", "во",
        "к", "из", "под", "без", "при", "после", "между", "над",
    }
    while words and len(" ".join(words)) > limit - 1:
        words.pop()
    while words and words[-1].lower().strip(".,:;()") in incomplete_tail:
        words.pop()
    return (" ".join(words).rstrip(" .;:-") + "…") if words else "Риск требует уточнения"


def presentation_maturity_style(score):
    """Return a clear red-yellow-green maturity scale for customer-facing slides."""
    value = max(0, min(100, int(score or 0)))
    if value < 40:
        return "#D92D20", "#FFFFFF"
    if value < 70:
        return "#F4B400", "#1F2937"
    return "#13877C", "#FFFFFF"


def presentation_evidence_for_key(semantic_key, results, context, item):
    """Use one questionnaire-grounded reason per recommendation card."""
    users = int(context.get("users", 0) or 0)
    servers = int(context.get("servers", 0) or 0)
    legacy_arm = int(results.get("ОС АРМ (Windows XP/Vista/7/8)", 0) or 0)
    legacy_servers = int(results.get("ОС Сервера (Windows Server 2008/2012 R2)", 0) or 0)
    wifi_users = results.get("_user_count") or users or "не указано"
    wifi_points = (
        results.get("WiFi Точки")
        or results.get("Wi-Fi Точки доступа")
        or results.get("Количество точек доступа")
        or "не указано"
    )
    wifi_controller = (
        results.get("WiFi Контроллер")
        or results.get("Wi-Fi Контроллер")
        or "Нет"
    )
    values = {
        "mfa": f"В анкете MFA: {results.get('MFA', 'Нет')}. Критичные доступы требуют подтвержденного покрытия вторым фактором.",
        "legacy_os": f"В анкете указаны устаревшие ОС: {legacy_arm} АРМ и {legacy_servers} серверов.",
        "siem_soc": f"В анкете SIEM: {results.get('SIEM', 'Нет')}. Централизованный контроль событий для критичных источников не подтвержден.",
        "patch": f"В анкете Patch Management: {results.get('Patch Management', 'Нет')}. Для {users} АРМ и {servers} серверов нужен управляемый цикл обновлений.",
        "endpoint_detection": (
            f"В анкете EPP: {results.get('EPP', 'Нет')}; "
            f"EDR/XDR/MDR: {results.get('EDR', 'Нет')}/{results.get('XDR', 'Нет')}/{results.get('MDR', 'Нет')}."
        ),
        "backup": f"В анкете резервное копирование: {results.get('Резервное копирование', 'Нет')}. RTO/RPO и результаты тестового восстановления не зафиксированы.",
        "web_waf": f"Публичные сервисы: {'есть' if context.get('has_public_web') else 'не указаны'}; WAF: {results.get('WAF', 'Нет')}.",
        "pam": f"В анкете PAM: {results.get('PAM', 'Нет')}; серверный контур: {servers} серверов.",
        "mail": f"Почтовая система: {results.get('1.5.1. Почтовая система', 'не указана')}; Mail Security: {results.get('Mail Security', 'Нет')}.",
        "appsec": f"Разработка: {'есть' if context.get('has_development') else 'не указана'}; SAST/DAST: {results.get('SAST', 'Нет')}/{results.get('DAST', 'Нет')}.",
        "wifi_capacity": (
            f"Пользователей: {wifi_users}; точек доступа: {wifi_points}; "
            f"Wi-Fi контроллер: {wifi_controller}."
        ),
        "network_performance": f"Основной канал: {results.get('Интернет канал (осн)', 'не указан')}; резервный: {results.get('Резервный канал', 'Нет')}; маршрутизация: {results.get('Маршрутизация', 'Нет')}.",
        "segmentation": "В анкете не приведены схема VLAN/VRF, ACL и матрица межсегментных потоков; OSPF сам по себе не подтверждает сегментацию.",
        "nac": f"В анкете NAC: {results.get('NAC', 'Нет')}. Требуется подтвердить контроль допуска проводных, Wi-Fi и неизвестных устройств.",
        "dlp": f"В анкете DLP: {results.get('DLP', 'Нет')}. Обработка персональных данных указана; контролируемые каналы передачи не подтверждены.",
        "itam": "В анкете не подтверждены единый реестр ПО, лицензий, владельцев активов и сроки поддержки.",
        "change_management": "В анкете не подтверждены единый процесс согласования изменений, тестирования и плана отката.",
        "it_monitoring": "В анкете не подтверждены единые метрики доступности, производительности и емкости для критичных сервисов.",
        "virtualization": f"В анкете виртуальных серверов: {results.get('Серверы (вирт)', 0)}; HA/DRS и резервы ресурсов не описаны.",
        "storage": "В анкете указаны СХД и RAID, но latency/IOPS, утилизация и запас емкости не раскрыты.",
        "dr": "В анкете не подтверждены согласованные RTO/RPO и результаты регулярного тестового восстановления.",
    }
    return presentation_action_text(values.get(semantic_key) or item.get("description") or item.get("impact"), 165)


def presentation_recommendation_key(item):
    normalized = {
        "risk": item.get("risk") or item.get("domain") or "",
        "description": item.get("description") or "",
        "impact": item.get("impact") or "",
        "recommendation": item.get("recommendation") or item.get("action") or "",
    }
    return risk_semantic_key(normalized)


def presentation_presales_profile(item):
    key = presentation_recommendation_key(item)
    profiles = {
        "legacy_os": {
            "title": "Устаревшие ОС требуют плана миграции",
            "impact": "Неподдерживаемые ОС повышают риск эксплуатации известных уязвимостей и ограничивают применение современных средств защиты.",
            "action": "Составить реестр устаревших ОС, согласовать миграцию и временно изолировать системы, которые нельзя обновить сразу.",
        },
        "wifi_capacity": {
            "title": "Wi-Fi требует проверки емкости и централизованного управления",
            "impact": "Недостаточная плотность точек доступа и отсутствие единого управления могут вызывать перегрузку, нестабильный роуминг и снижение производительности рабочих сервисов.",
            "action": "Провести радиообследование и замеры нагрузки, определить требуемую плотность точек доступа, затем внедрить централизованное WLAN-управление и контроль качества покрытия.",
        },
        "network_performance": {
            "title": "Производительность сети требует подтверждения измерениями",
            "impact": "Без замеров загрузки каналов, задержек и отказоустойчивости нельзя достоверно оценить запас производительности сети.",
            "action": "Провести аудит топологии и загрузки каналов, определить узкие места и подготовить целевую архитектуру сети.",
        },
        "change_management": {
            "title": "Изменения и конфигурации управляются неформально",
            "impact": "Несогласованные изменения повышают вероятность простоев, ошибок конфигурации и длительного восстановления критичных сервисов.",
            "action": "Ввести единый процесс запроса, согласования, тестирования, внедрения и отката изменений с назначенными владельцами.",
        },
        "it_monitoring": {
            "title": "Нет единого мониторинга доступности и производительности",
            "impact": "Без централизованных метрик команда поздно замечает деградацию сервисов и не может обоснованно планировать емкость.",
            "action": "Определить критичные метрики, пороги и владельцев, затем внедрить единый контроль доступности, производительности и емкости.",
        },
        "virtualization": {
            "title": "Виртуальной среде требуется подтвержденный запас ресурсов",
            "impact": "Высокая утилизация вычислительных ресурсов снижает запас на отказ хоста, рост нагрузки и обслуживание без простоя.",
            "action": "Провести capacity-анализ, проверить сценарий отказа хоста и утвердить план расширения вычислительных ресурсов.",
        },
        "storage": {
            "title": "Емкость и производительность СХД требуют измерения",
            "impact": "Без данных о latency, IOPS, утилизации и запасе емкости нельзя подтвердить устойчивость хранения критичных данных.",
            "action": "Провести health-check СХД, зафиксировать базовые метрики и подготовить план развития емкости и отказоустойчивости.",
        },
        "dr": {
            "title": "Тестирование восстановления и RTO/RPO не формализованы",
            "impact": "Наличие резервных копий не подтверждает, что критичные сервисы восстановятся в согласованные сроки после инцидента.",
            "action": "Согласовать RTO/RPO, провести контрольное восстановление и утвердить регулярные DR-учения с фиксацией результата.",
        },
        "mfa": {
            "title": "Критичные доступы не полностью защищены MFA",
            "impact": "Компрометация пароля может открыть доступ к почте, удаленным подключениям, административным консолям и бизнес-системам.",
            "action": "Проверить фактическое покрытие MFA и закрыть административные, удаленные и критичные доступы без второго фактора.",
        },
        "iam": {
            "title": "Жизненный цикл учетных записей требует централизованного управления",
            "impact": "Несвоевременное создание, изменение или отзыв прав повышает риск избыточного и несанкционированного доступа.",
            "action": "Определить процессы joiner/mover/leaver, владельцев ролей и интеграции; провести PoC IAM на ограниченной группе и подтвердить критерии масштабирования.",
        },
        "siem_soc": {
            "title": "События ИБ не собираются в единый контур",
            "impact": "Разрозненные журналы замедляют обнаружение атак и усложняют расследование инцидентов в критичных системах.",
            "action": "Определить критичные источники событий и сценарии контроля, затем выбрать поэтапную модель SIEM, SOC или MSSP.",
        },
        "endpoint_detection": {
            "title": "Endpoint-защита требует усиления обнаружения и реагирования",
            "impact": "Базовая антивирусная защита не дает полной телеметрии для расследования сложных атак и бокового перемещения.",
            "action": "Проверить покрытие EPP, провести пилот EDR или MDR на критичных группах и утвердить регламент реагирования.",
        },
        "web_waf": {
            "title": "Публичные веб-сервисы требуют прикладной защиты",
            "impact": "Интернет-магазин и личный кабинет подвержены прикладным атакам, бот-активности и нарушениям доступности.",
            "action": "Оценить веб-периметр, проверить действующие политики и внедрить WAF с контролем блокировок и доступности.",
        },
        "patch": {
            "title": "Уязвимости и обновления требуют управляемого цикла",
            "impact": "Без регулярной инвентаризации и SLA критичные уязвимости могут оставаться открытыми дольше допустимого срока.",
            "action": "Ввести регулярное сканирование, приоритизацию, SLA устранения и контроль исключений по критичным уязвимостям.",
        },
        "backup": {
            "title": "Восстановление критичных сервисов требует проверки",
            "impact": "Наличие резервных копий не гарантирует восстановление в требуемые сроки после сбоя или ransomware.",
            "action": "Согласовать RTO и RPO, изолировать критичные копии и регулярно подтверждать восстановление контрольными тестами.",
        },
        "itam": {
            "title": "Программные активы требуют единого учета",
            "impact": "Неполный учет версий, лицензий и сроков поддержки затрудняет обновления, бюджетирование и контроль технического долга.",
            "action": "Создать единый реестр ПО, лицензий, версий, владельцев и сроков поддержки с регулярной актуализацией.",
        },
        "pam": {
            "title": "Привилегированные доступы требуют отдельного контроля",
            "impact": "Общие или неконтролируемые административные учетные записи повышают риск несанкционированных изменений и компрометации.",
            "action": "Инвентаризировать привилегированные учетные записи, разделить персональные доступы и внедрить контроль критичных сессий.",
        },
        "nac": {
            "title": "Допуск устройств к сети не контролируется автоматически",
            "impact": "Без централизованного профилирования и политики допуска неизвестные или несоответствующие требованиям устройства могут попасть в корпоративную сеть.",
            "action": "Провести пилот NAC на Wi-Fi и одном проводном сегменте, настроить профилирование, проверку соответствия и изоляцию неизвестных устройств.",
        },
        "dlp": {
            "title": "Отсутствие DLP повышает риск утечки персональных данных",
            "impact": "Неконтролируемая передача данных может привести к утечке, регуляторным последствиям и репутационному ущербу.",
            "action": "Определить категории данных и каналы контроля, провести ограниченный пилот DLP и масштабировать подтвержденные политики.",
        },
        "iam": {
            "title": "Жизненный цикл учетных записей не автоматизирован",
            "impact": "Задержка отзыва доступа и накопление избыточных прав повышают вероятность несанкционированного доступа к бизнес-системам.",
            "action": "Описать прием, перевод и увольнение, определить владельцев ролей и провести PoC IAM на выбранных подразделениях и критичных системах.",
        },
        "segmentation": {
            "title": "Архитектуру сегментации необходимо подтвердить",
            "impact": "Без схемы VLAN, ACL и межсегментных политик нельзя достоверно оценить возможность бокового перемещения.",
            "action": "Проверить схему сегментов и матрицу потоков, подтвердить правила фильтрации и устранить выявленные разрывы.",
        },
        "mail": {
            "title": "Почтовый контур требует усиления защиты",
            "impact": "Фишинг и вредоносные вложения остаются одним из основных каналов компрометации учетных записей и рабочих мест.",
            "action": "Проверить текущие почтовые политики и усилить защиту от фишинга, подмены отправителя и вредоносных вложений.",
        },
        "appsec": {
            "title": "Проверки безопасности необходимо встроить в релизы",
            "impact": "Без автоматизированных проверок уязвимости приложений и зависимостей могут попадать в продуктивную среду.",
            "action": "Определить обязательные проверки кода, зависимостей и веб-приложений, затем встроить их в процесс выпуска релизов.",
        },
    }
    return key, profiles.get(key, {})


def presentation_success_metric(semantic_key):
    metrics = {
        "mfa": "100% критичных и удаленных учетных записей защищены MFA",
        "iam": "100% учетных записей имеют владельца и управляемый жизненный цикл",
        "legacy_os": "Нет рабочих мест на ОС без поддержки либо утвержден план миграции",
        "siem_soc": "Критичные источники подключены, SLA разбора событий утвержден",
        "patch": "Критичные уязвимости устраняются в согласованный SLA",
        "endpoint_detection": "Не менее 98% конечных точек передают телеметрию",
        "backup": "Тест восстановления подтверждает согласованные RTO и RPO",
        "web_waf": "Все публичные приложения защищены и проходят регулярную проверку",
        "pam": "Привилегированные учетные записи учтены и контролируются",
        "nac": "100% подключений идентифицируются; неизвестные устройства изолируются",
        "wifi_capacity": "Покрытие и емкость Wi-Fi подтверждены радиообследованием; пиковая загрузка точек остается в целевых пределах",
        "dlp": "Политики DLP контролируют согласованные каналы передачи персональных данных",
        "segmentation": "Матрица VLAN/ACL подтверждена тестом межсегментного доступа",
        "mail": "Защитные политики применены ко всем почтовым ящикам",
        "appsec": "Критичные релизы проходят обязательные проверки безопасности",
        "network_performance": "Емкость каналов подтверждена замерами и SLA",
        "itam": "Не менее 95% активов имеют владельца и актуальный статус",
        "change_management": "Все продуктивные изменения имеют согласование и план отката",
        "it_monitoring": "Критичные сервисы имеют метрики, пороги и владельцев реакции",
        "virtualization": "Запас ресурсов подтвержден для отказа одного хоста и прогнозируемого роста",
        "storage": "Емкость и производительность контролируются по утвержденным порогам и прогнозу",
        "dr": "Критичные сервисы проходят тест восстановления в пределах утвержденных RTO/RPO",
    }
    return metrics.get(semantic_key, "Владелец, срок и измеримый критерий результата утверждены")


def canonical_roadmap_action(item, phase):
    """Build one stage-appropriate action from a confirmed audit finding."""
    key = risk_semantic_key(item)
    actions = {
        "legacy_os": {
            "0-30 дней": "Изолировать или обновить устройства на неподдерживаемых ОС и утвердить срок полной миграции.",
            "31-60 дней": "Завершить миграцию приоритетных legacy-устройств и проверить совместимость прикладного ПО.",
            "61-90 дней": "Закрыть согласованный legacy-контур и включить контроль сроков поддержки ОС.",
        },
        "virtualization": {
            "0-30 дней": "Провести capacity-анализ виртуальной среды, проверить запас на отказ одного хоста и прогноз роста нагрузки.",
            "31-60 дней": "Утвердить план расширения вычислительных ресурсов по результатам capacity-анализа.",
            "61-90 дней": "Выполнить приоритетный этап расширения и включить регулярный контроль загрузки виртуальной среды.",
        },
        "storage": {
            "0-30 дней": "Провести health-check СХД и зафиксировать latency, IOPS, утилизацию и запас емкости.",
            "31-60 дней": "Утвердить целевую архитектуру и план развития емкости и отказоустойчивости СХД.",
            "61-90 дней": "Реализовать приоритетный этап развития СХД и включить контроль порогов емкости и производительности.",
        },
        "network_performance": {
            "0-30 дней": "Замерить загрузку WAN-каналов, проверить failover и согласовать требования к резервной полосе и SLA.",
            "31-60 дней": "Проверить целевую конфигурацию резервного канала и автоматическое переключение критичных сервисов.",
            "61-90 дней": "Ввести регулярный тест failover, контроль доступности каналов и отчетность по SLA.",
        },
        "wifi_capacity": {
            "0-30 дней": "Провести радиообследование Wi-Fi и замерить пиковую нагрузку, покрытие и качество роуминга.",
            "31-60 дней": "Провести пилот централизованного WLAN-управления и подтвердить целевой радиоплан.",
            "61-90 дней": "Масштабировать подтвержденную WLAN-архитектуру и контролировать загрузку, покрытие и роуминг.",
        },
        "change_management": {
            "0-30 дней": "Описать текущий поток изменений, владельцев, точки согласования и причины неуспешных изменений.",
            "31-60 дней": "Внедрить единый процесс согласования, тестирования и отката изменений с обязательной регистрацией.",
            "61-90 дней": "Закрепить процесс управления изменениями метриками качества и регулярным разбором отклонений.",
        },
        "itam": {
            "0-30 дней": "Определить владельцев, обязательные атрибуты активов и границы первого контура CMDB.",
            "31-60 дней": "Провести пилот CMDB и связать критичные активы с сервисами, изменениями и SLA.",
            "61-90 дней": "Расширить CMDB на целевой контур и включить регулярный контроль качества данных.",
        },
        "it_monitoring": {
            "0-30 дней": "Определить критичные сервисы, метрики доступности и производительности, пороги и владельцев реакции.",
            "31-60 дней": "Провести пилот единого ИТ-мониторинга на критичных сервисах и проверить правила оповещения.",
            "61-90 дней": "Расширить мониторинг на целевой контур и включить регулярную отчетность по доступности и емкости.",
        },
        "backup": {
            "0-30 дней": "Согласовать RTO/RPO и перечень критичных сервисов для контрольного восстановления.",
            "31-60 дней": "Провести тест восстановления критичных сервисов и зафиксировать фактические RTO/RPO.",
            "61-90 дней": "Утвердить регулярные тесты восстановления и контроль изолированных резервных копий.",
        },
        "dr": {
            "0-30 дней": "Определить критичные сервисы, зависимости, владельцев и требования к аварийному восстановлению.",
            "31-60 дней": "Провести ограниченное DR-учение и скорректировать runbook по фактическим результатам.",
            "61-90 дней": "Утвердить DR-runbook, периодичность учений и контроль выполнения целевых RTO/RPO.",
        },
        "pam": {
            "0-30 дней": "Инвентаризировать привилегированные доступы, критичные системы и границы пилота PAM.",
            "31-60 дней": "Провести пилот PAM и проверить vault, контроль сессий, аварийный доступ и интеграцию с SIEM.",
            "61-90 дней": "Расширить PAM на подтвержденный критичный контур и ввести регулярный пересмотр привилегий.",
        },
        "nac": {
            "0-30 дней": "Описать типы подключений, требования к NAC и пилотный контур проводной и беспроводной сети.",
            "31-60 дней": "Провести пилот NAC и проверить 802.1X, профилирование, соответствие и изоляцию устройств.",
            "61-90 дней": "Расширить подтвержденные политики NAC и включить контроль качества допуска устройств.",
        },
        "dlp": {
            "0-30 дней": "Определить категории данных, каналы контроля и измеримые критерии пилота DLP.",
            "31-60 дней": "Провести ограниченный пилот DLP и скорректировать политики по фактическим результатам.",
            "61-90 дней": "Масштабировать подтвержденные политики DLP и включить контроль инцидентов и исключений.",
        },
    }
    if key in actions:
        return actions[key][phase]

    recommendation = item.get("recommendation") or item.get("action") or item.get("description") or item.get("risk")
    if phase == "0-30 дней":
        return presentation_action_text(
            f"Подтвердить исходное состояние и границы меры: {recommendation}",
            220,
        )
    if phase == "31-60 дней":
        return presentation_action_text(recommendation, 220)
    return presentation_action_text(
        f"Масштабировать подтвержденную меру и включить контроль результата: {recommendation}",
        220,
    )


def build_canonical_report_roadmap(report_risks, results=None, context=None, max_items=6):
    """Create the Excel and PowerPoint roadmap from one fact-checked finding set."""
    phase_order = ("0-30 дней", "31-60 дней", "61-90 дней")
    preferred_phase = {
        "legacy_os": "0-30 дней",
        "virtualization": "0-30 дней",
        "storage": "0-30 дней",
        "network_performance": "0-30 дней",
        "segmentation": "0-30 дней",
        "change_management": "31-60 дней",
        "wifi_capacity": "31-60 дней",
        "patch": "31-60 дней",
        "itam": "31-60 дней",
        "mfa": "31-60 дней",
        "iam": "31-60 дней",
        "pam": "31-60 дней",
        "nac": "31-60 дней",
        "endpoint_detection": "31-60 дней",
        "mail": "31-60 дней",
        "web_waf": "31-60 дней",
        "backup": "61-90 дней",
        "dr": "61-90 дней",
        "it_monitoring": "61-90 дней",
        "siem_soc": "61-90 дней",
        "dlp": "61-90 дней",
        "appsec": "61-90 дней",
    }
    level_priority = {
        "CRITICAL": "P1", "КРИТИЧЕСКИЙ": "P1",
        "HIGH": "P1", "ВЫСОКИЙ": "P1",
        "MEDIUM": "P2", "СРЕДНИЙ": "P2",
        "LOW": "P3", "НИЗКИЙ": "P3",
    }
    domain_by_key = {
        "legacy_os": "Конечные устройства",
        "virtualization": "ИТ-инфраструктура",
        "storage": "Хранение данных",
        "network_performance": "Сеть и связь",
        "wifi_capacity": "Корпоративный Wi-Fi",
        "change_management": "ИТ-процессы",
        "itam": "Учет активов и сервисов",
        "it_monitoring": "ИТ-мониторинг",
        "backup": "Резервное копирование",
        "dr": "Непрерывность",
        "pam": "Привилегированный доступ",
        "nac": "Сетевой доступ",
        "dlp": "Защита данных",
    }
    phase_counts = {phase: 0 for phase in phase_order}
    roadmap = []
    seen = set()

    roadmap_key_order = {
        "legacy_os": 0,
        "network_performance": 1,
        "wifi_capacity": 2,
        "virtualization": 3,
        "storage": 4,
        "dr": 5,
        "it_monitoring": 6,
        "change_management": 7,
        "itam": 8,
    }
    ordered_risks = sorted(
        [item for item in (report_risks or []) if isinstance(item, dict)],
        key=lambda item: roadmap_key_order.get(risk_semantic_key(item), 50),
    )

    for raw_item in ordered_risks:
        if not isinstance(raw_item, dict):
            continue
        item = dict(raw_item)
        key = risk_semantic_key(item)
        if not key or key in seen:
            continue
        preferred = preferred_phase.get(key, "31-60 дней")
        preferred_index = phase_order.index(preferred)
        candidates = [
            *phase_order[preferred_index:],
            *reversed(phase_order[:preferred_index]),
        ]
        phase = next((candidate for candidate in candidates if phase_counts[candidate] < 2), None)
        if phase is None:
            break
        level = str(item.get("level", "MEDIUM")).strip().upper()
        priority = level_priority.get(level, "P2")
        area = str(item.get("area") or item.get("_ai_area") or "ИТ/ИБ").strip().upper()
        if area not in {"ИТ", "ИБ", "ИТ/ИБ"}:
            area = "ИТ" if key in {
                "legacy_os", "network_performance", "wifi_capacity", "virtualization",
                "storage", "dr", "it_monitoring", "change_management", "itam",
            } else "ИБ"
        metric = item.get("success_metric") or presentation_success_metric(key)
        rationale = item.get("impact") or item.get("description") or "Мера снижает подтвержденный риск аудита."
        roadmap.append({
            "phase": phase,
            "priority": priority,
            "domain": domain_by_key.get(key, area),
            "action": canonical_roadmap_action(item, phase),
            "rationale": presentation_action_text(rationale, 240),
            "owner": "ИТ" if area == "ИТ" else ("ИБ" if area == "ИБ" else "ИТ/ИБ"),
            "effort": "Высокая" if key in {"virtualization", "storage", "siem_soc", "pam", "iam"} else "Средняя",
            "result": presentation_action_text(metric, 155),
            "semantic_key": key,
        })
        seen.add(key)
        phase_counts[phase] += 1
        if len(roadmap) >= max_items:
            break

    # Keep a six-position management timeline readable when there are only five
    # distinct findings. The continuation must advance an existing action rather
    # than introduce an unconfirmed topic.
    if roadmap and len(roadmap) < max_items:
        for phase in phase_order:
            while phase_counts[phase] < 2 and len(roadmap) < max_items:
                candidate = next(
                    (
                        item for item in roadmap
                        if item["semantic_key"] not in {
                            existing["semantic_key"]
                            for existing in roadmap
                            if existing["phase"] == phase
                        }
                    ),
                    None,
                )
                if candidate is None:
                    break
                source = next(
                    item for item in report_risks
                    if isinstance(item, dict) and risk_semantic_key(item) == candidate["semantic_key"]
                )
                roadmap.append({
                    **candidate,
                    "phase": phase,
                    "action": canonical_roadmap_action(source, phase),
                })
                phase_counts[phase] += 1

    return sorted(
        roadmap,
        key=lambda item: (
            phase_order.index(item["phase"]),
            {"P1": 0, "P2": 1, "P3": 2}.get(item["priority"], 9),
        ),
    )


def presentation_severity_style(level):
    normalized = str(level or "MEDIUM").strip().upper()
    aliases = {
        "КРИТИЧЕСКИЙ": "CRITICAL",
        "ВЫСОКИЙ": "HIGH",
        "СРЕДНИЙ": "MEDIUM",
        "НИЗКИЙ": "LOW",
    }
    normalized = aliases.get(normalized, normalized)
    styles = {
        "CRITICAL": ("#D92D20", "#FFFFFF"),
        "HIGH": ("#EA580C", "#FFFFFF"),
        "MEDIUM": ("#F4B400", "#1F2937"),
        "LOW": ("#16A34A", "#FFFFFF"),
    }
    return normalized, *styles.get(normalized, styles["MEDIUM"])


def presentation_recommendation_entry(item, regulatory_profile=None, results=None, context=None):
    normalized = dict(item)
    for field in ("risk", "description", "impact", "recommendation", "action", "success_metric"):
        if field in normalized:
            normalized[field] = expand_regulatory_references(normalized[field])
    normalized["risk"] = normalized.get("risk") or normalized.get("domain") or "Рекомендация"
    normalized["recommendation"] = normalized.get("recommendation") or normalized.get("action") or normalized.get("description")
    semantic_key, profile = presentation_presales_profile(normalized)
    ai_authored = str(item.get("source") or item.get("_source") or "").strip().lower() in {
        "ии", "ai", "gemini", "groq"
    }
    generic_titles = {"ит", "иб", "ит/иб", "рекомендация"}
    title_by_key = {
        "mfa": "Многофакторная аутентификация",
        "iam": "Управление жизненным циклом учетных записей",
        "legacy_os": "Обновление устаревших операционных систем",
        "siem_soc": "Мониторинг событий и реагирование",
        "patch": "Управление уязвимостями и обновлениями",
        "endpoint_detection": "Защита конечных точек",
        "backup": "Резервное копирование и восстановление",
        "web_waf": "Защита публичных веб-сервисов",
        "pam": "Контроль привилегированных доступов",
        "nac": "Контроль допуска устройств к сети",
        "wifi_capacity": "Емкость и управляемость корпоративного Wi-Fi",
        "network_performance": "Управляемость и производительность сети",
        "itam": "Управление программными активами",
        "change_management": "Управление изменениями и конфигурациями",
        "it_monitoring": "Централизованный мониторинг ИТ",
    }
    raw_title = str(normalized["risk"] or "Рекомендация").strip()
    if profile.get("title") and not ai_authored:
        raw_title = profile["title"]
    elif raw_title.lower() in generic_titles:
        raw_title = title_by_key.get(semantic_key, "Практическая мера улучшения")
    title = presentation_title_text(raw_title, 88)
    action = (
        presentation_action_text(normalized["recommendation"], 190)
        if ai_authored
        else profile.get("action") or presentation_action_text(normalized["recommendation"], 190)
    )
    solution = presentation_text(solution_categories_for_report_item(normalized), 88)
    vendors = portfolio_manufacturers_for_report_item(normalized)
    if "матрице" in str(vendors).lower() or "нет подходящего" in str(vendors).lower():
        vendors = "Подбор после уточнения требований"
    else:
        vendor_limit = 5 if semantic_key == "web_waf" else 4
        vendors = ", ".join(split_portfolio_list(vendors)[:vendor_limit])
    if semantic_key in {"patch", "itam"} and "hcl" not in str(vendors).lower():
        vendors = ", ".join([*split_portfolio_list(vendors), "HCL BigFix"])
    vendors = presentation_text(vendors, 92)
    evidence_values = normalized.get("evidence", [])
    if not isinstance(evidence_values, list):
        evidence_values = [evidence_values] if evidence_values else []
    evidence = normalized.get("description") or "; ".join(
        str(value).strip() for value in evidence_values if str(value).strip()
    )
    if results is not None and context is not None and not evidence:
        evidence = presentation_evidence_for_key(semantic_key, results, context, item)
    elif not evidence:
        evidence = normalized.get("description") or normalized.get("impact") or "Основание приоритета требует уточнения"

    legal_ids = [
        value for value in item.get("legal_ids", [])
        if value in REGULATORY_CATALOG
    ]
    if legal_ids:
        legal = "; ".join(REGULATORY_CATALOG[value]["short"] for value in legal_ids[:2])
    elif regulatory_profile:
        legal = presentation_legal_basis(semantic_key, regulatory_profile)
        if not legal:
            legal = "Применимость подтверждается с учетом отрасли и роли организации"
    else:
        legal = "Применимость подтверждается с учетом отрасли и роли организации"

    raw_level, fill_color, text_color = presentation_severity_style(normalized.get("level"))
    return {
        "key": semantic_key,
        "level": risk_level_label(raw_level).upper(),
        "title": title,
        "action": action,
        "solution": solution,
        "vendors": vendors,
        "evidence": presentation_action_text(evidence, 165),
        "legal": presentation_text(legal, 130),
        "metric": presentation_text(item.get("success_metric") or presentation_success_metric(semantic_key), 125),
        "fill_color": fill_color,
        "text_color": text_color,
    }


def presentation_risk_entry(item):
    normalized = dict(item)
    for field in ("risk", "description", "impact", "recommendation", "action"):
        if field in normalized:
            normalized[field] = expand_regulatory_references(normalized[field])
    normalized["recommendation"] = normalized.get("recommendation") or normalized.get("action") or normalized.get("description")
    _, profile = presentation_presales_profile(normalized)
    raw_level, fill_color, text_color = presentation_severity_style(normalized.get("level"))
    return {
        "level": presentation_text(risk_level_label(raw_level), 16).upper(),
        "title": profile.get("title") or presentation_title_text(
            normalized.get("risk", "Риск требует внимания"), 88
        ),
        "impact": profile.get("impact") or presentation_action_text(
            normalized.get("impact") or normalized.get("description") or "Требуется уточнить влияние риска.",
            155,
        ),
        "action": profile.get("action") or presentation_action_text(normalized["recommendation"], 135),
        "fill_color": fill_color,
        "text_color": text_color,
    }


def presentation_focus_items(context, business_systems):
    users = int(context.get("users", 0) or 0)
    servers = int(context.get("servers", 0) or 0)
    continuity_scope = []
    if servers:
        continuity_scope.append(russian_count(servers, "сервер", "сервера", "серверов"))
    if business_systems:
        continuity_scope.append(russian_count(business_systems, "бизнес-система", "бизнес-системы", "бизнес-систем"))
    continuity_subject = " и ".join(continuity_scope) or "Критичные сервисы"

    focus = [
        (
            "Непрерывность сервисов",
            f"{continuity_subject} требуют согласованных RTO/RPO и регулярной проверки восстановления.",
        ),
        (
            "Управляемость среды",
            f"{users} рабочих мест и серверный контур требуют единого контроля обновлений, изменений и конфигураций.",
        ),
    ]
    if context.get("has_public_web"):
        focus.append((
            "Цифровой периметр",
            "Публичные сервисы требуют проверки WAF/DDoS, внешних доступов и мониторинга событий.",
        ))
    else:
        focus.append((
            "Контроль доступа",
            "Критичные и удаленные доступы требуют MFA, учета привилегий и регулярного пересмотра прав.",
        ))
    return focus


def build_audit_presentation_replacements(c_info, results, final_score, it_maturity_score):
    it_score_fill, it_score_text = presentation_maturity_style(it_maturity_score)
    security_score_fill, security_score_text = presentation_maturity_style(final_score)
    context = build_context(results, c_info)
    regulatory_profile = industry_regulatory_profile(c_info.get("Сфера деятельности", ""))
    regulatory_summary = presentation_regulatory_summary(
        c_info.get("Сфера деятельности", ""),
        regulatory_profile,
    )
    domain_scores = calculate_domain_scores(results)
    # Excel and PowerPoint consume one canonical, fact-checked audit set. AI remains
    # mandatory for the customer deck, while expert rules may complete its coverage.
    risk_sources = list(st.session_state.get("last_report_risk_sources", []))
    ai_narrative = sanitize_ai_audit_narrative(
        st.session_state.get("ai_audit_narrative", {}),
        results,
    )
    summary_items = []
    narrative_summary = [
        *ai_narrative.get("executive_summary", []),
        *[item.get("text", "") for item in ai_narrative.get("audit_observations", [])],
    ]
    for item in narrative_summary:
        clean_item = presentation_action_text(item, 220)
        if clean_item not in summary_items:
            summary_items.append(clean_item)
        if len(summary_items) >= 4:
            break

    if final_score < 35:
        summary_title = "Ключевые процессы защиты пока зависят от отдельных средств и ручных действий"
    elif final_score < 60:
        summary_title = "Базовые меры работают, но контроль и реагирование требуют систематизации"
    else:
        summary_title = "Защитный контур сформирован; следующий резерв — устойчивость и измеримость"

    profile_title, profile_text = infrastructure_profile(context)
    business_systems = sum(
        bool(context.get(key))
        for key in ("has_erp", "has_crm", "has_accounting", "has_hrm", "has_document_flow", "has_mail")
    )
    focus_items = presentation_focus_items(context, business_systems)

    normalized_risks = []
    for source_item in risk_sources:
        if not isinstance(source_item, dict):
            continue
        if risk_conflicts_with_answers(source_item, results):
            continue
        # The expert XLSX and presentation consume the same already normalized
        # finding. Re-running semantic enrichment here can change a WAN finding
        # into Backup or an IT-monitoring finding into Virtualization.
        item = dict(source_item)
        normalized_risks.append({
            "level": risk_level_label(item.get("level", "MEDIUM")),
            "risk": item.get("risk", "Риск"),
            "description": item.get("description") or item.get("impact") or "Требуется уточнение основания",
            "impact": item.get("impact") or item.get("description") or "Требуется уточнение влияния",
            "recommendation": item.get("recommendation") or item.get("action") or "Требуется план улучшений",
            "vendors": item.get("vendors", []),
            "area": item.get("area") or item.get("_ai_area") or "ИТ/ИБ",
            "legal_ids": item.get("legal_ids", []),
            "frameworks": item.get("frameworks", []),
            "evidence": item.get("evidence", []),
            "success_metric": item.get("success_metric", ""),
            "source": item.get("source", ""),
            "semantic_key": risk_semantic_key(item),
        })
    severity_order = {"CRITICAL": 0, "КРИТИЧЕСКИЙ": 0, "HIGH": 1, "ВЫСОКИЙ": 1, "MEDIUM": 2, "СРЕДНИЙ": 2, "LOW": 3, "НИЗКИЙ": 3}
    normalized_risks.sort(key=lambda item: severity_order.get(str(item.get("level", "MEDIUM")).upper(), 2))

    deduplicated_risks = []
    risk_keys = set()
    for item in normalized_risks:
        semantic_key, _ = presentation_presales_profile(item)
        dedupe_key = semantic_key or re.sub(r"\s+", " ", str(item.get("risk", "")).strip().lower())
        if dedupe_key in risk_keys:
            continue
        risk_keys.add(dedupe_key)
        deduplicated_risks.append(item)
    normalized_risks = deduplicated_risks[:12]

    for item in normalized_risks:
        candidate = presentation_action_text(
            f"{item.get('risk', '')}: {item.get('impact') or item.get('description') or ''}",
            220,
        )
        if candidate and candidate not in summary_items:
            summary_items.append(candidate)
        if len(summary_items) >= 4:
            break

    recommendation_items = []
    recommendation_keys = set()

    def add_recommendation(item):
        entry = presentation_recommendation_entry(
            item,
            regulatory_profile,
            results=results,
            context=context,
        )
        key = entry["key"] or re.sub(r"\s+", " ", entry["title"].lower())
        if key in recommendation_keys:
            return
        recommendation_items.append(entry)
        recommendation_keys.add(key)

    for item in normalized_risks:
        add_recommendation(item)

    canonical_roadmap = build_canonical_report_roadmap(
        normalized_risks,
        results=results,
        context=context,
    )
    roadmap_by_phase = {"0-30": [], "31-60": [], "61-90": []}
    for item in canonical_roadmap:
        phase_key = next(
            (key for key in roadmap_by_phase if key in str(item.get("phase", ""))),
            None,
        )
        if not phase_key:
            continue
        roadmap_by_phase[phase_key].append({
            "action": presentation_action_text(item.get("action"), 145),
            "result": presentation_action_text(item.get("result"), 110),
            "key": item.get("semantic_key"),
        })

    enabled_controls, _ = security_control_snapshot(results)
    strengths = [presentation_text(item, 105) for item in enabled_controls[:4]]
    if context.get("has_backup"):
        strengths.append("В инфраструктуре используется резервное копирование")
    if context.get("has_virtualization"):
        strengths.append("Серверная нагрузка использует виртуализацию")
    while len(strengths) < 4:
        strengths.append("Сильная сторона требует подтверждения на рабочей сессии")

    replacements = {
        "COMPANY": presentation_text(c_info.get("Наименование компании", "Компания"), 54),
        "INDUSTRY": presentation_text(c_info.get("Сфера деятельности", ""), 38),
        "CITY": presentation_text(c_info.get("Город", ""), 28),
        "SCORE": str(int(final_score)),
        "IT_SCORE": str(int(it_maturity_score)),
        "IT_SCORE_FILL": it_score_fill,
        "IT_SCORE_TEXT": it_score_text,
        "SCORE_FILL": security_score_fill,
        "SCORE_TEXT": security_score_text,
        "DATE": datetime.now().strftime("%d.%m.%Y"),
        "SUMMARY_TITLE": presentation_text(summary_title, 120),
        "USERS": str(context.get("users", 0)),
        "SERVERS": str(context.get("servers", 0)),
        "PUBLIC": "Есть" if context.get("has_public_web") else "Нет",
        "BUSINESS": str(business_systems),
        "PROFILE": presentation_text(f"{profile_title}. {profile_text}", 240),
        "SEC_LEVEL": get_maturity_level(final_score)[0],
        "IT_LEVEL": get_maturity_level(it_maturity_score)[0],
        "FRAMEWORKS": presentation_text(", ".join(regulatory_profile.get("frameworks", [])), 220),
        "FRAMEWORK_NOTE": (
            "Дополнительные стандарты применяются только после подтверждения соответствующих операций, данных и договорных обязательств."
        ),
        "REG_TITLE": presentation_text(regulatory_summary["title"], 150),
        "REG_APPLICABILITY": presentation_text(regulatory_summary["applicability"], 260),
        "REG_EXPECTATIONS": presentation_text(regulatory_summary["expectations"], 260),
        "REG_IMPLEMENTATION": presentation_text(regulatory_summary["implementation"], 260),
        "REG_ANCHORS": presentation_text(regulatory_summary["anchors"], 220),
    }

    threat_domains = [
        ("Сеть", domain_scores.get("Сетевая безопасность", 0)),
        ("Endpoint", domain_scores.get("Защита конечных точек", 0)),
        ("Доступы", domain_scores.get("Идентификация и доступ", 0)),
        ("Мониторинг", domain_scores.get("Мониторинг и SOC", 0)),
        ("Восстановление", domain_scores.get("Резервное копирование", 0)),
        ("Инфраструктура", domain_scores.get("Инфраструктура", 0)),
    ]
    coverage_values = [max(0, min(100, int(score or 0))) for _, score in threat_domains]
    replacements["COVERAGE_AVERAGE"] = str(round(sum(coverage_values) / max(1, len(coverage_values))))
    strongest_label, strongest_score = max(threat_domains, key=lambda item: int(item[1] or 0))
    weakest_label, weakest_score = min(threat_domains, key=lambda item: int(item[1] or 0))
    replacements["COVERAGE_INSIGHT"] = presentation_text(
        f"Сильнейший домен: {strongest_label} — {int(strongest_score or 0)}%. "
        f"Главный резерв улучшения: {weakest_label} — {int(weakest_score or 0)}%.",
        150,
    )
    for index, (label, score) in enumerate(threat_domains, start=1):
        coverage = max(0, min(100, int(score or 0)))
        if coverage < 40:
            fill = "#D92D20"
        elif coverage < 70:
            fill = "#F4B400"
        else:
            fill = "#13877C"
        replacements[f"THREAT_{index}_LABEL"] = label
        replacements[f"THREAT_{index}_VALUE"] = str(coverage)
        replacements[f"THREAT_{index}_FILL"] = fill
    for index, (title, text) in enumerate(focus_items, start=1):
        replacements[f"FOCUS_{index}_TITLE"] = presentation_text(title, 42)
        replacements[f"FOCUS_{index}_TEXT"] = presentation_text(text, 145)

    for index, strength in enumerate(strengths[:4], start=1):
        replacements[f"STRENGTH_{index}"] = strength

    laws = list(regulatory_profile.get("laws", []))
    while len(laws) < 4:
        laws.append({
            "title": "Дополнительные отраслевые требования",
            "short": "Применимость требует подтверждения",
            "scope": "Уточняется по лицензиям, роли организации и обрабатываемым данным.",
            "status": "Требует подтверждения",
        })
    for index, law in enumerate(laws[:4], start=1):
        replacements[f"LAW_{index}_TITLE"] = presentation_text(law.get("short") or law.get("title"), 90)
        replacements[f"LAW_{index}_SCOPE"] = presentation_text(law.get("scope", ""), 150)
        replacements[f"LAW_{index}_STATUS"] = presentation_text(law.get("status", ""), 48)

    summary_fallbacks = [
        f"Профиль инфраструктуры: {profile_title.lower()}.",
        f"Приоритетный резерв улучшения: {normalized_risks[0]['risk']}." if normalized_risks else "Критичные дополнительные разрывы по анкете не подтверждены.",
        "Очередность мер учитывает влияние на непрерывность сервисов и подтвержденные факты анкеты.",
    ]
    for fallback in summary_fallbacks:
        clean_fallback = presentation_action_text(fallback, 220)
        if clean_fallback and clean_fallback not in summary_items:
            summary_items.append(clean_fallback)

    for index in range(6):
        replacements[f"SUMMARY_{index + 1}"] = (
            summary_items[index]
            if index < len(summary_items)
            else "Дополнительный вывод не требуется: существенный отдельный разрыв не подтвержден."
        )
        if index < len(normalized_risks):
            risk = normalized_risks[index]
        elif index < len(recommendation_items):
            entry = recommendation_items[index]
            risk = {
                "level": entry["level"],
                "risk": entry["title"],
                "impact": entry["evidence"],
                "recommendation": entry["action"],
            }
        else:
            risk = {
                "level": "LOW",
                "risk": "Дополнительный критичный риск не подтвержден",
                "impact": "По данным анкеты оснований для отдельного вывода не выявлено.",
                "recommendation": "Сохранять текущий контроль и периодически пересматривать остаточные риски.",
            }
        risk_entry = presentation_risk_entry(risk)
        replacements[f"RISK_{index + 1}_LEVEL"] = risk_entry["level"]
        replacements[f"RISK_{index + 1}_TITLE"] = risk_entry["title"]
        replacements[f"RISK_{index + 1}_IMPACT"] = risk_entry["impact"]
        replacements[f"RISK_{index + 1}_RECOMMENDATION"] = risk_entry["action"]
        replacements[f"RISK_{index + 1}_FILL"] = risk_entry["fill_color"]
        replacements[f"RISK_{index + 1}_TEXT"] = risk_entry["text_color"]
        if index < len(ai_narrative.get("management_decisions", [])):
            replacements[f"DECISION_{index + 1}"] = presentation_text(
                ai_narrative["management_decisions"][index],
                205,
            )

    fallback_decisions = [
        presentation_action_text(
            f"Утвердить план по направлению «{entry['title']}»: {entry['action']}",
            205,
        )
        for entry in recommendation_items[:4]
    ]
    for index in range(4):
        replacements.setdefault(
            f"DECISION_{index + 1}",
            fallback_decisions[index]
            if index < len(fallback_decisions)
            else "Утвердить владельца контроля, целевой показатель и периодичность проверки результата.",
        )

    replacements["__RECOMMENDATION_COUNT__"] = len(recommendation_items)
    replacements["__RISK_COUNT__"] = len(normalized_risks)
    recommendation_fields = (
        "LEVEL", "TITLE", "ACTION", "SOLUTION", "VENDORS",
        "EVIDENCE", "LEGAL", "METRIC",
    )
    for index in range(1, max(8, len(recommendation_items)) + 1):
        for field in recommendation_fields:
            replacements[f"REC_{index}_{field}"] = ""
        replacements[f"REC_{index}_FILL"] = "#FFFFFF"
        replacements[f"REC_{index}_TEXT"] = "#FFFFFF"

    for index, entry in enumerate(recommendation_items, start=1):
        replacements[f"REC_{index}_LEVEL"] = entry["level"]
        replacements[f"REC_{index}_TITLE"] = entry["title"]
        replacements[f"REC_{index}_ACTION"] = entry["action"]
        replacements[f"REC_{index}_SOLUTION"] = entry["solution"]
        replacements[f"REC_{index}_VENDORS"] = entry["vendors"]
        replacements[f"REC_{index}_EVIDENCE"] = entry["evidence"]
        replacements[f"REC_{index}_LEGAL"] = entry["legal"]
        replacements[f"REC_{index}_METRIC"] = entry["metric"]
        replacements[f"REC_{index}_FILL"] = entry["fill_color"]
        replacements[f"REC_{index}_TEXT"] = entry["text_color"]

    for index in range(1, 5):
        replacements[f"OUTCOME_{index}_TITLE"] = "Поддержание зрелости"
        replacements[f"OUTCOME_{index}_FROM"] = "Критичный дополнительный разрыв не подтвержден"
        replacements[f"OUTCOME_{index}_TO"] = "Контроль регулярно пересматривается"
    for index, entry in enumerate(recommendation_items[:4], start=1):
        replacements[f"OUTCOME_{index}_TITLE"] = presentation_text(entry["title"], 70)
        replacements[f"OUTCOME_{index}_FROM"] = presentation_text(entry["evidence"], 145)
        replacements[f"OUTCOME_{index}_TO"] = presentation_text(entry["metric"], 125)

    for phase_index, phase in enumerate(("0-30", "31-60", "61-90"), start=1):
        while len(roadmap_by_phase[phase]) < 2:
            roadmap_by_phase[phase].append({
                "action": "Поддерживать действующий контроль и подтвердить его эффективность.",
                "result": "Контроль проверен, владелец и периодичность пересмотра утверждены.",
            })
        for item_index, item in enumerate(roadmap_by_phase[phase][:2], start=1):
            replacements[f"ROADMAP_{phase_index}_{item_index}"] = item["action"]
            replacements[f"ROADMAP_{phase_index}_{item_index}_RESULT"] = item["result"]
    return replacements


def render_audit_presentation_template(template_path, replacements):
    import zipfile
    import xml.etree.ElementTree as ET
    from xml.sax.saxutils import escape

    recommendation_count = max(0, int(replacements.get("__RECOMMENDATION_COUNT__", 8)))
    risk_count = max(0, int(replacements.get("__RISK_COUNT__", 0)))
    recommendation_slide_count = (recommendation_count + 1) // 2

    presentation_ns = "http://schemas.openxmlformats.org/presentationml/2006/main"
    relationships_ns = "http://schemas.openxmlformats.org/package/2006/relationships"
    office_relationships_ns = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
    drawing_ns = "http://schemas.openxmlformats.org/drawingml/2006/main"
    content_types_ns = "http://schemas.openxmlformats.org/package/2006/content-types"
    ET.register_namespace("p", presentation_ns)
    ET.register_namespace("a", drawing_ns)
    ET.register_namespace("r", office_relationships_ns)

    output = BytesIO()
    with zipfile.ZipFile(template_path, "r") as source:
        source_items = source.infolist()
        package_files = {item.filename: source.read(item.filename) for item in source_items}

    presentation_rels = ET.fromstring(package_files["ppt/_rels/presentation.xml.rels"])
    presentation_root = ET.fromstring(package_files["ppt/presentation.xml"])
    slide_id_list = presentation_root.find(f"{{{presentation_ns}}}sldIdLst")
    content_types_root = ET.fromstring(package_files["[Content_Types].xml"])

    relationship_elements = presentation_rels.findall(f"{{{relationships_ns}}}Relationship")
    relationship_targets = {
        rel.get("Id"): rel.get("Target", "")
        for rel in relationship_elements
    }
    slide10_rel_id = next(
        (
            rel_id for rel_id, target in relationship_targets.items()
            if target.endswith("slides/slide10.xml")
        ),
        None,
    )
    slide10_position = None
    if slide_id_list is not None and slide10_rel_id:
        for position, slide_id in enumerate(list(slide_id_list)):
            if slide_id.get(f"{{{office_relationships_ns}}}id") == slide10_rel_id:
                slide10_position = position
                break

    recommendation_slide_numbers = [7, 8, 9, 10]
    extra_slide_paths = []
    extra_pairs = max(0, recommendation_slide_count - 4)
    existing_slide_numbers = [
        int(match.group(1))
        for path in package_files
        for match in [re.match(r"ppt/slides/slide(\d+)\.xml$", path)]
        if match
    ]
    next_slide_number = max(existing_slide_numbers, default=13) + 1
    numeric_rel_ids = [
        int(match.group(1))
        for rel in relationship_elements
        for match in [re.match(r"rId(\d+)$", rel.get("Id", ""))]
        if match
    ]
    next_rel_number = max(numeric_rel_ids, default=0) + 1
    existing_slide_ids = [
        int(slide_id.get("id", "0") or 0)
        for slide_id in list(slide_id_list or [])
    ]
    next_slide_id = max(existing_slide_ids, default=255) + 1

    for extra_index in range(extra_pairs):
        new_slide_number = next_slide_number + extra_index
        first_recommendation = 9 + extra_index * 2
        second_recommendation = first_recommendation + 1
        new_slide_path = f"ppt/slides/slide{new_slide_number}.xml"
        clone_xml = package_files["ppt/slides/slide10.xml"].decode("utf-8")
        clone_xml = clone_xml.replace("REC_7_", f"REC_{first_recommendation}_")
        clone_xml = clone_xml.replace("REC_8_", f"REC_{second_recommendation}_")
        clone_xml = clone_xml.replace("B10007", f"B1{first_recommendation:04d}")
        clone_xml = clone_xml.replace("B20007", f"B2{first_recommendation:04d}")
        clone_xml = clone_xml.replace("B10008", f"B1{second_recommendation:04d}")
        clone_xml = clone_xml.replace("B20008", f"B2{second_recommendation:04d}")
        clone_xml = clone_xml.replace(">07<", f">{first_recommendation:02d}<")
        clone_xml = clone_xml.replace(">08<", f">{second_recommendation:02d}<")
        package_files[new_slide_path] = clone_xml.encode("utf-8")
        extra_slide_paths.append(new_slide_path)

        source_rels_path = "ppt/slides/_rels/slide10.xml.rels"
        if source_rels_path in package_files:
            new_rels_path = f"ppt/slides/_rels/slide{new_slide_number}.xml.rels"
            package_files[new_rels_path] = package_files[source_rels_path]
            extra_slide_paths.append(new_rels_path)

        new_rel_id = f"rId{next_rel_number + extra_index}"
        ET.SubElement(
            presentation_rels,
            f"{{{relationships_ns}}}Relationship",
            {
                "Id": new_rel_id,
                "Type": "http://schemas.openxmlformats.org/officeDocument/2006/relationships/slide",
                "Target": f"slides/slide{new_slide_number}.xml",
            },
        )
        if slide_id_list is not None:
            new_slide_id = ET.Element(
                f"{{{presentation_ns}}}sldId",
                {
                    "id": str(next_slide_id + extra_index),
                    f"{{{office_relationships_ns}}}id": new_rel_id,
                },
            )
            insert_at = (slide10_position + 1 + extra_index) if slide10_position is not None else len(slide_id_list)
            slide_id_list.insert(insert_at, new_slide_id)
        ET.SubElement(
            content_types_root,
            f"{{{content_types_ns}}}Override",
            {
                "PartName": f"/ppt/slides/slide{new_slide_number}.xml",
                "ContentType": "application/vnd.openxmlformats-officedocument.presentationml.slide+xml",
            },
        )
        recommendation_slide_numbers.append(new_slide_number)

    active_recommendation_slides = set(
        recommendation_slide_numbers[:recommendation_slide_count]
    )
    unused_recommendation_slides = set(recommendation_slide_numbers) - active_recommendation_slides
    partial_recommendation_slide = (
        recommendation_slide_numbers[recommendation_slide_count - 1]
        if recommendation_count % 2 and recommendation_slide_count
        else None
    )

    relationship_targets = {
        rel.get("Id"): rel.get("Target", "")
        for rel in presentation_rels.findall(f"{{{relationships_ns}}}Relationship")
    }
    if slide_id_list is not None:
        for slide_id in list(slide_id_list):
            rel_id = slide_id.get(f"{{{office_relationships_ns}}}id")
            target = relationship_targets.get(rel_id, "")
            match = re.search(r"slides/slide(\d+)\.xml$", target)
            if match and int(match.group(1)) in unused_recommendation_slides:
                slide_id_list.remove(slide_id)
            elif risk_count == 0 and match and int(match.group(1)) == 5:
                slide_id_list.remove(slide_id)
    package_files["ppt/presentation.xml"] = ET.tostring(
        presentation_root,
        encoding="utf-8",
        xml_declaration=True,
    )
    ET.register_namespace("", relationships_ns)
    package_files["ppt/_rels/presentation.xml.rels"] = ET.tostring(
        presentation_rels,
        encoding="utf-8",
        xml_declaration=True,
    )
    ET.register_namespace("", content_types_ns)
    package_files["[Content_Types].xml"] = ET.tostring(
        content_types_root,
        encoding="utf-8",
        xml_declaration=True,
    )

    active_slide_numbers = []
    if slide_id_list is not None:
        for slide_id in slide_id_list:
            rel_id = slide_id.get(f"{{{office_relationships_ns}}}id")
            target = relationship_targets.get(rel_id, "")
            match = re.search(r"slides/slide(\d+)\.xml$", target)
            if match:
                active_slide_numbers.append(int(match.group(1)))

    for visible_number, slide_number in enumerate(active_slide_numbers, start=1):
        slide_path = f"ppt/slides/slide{slide_number}.xml"
        if slide_path not in package_files:
            continue
        slide_root = ET.fromstring(package_files[slide_path])
        shape_tree = slide_root.find(f".//{{{presentation_ns}}}spTree")
        if shape_tree is not None and slide_number == 5:
            visible_risks = min(risk_count, 6)
            risk_rows = (
                (1350000, 2860000),
                (2900000, 4420000),
                (4460000, 6000000),
            )
            for shape in list(shape_tree):
                offset = shape.find(f".//{{{drawing_ns}}}xfrm/{{{drawing_ns}}}off")
                if offset is None:
                    continue
                x_position = int(offset.get("x", "0") or 0)
                y_position = int(offset.get("y", "0") or 0)
                row_index = next(
                    (index for index, (top, bottom) in enumerate(risk_rows) if top <= y_position < bottom),
                    None,
                )
                if row_index is None:
                    continue
                card_index = row_index * 2 + (2 if x_position >= 6000000 else 1)
                if card_index > visible_risks:
                    shape_tree.remove(shape)

        if shape_tree is not None and slide_number == partial_recommendation_slide:
            for shape in list(shape_tree):
                offset = shape.find(f".//{{{drawing_ns}}}xfrm/{{{drawing_ns}}}off")
                if offset is None:
                    continue
                y_position = int(offset.get("y", "0") or 0)
                if 3700000 <= y_position < 6300000:
                    shape_tree.remove(shape)

        if shape_tree is not None:
            for shape in list(shape_tree):
                offset = shape.find(f".//{{{drawing_ns}}}xfrm/{{{drawing_ns}}}off")
                if offset is None:
                    continue
                x_position = int(offset.get("x", "0") or 0)
                y_position = int(offset.get("y", "0") or 0)
                if x_position >= 11000000 and y_position >= 6500000:
                    text_nodes = shape.findall(f".//{{{drawing_ns}}}t")
                    if text_nodes:
                        text_nodes[0].text = str(visible_number)

        package_files[slide_path] = ET.tostring(
            slide_root,
            encoding="utf-8",
            xml_declaration=True,
        )

    with zipfile.ZipFile(output, "w", compression=zipfile.ZIP_DEFLATED) as target:
        written_paths = set()
        for item in source_items:
            content = package_files[item.filename]
            if item.filename.endswith(".xml"):
                xml_text = content.decode("utf-8")
                for key, value in replacements.items():
                    xml_text = xml_text.replace(f"{{{{{key}}}}}", escape(str(value)))
                for index in range(1, 7):
                    xml_text = xml_text.replace(f"A1000{index}", replacements.get(f"RISK_{index}_FILL", "#F4B400").lstrip("#"))
                    xml_text = xml_text.replace(f"A2000{index}", replacements.get(f"RISK_{index}_TEXT", "#1F2937").lstrip("#"))
                for index in range(1, max(8, recommendation_count) + 1):
                    xml_text = xml_text.replace(f"B1{index:04d}", replacements.get(f"REC_{index}_FILL", "#F4B400").lstrip("#"))
                    xml_text = xml_text.replace(f"B2{index:04d}", replacements.get(f"REC_{index}_TEXT", "#1F2937").lstrip("#"))
                for index in range(1, 7):
                    xml_text = xml_text.replace(f"B3000{index}", replacements.get(f"THREAT_{index}_FILL", "#F4B400").lstrip("#"))
                xml_text = xml_text.replace("C10001", replacements.get("IT_SCORE_FILL", "#13877C").lstrip("#"))
                xml_text = xml_text.replace("C20001", replacements.get("IT_SCORE_TEXT", "#FFFFFF").lstrip("#"))
                xml_text = xml_text.replace("C10002", replacements.get("SCORE_FILL", "#13877C").lstrip("#"))
                xml_text = xml_text.replace("C20002", replacements.get("SCORE_TEXT", "#FFFFFF").lstrip("#"))
                xml_text = xml_text.replace(
                    "PCI DSS и GDPR применяются только при наличии соответствующих данных и операций.",
                    escape(str(replacements.get("FRAMEWORK_NOTE", ""))),
                )
                xml_text = xml_text.replace("ЧТО ДАСТ ЭТАП", "РЕЗУЛЬТАТ ЭТАПА")
                content = xml_text.encode("utf-8")
            target.writestr(item, content)
            written_paths.add(item.filename)

        for path, content in package_files.items():
            if path in written_paths:
                continue
            if path.endswith(".xml"):
                xml_text = content.decode("utf-8")
                for key, value in replacements.items():
                    xml_text = xml_text.replace(f"{{{{{key}}}}}", escape(str(value)))
                for index in range(1, max(8, recommendation_count) + 1):
                    xml_text = xml_text.replace(f"B1{index:04d}", replacements.get(f"REC_{index}_FILL", "#F4B400").lstrip("#"))
                    xml_text = xml_text.replace(f"B2{index:04d}", replacements.get(f"REC_{index}_TEXT", "#1F2937").lstrip("#"))
                for index in range(1, 7):
                    xml_text = xml_text.replace(f"B3000{index}", replacements.get(f"THREAT_{index}_FILL", "#F4B400").lstrip("#"))
                xml_text = xml_text.replace(
                    "PCI DSS и GDPR применяются только при наличии соответствующих данных и операций.",
                    escape(str(replacements.get("FRAMEWORK_NOTE", ""))),
                )
                xml_text = xml_text.replace("ЧТО ДАСТ ЭТАП", "РЕЗУЛЬТАТ ЭТАПА")
                content = xml_text.encode("utf-8")
            target.writestr(path, content)

    presentation_bytes = output.getvalue()
    with zipfile.ZipFile(BytesIO(presentation_bytes), "r") as check_zip:
        unresolved = []
        active_slide_paths = {f"ppt/slides/slide{number}.xml" for number in active_slide_numbers}
        for name in check_zip.namelist():
            if not name.startswith("ppt/slides/slide") or not name.endswith(".xml"):
                continue
            if name not in active_slide_paths:
                continue
            slide_xml = check_zip.read(name).decode("utf-8")
            unresolved.extend(re.findall(r"\{\{[A-Z0-9_]+\}\}", slide_xml))
        if unresolved:
            raise ValueError(f"В презентации остались незаполненные поля: {sorted(set(unresolved))}")
    return presentation_bytes


def make_audit_presentation(c_info, results, final_score, it_maturity_score):
    brand_key = presentation_brand_key()
    template_path = os.path.join(
        os.path.dirname(__file__),
        "static",
        f"audit_presentation_{brand_key}.pptx",
    )
    if not os.path.exists(template_path):
        raise FileNotFoundError(f"Не найден шаблон презентации: {template_path}")
    replacements = build_audit_presentation_replacements(
        c_info,
        results,
        final_score,
        it_maturity_score,
    )
    return render_audit_presentation_template(template_path, replacements)


# --- Отчет ---
def make_expert_excel(c_info, results, final_score):
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    output = BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "01 Заключение"
    # =========================
    # CORPORATE STYLES
    # =========================

    dark_blue_fill = PatternFill(start_color="1F3A5F", end_color="1F3A5F", fill_type="solid")
    light_blue_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
    gray_fill = PatternFill(start_color="F3F6F9", end_color="F3F6F9", fill_type="solid")
    critical_fill = PatternFill(start_color="FDE9E7", end_color="FDE9E7", fill_type="solid")
    medium_fill = PatternFill(start_color="FFF4CC", end_color="FFF4CC", fill_type="solid")

    white_font = Font(color="FFFFFF", bold=True)
    section_font = Font(size=14, bold=True, color="1F1F1F")
    normal_font = Font(size=11)

    border = Border(
        left=Side(style='thin', color="D9D9D9"),
        right=Side(style='thin', color="D9D9D9"),
        top=Side(style='thin', color="D9D9D9"),
        bottom=Side(style='thin', color="D9D9D9")
    )
    maturity, maturity_icon = get_maturity_level(final_score)
    domain_scores = calculate_domain_scores(results)
    context = build_context(results, c_info)
    profile_title, profile_text = infrastructure_profile(context)
    it_assets_text, it_focus_text = it_context_summary(results, context)
    report_risks, ai_used = build_report_risk_set(c_info, results, context)
    ai_narrative = st.session_state.get("ai_audit_narrative", {}) if ai_used else {}
    ai_narrative = sanitize_ai_audit_narrative(ai_narrative, results)
    roadmap_items = build_canonical_report_roadmap(
        report_risks,
        results=results,
        context=context,
    )
    st.session_state.last_report_roadmap = list(roadmap_items)
    expert_conclusion = build_expert_conclusion(
        results,
        context,
        final_score,
        domain_scores,
        roadmap_items
    )

    # =========================
    # EXECUTIVE SUMMARY
    # =========================

    ws.merge_cells('A1:D1')
    ws['A1'] = "ЭКСПЕРТНЫЙ ОТЧЕТ ПО ИТ И ИБ"
    ws['A1'].font = Font(bold=True, size=20, color="1F1F1F")

    ws.merge_cells('A3:D3')
    ws['A3'] = "ПАСПОРТ КОМПАНИИ И АУДИТА"
    ws['A3'].font = white_font
    ws['A3'].fill = dark_blue_fill
    ws['A3'].alignment = Alignment(horizontal='center')

    company_rows = [
        ("Компания", c_info.get('Наименование компании', '-'), "Дата аудита", datetime.now().strftime('%d.%m.%Y')),
        ("Сфера", c_info.get('Сфера деятельности', '-'), "Город", c_info.get('Город', '-')),
        ("Сайт", c_info.get('Сайт компании', '-'), "Контакт", c_info.get('ФИО контактного лица', '-')),
        ("Зрелость ИБ", f"{maturity_icon} {maturity} / {final_score}%", "Профиль инфраструктуры", profile_title),
        ("Масштаб", profile_text, "Формат", "Автоматический экспертный аудит"),
        ("ИТ-контекст", it_assets_text, "Фокус эксплуатации", it_focus_text),
    ]

    for row_idx, row_values in enumerate(company_rows, start=4):
        for col_idx, value in enumerate(row_values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            cell.alignment = Alignment(wrap_text=True, vertical='top')
            if col_idx in [1, 3]:
                cell.font = Font(bold=True)
                cell.fill = light_blue_fill
            else:
                cell.fill = gray_fill

    # Executive Summary Block
    ws.merge_cells('A11:D11')
    ws['A11'] = "УПРАВЛЕНЧЕСКОЕ РЕЗЮМЕ"
    ws['A11'].font = Font(bold=True, size=16, color="FFFFFF")
    ws['A11'].fill = dark_blue_fill
    ws['A11'].alignment = Alignment(horizontal='center')

    summary_text = []
    if ai_narrative.get("executive_summary"):
        summary_text = [
            f"• {item}"
            for item in ai_narrative.get("executive_summary", [])[:7]
            if str(item).strip()
        ]
    else:
        summary_text.append(f"• Профиль инфраструктуры: {profile_text}")
        summary_text.append("• Общий вывод: текущий уровень зрелости требует не точечных закупок, а короткой программы стабилизации ИТ и ИБ с владельцами, сроками и метриками контроля.")

        if results.get("NGFW") != "Нет":
            summary_text.append("• Сильная сторона: используется NGFW, его нужно включить в единый контур мониторинга, журналирования и реагирования.")

        if results.get("MFA") == "Нет":
            summary_text.append("• Критичный разрыв: MFA не указана для критичных доступов, поэтому компрометация пароля остается одним из наиболее вероятных сценариев инцидента.")

        if results.get("SIEM") == "Нет":
            if context.get("small_company"):
                summary_text.append("• Для текущего масштаба приоритетнее сбор критичных журналов и регламент реакции, чем тяжелый SIEM-проект")
            else:
                summary_text.append("• Требуется централизованный мониторинг событий по минимальному scope: NGFW, учетные записи, серверы, endpoint, backup, почта и web.")

        if results.get("Резервное копирование") == "Нет":
            summary_text.append("• Не обнаружено централизованное резервное копирование")
        else:
            summary_text.append("• Backup заявлен, но управленчески важно подтвердить RTO/RPO, immutable/offline-копии и результаты тестового восстановления.")

        if results.get("_user_count", 0) > 100:
            summary_text.append("• Масштаб 100+ АРМ требует формализации patch management, endpoint response, мониторинга инфраструктуры и регулярного отчета по остаточным рискам.")

        if not summary_text:
            summary_text.append("• Базовые меры информационной безопасности реализованы")

    ws.merge_cells('A12:D19')
    ws['A12'] = "\n".join(summary_text)
    ws['A12'].alignment = Alignment(wrap_text=True, vertical='top')

    for row in range(12, 20):
        for col in range(1, 5):
            ws.cell(row=row, column=col).fill = gray_fill
            ws.cell(row=row, column=col).border = border

    ws['A12'].font = Font(size=12)

    current_row = 21

    # =========================
    # EXPERT CONCLUSION
    # =========================

    ws.merge_cells(f'A{current_row}:D{current_row}')
    conclusion_header = ws.cell(row=current_row, column=1, value="ЭКСПЕРТНОЕ ЗАКЛЮЧЕНИЕ")
    conclusion_header.font = white_font
    conclusion_header.fill = dark_blue_fill
    conclusion_header.alignment = Alignment(horizontal='center')
    current_row += 1

    conclusion_end_row = current_row + 8
    ws.merge_cells(start_row=current_row, start_column=1, end_row=conclusion_end_row, end_column=4)
    conclusion_cell = ws.cell(row=current_row, column=1, value="\n\n".join(expert_conclusion))
    conclusion_cell.alignment = Alignment(wrap_text=True, vertical='top')
    conclusion_cell.font = Font(size=11)

    for row in range(current_row, conclusion_end_row + 1):
        for col in range(1, 5):
            ws.cell(row=row, column=col).fill = gray_fill
            ws.cell(row=row, column=col).border = border

    observations = ai_narrative.get("audit_observations") or build_audit_observations(results, context, domain_scores, report_risks)
    target_model = build_target_operating_model(results, context)
    management_decisions = ai_narrative.get("management_decisions") or build_management_decisions(results, context)

    current_row = conclusion_end_row + 2

    # =========================
    # AUDITOR OBSERVATIONS
    # =========================
    ws.merge_cells(f'A{current_row}:D{current_row}')
    obs_header = ws.cell(row=current_row, column=1, value="КЛЮЧЕВЫЕ НАБЛЮДЕНИЯ АУДИТОРА")
    obs_header.font = white_font
    obs_header.fill = dark_blue_fill
    obs_header.alignment = Alignment(horizontal='center')
    current_row += 1

    for title, body in observations:
        ws.cell(row=current_row, column=1, value=title).font = Font(bold=True)
        ws.cell(row=current_row, column=1).fill = light_blue_fill
        ws.cell(row=current_row, column=1).border = border
        ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=4)
        body_cell = ws.cell(row=current_row, column=2, value=body)
        body_cell.alignment = Alignment(wrap_text=True, vertical='top')
        body_cell.border = border
        body_cell.fill = gray_fill
        for col in range(3, 5):
            ws.cell(row=current_row, column=col).border = border
            ws.cell(row=current_row, column=col).fill = gray_fill
        current_row += 1

    current_row += 1

    # =========================
    # TARGET OPERATING MODEL
    # =========================
    ws.merge_cells(f'A{current_row}:D{current_row}')
    tom_header = ws.cell(row=current_row, column=1, value="ЦЕЛЕВАЯ МОДЕЛЬ УЛУЧШЕНИЙ")
    tom_header.font = white_font
    tom_header.fill = dark_blue_fill
    tom_header.alignment = Alignment(horizontal='center')
    current_row += 1

    for title, body in target_model:
        ws.cell(row=current_row, column=1, value=title).font = Font(bold=True)
        ws.cell(row=current_row, column=1).fill = light_blue_fill
        ws.cell(row=current_row, column=1).border = border
        ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=4)
        body_cell = ws.cell(row=current_row, column=2, value=body)
        body_cell.alignment = Alignment(wrap_text=True, vertical='top')
        body_cell.border = border
        body_cell.fill = gray_fill
        for col in range(3, 5):
            ws.cell(row=current_row, column=col).border = border
            ws.cell(row=current_row, column=col).fill = gray_fill
        current_row += 1

    current_row += 1

    # =========================
    # MANAGEMENT DECISIONS
    # =========================
    ws.merge_cells(f'A{current_row}:D{current_row}')
    dec_header = ws.cell(row=current_row, column=1, value="ПЕРВЫЕ УПРАВЛЕНЧЕСКИЕ РЕШЕНИЯ")
    dec_header.font = white_font
    dec_header.fill = dark_blue_fill
    dec_header.alignment = Alignment(horizontal='center')
    current_row += 1

    for idx, decision in enumerate(management_decisions, start=1):
        ws.cell(row=current_row, column=1, value=idx).font = Font(bold=True)
        ws.cell(row=current_row, column=1).alignment = Alignment(horizontal='center')
        ws.cell(row=current_row, column=1).fill = light_blue_fill
        ws.cell(row=current_row, column=1).border = border
        ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=4)
        decision_cell = ws.cell(row=current_row, column=2, value=decision)
        decision_cell.alignment = Alignment(wrap_text=True, vertical='top')
        decision_cell.fill = gray_fill
        decision_cell.border = border
        for col in range(3, 5):
            ws.cell(row=current_row, column=col).border = border
            ws.cell(row=current_row, column=col).fill = gray_fill
        current_row += 1

    current_row += 2

    # =========================
    # 7 / 30 / 90 DAY ACTION PLAN
    # =========================
    ws.merge_cells(f'A{current_row}:D{current_row}')
    plan_header = ws.cell(row=current_row, column=1, value="ПЛАН ДЕЙСТВИЙ: 7 / 30 / 90 ДНЕЙ")
    plan_header.font = white_font
    plan_header.fill = dark_blue_fill
    plan_header.alignment = Alignment(horizontal='center')
    current_row += 1

    action_plan = [
        (
            "Первые 7 дней",
            "Подтвердить владельцев рисков, критичные сервисы, RTO/RPO, администраторские доступы и текущий порядок обновлений.",
            "Появляется единый список приоритетов и ответственных, без запуска тяжелого проекта."
        ),
        (
            "До 30 дней",
            "; ".join(item["action"] for item in roadmap_items if item["phase"] == "0-30 дней") or
            "Закрыть быстрые меры: MFA для критичных доступов, проверка backup, инвентаризация активов и критичных уязвимостей.",
            "Снижаются наиболее вероятные сценарии инцидентов и появляется измеримый baseline."
        ),
        (
            "До 90 дней",
            "; ".join(item["action"] for item in roadmap_items if item["phase"] in {"31-60 дней", "61-90 дней"})[:420] or
            "Перейти от точечных мер к управляемой модели мониторинга, обновлений, восстановления и реагирования.",
            "Формируется программа улучшений с метриками контроля и понятным бюджетированием."
        ),
    ]

    for period, action, result in action_plan:
        ws.cell(row=current_row, column=1, value=period).font = Font(bold=True)
        ws.cell(row=current_row, column=1).fill = light_blue_fill
        ws.cell(row=current_row, column=1).border = border
        ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=3)
        action_cell = ws.cell(row=current_row, column=2, value=action)
        action_cell.alignment = Alignment(wrap_text=True, vertical='top')
        action_cell.fill = gray_fill
        action_cell.border = border
        result_cell = ws.cell(row=current_row, column=4, value=result)
        result_cell.alignment = Alignment(wrap_text=True, vertical='top')
        result_cell.fill = gray_fill
        result_cell.border = border
        for col in range(3, 4):
            ws.cell(row=current_row, column=col).fill = gray_fill
            ws.cell(row=current_row, column=col).border = border
        current_row += 1

    current_row += 2

    # =========================
    # TOP RISKS OVERVIEW
    # =========================
    ws.merge_cells(f'A{current_row}:D{current_row}')

    risk_header = ws.cell(
        row=current_row,
        column=1,
        value="КЛЮЧЕВЫЕ РИСКИ ИТ И ИБ"
    )

    risk_header.font = white_font
    risk_header.fill = dark_blue_fill
    risk_header.alignment = Alignment(horizontal='center')

    current_row += 1

    headers = ["#", "Риск", "Критичность", "Влияние на бизнес"]

    for col_num, header in enumerate(headers, 1):

        cell = ws.cell(
            row=current_row,
            column=col_num,
            value=header
        )

        cell.font = white_font
        cell.fill = dark_blue_fill
        cell.border = border

    current_row += 1

    for idx, risk in enumerate(report_risks[:5], start=1):

        level = risk.get("level", "MEDIUM")
        impact = risk.get("impact", "-")

        ws.cell(row=current_row, column=1, value=idx).border = border
        ws.cell(row=current_row, column=2, value=risk.get("risk", "-")).border = border
        ws.cell(row=current_row, column=3, value=risk_level_label(level)).border = border
        ws.cell(row=current_row, column=4, value=impact).border = border

        if "CRITICAL" in str(level).upper():
            fill = critical_fill
        else:
            fill = medium_fill

        for c in range(1, 5):
            ws.cell(row=current_row, column=c).fill = fill

        current_row += 1

    current_row += 2

    # =========================
    # Оценка доменов безопасности
    # =========================

    ws.merge_cells(f'A{current_row}:D{current_row}')
    dom_cell = ws.cell(row=current_row, column=1, value="ОЦЕНКА ДОМЕНОВ БЕЗОПАСНОСТИ")
    dom_cell.font = white_font
    dom_cell.fill = dark_blue_fill
    dom_cell.alignment = Alignment(horizontal='center')

    current_row += 1

    headers = ["Домен безопасности", "Оценка", "Статус"]

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=current_row, column=col_num, value=header)
        cell.font = white_font
        cell.fill = dark_blue_fill
        cell.border = border

    current_row += 1

    for domain, score in domain_scores.items():

        if score >= 80:
            status = "🟢 Сильный"
            fill = light_blue_fill

        elif score >= 50:
            status = "🟠 Средний"
            fill = medium_fill

        else:
            status = "🔴 Слабый"
            fill = critical_fill

        ws.cell(row=current_row, column=1, value=domain).border = border
        ws.cell(row=current_row, column=2, value=f"{score}%").border = border
        ws.cell(row=current_row, column=3, value=status).border = border

        for c in range(1, 4):
            ws.cell(row=current_row, column=c).fill = fill

        current_row += 1

    current_row += 2

    curr_row = current_row
    ws.merge_cells(f'A{curr_row}:B{curr_row}')
    ws.cell(row=curr_row, column=1, value="ВЫЯВЛЕННЫЕ РИСКИ И РЕКОМЕНДАЦИИ").font = Font(bold=True, size=14)
    curr_row += 1

    ai_data = report_risks

    if ai_data:
        for item in ai_data:

            # Уровень и Название
            lvl = item.get('level', 'СРЕДНИЙ')

            ws.merge_cells(f'A{curr_row}:B{curr_row}')

            cell = ws.cell(
                row=curr_row,
                column=1,
                value=f"[{risk_level_label(lvl)}] {item.get('risk', 'Риск')}"
            )

            cell.font = Font(bold=True)

            if "КРИТ" in str(lvl).upper() or "CRITICAL" in str(lvl).upper():
                cell.fill = critical_fill
            else:
                cell.fill = medium_fill

            curr_row += 1

            # Описание, Влияние, Рекомендация
            fields = [
                ("Описание", item.get('description', '-')),
                ("Влияние", item.get('impact', '-')),
                ("Рекомендация", item.get('recommendation', '-')),
                (
                    "Регуляторы",
                    ", ".join(item.get('regulators', []))
                    if isinstance(item.get('regulators'), list)
                    else "-"
                ),
                (
                    "Решения",
                    solution_categories_for_report_item(item)
                ),
                (
                    "Производители",
                    portfolio_manufacturers_for_report_item(item)
                )
            ]

            for f_label, f_val in fields:

                ws.cell(row=curr_row, column=1, value=f_label).font = Font(italic=True)
                ws.merge_cells(start_row=curr_row, start_column=2, end_row=curr_row, end_column=4)

                ws.cell(
                    row=curr_row,
                    column=2,
                    value=f_val
                ).alignment = Alignment(wrap_text=True)

                for col in range(1, 5):
                    ws.cell(row=curr_row, column=col).border = border

                curr_row += 1

            curr_row += 1

    curr_row += 1

    # =========================
    # ROADMAP PREVIEW AT THE END
    # =========================

    ws.merge_cells(f'A{curr_row}:D{curr_row}')
    roadmap_preview_header = ws.cell(
        row=curr_row,
        column=1,
        value="ROADMAP НА 90 ДНЕЙ: ПРИОРИТЕТНЫЕ ШАГИ"
    )
    roadmap_preview_header.font = white_font
    roadmap_preview_header.fill = dark_blue_fill
    roadmap_preview_header.alignment = Alignment(horizontal='center')
    curr_row += 1

    ws.merge_cells(f'A{curr_row}:D{curr_row}')
    roadmap_hint = ws.cell(
        row=curr_row,
        column=1,
        value="Дорожная карта размещена в конце заключения: сначала оценка и риски, затем практический план действий."
    )
    roadmap_hint.alignment = Alignment(wrap_text=True)
    roadmap_hint.fill = light_blue_fill
    roadmap_hint.border = border
    curr_row += 1

    preview_headers = ["Период", "Приоритет", "Домен", "Действие"]
    for col_num, header in enumerate(preview_headers, 1):
        cell = ws.cell(row=curr_row, column=col_num, value=header)
        cell.font = white_font
        cell.fill = dark_blue_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
    curr_row += 1

    for item in roadmap_items[:8]:
        ws.cell(row=curr_row, column=1, value=item["phase"]).border = border
        ws.cell(row=curr_row, column=2, value=item["priority"]).border = border
        ws.cell(row=curr_row, column=3, value=item["domain"]).border = border
        action_cell = ws.cell(row=curr_row, column=4, value=item["action"])
        action_cell.border = border
        action_cell.alignment = Alignment(wrap_text=True, vertical='top')
        if item["priority"] == "P1":
            fill = critical_fill
        elif item["priority"] == "P2":
            fill = medium_fill
        else:
            fill = light_blue_fill
        for col in range(1, 5):
            ws.cell(row=curr_row, column=col).fill = fill
        curr_row += 1

    ws.column_dimensions['A'].width = 38
    ws.column_dimensions['B'].width = 95
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20

    # =========================
    # ROADMAP SHEET
    # =========================

    roadmap_ws = wb.create_sheet("02 Roadmap 90 дней")
    roadmap_ws.merge_cells('A1:G1')
    roadmap_ws['A1'] = "ДОРОЖНАЯ КАРТА УЛУЧШЕНИЙ НА 90 ДНЕЙ"
    roadmap_ws['A1'].font = Font(bold=True, size=18, color="FFFFFF")
    roadmap_ws['A1'].fill = dark_blue_fill
    roadmap_ws['A1'].alignment = Alignment(horizontal='center')

    roadmap_ws.merge_cells('A3:G3')
    roadmap_ws['A3'] = f"{profile_title}. {profile_text}"
    roadmap_ws['A3'].alignment = Alignment(wrap_text=True, vertical='top')
    roadmap_ws['A3'].fill = gray_fill

    roadmap_headers = [
        "Период",
        "Приоритет",
        "Домен",
        "Что сделать",
        "Почему это важно",
        "Ответственный",
        "Сложность",
    ]

    roadmap_row = 5
    for col_num, header in enumerate(roadmap_headers, 1):
        cell = roadmap_ws.cell(row=roadmap_row, column=col_num, value=header)
        cell.font = white_font
        cell.fill = dark_blue_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    roadmap_row += 1

    priority_fills = {
        "P1": critical_fill,
        "P2": medium_fill,
        "P3": light_blue_fill,
    }

    for item in roadmap_items:
        values = [
            item["phase"],
            item["priority"],
            item["domain"],
            item["action"],
            item["rationale"],
            item["owner"],
            item["effort"],
        ]

        fill = priority_fills.get(item["priority"], gray_fill)
        for col_num, value in enumerate(values, 1):
            cell = roadmap_ws.cell(row=roadmap_row, column=col_num, value=value)
            cell.border = border
            cell.alignment = Alignment(wrap_text=True, vertical='top')
            if col_num <= 2:
                cell.fill = fill

        roadmap_row += 1

    roadmap_ws.freeze_panes = "A6"
    roadmap_ws.auto_filter.ref = f"A5:G{max(roadmap_row - 1, 5)}"

    roadmap_row += 2
    roadmap_ws.merge_cells(start_row=roadmap_row, start_column=1, end_row=roadmap_row, end_column=7)
    note_cell = roadmap_ws.cell(row=roadmap_row, column=1, value=(
        "Примечание: roadmap построен с учетом размера инфраструктуры, наличия серверов, "
        "публичных сервисов, бизнес-систем и текущих средств ИБ. Для малой инфраструктуры "
        "приоритет отдается быстрым управляемым мерам, а не тяжелым enterprise-платформам."
    ))
    note_cell.alignment = Alignment(wrap_text=True, vertical='top')
    note_cell.fill = gray_fill
    note_cell.border = border

    roadmap_ws.column_dimensions['A'].width = 16
    roadmap_ws.column_dimensions['B'].width = 12
    roadmap_ws.column_dimensions['C'].width = 22
    roadmap_ws.column_dimensions['D'].width = 48
    roadmap_ws.column_dimensions['E'].width = 58
    roadmap_ws.column_dimensions['F'].width = 18
    roadmap_ws.column_dimensions['G'].width = 14

    wb.save(output)
    return output.getvalue()


def make_internal_sales_excel(c_info, results, final_score, client_report_bytes=None):
    from io import BytesIO
    from openpyxl import load_workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    output = BytesIO()
    client_report_bytes = client_report_bytes or make_expert_excel(c_info, results, final_score)
    wb = load_workbook(BytesIO(client_report_bytes))
    ws = wb.create_sheet("03 Что продавать")

    dark_fill = PatternFill(start_color="0F766E", end_color="0F766E", fill_type="solid")
    light_blue_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
    gray_fill = PatternFill(start_color="F3F6F9", end_color="F3F6F9", fill_type="solid")
    critical_fill = PatternFill(start_color="FDE9E7", end_color="FDE9E7", fill_type="solid")
    medium_fill = PatternFill(start_color="FFF4CC", end_color="FFF4CC", fill_type="solid")
    white_font = Font(color="FFFFFF", bold=True)
    border = Border(
        left=Side(style='thin', color="D9D9D9"),
        right=Side(style='thin', color="D9D9D9"),
        top=Side(style='thin', color="D9D9D9"),
        bottom=Side(style='thin', color="D9D9D9")
    )

    context = build_context(results, c_info)
    domain_scores = calculate_domain_scores(results)
    rule_risks = generate_rule_based_risks(results, context)
    roadmap_items = build_contextual_roadmap(results, context, domain_scores, rule_risks)
    risk_sources = st.session_state.get("last_report_risk_sources", [])
    sales_opportunities = (
        build_ai_first_sales_opportunities(risk_sources, results, context)
        or build_sales_opportunities(results, context, roadmap_items)
    )
    sales_opportunities = ensure_sales_playbook_priorities(
        sales_opportunities,
        results,
        context
    )
    sales_pack = build_sales_conversation_pack(
        c_info,
        results,
        context,
        roadmap_items,
        sales_opportunities
    )

    def style_sales_header(sheet, title, end_col):
        sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=end_col)
        sheet.cell(row=1, column=1, value=title)
        sheet.cell(row=1, column=1).font = Font(bold=True, size=18, color="FFFFFF")
        sheet.cell(row=1, column=1).fill = dark_fill
        sheet.cell(row=1, column=1).alignment = Alignment(horizontal='center')

    def write_table(sheet, start_row, headers, rows, widths):
        for col_num, header in enumerate(headers, 1):
            cell = sheet.cell(row=start_row, column=col_num, value=header)
            cell.font = white_font
            cell.fill = dark_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        row_idx = start_row + 1
        for row_values in rows:
            for col_num, value in enumerate(row_values, 1):
                cell = sheet.cell(row=row_idx, column=col_num, value=value)
                cell.border = border
                cell.alignment = Alignment(wrap_text=True, vertical='top')
                if row_idx % 2 == 0:
                    cell.fill = gray_fill
            row_idx += 1
        for col_letter, width in widths.items():
            sheet.column_dimensions[col_letter].width = width
        sheet.freeze_panes = f"A{start_row + 1}"
        sheet.auto_filter.ref = f"A{start_row}:{chr(64 + len(headers))}{max(row_idx - 1, start_row)}"
        return row_idx

    ws.merge_cells('A1:J1')
    ws['A1'] = "ВНУТРЕННИЙ SALES PLAYBOOK ПО ИТОГАМ АУДИТА"
    ws['A1'].font = Font(bold=True, size=18, color="FFFFFF")
    ws['A1'].fill = dark_fill
    ws['A1'].alignment = Alignment(horizontal='center')

    ws.merge_cells('A3:J3')
    ws['A3'] = (
        f"Компания: {c_info.get('Наименование компании', '-')} | "
        f"Город: {c_info.get('Город', '-')} | "
        f"Сфера: {c_info.get('Сфера деятельности', '-')} | "
        f"Зрелость ИБ: {final_score}%"
    )
    ws['A3'].alignment = Alignment(wrap_text=True)
    ws['A3'].fill = gray_fill
    ws['A3'].border = border

    info_rows = [
        ("Компания", c_info.get("Наименование компании", "-")),
        ("Город", c_info.get("Город", "-")),
        ("Сфера деятельности", c_info.get("Сфера деятельности", "-")),
        ("Сайт", c_info.get("Сайт компании", "-")),
        ("Email", c_info.get("Email", "-")),
        ("Контактное лицо", c_info.get("ФИО контактного лица", "-")),
        ("Должность", c_info.get("Должность", "-")),
        ("Телефон", c_info.get("Контактный телефон", "-")),
        ("Зрелость ИБ", f"{final_score}%"),
    ]

    ws.merge_cells('A5:J5')
    ws['A5'] = "ИНФОРМАЦИЯ О КОМПАНИИ"
    ws['A5'].font = Font(bold=True, color="FFFFFF")
    ws['A5'].fill = dark_fill
    ws['A5'].alignment = Alignment(horizontal='center')

    info_row = 6
    for label, value in info_rows:
        ws.cell(row=info_row, column=1, value=label).font = Font(bold=True)
        ws.cell(row=info_row, column=1).fill = gray_fill
        ws.cell(row=info_row, column=1).border = border
        ws.merge_cells(start_row=info_row, start_column=2, end_row=info_row, end_column=10)
        value_cell = ws.cell(row=info_row, column=2, value=value)
        value_cell.border = border
        value_cell.alignment = Alignment(wrap_text=True, vertical='top')
        for col_num in range(3, 11):
            ws.cell(row=info_row, column=col_num).border = border
        info_row += 1

    headers = [
        "Приоритет",
        "Гипотеза возможности",
        "Факт / что подтвердить",
        "Что предложить",
        "Ценность для клиента",
        "Кого подключить",
        "Производители из портфеля",
        "Дистрибьюторы",
        "Следующий шаг сейла",
        "Источник",
    ]
    nav_row = info_row + 1
    ws.merge_cells(start_row=nav_row, start_column=1, end_row=nav_row, end_column=10)
    nav_cell = ws.cell(
        row=nav_row,
        column=1,
        value=(
            "Навигация: основная таблица ниже отфильтровывается по приоритету, источнику и решениям; "
            "первые три колонки и заголовок закреплены. Формулировки в колонке «Факт / что подтвердить» "
            "нужно валидировать на встрече до подготовки спецификации."
        )
    )
    nav_cell.fill = light_blue_fill
    nav_cell.border = border
    nav_cell.alignment = Alignment(wrap_text=True, vertical='top')

    header_row = nav_row + 1
    row = header_row
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_num, value=header)
        cell.font = white_font
        cell.fill = dark_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    row += 1
    first_data_row = row
    priority_fills = {"P1": critical_fill, "P2": medium_fill, "P3": gray_fill}
    for item in sales_opportunities:
        guidance = sales_account_guidance(item)
        values = [
            item["priority"],
            item["problem"],
            item["trigger"],
            item["offer"],
            guidance["business_value"],
            guidance["stakeholders"],
            item["vendors"],
            verified_distributors_for_vendors(item["vendors"]),
            item["next_step"],
            item.get("source", "Базовые правила"),
        ]
        fill = priority_fills.get(item["priority"], gray_fill)
        for col_num, value in enumerate(values, 1):
            cell = ws.cell(row=row, column=col_num, value=value)
            cell.border = border
            cell.alignment = Alignment(wrap_text=True, vertical='top')
            if col_num == 1:
                cell.fill = fill
        row += 1

    if not sales_opportunities:
        ws.merge_cells(start_row=first_data_row, start_column=1, end_row=first_data_row, end_column=10)
        cell = ws.cell(
            row=first_data_row,
            column=1,
            value="Явных продуктовых триггеров по анкете мало. Нужно назначить короткий экспертный созвон и уточнить детали инфраструктуры."
        )
        cell.alignment = Alignment(wrap_text=True)
        cell.fill = gray_fill
        cell.border = border
        row = first_data_row + 1

    ws.freeze_panes = f"D{first_data_row}"
    ws.auto_filter.ref = f"A{header_row}:J{max(row - 1, header_row)}"
    ws.sheet_view.zoomScale = 75
    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 34
    ws.column_dimensions['C'].width = 48
    ws.column_dimensions['D'].width = 48
    ws.column_dimensions['E'].width = 42
    ws.column_dimensions['F'].width = 40
    ws.column_dimensions['G'].width = 34
    ws.column_dimensions['H'].width = 32
    ws.column_dimensions['I'].width = 48
    ws.column_dimensions['J'].width = 16

    source_row = row + 2
    ws.merge_cells(start_row=source_row, start_column=1, end_row=source_row, end_column=10)
    ws.cell(row=source_row, column=1, value="ИСТОЧНИКИ РИСКОВ И РЕКОМЕНДАЦИЙ").font = white_font
    ws.cell(row=source_row, column=1).fill = dark_fill
    ws.cell(row=source_row, column=1).alignment = Alignment(horizontal='center')
    source_row += 1

    source_headers = ["Критичность", "Риск", "Рекомендация", "Источник"]
    for col_num, header in enumerate(source_headers, 1):
        cell = ws.cell(row=source_row, column=col_num, value=header)
        cell.font = white_font
        cell.fill = dark_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    source_row += 1

    risk_sources = st.session_state.get("last_report_risk_sources", [])
    if risk_sources:
        for item in risk_sources:
            values = [
                item.get("level", "-"),
                item.get("risk", "-"),
                item.get("recommendation", "-"),
                item.get("source", "-"),
            ]
            for col_num, value in enumerate(values, 1):
                cell = ws.cell(row=source_row, column=col_num, value=value)
                cell.border = border
                cell.alignment = Alignment(wrap_text=True, vertical='top')
                if item.get("source") == "ИИ":
                    cell.fill = light_blue_fill
                elif source_row % 2 == 0:
                    cell.fill = gray_fill
            source_row += 1
    else:
        ws.merge_cells(start_row=source_row, start_column=1, end_row=source_row, end_column=4)
        cell = ws.cell(row=source_row, column=1, value="Источники рисков не зафиксированы для текущей генерации.")
        cell.fill = gray_fill
        cell.border = border
        cell.alignment = Alignment(wrap_text=True)

    ws.sheet_view.showGridLines = False

    pains_ws = wb.create_sheet("04 Боли и гипотезы")
    style_sales_header(pains_ws, "КАРТА ГИПОТЕЗ И КВАЛИФИКАЦИИ", 7)
    pains_rows = []
    for item in sales_pack["pains"]:
        guidance = sales_account_guidance({
            "problem": item["pain"],
            "trigger": item["evidence"],
            "offer": item["commercial_angle"],
        })
        pains_rows.append([
            item["priority"],
            item["pain"],
            item["evidence"],
            guidance["business_value"],
            guidance["stakeholders"],
            item["discovery_question"],
            guidance["meeting_goal"],
        ])
    write_table(
        pains_ws,
        3,
        ["Приоритет", "Гипотеза", "Факт из анкеты", "Ожидаемая ценность", "Карта участников", "Вопрос для подтверждения", "Цель первой встречи"],
        pains_rows,
        {"A": 12, "B": 40, "C": 42, "D": 42, "E": 40, "F": 52, "G": 48}
    )

    script_ws = wb.create_sheet("05 Сценарий звонка")
    style_sales_header(script_ws, "СЦЕНАРИЙ ПЕРВОГО ЗВОНКА", 4)
    script_rows = [
        [idx, item["stage"], item["goal"], item["talk_track"]]
        for idx, item in enumerate(sales_pack["call_script"], start=1)
    ]
    write_table(
        script_ws,
        3,
        ["#", "Этап", "Цель этапа", "Что сказать / сделать"],
        script_rows,
        {"A": 8, "B": 26, "C": 38, "D": 100}
    )

    questions_ws = wb.create_sheet("06 Вопросы")
    style_sales_header(questions_ws, "ВОПРОСЫ ДЛЯ КВАЛИФИКАЦИИ ВОЗМОЖНОСТИ", 3)
    write_table(
        questions_ws,
        3,
        ["Тема", "Вопрос", "Зачем спрашиваем"],
        [[topic, question, purpose] for topic, question, purpose in sales_pack["questions"]],
        {"A": 28, "B": 90, "C": 55}
    )

    objections_ws = wb.create_sheet("07 Возражения")
    style_sales_header(objections_ws, "ТИПОВЫЕ ВОЗРАЖЕНИЯ И ОТРАБОТКА", 3)
    write_table(
        objections_ws,
        3,
        ["Возражение", "Как отвечать", "Вопрос после ответа"],
        [[objection, answer, follow_up] for objection, answer, follow_up in sales_pack["objections"]],
        {"A": 40, "B": 80, "C": 60}
    )

    next_ws = wb.create_sheet("08 Следующие шаги")
    style_sales_header(next_ws, "ПЛАН РАЗВИТИЯ ВОЗМОЖНОСТИ", 7)
    next_rows = [
        [
            item["priority"],
            item["offer"],
            item["step"],
            item["stakeholders"],
            item["meeting_goal"],
            item["seller_artifact"],
            item["success_criteria"],
        ]
        for item in sales_pack["next_steps"]
    ]
    write_table(
        next_ws,
        3,
        ["Приоритет", "Возможность", "Действие сейла", "Участники клиента", "Цель встречи", "Что подготовить сейлу", "Критерий перехода"],
        next_rows,
        {"A": 12, "B": 42, "C": 55, "D": 40, "E": 48, "F": 52, "G": 60}
    )

    answers_ws = wb.create_sheet("09 Заполненная анкета")
    answers_ws.merge_cells('A1:B1')
    answers_ws['A1'] = "ЗАПОЛНЕННАЯ АНКЕТА / ДАННЫЕ ДЛЯ СЕЙЛА И ПРЕСЕЙЛА"
    answers_ws['A1'].font = Font(bold=True, size=18, color="FFFFFF")
    answers_ws['A1'].fill = dark_fill
    answers_ws['A1'].alignment = Alignment(horizontal='center')

    answers_ws.cell(row=3, column=1, value="Поле").font = white_font
    answers_ws.cell(row=3, column=2, value="Значение").font = white_font
    for col_num in range(1, 3):
        cell = answers_ws.cell(row=3, column=col_num)
        cell.fill = dark_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center')

    answer_row = 4
    for key, value in results.items():
        if str(key).startswith("_"):
            continue
        answers_ws.cell(row=answer_row, column=1, value=key).border = border
        value_cell = answers_ws.cell(row=answer_row, column=2, value=str(value))
        value_cell.border = border
        value_cell.alignment = Alignment(wrap_text=True, vertical='top')
        if answer_row % 2 == 0:
            answers_ws.cell(row=answer_row, column=1).fill = gray_fill
            answers_ws.cell(row=answer_row, column=2).fill = gray_fill
        answer_row += 1

    answers_ws.freeze_panes = "A4"
    answers_ws.auto_filter.ref = f"A3:B{max(answer_row - 1, 3)}"
    answers_ws.column_dimensions['A'].width = 42
    answers_ws.column_dimensions['B'].width = 100

    wb.save(output)
    return output.getvalue(), sales_opportunities


def build_report_results(
    data,
    main_speed,
    back_speed,
    total_arm,
    ap_cnt,
    selected_routing,
    ngfw_vendor,
    phys_count,
    virt_count,
    v_n_b
):
    results = data.copy()
    results.update({
        "Интернет канал (осн)": f"{main_speed} Mbit/s",
        "Резервный канал": f"{back_speed} Mbit/s",
        "_main_speed": main_speed,
        "_back_speed": back_speed,
        "_user_count": total_arm,
        "WiFi Точки": ap_cnt,
        "WiFi Контроллер": data.get('Wi-Fi Контроллер', "Нет"),
        "Маршрутизация": ", ".join(selected_routing) if selected_routing else "Нет",
        "NGFW": ngfw_vendor if ngfw_vendor else "Нет",
        "Серверы (физ)": phys_count,
        "Серверы (вирт)": virt_count,
        "Резервное копирование": v_n_b if v_n_b else "Нет",
    })

    results["MFA"] = results.get("Блок 2. MFA", "Нет")
    results["SIEM"] = results.get("Блок 2. SIEM", "Нет")
    results["WAF"] = results.get("Блок 2. WAF", "Нет")
    results["Anti-DDoS"] = results.get("Блок 2. Anti-DDoS", "Нет")
    results["EPP"] = results.get("Блок 2. EPP", "Нет")
    results["Антивирус"] = results.get("Блок 2. EPP", "Нет")
    results["EDR"] = results.get("Блок 2. EDR", "Нет")
    results["XDR"] = results.get("Блок 2. XDR", "Нет")
    results["MDR"] = results.get("Блок 2. MDR", "Нет")
    results["DLP"] = results.get("Блок 2. DLP", "Нет")
    results["Mail Security"] = results.get("Блок 2. Mail Security", "Нет")
    results["CASB"] = results.get("Блок 2. CASB", "Нет")
    results["IDS/IPS"] = results.get("Блок 2. IDS/IPS", "Нет")
    results["NAC"] = results.get("Блок 2. NAC", "Нет")
    results["ZTNA"] = results.get("Блок 2. ZTNA", "Нет")
    results["SAST"] = results.get("Блок 2. SAST", "Нет")
    results["DAST"] = results.get("Блок 2. DAST", "Нет")
    results["IAM"] = results.get("Блок 2. IAM", "Нет")
    results["PAM"] = results.get("Блок 2. PAM", "Нет")
    results["SOAR"] = results.get("Блок 2. SOAR", "Нет")
    results["NAD"] = results.get("Блок 2. NAD", "Нет")
    results["Patch Management"] = results.get("Блок 2. Patch Management", "Нет")
    for control in ("WAF", "EDR", "MFA"):
        if not is_enabled(results.get(control)) and control_confirmed_in_results(results, control):
            results[control] = "Есть (подтверждено в примечании анкеты)"
    results["Виртуализация"] = "Да" if virt_count > 0 else "Нет"
    results["СХД"] = "Да" if any(str(key).startswith("1.4.") for key in results) else "Нет"
    results["Резервный канал"] = f"{back_speed} Mbit/s" if back_speed > 0 else "Нет"
    return results


def _status_label(score_value):
    if score_value >= 80:
        return "Сильный"
    if score_value >= 50:
        return "Средний"
    return "Слабый"


def _quick_win_signals(results):
    signals = []

    if results.get("MFA") == "Нет":
        signals.append(("critical", "MFA", "Внедрение MFA обычно быстрее всего снижает риск компрометации учетных записей."))

    if results.get("SIEM") == "Нет":
        signals.append(("warn", "SIEM / SOC", "Нет централизованного мониторинга: инциденты сложнее обнаруживать и расследовать."))

    if results.get("EDR") == "Нет":
        signals.append(("warn", "Endpoint", "Без EDR/XDR защита рабочих мест остается слабее против сложных атак."))

    if results.get("Резервное копирование") == "Нет":
        signals.append(("critical", "Backup", "Не указан backup-контур: это критично для устойчивости к ransomware и сбоям."))

    if results.get("Patch Management") == "Нет":
        signals.append(("warn", "Patch Management", "Без централизованных обновлений растет риск эксплуатации известных уязвимостей."))

    legacy_arm = results.get("ОС АРМ (Windows XP/Vista/7/8)", 0)
    legacy_srv = results.get("ОС Сервера (Windows Server 2008/2012 R2)", 0)
    if legacy_arm or legacy_srv:
        signals.append(("critical", "Legacy OS", "Обнаружены устаревшие ОС: нужен план миграции или изоляции."))

    return signals[:5]


def section_status(enabled, complete):
    if not enabled:
        return "disabled"
    return "complete" if complete else "missing"


def unpack_section_status(section_item):
    if len(section_item) == 3:
        return section_item

    step, status = section_item
    return step, status, ""


def calculate_audit_readiness(client_info, validation_errors, section_statuses):
    required_fields = [
        client_info.get('Город'),
        client_info.get('Наименование компании'),
        client_info.get('Сфера деятельности'),
        client_info.get('Сайт компании'),
        client_info.get('Email'),
        client_info.get('ФИО контактного лица'),
        client_info.get('Должность'),
        client_info.get('Контактный телефон'),
    ]

    filled_required = sum(1 for value in required_fields if value)
    required_readiness = (filled_required / len(required_fields)) * 40

    active_sections = []
    for section_item in section_statuses:
        _, status, _ = unpack_section_status(section_item)
        if status != "disabled":
            active_sections.append(status)
    completed_sections = sum(
        1 for status in active_sections
        if status == "complete"
    )
    section_readiness = (
        (completed_sections / len(active_sections)) * 45
        if active_sections
        else 0
    )
    validation_readiness = 15 if not validation_errors else 0

    return int(round(
        min(100, required_readiness + section_readiness + validation_readiness)
    ))


def render_audit_cockpit(client_info, results, validation_errors, final_score, it_maturity_score, section_statuses):
    readiness = calculate_audit_readiness(
        client_info,
        validation_errors,
        section_statuses
    )
    domain_scores = calculate_domain_scores(results)
    risk_signals = _quick_win_signals(results)
    maturity, maturity_icon = get_maturity_level(final_score)
    it_maturity, it_maturity_icon = get_maturity_level(it_maturity_score)

    st.sidebar.markdown("### Навигатор аудита")
    st.sidebar.progress(readiness, text=f"Готовность анкеты: {readiness}%")
    st.sidebar.metric("Зрелость ИБ", f"{final_score}%")
    st.sidebar.caption(f"{maturity_icon} {maturity}")
    st.sidebar.metric("Зрелость ИТ", f"{it_maturity_score}%")
    st.sidebar.caption(f"{it_maturity_icon} {it_maturity}")
    st.sidebar.metric("Ошибки заполнения", len(validation_errors))

    st.sidebar.markdown("#### Разделы")
    status_styles = {
        "complete": ("green", "заполнен"),
        "missing": ("red", "не заполнен"),
        "disabled": ("gray", "не включен"),
    }

    for section_item in section_statuses:
        step, status, anchor_id = unpack_section_status(section_item)
        dot_class, title = status_styles.get(status, status_styles["disabled"])
        anchor_attr = f' href="#{html.escape(anchor_id)}"' if anchor_id else ""
        st.sidebar.markdown(
            f'<a class="sidebar-step sidebar-step-link"{anchor_attr} title="{html.escape(title)}">'
            f'<span class="sidebar-dot {dot_class}"></span><span>{html.escape(step)}</span>'
            f'</a>',
            unsafe_allow_html=True
        )

    if validation_errors:
        st.sidebar.markdown("#### Требует внимания")
        for err in list(dict.fromkeys(validation_errors))[:6]:
            anchor_id = get_error_anchor(err)
            st.sidebar.markdown(
                f'<a class="sidebar-error-link" href="#{html.escape(anchor_id)}">'
                f'<span>Исправить</span>{html.escape(err)}'
                f'</a>',
                unsafe_allow_html=True
            )


def render_analysis_preview(results, final_score):
    domain_scores = calculate_domain_scores(results)
    risk_signals = _quick_win_signals(results)
    maturity, maturity_icon = get_maturity_level(final_score)

    st.markdown("### Сводка аудита")
    metric_cols = st.columns(4)
    metric_values = [
        ("Зрелость ИБ", f"{final_score}%", f"{maturity_icon} {maturity}"),
        ("Доменов", str(len(domain_scores)), "в текущей оценке"),
        ("Рисков", str(len(risk_signals)), "быстрые сигналы"),
        ("Статус", "Предпросмотр", "перед отчетом"),
    ]

    for col, (label, value, hint) in zip(metric_cols, metric_values):
        with col:
            st.markdown(f"""
            <div class="metric-card">
                <div class="label">{html.escape(label)}</div>
                <div class="value">{html.escape(value)}</div>
                <div class="hint">{html.escape(hint)}</div>
            </div>
            """, unsafe_allow_html=True)

    domain_cards = ''.join(
        f'<div class="domain-card">'
        f'<strong>{html.escape(domain)}</strong>'
        f'<div class="domain-score">{score_value}%</div>'
        f'<div class="hint">{html.escape(_status_label(score_value))}</div>'
        f'</div>'
        for domain, score_value in domain_scores.items()
    )
    st.markdown(f'<div class="domain-row">{domain_cards}</div>', unsafe_allow_html=True)

    if risk_signals:
        st.markdown("#### Быстрые улучшения")
        for kind, title, body in risk_signals:
            st.markdown(f"""
            <div class="risk-chip {html.escape(kind)}">
                <strong>{html.escape(title)}</strong><br>
                {html.escape(body)}
            </div>
            """, unsafe_allow_html=True)


def render_analysis_teaser(results, final_score, validation_errors, section_statuses, client_info):
    domain_scores = calculate_domain_scores(results)
    risk_signals = _quick_win_signals(results)
    maturity, maturity_icon = get_maturity_level(final_score)
    readiness = calculate_audit_readiness(
        client_info,
        validation_errors,
        section_statuses
    )

    weak_domains = sorted(
        domain_scores.items(),
        key=lambda item: item[1]
    )[:3]
    top_signals = risk_signals[:2]

    if top_signals:
        signal_html = "".join(
            f'<div class="mini-signal {html.escape(kind)}">'
            f'<strong>{html.escape(title)}</strong><br>'
            f'{html.escape(body)}'
            f'</div>'
            for kind, title, body in top_signals
        )
    else:
        signal_html = (
            '<div class="mini-signal">'
            '<strong>Быстрые улучшения появятся здесь</strong><br>'
            'Заполните ключевые блоки, и система покажет первые практические рекомендации.'
            '</div>'
        )

    domain_html = "".join(
        f'<div class="mini-domain">'
        f'<span>{html.escape(domain)}</span>'
        f'<strong>{score_value}%</strong>'
        f'</div>'
        for domain, score_value in weak_domains
    )

    st.markdown(f"""
    <div class="analysis-teaser">
        <div class="analysis-teaser-head">
            <div>
                <div class="analysis-teaser-title">Предварительная аналитика уже собирается</div>
                <div class="analysis-teaser-copy">
                    По мере заполнения анкеты здесь появляется ранняя оценка защиты,
                    слабые домены и быстрые улучшения. Полную сводку можно раскрыть ниже.
                </div>
            </div>
            <div class="analysis-pill">обновляется автоматически</div>
        </div>
        <div class="teaser-grid">
            <div class="teaser-card">
                <div class="label">Зрелость ИБ</div>
                <div class="value">{final_score}%</div>
                <div class="hint">{html.escape(maturity_icon)} {html.escape(maturity)}</div>
            </div>
            <div class="teaser-card">
                <div class="label">Готовность анкеты</div>
                <div class="value">{readiness}%</div>
                <div class="hint">до формирования отчета</div>
            </div>
            <div class="teaser-card">
                <div class="label">Быстрые сигналы</div>
                <div class="value">{len(risk_signals)}</div>
                <div class="hint">первые практические подсказки</div>
            </div>
        </div>
        <div class="teaser-columns">
            <div>
                <div class="label">Что система уже видит</div>
                {signal_html}
            </div>
            <div>
                <div class="label">Домены, требующие внимания</div>
                {domain_html}
            </div>
        </div>
        <div class="analysis-teaser-note">
            Раскройте полный блок ниже, чтобы увидеть все домены, статусы и быстрые улучшения.
        </div>
    </div>
    """, unsafe_allow_html=True)

# --- ИНИЦИАЛИЗАЦИЯ И СТЕК СОСТОЯНИЙ (в самом начале финального блока) ---
if "generation_state" not in st.session_state:
    st.session_state.generation_state = "idle"  # Может быть: idle, preparing, heavy_ai, finalized
if "cached_report_bytes" not in st.session_state:
    st.session_state.cached_report_bytes = None
if "cached_sales_report_bytes" not in st.session_state:
    st.session_state.cached_sales_report_bytes = None
if "cached_presentation_bytes" not in st.session_state:
    st.session_state.cached_presentation_bytes = None
if "presentation_status" not in st.session_state:
    st.session_state.presentation_status = ""
if "telegram_status" not in st.session_state:
    st.session_state.telegram_status = ""
if "generation_attempt_started_at" not in st.session_state:
    st.session_state.generation_attempt_started_at = None
if "generation_error_message" not in st.session_state:
    st.session_state.generation_error_message = ""
if "report_shortened_last" not in st.session_state:
    st.session_state.report_shortened_last = False
if "last_report_risk_sources" not in st.session_state:
    st.session_state.last_report_risk_sources = []
if "telegram_generation_started_sent" not in st.session_state:
    st.session_state.telegram_generation_started_sent = False


def render_generation_failure_state():
    render_generation_guard(False)
    message = st.session_state.get("generation_error_message") or (
        "Сервис формирования экспертного заключения временно недоступен. "
        "Презентация не сформирована. "
        "Попробуйте повторить позже."
    )
    st.error(message)
    if st.button("Повторить формирование", key="presentation_retry_after_error"):
        st.session_state.generation_error_message = ""
        st.session_state.telegram_generation_started_sent = False
        st.session_state.generation_attempt_started_at = None
        st.session_state.generation_state = "preparing"
        st.rerun()
    st.stop()

render_generation_guard(
    st.session_state.generation_state in {"preparing", "heavy_ai"}
)

preview_results = build_report_results(
    data,
    main_speed,
    back_speed,
    total_arm,
    ap_cnt,
    selected_routing,
    ngfw_vendor,
    phys_count,
    virt_count,
    v_n_b
)
audit_started = any(str(value).strip() for value in client_info.values()) or total_arm > 0 or any([
    net_active,
    server_active,
    storage_active,
    is_active,
    enable_security,
    dev_active,
])
security_controls = [
    (epp, epp_v, 8),
    (edr, edr_v, 12),
    (xdr, xdr_v, 8),
    (mdr, mdr_v, 8),
    (dlp, dlp_v, 6),
    (mail_sec, mail_v, 5),
    (casb, casb_v, 4),
    (waf, waf_v, 6),
    (ddos, ddos_v, 4),
    (ids, ids_v, 6),
    (nac, nac_v, 5),
    (ztna, ztna_v, 5),
    (sast, sast_v, 5),
    (dast, dast_v, 5),
    (iam, iam_v, 5),
    (mfa, mfa_v, 10),
    (pam, pam_v, 8),
    (siem, siem_v, 10),
    (soar, soar_v, 4),
    (vuln, vuln_v, 5),
    (patch, patch_v, 6),
    (nad, nad_v, 5),
]
preview_score = calculate_weighted_security_score(enable_security, security_controls)

def has_validation_error(*markers):
    return any(
        any(marker in err for marker in markers)
        for err in validation_errors
    )


general_complete = all([
    client_info.get('Город'),
    client_info.get('Наименование компании'),
    client_info.get('Сфера деятельности'),
    client_info.get('Сайт компании'),
    client_info.get('Email'),
    client_info.get('ФИО контактного лица'),
    client_info.get('Должность'),
    client_info.get('Контактный телефон'),
])
endpoint_complete = total_arm > 0 and bool(selected_os_arm) and sum_os_arm == total_arm
network_complete = net_active and bool(selected_routing) and not has_validation_error(
    "маршрутизации",
    "маршрутизаторов",
    "коммутаторов",
    "Core-уровня",
    "уровня распределения",
    "уровня доступа",
    "Wi-Fi",
    "NGFW",
)
server_complete = server_active and (phys_count > 0 or virt_count > 0) and not has_validation_error(
    "серверов",
    "хостов",
    "резервного копирования",
)
storage_complete = storage_active and not has_validation_error("СХД", "RAID")
systems_complete = is_active and not has_validation_error("название и версию", "версию")
security_complete = enable_security and not has_validation_error("производитель")
web_complete = web_active
dev_complete = dev_active and not has_validation_error("разработчиков", "языки разработки")

it_maturity_score = calculate_it_maturity_score(
    total_arm,
    selected_os_arm,
    sum_os_arm,
    net_active,
    main_speed,
    back_speed,
    selected_routing,
    ap_cnt,
    ngfw_vendor,
    server_active,
    phys_count,
    virt_count,
    selected_virt_sys,
    v_n_b,
    storage_active,
    st_media_sel,
    cnt_hdd,
    cnt_ssd,
    raid_selected,
    is_active,
    web_active,
    dev_active,
    dev_count,
    sel_langs,
    cicd_active,
    wifi_enabled=wifi_enabled,
    wifi_ctrl_enabled=wifi_ctrl_enabled,
    operational_notes=[
        st.session_state.get("note_1_1", ""),
        st.session_state.get("note_1_2", ""),
        st.session_state.get("note_1_3", ""),
        st.session_state.get("note_1_4", ""),
        st.session_state.get("note_1_5", ""),
    ],
)

section_statuses = [
    ("Компания", section_status(True, general_complete), "section-company"),
    ("Конечные точки", section_status(True, endpoint_complete), "section-endpoints"),
    ("Сеть", section_status(net_active, network_complete), "section-network"),
    ("Серверы", section_status(server_active, server_complete), "section-servers"),
    ("СХД", section_status(storage_active, storage_complete), "section-storage"),
    ("ИС", section_status(is_active, systems_complete), "section-systems"),
    ("ИБ", section_status(enable_security, security_complete), "section-security"),
    ("Веб", section_status(web_active, web_complete), "section-web"),
    ("Разработка", section_status(dev_active, dev_complete), "section-dev"),
]

render_audit_cockpit(
    client_info,
    preview_results,
    validation_errors,
    preview_score,
    it_maturity_score,
    section_statuses
)

render_analysis_teaser(
    preview_results,
    preview_score,
    validation_errors,
    section_statuses,
    client_info
)

with st.expander("Открыть полный предварительный анализ", expanded=False):
    render_analysis_preview(preview_results, preview_score)

# --- ФИНАЛ ---
st.divider()
if validation_errors:
    render_validation_summary(validation_errors)

# КНОПКА ЗАПУСКА ПРОЦЕССА
# Она активна только тогда, когда процесс еще не запущен
st.markdown("""
Нажимая «Сформировать презентацию аудита», вы даете согласие
на обработку персональных данных в соответствии с
<a href="https://drive.google.com/file/d/1ypEIH9_ePGo3elkR2ifLFBulD5CAFOfs/view?usp=sharing" target="_blank">
Политикой конфиденциальности
</a>.
""", unsafe_allow_html=True)
if st.session_state.generation_state == "idle":
    if st.session_state.generation_error_message:
        st.error(st.session_state.generation_error_message)
    if st.button(
        "Сформировать презентацию аудита",
        disabled=len(validation_errors) > 0,
        key="presentation_generate",
        type="primary",
        use_container_width=False,
    ):
        st.session_state.generation_error_message = ""
        st.session_state.telegram_generation_started_sent = False
        st.session_state.generation_state = "preparing"
        st.rerun()
        render_generation_guard(True)
        alert_placeholder = st.empty()
        console_placeholder = st.empty()
        progress_bar = st.progress(0)

        alert_placeholder.markdown(
            """
            <div class="analysis-status-panel">
                <div class="analysis-status-title">Формируется презентация аудита</div>
                Выполняется нормализация данных, расчет зрелости и сборка рекомендаций. Это может занять до 4 минут.
                <div class="page-lock-note">Не закрывайте и не обновляйте страницу до завершения формирования.</div>
            </div>
            """,
            unsafe_allow_html=True
        )

        steps = [

            "Инициализация ядра аудита...",
            "Проверка обязательных полей...",
            "Нормализация инфраструктурных данных...",
            "Анализ сетевого периметра...",
            "Анализ защищенности конечных точек...",
            "Проверка устойчивости резервного копирования...",
            "Расчет зрелости киберзащиты...",
            "Построение доменов безопасности...",
            "Глубокий анализ рисков...",
            "Формирование управленческого резюме...",
            "Генерация экспертной презентации...",
            "Финализация артефактов..."
        ]

        active_logs = []

        progress = 0

        for step in steps:

            active_logs.append(
                f"[{time.strftime('%H:%M:%S')}] {step}"
            )

            if len(active_logs) > 4:
                active_logs.pop(0)

            console_placeholder.markdown(
                '<div class="analysis-log">' +
                "".join([f'<div>▶ {line}</div>' for line in active_logs]) +
                '</div>',
                unsafe_allow_html=True
            )

            progress += random.randint(5, 9)

            progress_bar.progress(min(progress, 88))

            time.sleep(random.uniform(0.7, 1.4))



        # При клике мы просто меняем статус в памяти на "подготовка" и мгновенно перезапускаем страницу
        st.session_state.generation_state = "preparing"
        st.rerun()

# --- СЦЕНАРИЙ 1: ЭКРАН ОЖИДАНИЯ С ФАКТАМИ ИБ (Показывается СРАЗУ же после клика) ---
if st.session_state.generation_state == "preparing":

    # 1. Сразу жестко выводим на экран поле логов и факты информационной безопасности
    st.markdown("#### 🛠️ Ход выполнения анализа:")
    render_generation_live_panel("Подготовка экспертного анализа", active_step=1)

    # Имитируем лог-систему, как вы просили
    st.info(f"⚙️ `[СИСТЕМА]`: Инициализация аналитического ядра Khalil Consulting {APP_VERSION}...")
    st.success("⚙️ `[МАТРИЦА]`: Агрегация параметров ИТ-инфраструктуры успешно завершена.")

    st.markdown("---")
    st.markdown("#### Полезные факты и рекомендации по ИБ:")

    # Выводим на экран массив фактов в красивом поле, который пользователь будет читать все 3 минуты
    st.markdown("""
    <div class="facts-panel">
        <p><strong>Многофакторная аутентификация:</strong> внедрение MFA блокирует до 99.9% автоматизированных атак на корпоративные учетные записи.</p>
        <p><strong>Защита рабочих мест:</strong> обычного антивируса (EPP) в 2026 году уже недостаточно. Решения класса EDR/XDR критически необходимы для выявления скрытых бесфайловых угроз.</p>
        <p><strong>Безопасность архивов:</strong> резервные копии должны быть изолированы от основной сети. Принцип Immutable Backup снижает риск шифрования бэкапов.</p>
        <p><strong>Сетевой периметр:</strong> сетевая сегментация, VLAN и Zero Trust помогают остановить распространение атаки внутри компании.</p>
        <p><strong>Человеческий фактор:</strong> значительная часть успешных кибератак начинается с фишингового письма. Регулярное обучение снижает этот риск.</p>
    </div>
    """, unsafe_allow_html=True)

    # Делаем маленькую паузу в 1.5 секунды, чтобы Streamlit успел железно отправить этот интерфейс в браузер клиента
    time.sleep(1.5)

    # Меняем статус на "Запуск тяжелого ИИ" и перезапускаем страницу.
    # Теперь этот красивый экран останется висеть в браузере, пока ИИ думает!
    st.session_state.generation_state = "heavy_ai"
    st.rerun()

# --- СЦЕНАРИЙ 2: ЗАПУСК ТЯЖЕЛОГО ИИ И СБОРКИ EXCEL ---
if st.session_state.generation_state == "heavy_ai":
    ai_failure_notified = False
    if st.session_state.generation_attempt_started_at is None:
        st.session_state.generation_attempt_started_at = time.time()
    elif time.time() - st.session_state.generation_attempt_started_at > 300:
        st.session_state.generation_state = "ai_failed"
        st.session_state.generation_attempt_started_at = None
        st.session_state.generation_error_message = (
            "Формирование презентации превысило допустимое время. Попробуйте повторить позже."
        )
        render_generation_failure_state()

    generation_panel = st.empty()
    with generation_panel.container():
        render_generation_live_panel("Идет глубокий анализ и сборка презентации", active_step=4)

    if not st.session_state.telegram_generation_started_sent:
        st.session_state.telegram_status = send_internal_telegram_message(
            build_telegram_generation_started_text(client_info, preview_score)
        )
        if st.session_state.telegram_status == "ok":
            st.session_state.telegram_generation_started_sent = True

    # Этот текст и анимация будут гореть параллельно с фактами сверху
    with st.container():
        try:
            # Подготовка данных перед передачей
            results = build_report_results(
                data,
                main_speed,
                back_speed,
                total_arm,
                ap_cnt,
                selected_routing,
                ngfw_vendor,
                phys_count,
                virt_count,
                v_n_b
            )
            f_score = preview_score

            st.session_state.ai_last_error = ""
            st.session_state.ai_model_used = ""
            st.session_state.ai_provider_used = ""
            st.session_state.ai_used_in_last_report = False
            st.session_state.ai_analysis_succeeded = False
            st.session_state.ai_audit_narrative = {}

            # Внутренний XLSX нужен только как источник листов для sales playbook.
            report_bytes = make_expert_excel(client_info, results, f_score)
            ai_report_ready = (
                bool(st.session_state.get("ai_used_in_last_report"))
                and not st.session_state.get("ai_last_error")
            )
            st.session_state.report_shortened_last = False
            if not ai_report_ready:
                ai_failure_detail = st.session_state.get(
                    "ai_last_error",
                    "AI analysis did not return recommendations",
                )
                st.session_state.telegram_status = send_internal_telegram_message(
                    build_telegram_ai_failure_text(
                        client_info,
                        f_score,
                        ai_failure_detail,
                    )
                )
                ai_failure_notified = True
                st.session_state.cached_report_bytes = None
                st.session_state.cached_sales_report_bytes = None
                st.session_state.cached_presentation_bytes = None
                st.session_state.presentation_status = "error"
                raise RuntimeError("AI quality gate rejected the customer presentation")

            sales_report_bytes, telegram_sales = make_internal_sales_excel(
                client_info,
                results,
                f_score,
                report_bytes
            )
            st.session_state.cached_report_bytes = report_bytes
            st.session_state.cached_sales_report_bytes = sales_report_bytes
            try:
                st.session_state.cached_presentation_bytes = make_audit_presentation(
                    client_info,
                    results,
                    f_score,
                    it_maturity_score,
                )
                st.session_state.presentation_status = "ok"
            except Exception as presentation_exc:
                st.session_state.cached_presentation_bytes = None
                st.session_state.presentation_status = "error"
                send_internal_telegram_message(
                    f"[{get_app_instance_label()}] Презентация не сформирована: "
                    f"{redact_secret(presentation_exc, TOKEN)}"
                )
                raise RuntimeError("Не удалось сформировать клиентскую презентацию") from presentation_exc
        except Exception as exc:
            if not ai_failure_notified:
                st.session_state.telegram_status = send_internal_telegram_message(
                    build_telegram_generation_error_text(
                        client_info,
                        preview_score,
                        redact_secret(exc, TOKEN)
                    )
                )
            st.session_state.generation_state = "ai_failed"
            st.session_state.generation_attempt_started_at = None
            st.session_state.generation_error_message = (
                "Сервис формирования экспертного заключения временно недоступен. "
                "Презентация не сформирована. "
                "Попробуйте повторить позже."
            )
            generation_panel.empty()
            render_generation_failure_state()

    # Тихо отправляем в ТГ без создания задержек на экране
    st.session_state.telegram_status = ""
    if TOKEN and CHAT_ID:
        try:
            brand_file_label = "BTG" if presentation_brand_key() == "btg" else "Khalil"
            sales_lines = []
            for idx, item in enumerate(telegram_sales[:3], start=1):
                sales_lines.append(
                    f"{idx}. {item['offer']} | {item['vendors']}"
                )
            sales_digest = "\n".join(sales_lines) if sales_lines else "Нет явных продуктовых триггеров, нужен экспертный разбор."

            telegram_text = build_telegram_lead_text(
                client_info,
                f_score,
                sales_digest
            )
            telegram_send_node(
                TOKEN,
                "sendMessage",
                {"chat_id": CHAT_ID, "text": telegram_text},
                timeout_seconds=8
            )

            telegram_send_node(
                TOKEN,
                "sendDocument",
                {
                    "chat_id": CHAT_ID,
                    "caption": f"[{get_app_instance_label()}] Sales playbook: {client_info['Наименование компании']}"
                },
                files=[{
                    "field": "document",
                    "filename": f"Sales_Playbook_{client_info['Наименование компании']}.xlsx",
                    "bytes": sales_report_bytes,
                    "suffix": ".xlsx",
                }],
                timeout_seconds=15
            )

            telegram_send_node(
                TOKEN,
                "sendDocument",
                {
                    "chat_id": CHAT_ID,
                    "caption": f"[{get_app_instance_label()}] Клиентская презентация: {client_info['Наименование компании']}"
                },
                files=[{
                    "field": "document",
                    "filename": f"Audit_Presentation_{brand_file_label}_{client_info['Наименование компании']}.pptx",
                    "bytes": st.session_state.cached_presentation_bytes,
                    "suffix": ".pptx",
                }],
                timeout_seconds=20
            )
            st.session_state.telegram_status = "ok"
        except Exception as exc:
            st.session_state.telegram_status = f"Telegram не отправлен: {redact_secret(exc, TOKEN)}"
    else:
        st.session_state.telegram_status = "Telegram не отправлен: не найдены TELEGRAM_TOKEN или TELEGRAM_CHAT_ID."

    # Переключаем статус в финал
    st.session_state.generation_state = "finalized"
    st.session_state.generation_attempt_started_at = None
    st.session_state.generation_error_message = ""
    st.rerun()

# --- СЦЕНАРИЙ 3: НЕУДАЧНАЯ СБОРКА БЕЗ КЛИЕНТСКОГО ОТЧЕТА ---
if st.session_state.generation_state == "ai_failed":
    render_generation_failure_state()

# --- СЦЕНАРИЙ 4: ВЫВОД ГОТОВОГО РЕЗУЛЬТАТА ---
if st.session_state.generation_state == "finalized":

    st.success("🎉 Экспертная презентация сформирована и проверена системой контроля качества Khalil Consulting!")

    brand_file_label = "BTG" if presentation_brand_key() == "btg" else "Khalil"
    if st.session_state.cached_presentation_bytes:
        st.download_button(
            label="Скачать заключение по аудиту",
            data=st.session_state.cached_presentation_bytes,
            file_name=f"Audit_Presentation_{brand_file_label}_{client_info['Наименование компании']}.pptx",
            mime="application/vnd.openxmlformats-officedocument.presentationml.presentation",
            key="presentation_download",
            type="secondary",
            use_container_width=False,
        )
    elif st.session_state.presentation_status == "error":
        st.error("Не удалось сформировать презентацию. Попробуйте повторить позже.")

    # Кнопка для сброса состояния, если пользователь захочет перегенерировать отчет
    if st.button("Сформировать новую презентацию"):
        st.session_state.generation_state = "idle"
        st.session_state.cached_report_bytes = None
        st.session_state.cached_sales_report_bytes = None
        st.session_state.cached_presentation_bytes = None
        st.session_state.presentation_status = ""
        st.session_state.telegram_status = ""
        st.session_state.ai_last_error = ""
        st.session_state.report_shortened_last = False
        st.session_state.generation_attempt_started_at = None
        st.session_state.generation_error_message = ""
        st.session_state.telegram_generation_started_sent = False
        st.rerun()

st.info(f"Khalil Audit System {APP_VERSION} | by Ivan Rudoy | Алматы 2026")
